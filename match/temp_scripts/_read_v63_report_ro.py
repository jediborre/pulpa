from openpyxl import load_workbook

p = r"training/model_outputs_v6_3/Q4_ROI_match_by_match_v6_2_vs_v6_3_updated.xlsx"
wb = load_workbook(p, data_only=True, read_only=True)

ws = wb['summary']
rows = ws.iter_rows(values_only=True)
headers = next(rows)
idx = {k:i for i,k in enumerate(headers)}
cols = ['modelo','apuestas','ganadas','perdidas','efectividad','ganancia','roi_bank','minuto_base_apuesta']
print('SUMMARY')
print(' | '.join(cols))
for r in rows:
    print(' | '.join(str(r[idx[c]]) for c in cols))

wl = wb['efectividad_liga']
rows2 = wl.iter_rows(values_only=True)
h2 = next(rows2)
j = {k:i for i,k in enumerate(h2)}
cache = list(rows2)
for model_name in ['v6.2', 'v6.3_raw_no_filter', 'v6.3_iso_no_filter', 'v6.3_platt_no_filter']:
    rr = [x for x in cache if x[j['modelo']] == model_name and isinstance(x[j['ganancia']], (int,float))]
    rr.sort(key=lambda x: x[j['ganancia']], reverse=True)
    print('\nTOP_LIGAS_' + model_name)
    print('liga | matches_apostados | efectividad | ganancia')
    for x in rr[:5]:
        print(str(x[j['liga']]) + ' | ' + str(x[j['matches_apostados']]) + ' | ' + str(x[j['efectividad']]) + ' | ' + str(x[j['ganancia']]))
