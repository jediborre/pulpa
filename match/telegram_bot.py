import asyncio
import csv
from datetime import datetime, timedelta
import html
import importlib
import json
import logging
import os
import re
import sqlite3
import sys
import tempfile
import threading
import time
import unicodedata
from pathlib import Path

from dotenv import load_dotenv
from telegram import (
    BotCommand,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    MenuButtonCommands,
    Message,
    ReplyKeyboardMarkup,
    Update,
)
from telegram.error import BadRequest, NetworkError, TimedOut
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)
from telegram.request import HTTPXRequest

import db as db_mod
import ml_tools as ml_mod
import scraper as scraper_mod
import bet_monitor as bet_monitor_mod


class CustomHTTPXRequest(HTTPXRequest):
    """HTTPXRequest with extended timeouts for slow networks and webhooks."""
    def __init__(self, *args, **kwargs):
        # Set extended connect/pool/read/write timeouts (60 seconds each)
        kwargs.setdefault('connect_timeout', 60.0)
        kwargs.setdefault('pool_timeout', 60.0)
        kwargs.setdefault('read_timeout', 60.0)
        kwargs.setdefault('write_timeout', 60.0)
        super().__init__(*args, **kwargs)


BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env")

BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
_raw_db_path = os.getenv("MATCH_DB_PATH", "").strip()
if _raw_db_path:
    _db_p = Path(_raw_db_path)
    DB_PATH = str(_db_p if _db_p.is_absolute() else BASE_DIR / _db_p)
else:
    DB_PATH = str(BASE_DIR / "matches.db")
ALLOWED_CHAT_IDS_RAW = os.getenv("TELEGRAM_ALLOWED_CHAT_IDS", "").strip()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger(__name__)
logging.getLogger("telegram").setLevel(logging.WARNING)
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("httpcore").setLevel(logging.WARNING)

DATE_PAGE_SIZE = 8
MATCH_PAGE_SIZE = 8
LIVE_PAGE_SIZE = 8
AWAITING_MATCH_ID_KEY = "awaiting_match_id"
AWAITING_FETCH_DATE_KEY = "awaiting_fetch_date"
LIVE_RESULT_KEY = "live_result"

RETRAIN_LOCK = asyncio.Lock()
TRAIN_STATUS: dict[str, object] = {
    "running": False,
    "owner_chat_id": None,
    "last_exit_code": None,
    "last_finished_utc": None,
    "last_error_tail": "",
    "current_task": "",
    "progress_done": 0,
    "progress_total": 0,
}
MENU_BUTTON_TEXT    = "Menu"
STATS_BUTTON_TEXT   = "📊 Stats"
MONTHLY_BUTTON_TEXT = "📆 Mes"
ID_BUTTON_TEXT      = "🔍 ID"
UTC_OFFSET_HOURS = -6
DATE_INGEST_PROGRESS_EVERY = 10
DATE_INGEST_STATUS_INTERVAL_SECONDS = 4
REFRESH_DATE_STATUS_INTERVAL_SECONDS = 4
DATE_INGEST_JOBS: dict[int, dict[str, object]] = {}
REFRESH_JOBS: dict[int, dict[str, object]] = {}
FOLLOW_JOBS: dict[int, dict[str, object]] = {}
# Monitor runs in its own OS thread with a dedicated asyncio loop so it
# never competes with the bot's event loop.
_MONITOR_THREAD: threading.Thread | None = None
_MONITOR_LOOP: asyncio.AbstractEventLoop | None = None
_MONITOR_STOP_REF: list = [None]  # holds the asyncio.Event created inside the thread
# chat_ids subscribed to monitor notifications.
# Value: dict with {"signal_type": "all"|"bet_only", "quarters": ["q3", "q4"]}
_MONITOR_SUBSCRIBERS: dict[int, dict] = {}
_SUBSCRIBERS_SETTING_KEY = "monitor_subscribers"  # JSON stored in settings table


def _persist_subscribers() -> None:
    """Save current _MONITOR_SUBSCRIBERS to the settings table."""
    import json as _json
    conn = _open_conn()
    db_mod.set_setting(conn, _SUBSCRIBERS_SETTING_KEY, _json.dumps(_MONITOR_SUBSCRIBERS))
    conn.close()


def _load_subscribers() -> None:
    """Load _MONITOR_SUBSCRIBERS from the settings table (called on startup)."""
    import json as _json
    global _MONITOR_SUBSCRIBERS
    conn = _open_conn()
    raw = db_mod.get_setting(conn, _SUBSCRIBERS_SETTING_KEY)
    conn.close()
    if raw:
        try:
            data = _json.loads(raw)
            if isinstance(data, dict):
                for k, v in data.items():
                    chat_id = int(k)
                    # Handle old format (string) and new format (dict)
                    if isinstance(v, str):
                        # Convert old "all"/"bet_only" to new format
                        _MONITOR_SUBSCRIBERS[chat_id] = {
                            "signal_type": v,
                            "quarters": ["q3", "q4"]
                        }
                    elif isinstance(v, dict):
                        # Ensure fields exist
                        _MONITOR_SUBSCRIBERS[chat_id] = {
                            "signal_type": v.get("signal_type", "all"),
                            "quarters": v.get("quarters", ["q3", "q4"])
                        }
        except Exception:
            pass


def _get_subscriber_pref(chat_id: int | None) -> tuple[str, list[str]]:
    """Get subscriber preference for signal type and quarters.
    Returns (signal_type, quarters) tuple.
    """
    if chat_id is None:
        return "all", ["q3", "q4"]
    pref = _MONITOR_SUBSCRIBERS.get(chat_id)
    if pref is None:
        return "all", ["q3", "q4"]
    return pref.get("signal_type", "all"), pref.get("quarters", ["q3", "q4"])
MODEL_OUTPUTS_V4_DIR = BASE_DIR / "training" / "model_outputs_v4"
MODEL_OUTPUTS_V2_DIR = BASE_DIR / "training" / "model_outputs_v2"
MODEL_OUTPUTS_V6_DIR = BASE_DIR / "training" / "model_outputs_v6"
MODEL_OUTPUTS_V6_1_DIR = BASE_DIR / "training" / "model_outputs_v6_1"
MODEL_OUTPUTS_V9_DIR = BASE_DIR / "training" / "model_outputs_v9"
MODEL_OUTPUTS_V12_DIR = BASE_DIR / "training" / "v12" / "model_outputs"
V12_INFERENCE_SCRIPT = BASE_DIR / "training" / "v12" / "infer_match_v12.py"
V12_LIVE_SCRIPT = BASE_DIR / "training" / "v12" / "live_engine" / "virtual_bookmaker.py"
MODEL_OUTPUTS_V13_DIR = BASE_DIR / "training" / "v13" / "model_outputs"
MODEL_OUTPUTS_V15_DIR = BASE_DIR / "training" / "v15" / "model_outputs"
MODEL_OUTPUTS_V16_DIR = BASE_DIR / "training" / "v16" / "model_outputs"
MODEL_OUTPUTS_V17_DIR = BASE_DIR / "training" / "v17" / "model_outputs"
V13_INFERENCE_SCRIPT = BASE_DIR / "training" / "v13" / "infer_match_v13.py"
FOLLOW_REFRESH_SECONDS = 45
FOLLOW_STALE_CYCLES_LIMIT = 8

# Active model per quarter — changeable at runtime via bot menu
# MODEL_CONFIG  → used for predictions in the regular bot flow (match detail page)
# MONITOR_MODEL_CONFIG → used by the bet_monitor daemon for automated bet evaluation
MODEL_CONFIG: dict[str, str] = {"q3": "v4", "q4": "v4"}
MONITOR_MODEL_CONFIG: dict[str, str] = {"q3": "v4", "q4": "v4"}
AVAILABLE_MODELS: list[str] = [
    "v2",
    "v4",
    "v6",
    "v6_1",
    "v9",
    "v12",
    "v13",
    "v15",
    "v16",
    "v17",
]


def _model_version_callback_token(version: str) -> str:
    """Token corto para callback_data (límite 64 B de Telegram; evita ':' dentro del slug)."""
    try:
        return str(AVAILABLE_MODELS.index(version))
    except ValueError:
        return version


def _model_version_from_callback_token(tok: str) -> str | None:
    """Índice numérico o slug legacy (p. ej. v6_1) → versión en AVAILABLE_MODELS."""
    if not tok:
        return None
    if tok.isdigit():
        i = int(tok)
        if 0 <= i < len(AVAILABLE_MODELS):
            return AVAILABLE_MODELS[i]
        return None
    if tok in AVAILABLE_MODELS:
        return tok
    return None


def _log_raw_inference_result(match_id: str, source: str, infer_result: object) -> None:
    """Log raw inference payload to terminal for full model-output visibility."""
    try:
        payload = json.dumps(infer_result, ensure_ascii=False, default=str, indent=2)
    except Exception:
        payload = str(infer_result)
    logger.info("[MODEL_RAW] source=%s match_id=%s", source, match_id)
    logger.info("[MODEL_RAW] payload=\n%s", payload)


def _parse_allowed_chat_ids(raw: str) -> set[int]:
    if not raw:
        return set()
    out: set[int] = set()
    for chunk in raw.split(","):
        value = chunk.strip()
        if not value:
            continue
        try:
            out.add(int(value))
        except ValueError:
            continue
    return out


ALLOWED_CHAT_IDS = _parse_allowed_chat_ids(ALLOWED_CHAT_IDS_RAW)


# Cache for V12 MAE metrics (loaded once at module startup)
_V12_MAE_CACHE: dict[str, float] = {}
_V12_MAE_CACHE_LOADED = False


def _load_v12_mae_metrics() -> dict[str, float]:
    """Load V12 MAE metrics from all_metrics.csv (cached after first load)."""
    global _V12_MAE_CACHE, _V12_MAE_CACHE_LOADED
    
    if _V12_MAE_CACHE_LOADED:
        return _V12_MAE_CACHE
    
    metrics: dict[str, float] = {}
    csv_path = MODEL_OUTPUTS_V12_DIR / "all_metrics.csv"
    
    if csv_path.exists():
        try:
            with csv_path.open("r", encoding="utf-8", newline="") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    model_name = str(row.get("model", "") or "")
                    mae = float(row.get("mae", 0))
                    metrics[model_name] = mae
            logger.info("[V12_MAE] Loaded %d metrics from all_metrics.csv", len(metrics))
        except Exception as exc:
            logger.warning("[V12_MAE] Failed to load metrics: %s", exc)
    
    _V12_MAE_CACHE = metrics
    _V12_MAE_CACHE_LOADED = True
    
    return metrics


# Preload V12 MAE metrics at module startup
try:
    _load_v12_mae_metrics()
except Exception:
    pass


def _get_v12_mae_for_target(target: str, metric_type: str = "home") -> float:
    """Get actual V12 MAE for a specific target and type."""
    metrics = _load_v12_mae_metrics()
    
    # Map target and type to the model name in all_metrics.csv
    # Format: q{3|4}_{home|away|total}_men_or_open
    key = f"{target}_{metric_type}_men_or_open"
    
    if key in metrics:
        return metrics[key]
    
    # Fallback to women's league if men_or_open not found
    key_women = f"{target}_{metric_type}_women"
    if key_women in metrics:
        return metrics[key_women]
    
    # Fallback to default if not found
    fallbacks = {
        "q3_home": 3.63,
        "q3_away": 3.60,
        "q3_total": 5.33,
        "q4_home": 3.63,
        "q4_away": 3.60,
        "q4_total": 5.33,
    }
    
    return fallbacks.get(key, 5.0)


def _run_v12_inference(match_id: str, target: str = "q4") -> dict:
    """Run V12 inference for winner prediction."""
    try:
        import importlib
        v12_infer_mod = importlib.import_module("training.v12.infer_match_v12")
        v12_result = v12_infer_mod.run_inference(
            match_id=match_id,
            target=target,
            fetch_missing=False,
        )

        if isinstance(v12_result, dict) and not v12_result.get("ok", True):
            return {"ok": False, "reason": v12_result.get("reason", "Error")}
        
        pred = v12_result
        
        # Load actual V12 MAE from trained metrics
        mae_total = _get_v12_mae_for_target(target, "total")
        mae_home = _get_v12_mae_for_target(target, "home")
        mae_away = _get_v12_mae_for_target(target, "away")
        
        return {
            "ok": True,
            "predictions": {
                target: {
                    "available": True,
                    "predicted_winner": pred.winner_pick,
                    "confidence": pred.winner_confidence,
                    "bet_signal": pred.winner_signal,
                    "final_recommendation": pred.final_signal,
                    "predicted_total": pred.predicted_total,
                    "predicted_home": pred.predicted_home,
                    "predicted_away": pred.predicted_away,
                    "over_under_signal": pred.over_under_signal,
                    "league_quality": pred.league_quality,
                    "league_bettable": pred.league_bettable,
                    "volatility_index": pred.volatility_index,
                    "data_quality": pred.data_quality,
                    "risk_level": pred.risk_level,
                    "reasoning": pred.reasoning,
                    "mae": mae_total,
                    "mae_home": mae_home,
                    "mae_away": mae_away,
                }
            }
        }
    except Exception as exc:
        return {"ok": False, "reason": f"V12 error: {exc}"}


def _run_v12_live_analysis(
    match_id: str,
    quarter: str,
    qtr_home_score: int,
    qtr_away_score: int,
    total_home_score: int,
    total_away_score: int,
    elapsed_minutes: float,
    graph_points: list,
    pbp_events: list,
) -> dict:
    """Run V12 LIVE virtual bookmaker analysis."""
    try:
        import importlib
        v12_live_mod = importlib.import_module("training.v12.live_engine.virtual_bookmaker")
        
        analysis = v12_live_mod.analizar_quarter_en_vivo(
            match_id=match_id,
            quarter=quarter,
            qtr_home_score=qtr_home_score,
            qtr_away_score=qtr_away_score,
            total_home_score=total_home_score,
            total_away_score=total_away_score,
            elapsed_minutes=elapsed_minutes,
            graph_points=graph_points,
            pbp_events=pbp_events,
        )
        
        if analysis is None:
            return {"ok": False, "reason": "No se pudo analizar"}
        
        markets_text = []
        markets_raw = []
        for m in analysis.markets:
            markets_text.append(
                f"{m.description}\n"
                f"  Prob: {m.our_probability:.1%} | Fair: {m.fair_odds:.2f}\n"
                f"  → Si tu casa ofrece > {m.fair_odds * 1.15:.2f} → VALUE"
            )
            markets_raw.append({
                "description": m.description,
                "prob": m.our_probability,
                "fair": m.fair_odds,
                "value_threshold": round(m.fair_odds * 1.15, 2),
            })
        
        return {
            "ok": True,
            "score_diff": analysis.score_diff,
            "elapsed": analysis.elapsed_minutes,
            "remaining": analysis.minutes_remaining,
            "momentum": analysis.graph_momentum,
            "projections": {
                "home": analysis.projected_home_pts,
                "away": analysis.projected_away_pts,
                "total": analysis.projected_total_pts,
                "diff": analysis.projected_diff,
            },
            "markets_text": "\n\n".join(markets_text),
            "markets": markets_raw,
            "recommendation": analysis.overall_recommendation,
            "best_market": analysis.best_market,
            "reasoning": analysis.reasoning,
        }
    except Exception as exc:
        return {"ok": False, "reason": f"V12 LIVE error: {exc}"}


def _run_v13_inference(match_id: str, target: str = "q4") -> dict:
    """Run V13 inference for winner prediction (pace-aware ensemble model)."""
    try:
        v13_infer_mod = importlib.import_module("training.v13.infer_match_v13")
        v13_result = v13_infer_mod.run_inference(match_id=match_id, target=target)

        if not v13_result.get("ok", False):
            return {"ok": False, "reason": v13_result.get("reason", "V13 no disponible")}

        pred = v13_result["prediction"]

        return {
            "ok": True,
            "predictions": {
                target: {
                    "available": True,
                    "predicted_winner": getattr(pred, "winner_pick", None),
                    "confidence": getattr(pred, "winner_confidence", None),
                    "bet_signal": getattr(pred, "winner_signal", None),
                    "final_recommendation": getattr(pred, "final_signal", None),
                    "predicted_total": getattr(pred, "predicted_total", None),
                    "predicted_home": getattr(pred, "predicted_home", None),
                    "predicted_away": getattr(pred, "predicted_away", None),
                    "reasoning": getattr(pred, "reasoning", None),
                    "mae": getattr(pred, "mae", None),
                    "mae_home": getattr(pred, "mae_home", None),
                    "mae_away": getattr(pred, "mae_away", None),
                    "league_quality": getattr(pred, "league_quality", None),
                    "league_bettable": getattr(pred, "league_bettable", None),
                    "volatility_index": getattr(pred, "volatility_index", None),
                    "data_quality": getattr(pred, "data_quality", None),
                    # V13-specific quality metadata
                    "model_quality": getattr(pred, "model_quality", None),
                    "model_samples": getattr(pred, "model_samples", None),
                    "model_gap": getattr(pred, "model_gap", None),
                    "model_f1": getattr(pred, "model_f1", None),
                    "fallback_used": getattr(pred, "fallback_used", False),
                }
            },
        }
    except Exception as exc:
        return {"ok": False, "reason": f"V13: {exc}"}


# ─── V15 / V16 / V17 engine caches (loaded once, reused across predictions) ─────────
_V15_ENGINE_INSTANCE: object | None = None
_V16_ENGINE_INSTANCE: object | None = None
_V17_ENGINE_INSTANCE: object | None = None


def _get_v15_engine():
    global _V15_ENGINE_INSTANCE
    if _V15_ENGINE_INSTANCE is None:
        v15_mod = importlib.import_module("training.v15.inference")
        _V15_ENGINE_INSTANCE = v15_mod.V15Engine.load()
    return _V15_ENGINE_INSTANCE


def _get_v16_engine():
    global _V16_ENGINE_INSTANCE
    if _V16_ENGINE_INSTANCE is None:
        # V16's inference.py also names its main class V15Engine
        v16_mod = importlib.import_module("training.v16.inference")
        _V16_ENGINE_INSTANCE = v16_mod.V15Engine.load()
    return _V16_ENGINE_INSTANCE


def _get_v17_engine():
    global _V17_ENGINE_INSTANCE
    if _V17_ENGINE_INSTANCE is None:
        # V17's inference.py also names its main class V15Engine
        v17_mod = importlib.import_module("training.v17.inference")
        _V17_ENGINE_INSTANCE = v17_mod.V15Engine.load()
    return _V17_ENGINE_INSTANCE


def _extract_v15v16_inputs(data: dict, match_id: str, target: str):
    """Return (league, quarter_scores, graph_points, pbp_events) from a data dict."""
    m = data.get("match", {})
    s = data.get("score", {})
    quarters = s.get("quarters", {})
    gp: list[dict] = data.get("graph_points") or []
    # Flatten quarter-keyed PBP dict → flat list with 'quarter' key injected
    pbp: list[dict] = []
    for q_code, evts in (data.get("play_by_play") or {}).items():
        for e in (evts or []):
            pbp.append({**e, "quarter": q_code})
    league = str(m.get("league") or "Unknown")

    def _qsc(q_code: str, side: str) -> int:
        return int((quarters.get(q_code) or {}).get(side) or 0)

    quarter_scores: dict[str, int] = {
        "q1_home": _qsc("Q1", "home"), "q1_away": _qsc("Q1", "away"),
        "q2_home": _qsc("Q2", "home"), "q2_away": _qsc("Q2", "away"),
    }
    if target == "q4":
        quarter_scores["q3_home"] = _qsc("Q3", "home")
        quarter_scores["q3_away"] = _qsc("Q3", "away")
    return league, quarter_scores, gp, pbp


def _v15v16_pred_to_dict(pred: object, target: str) -> dict:
    """Normalise a V15/V16 Prediction dataclass to the standard predictions dict."""
    sig = getattr(pred, "signal", "NO_BET")
    d = getattr(pred, "debug", None)
    winner_pick = "home" if sig == "BET_HOME" else ("away" if sig == "BET_AWAY" else None)
    return {
        "available": True,
        "predicted_winner": winner_pick,
        "confidence": getattr(pred, "confidence", None),
        "bet_signal": sig,
        "final_recommendation": sig,
        "predicted_total": getattr(d, "pred_total", None) if d else None,
        "predicted_home": getattr(d, "pred_home", None) if d else None,
        "predicted_away": getattr(d, "pred_away", None) if d else None,
        "reasoning": getattr(pred, "reason", ""),
        "mae": getattr(d, "reg_mae_total", None) if d else None,
        "mae_home": None,
        "mae_away": None,
        "league_quality": None,
        "league_bettable": sig != "NO_BET",
        "volatility_index": None,
        "data_quality": None,
        # V15/V16-specific extras for display
        "model_found": getattr(d, "model_found", None) if d else None,
        "gp_count": getattr(d, "gp_count", None) if d else None,
        "threshold": getattr(pred, "threshold", None),
        "probability": getattr(pred, "probability", None),
    }


def _run_v15_inference(match_id: str, target: str = "q4") -> dict:
    """Run V15 (per-league gradient boosting) inference."""
    try:
        data = _get_match_detail(match_id)
        if not data:
            return {"ok": False, "reason": "V15: match no encontrado en DB"}
        league, quarter_scores, gp, pbp = _extract_v15v16_inputs(data, match_id, target)
        engine = _get_v15_engine()
        pred = engine.predict(
            match_id=match_id, target=target, league=league,
            quarter_scores=quarter_scores, graph_points=gp, pbp_events=pbp,
        )
        return {"ok": True, "predictions": {target: _v15v16_pred_to_dict(pred, target)}}
    except Exception as exc:
        return {"ok": False, "reason": f"V15 error: {exc}"}


def _run_v16_inference(match_id: str, target: str = "q4") -> dict:
    """Run V16 (per-league + TFM features) inference."""
    try:
        data = _get_match_detail(match_id)
        if not data:
            return {"ok": False, "reason": "V16: match no encontrado en DB"}
        league, quarter_scores, gp, pbp = _extract_v15v16_inputs(data, match_id, target)
        engine = _get_v16_engine()
        pred = engine.predict(
            match_id=match_id, target=target, league=league,
            quarter_scores=quarter_scores, graph_points=gp, pbp_events=pbp,
        )
        return {"ok": True, "predictions": {target: _v15v16_pred_to_dict(pred, target)}}
    except Exception as exc:
        return {"ok": False, "reason": f"V16 error: {exc}"}


def _run_v17_inference(match_id: str, target: str = "q4") -> dict:
    """Run V17 inference."""
    try:
        data = _get_match_detail(match_id)
        if not data:
            return {"ok": False, "reason": "V17: match no encontrado en DB"}
        league, quarter_scores, gp, pbp = _extract_v15v16_inputs(data, match_id, target)
        engine = _get_v17_engine()
        pred = engine.predict(
            match_id=match_id, target=target, league=league,
            quarter_scores=quarter_scores, graph_points=gp, pbp_events=pbp,
        )
        return {"ok": True, "predictions": {target: _v15v16_pred_to_dict(pred, target)}}
    except Exception as exc:
        return {"ok": False, "reason": f"V17 error: {exc}"}


# V12 LIVE tracking state
V12_LIVE_JOBS: dict[str, dict] = {}  # key: "{chat_id}:{msg_id}" -> job info
V12_LIVE_POLL_SECONDS = 30


def _build_v12_live_message(data: dict, result: dict, quarter_label: str) -> str:
    """Build a Telegram-friendly V12 LIVE message."""
    m = data.get("match", {})
    s = data.get("score", {})
    quarters = s.get("quarters", {})
    gp = data.get("graph_points", [])
    minute_est = None
    if gp:
        try:
            minute_est = int(gp[-1].get("minute", 0))
        except (TypeError, ValueError):
            pass

    home_team = str(m.get("home_team", "?"))
    away_team = str(m.get("away_team", "?"))
    total_home = s.get("home", "?")
    total_away = s.get("away", "?")
    status = str(m.get("status_type", "?"))

    # Quarter scores
    q_order = ["Q1", "Q2", "Q3", "Q4"]
    q_lines = []
    for q in q_order:
        qd = quarters.get(q, {})
        if qd:
            qh = qd.get("home", "-")
            qa = qd.get("away", "-")
            emoji = "✅" if q == quarter_label else "·"
            q_lines.append(f"  {emoji} {q}: {qh} - {qa}")

    # Projections
    proj = result.get("projections", {})
    rec = result.get("recommendation", "?")
    rec_emoji = {"WATCH": "✅", "SKIP": "🚫", "LEAN": "⚠️"}.get(rec, "❓")
    best = result.get("best_market", "N/A")

    # Momentum
    mom = result.get("momentum", 0)
    mom_arrow = "↑" if mom > 0.1 else "↓" if mom < -0.1 else "→"

    lines = [
        f"🏀 V12 LIVE — {quarter_label}",
        f"{home_team} vs {away_team}",
        f"⏱ Min: {minute_est if minute_est is not None else '?'} | Estado: {status}",
        "",
        f"📊 Marcador: {total_home} - {total_away}",
    ]
    lines.extend(q_lines)
    lines += [
        "",
        f"📈 Proyeccion: {int(proj.get('home', 0))} - {int(proj.get('away', 0))}",
        f"   Diff: {'+' if proj.get('diff', 0) >= 0 else ''}{int(proj.get('diff', 0))} pts",
        f"   Total: {int(proj.get('total', 0))} pts",
        f"   Momentum: {mom_arrow} {mom:+.1f}",
        "",
        f"{rec_emoji} Recomendacion: {rec}",
        f"🎯 Mejor mercado: {best}",
    ]

    # Markets with clear formatting
    _num_emojis = ["1️⃣", "2️⃣", "3️⃣"]
    _markets_raw = result.get("markets", [])
    if _markets_raw:
        lines.append("")
        lines.append("📋 Mercados:")
        for _i, _mkt in enumerate(_markets_raw[:3]):
            _prob_pct = f"{_mkt['prob']:.0%}"
            _fair = _mkt['fair']
            _val_thr = _mkt.get('value_threshold', round(_fair * 1.15, 2))
            _num = _num_emojis[_i] if _i < len(_num_emojis) else f"{_i+1}."
            lines += [
                "",
                f"  {_num} {_mkt['description']}",
                f"     Prob: {_prob_pct}  |  Cuota justa: {_fair:.2f}",
                f"     ✅ VALUE si ofrecen > {_val_thr:.2f}",
            ]

    return "\n".join(lines)


async def _v12_live_poll(chat_id: int, msg_id: int, match_id: str, page: int, app: Application):
    """Background polling for V12 LIVE tracking."""
    job_key = f"{chat_id}:{msg_id}"
    job = V12_LIVE_JOBS.get(job_key)
    if not job:
        return

    poll_interval = V12_LIVE_POLL_SECONDS
    consecutive_errors = 0

    try:
        while V12_LIVE_JOBS.get(job_key):
            # Check if job was cancelled
            if job.get("cancelled"):
                break

            try:
                # Fetch fresh data
                fresh_data = await asyncio.to_thread(_refresh_match_data, match_id)
                if not fresh_data:
                    fresh_data = _get_match_detail(match_id)
                if not fresh_data:
                    raise RuntimeError("match no encontrado")

                # Get quarter info
                quarters = fresh_data.get("score", {}).get("quarters", {})
                pbp = fresh_data.get("play_by_play", {})
                gp = fresh_data.get("graph_points", [])

                q = quarters.get("Q4") or quarters.get("Q3")
                if not q:
                    raise RuntimeError("no hay datos de quarters")

                quarter_label = "Q4" if "Q4" in quarters else "Q3"
                q_home = int(q.get("home", 0))
                q_away = int(q.get("away", 0))

                q_order = ["Q1", "Q2", "Q3", "Q4"]
                q_idx = q_order.index(quarter_label)
                total_home = total_away = 0
                for i in range(q_idx + 1):
                    qd = quarters.get(q_order[i], {})
                    total_home += int(qd.get("home", 0))
                    total_away += int(qd.get("away", 0))

                elapsed = 6.0
                cutoff = (q_idx * 12) + elapsed
                gp_filtered = [p for p in gp if int(p.get("minute", 0)) <= cutoff]

                result = _run_v12_live_analysis(
                    match_id=match_id,
                    quarter=quarter_label,
                    qtr_home_score=q_home // 2,
                    qtr_away_score=q_away // 2,
                    total_home_score=total_home,
                    total_away_score=total_away,
                    elapsed_minutes=elapsed,
                    graph_points=gp_filtered,
                    pbp_events=pbp.get(quarter_label, [])[:10],
                )

                if not result.get("ok"):
                    raise RuntimeError(result.get("reason", "error"))

                # Build message
                message = _build_v12_live_message(fresh_data, result, quarter_label)

                # Keyboard
                keyboard = [
                    [InlineKeyboardButton("🔄 Refresh", callback_data=f"v12live:refresh:{match_id}:_:{page}")],
                    [InlineKeyboardButton("🔴 Seguir en vivo", callback_data=f"v12live:track:{match_id}:_:{page}")],
                    [InlineKeyboardButton("⬅️ Volver", callback_data=f"match:{match_id}:_:{page}")],
                ]

                # Update message (caption for photo messages, text for text messages)
                cap = message if len(message) <= 1024 else message[:1021] + "..."
                try:
                    if job.get("is_photo"):
                        await app.bot.edit_message_caption(
                            chat_id=chat_id,
                            message_id=msg_id,
                            caption=cap,
                            reply_markup=InlineKeyboardMarkup(keyboard),
                        )
                    else:
                        await app.bot.edit_message_text(
                            chat_id=chat_id,
                            message_id=msg_id,
                            text=message,
                            reply_markup=InlineKeyboardMarkup(keyboard),
                        )
                except Exception as e:
                    # Message might be too long or unchanged
                    if "message is not modified" not in str(e).lower():
                        logger.warning("[V12 LIVE] Edit failed: %s", e)

                consecutive_errors = 0

                # Stop if match finished
                status = str(fresh_data.get("match", {}).get("status_type", "")).lower()
                if status == "finished":
                    job["stop_reason"] = "match_finished"
                    break

            except Exception as exc:
                consecutive_errors += 1
                if consecutive_errors >= 5:
                    job["stop_reason"] = "errors"
                    _err_msg = f"❌ Error V12 LIVE: {str(exc)[:100]}\nSeguimiento detenido."
                    _err_kb = InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Volver", callback_data=f"match:{match_id}:_:{page}")]])
                    try:
                        if job.get("is_photo"):
                            await app.bot.edit_message_caption(chat_id=chat_id, message_id=msg_id, caption=_err_msg, reply_markup=_err_kb)
                        else:
                            await app.bot.edit_message_text(chat_id=chat_id, message_id=msg_id, text=_err_msg, reply_markup=_err_kb)
                    except Exception:
                        pass
                    break

            await asyncio.sleep(poll_interval)

    finally:
        V12_LIVE_JOBS.pop(job_key, None)


async def _handle_v12_live_analysis(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    query,
    match_id: str,
    page: int,
    refresh: bool = False,
):
    """Handle V12 LIVE Bookmaker analysis."""
    try:
        await query.answer("Analizando V12 LIVE...")

        data = await _get_match_detail_async(match_id, update)
        if not data:
            await query.edit_message_text(
                text="Match no encontrado.",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("Menu principal", callback_data="nav:main")]
                ]),
            )
            return

        status_type = str(data.get("match", {}).get("status_type", "") or "").lower()
        if status_type not in ["inprogress", "finished"]:
            await query.edit_message_text(
                text="V12 LIVE solo funciona para partidos en vivo o terminados.\n"
                     "Estado: " + status_type,
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("Volver", callback_data=f"match:{match_id}:_:{page}")]
                ]),
            )
            return

        quarters = data.get("score", {}).get("quarters", {})
        pbp = data.get("play_by_play", {})
        gp = data.get("graph_points", [])

        q = quarters.get("Q4")
        if not q:
            q = quarters.get("Q3")

        if not q:
            await query.edit_message_text(
                text="No hay datos de quarters.",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("Volver", callback_data=f"match:{match_id}:_:{page}")]
                ]),
            )
            return

        quarter_label = "Q4" if "Q4" in quarters else "Q3"
        q_home = int(q.get("home", 0))
        q_away = int(q.get("away", 0))

        q_order = ["Q1", "Q2", "Q3", "Q4"]
        q_idx = q_order.index(quarter_label)
        total_home = 0
        total_away = 0
        for i in range(q_idx + 1):
            q_data = quarters.get(q_order[i], {})
            total_home += int(q_data.get("home", 0))
            total_away += int(q_data.get("away", 0))

        elapsed = 6.0
        cutoff = (q_idx * 12) + elapsed
        gp_filtered = [p for p in gp if int(p.get("minute", 0)) <= cutoff]

        result = _run_v12_live_analysis(
            match_id=match_id,
            quarter=quarter_label,
            qtr_home_score=q_home // 2,
            qtr_away_score=q_away // 2,
            total_home_score=total_home,
            total_away_score=total_away,
            elapsed_minutes=elapsed,
            graph_points=gp_filtered,
            pbp_events=pbp.get(quarter_label, [])[:10],
        )

        if not result.get("ok"):
            await query.edit_message_text(
                text="Error V12 LIVE: " + str(result.get("reason", "Unknown")),
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("Volver", callback_data=f"match:{match_id}:_:{page}")]
                ]),
            )
            return

        message = _build_v12_live_message(data, result, quarter_label)

        keyboard = [
            [InlineKeyboardButton("🔄 Refresh", callback_data=f"v12live:refresh:{match_id}:_:{page}")],
            [InlineKeyboardButton("🔴 Seguir en vivo", callback_data=f"v12live:track:{match_id}:_:{page}")],
            [InlineKeyboardButton("⬅️ Volver", callback_data=f"match:{match_id}:_:{page}")],
        ]

        # Send as photo with analysis as caption
        chat_id = query.message.chat_id
        # Cancel any existing running V12 LIVE jobs for this chat (they reference the old message)
        for _jk in list(V12_LIVE_JOBS.keys()):
            if V12_LIVE_JOBS[_jk].get("chat_id") == chat_id:
                V12_LIVE_JOBS[_jk]["cancelled"] = True
        try:
            await query.message.delete()
        except Exception:
            pass
        caption = message if len(message) <= 1024 else message[:1021] + "..."
        try:
            image_path = _build_graph_image(match_id, data)
            await context.bot.send_photo(
                chat_id=chat_id,
                photo=open(image_path, "rb"),
                caption=caption,
                reply_markup=InlineKeyboardMarkup(keyboard),
            )
        except Exception:
            await context.bot.send_message(
                chat_id=chat_id,
                text=message,
                reply_markup=InlineKeyboardMarkup(keyboard),
            )

    except Exception as exc:
        logger.error("[V12 LIVE] Error: %s", exc, exc_info=True)
        try:
            chat_id = query.message.chat_id if hasattr(query.message, "chat_id") else None
            if chat_id:
                await context.bot.send_message(
                    chat_id=chat_id,
                    text="Error V12 LIVE: " + str(exc)[:200],
                    reply_markup=InlineKeyboardMarkup([
                        [InlineKeyboardButton("Volver", callback_data=f"match:{match_id}:_:{page}")]
                    ]),
                )
        except Exception:
            pass


async def _handle_v12_live_graph(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    query,
    match_id: str,
    page: int,
):
    """Show graph for V12 LIVE match."""
    await query.answer("Generando grafica...")
    data = await _get_match_detail_async(match_id, update)
    if not data:
        await query.answer("Match no encontrado", show_alert=True)
        return

    try:
        image_path = _build_graph_image(match_id, data)
        keyboard = [
            [InlineKeyboardButton("⬅️ Volver a V12 LIVE", callback_data=f"v12live:refresh:{match_id}:_:{page}")],
        ]
        await context.bot.send_photo(
            chat_id=query.message.chat_id,
            photo=open(image_path, "rb"),
            caption=f"Grafica - {data.get('match', {}).get('home_team', '?')} vs {data.get('match', {}).get('away_team', '?')}",
            reply_markup=InlineKeyboardMarkup(keyboard),
        )
    except Exception as exc:
        await query.answer(f"Error generando grafica: {exc}", show_alert=True)


def _open_conn() -> sqlite3.Connection:
    return db_mod.get_conn(DB_PATH)


def _winner_from_scores(home: int | None, away: int | None) -> str:
    if home is None or away is None:
        return "-"
    if home == away:
        return "push"
    return "home" if home > away else "away"


def _winner_short(home: int | None, away: int | None) -> str:
    winner = _winner_from_scores(home, away)
    if winner == "home":
        return "🏠"
    if winner == "away":
        return "✈️"
    if winner == "push":
        return "P"
    return "-"


def _outcome_emoji(outcome: str | None) -> str:
    value = (outcome or "").lower()
    if value == "hit":
        return "✅"
    if value == "miss":
        return "❌"
    if value == "push":
        return "➖"
    return ""


def _abbr_team(name: str, max_len: int = 14) -> str:
    cleaned = " ".join((name or "").strip().split())
    if len(cleaned) <= max_len:
        return cleaned

    parts = [p for p in cleaned.split(" ") if p]
    if len(parts) >= 2:
        initials = "".join(p[0].upper() for p in parts[:4])
        if 2 <= len(initials) <= max_len:
            return initials

    return cleaned[: max_len - 1] + "~"


def _safe_int(value: object) -> int | None:
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _to_local_datetime(date_str: str, time_str: str) -> datetime | None:
    try:
        dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
    except ValueError:
        return None
    return dt + timedelta(hours=UTC_OFFSET_HOURS)


def _fetch_date_summaries() -> list[dict]:
    conn = _open_conn()
    rows = conn.execute(
                f"""
        SELECT
                        date(datetime(m.date || ' ' || m.time, '{UTC_OFFSET_HOURS} hours')) AS event_date,
            COUNT(*) AS total_matches,
            SUM(CASE WHEN q3.home IS NOT NULL AND q3.away IS NOT NULL AND q3.home <> q3.away THEN 1 ELSE 0 END) AS q3_won,
            SUM(CASE WHEN q4.home IS NOT NULL AND q4.away IS NOT NULL AND q4.home <> q4.away THEN 1 ELSE 0 END) AS q4_won
        FROM matches m
        LEFT JOIN quarter_scores q3
          ON q3.match_id = m.match_id AND q3.quarter = 'Q3'
        LEFT JOIN quarter_scores q4
          ON q4.match_id = m.match_id AND q4.quarter = 'Q4'
                GROUP BY date(datetime(m.date || ' ' || m.time, '{UTC_OFFSET_HOURS} hours'))
                ORDER BY event_date DESC
        """
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def _fetch_dates_pred_stats() -> dict[str, dict]:
    """Returns per-date outcome stats in 2 DB queries instead of N+1."""
    conn = _open_conn()
    try:
        match_rows = conn.execute(
            f"""
            SELECT match_id,
                   date(datetime(date || ' ' || time, '{UTC_OFFSET_HOURS} hours')) AS event_date
            FROM matches
            """
        ).fetchall()
        if not match_rows:
            return {}
        mid_to_date: dict[str, str] = {
            str(r["match_id"]): str(r["event_date"] or "")
            for r in match_rows
        }
        tags = _all_result_tags(conn)
        if not tags:
            return {}
        all_ids = list(mid_to_date.keys())
        placeholders = ",".join("?" for _ in all_ids)
        eval_rows = conn.execute(
            f"""
            SELECT * FROM eval_match_results
            WHERE match_id IN ({placeholders})
            ORDER BY match_id, updated_at DESC
            """,
            tuple(all_ids),
        ).fetchall()
    finally:
        conn.close()

    stats: dict[str, dict] = {}
    seen: set[str] = set()
    for row in eval_rows:
        mid = str(row["match_id"])
        if mid in seen:
            continue
        seen.add(mid)
        event_date = mid_to_date.get(mid, "")
        if not event_date:
            continue
        for tag in tags:
            q3_av_key = f"q3_available__{tag}"
            if q3_av_key not in row.keys():
                continue
            q3_av = int(row[q3_av_key] or 0)
            q4_av = int(row[f"q4_available__{tag}"] or 0) if f"q4_available__{tag}" in row.keys() else 0
            if not (q3_av or q4_av):
                continue
            q3o = str(_row_value(row, f"q3_outcome__{tag}") or "").lower() if q3_av else ""
            q4o = str(_row_value(row, f"q4_outcome__{tag}") or "").lower() if q4_av else ""
            c = stats.setdefault(event_date, {"q3_hit": 0, "q3_miss": 0, "q3_push": 0, "q4_hit": 0, "q4_miss": 0, "q4_push": 0})
            if q3o == "hit":
                c["q3_hit"] += 1
            elif q3o == "miss":
                c["q3_miss"] += 1
            elif q3o == "push":
                c["q3_push"] += 1
            if q4o == "hit":
                c["q4_hit"] += 1
            elif q4o == "miss":
                c["q4_miss"] += 1
            elif q4o == "push":
                c["q4_push"] += 1
            break  # first tag with availability wins
    return stats


def _short_date(date_str: str) -> str:
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").strftime("%b %d")
    except ValueError:
        return date_str


def _fetch_matches_for_date(event_date: str) -> list[dict]:
    conn = _open_conn()
    rows = conn.execute(
        f"""
        SELECT
            m.match_id,
            m.date,
            m.time,
            strftime('%H:%M', datetime(m.date || ' ' || m.time, '{UTC_OFFSET_HOURS} hours')) AS display_time,
            m.status_type,
            m.home_team,
            m.away_team,
            m.league,
            m.home_score,
            m.away_score,
            m.event_slug,
            m.custom_id,
            m.home_slug,
            m.away_slug,
            q3.home AS q3_home,
            q3.away AS q3_away,
            q4.home AS q4_home,
            q4.away AS q4_away
        FROM matches m
        LEFT JOIN quarter_scores q3
          ON q3.match_id = m.match_id AND q3.quarter = 'Q3'
        LEFT JOIN quarter_scores q4
          ON q4.match_id = m.match_id AND q4.quarter = 'Q4'
                WHERE date(datetime(m.date || ' ' || m.time, '{UTC_OFFSET_HOURS} hours')) = ?
                ORDER BY datetime(m.date || ' ' || m.time, '{UTC_OFFSET_HOURS} hours') DESC, m.match_id DESC
        """,
        (event_date,),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def _fetch_monitor_log_pred(event_date: str) -> dict[str, dict]:
    """Build a pred_map from bet_monitor_log for a given date."""
    conn = _open_conn()
    try:
        log_rows = conn.execute(
            """
            SELECT match_id, target, signal, pick, result, confidence
            FROM bet_monitor_log
            WHERE event_date = ?
            ORDER BY match_id, id ASC
            """,
            (event_date,),
        ).fetchall()
    except Exception:
        log_rows = []
    conn.close()
    result_map = {"win": "hit", "loss": "miss", "push": "push"}
    out: dict[str, dict] = {}
    for row in log_rows:
        mid = str(row["match_id"])
        target = str(row["target"] or "").lower()
        if target not in ("q3", "q4"):
            continue
        signal = str(row["signal"] or "").strip().upper()
        pick = str(row["pick"] or "").lower()
        db_result = str(row["result"] or "pending").lower()
        outcome = result_map.get(db_result)  # None = pending
        confidence = float(row["confidence"] or 0.0)
        if mid not in out:
            out[mid] = {
                "q3_available": False, "q4_available": False,
                "q3_signal": None, "q4_signal": None,
                "q3_outcome": None, "q4_outcome": None,
                "q3_pick": None, "q4_pick": None,
                "q3_confidence": 0.0, "q4_confidence": 0.0,
            }
        out[mid][f"{target}_available"] = True
        out[mid][f"{target}_signal"] = signal
        out[mid][f"{target}_pick"] = pick
        out[mid][f"{target}_outcome"] = outcome
        out[mid][f"{target}_confidence"] = confidence
    return out


def _fetch_date_pred_outcomes(event_date: str) -> dict[str, dict]:
    match_rows = _fetch_matches_for_date(event_date)
    match_ids = [str(row.get("match_id", "") or "") for row in match_rows]
    match_ids = [mid for mid in match_ids if mid]
    if not match_ids:
        return {}

    conn = _open_conn()
    tags = _all_result_tags(conn)
    if not tags:
        conn.close()
        return {}
    placeholders = ",".join("?" for _ in match_ids)
    rows = conn.execute(
        f"""
        SELECT *
        FROM eval_match_results
        WHERE match_id IN ({placeholders})
        ORDER BY match_id, updated_at DESC
        """,
        tuple(match_ids),
    ).fetchall()
    conn.close()

    result: dict[str, dict] = {}
    seen: set[str] = set()
    for row in rows:
        mid = str(row["match_id"])
        if mid in seen:
            continue
        best_candidate: dict | None = None
        fallback_candidate: dict | None = None

        for tag in tags:
            q3_av_key = f"q3_available__{tag}"
            if q3_av_key not in row.keys():
                continue
            q3_av = int(row[q3_av_key] or 0)
            q4_av_key = f"q4_available__{tag}"
            q4_av = int(row[q4_av_key] or 0)
            candidate = {
                "q3_available": bool(q3_av),
                "q4_available": bool(q4_av),
                "q3_signal": _row_value(row, f"q3_signal__{tag}"),
                "q4_signal": _row_value(row, f"q4_signal__{tag}"),
                "q3_outcome": _row_value(row, f"q3_outcome__{tag}") if q3_av else None,
                "q4_outcome": _row_value(row, f"q4_outcome__{tag}") if q4_av else None,
                "q3_pick": _row_value(row, f"q3_pick__{tag}") if q3_av else None,
                "q4_pick": _row_value(row, f"q4_pick__{tag}") if q4_av else None,
                "q3_confidence": float(_row_value(row, f"q3_confidence__{tag}") or 0.0) if q3_av else 0.0,
                "q4_confidence": float(_row_value(row, f"q4_confidence__{tag}") or 0.0) if q4_av else 0.0,
            }

            if q3_av or q4_av:
                best_candidate = candidate
                break
            if fallback_candidate is None:
                fallback_candidate = candidate

        selected = best_candidate or fallback_candidate
        if selected is not None:
            result[mid] = selected
            seen.add(mid)

    # Merge live monitor signals:
    # - Matches not in eval_match_results: add them directly from monitor log
    # - Matches already in eval_match_results: only override quarters where the
    #   monitor placed an ACTUAL bet (not NO BET / UNAVAILABLE).  This keeps
    #   the recalculated (v13) signal for NO-BET quarters so the date list
    #   stays consistent with the match detail view.
    _NO_BET_SIGNALS = {"NO BET", "NO_BET", "UNAVAILABLE", "ERROR", "window_missed", "no_data", "no_graph"}
    monitor_map = _fetch_monitor_log_pred(event_date)
    for mid, pred in monitor_map.items():
        if mid not in result:
            result[mid] = pred
        else:
            for q in ("q3", "q4"):
                sig = str(pred.get(f"{q}_signal") or "").upper().replace("_", " ")
                if pred.get(f"{q}_available") and sig not in _NO_BET_SIGNALS:
                    result[mid].update({
                        f"{q}_available": True,
                        f"{q}_signal": pred[f"{q}_signal"],
                        f"{q}_pick": pred[f"{q}_pick"],
                        f"{q}_outcome": pred[f"{q}_outcome"],
                        # Prefer monitor confidence if > 0; otherwise keep eval confidence
                        f"{q}_confidence": pred.get(f"{q}_confidence") or result[mid].get(f"{q}_confidence", 0.0),
                    })
    return result


def _pred_stats_text(pred_map: dict, total_matches: int, match_rows: list[dict] | None = None) -> str:
    # ── League exclusion list (same as monitor/signals) ───────────────
    _EXCLUDED_LEAGUE_PATTERNS = [
        "WNBA", "Women", "women", "Feminina", "Femenina",
        "Playoff", "PLAY OFF", "U21 Espoirs Elite", "Liga Femenina",
        "LF Challenge", "Polish Basketball League",
        "SuperSport Premijer Liga", "Prvenstvo Hrvatske za d",
        "ABA Liga", "Argentina Liga Nacional", "Basketligaen",
        "lite 2", "EYBL", "I B MCKL", "Liga 1 Masculin",
        "Liga Nationala", "NBL1", "PBA Commissioner",
        "Rapid League", "Stoiximan GBL", "Playout",
        "Superleague", "Superliga", "Swedish Basketball Superettan",
        "Swiss Cup", "Финал", "Turkish Basketball Super League",
        "NBA",
        "Big V", "Egyptian Basketball Super League", "Lega A Basket",
        "Liga e Par", "Liga Ouro", "Señal", "LNB",
        "Meridianbet KLS", "MPBL", "Nationale 1", "Poland 2nd Basketball League",
        "Portugal LBP", "Portugal Proliga", "Saku I liiga", "Serie A2",
        "Slovenian Second Basketball", "Super League", "United Cup", "United League",
    ]

    def _league_excluded(league: str) -> bool:
        return any(p in league for p in _EXCLUDED_LEAGUE_PATTERNS)
    def _pct(count: int, total: int) -> str:
        if total <= 0:
            return "0.0%"
        return f"{(count * 100.0 / total):.1f}%"

    def _table(headers: list[str], rows: list[list[str]]) -> list[str]:
        widths = [len(h) for h in headers]
        for row in rows:
            for i, cell in enumerate(row):
                widths[i] = max(widths[i], len(cell))
        head = " | ".join(h.ljust(widths[i]) for i, h in enumerate(headers))
        sep = "-+-".join("-" * w for w in widths)
        out = [head, sep]
        for row in rows:
            out.append(" | ".join(cell.ljust(widths[i]) for i, cell in enumerate(row)))
        return out

    # Read configured odds/stake/bank from DB settings
    _conn = _open_conn()
    _stake = float(db_mod.get_setting(_conn, "sig_bet_size", "100") or 100)
    _odds  = float(db_mod.get_setting(_conn, "sig_odds",     "1.4")  or 1.4)
    _bank  = float(db_mod.get_setting(_conn, "sig_bank",     "1000") or 1000)
    _conn.close()

    def _profit_text(hit: int, miss: int, push: int) -> str:
        net = hit * (_stake * (_odds - 1.0)) - (miss + push) * _stake
        bank_end = _bank + net
        sign = "+" if net >= 0 else ""
        return f"{sign}${net:.0f} (${bank_end:.0f})"

    def _resolve_outcome(pred: dict, quarter: str, row: dict) -> str:
        """Return 'hit'|'miss'|'push'|'pending'|'no_bet'."""
        available = bool(pred.get(f"{quarter}_available"))
        if not available:
            return "no_bet"
        signal = str(pred.get(f"{quarter}_signal") or "").strip().upper().replace("_", " ")
        if signal in ("NO BET", "NO_BET"):
            return "no_bet"
        pick = str(pred.get(f"{quarter}_pick") or "").lower()
        status_type = str(row.get("status_type", "") or "").strip().lower()
        is_finished = status_type == "finished"
        q_home = _safe_int(row.get("q3_home" if quarter == "q3" else "q4_home"))
        q_away = _safe_int(row.get("q3_away" if quarter == "q3" else "q4_away"))
        if is_finished and q_home is not None and q_away is not None and pick in ("home", "away"):
            actual_winner = _winner_from_scores(q_home, q_away)
            if actual_winner == "push":
                return "push"
            return "hit" if actual_winner == pick else "miss"
        outcome = str(pred.get(f"{quarter}_outcome") or "pending").lower()
        return outcome if outcome in ("hit", "miss", "push") else "pending"

    def _nobet_would_win(pred: dict, quarter: str, row: dict) -> bool | None:
        """For NO BET signals with a tendency, check if that tendency won.
        Returns True/False/None (None = no tendency or match not finished)."""
        available = bool(pred.get(f"{quarter}_available"))
        signal = str(pred.get(f"{quarter}_signal") or "").strip().upper().replace("_", " ")
        if available and signal not in ("NO BET", "NO_BET"):
            return None  # was a real bet, not a no-bet
        tendency = str(pred.get(f"{quarter}_pick") or "").lower()
        if tendency not in ("home", "away"):
            return None
        status_type = str(row.get("status_type", "") or "").strip().lower()
        if status_type != "finished":
            return None
        q_home = _safe_int(row.get("q3_home" if quarter == "q3" else "q4_home"))
        q_away = _safe_int(row.get("q3_away" if quarter == "q3" else "q4_away"))
        if q_home is None or q_away is None:
            return None
        actual_winner = _winner_from_scores(q_home, q_away)
        if actual_winner == "push":
            return None
        return actual_winner == tendency

    stats = {
        "q3": {"hit": 0, "miss": 0, "push": 0, "pending": 0, "no_bet": 0},
        "q4": {"hit": 0, "miss": 0, "push": 0, "pending": 0, "no_bet": 0},
    }
    # league_stats[league][quarter] = {"hit":0, "miss":0, "push":0}
    league_stats: dict[str, dict[str, dict]] = {}
    # nobet_stats: counts for NO BET tendency phantom breakdown
    nobet_stats = {
        "q3": {"hit": 0, "miss": 0},
        "q4": {"hit": 0, "miss": 0},
    }

    # Build a lookup from match_rows for resolving outcomes dynamically
    match_lookup = {}
    if match_rows:
        for row in match_rows:
            mid = str(row.get("match_id", ""))
            match_lookup[mid] = row

    _NO_BET_SIGS = {"NO BET", "NO_BET"}

    for match_id, pred in pred_map.items():
        row = match_lookup.get(match_id, {})

        # If match_rows was provided, only count matches present in it
        # (any match_id missing from match_lookup was filtered out, e.g. excluded league)
        if match_rows is not None and match_id not in match_lookup:
            continue

        league = str(row.get("league", "") or "").strip()

        # Skip excluded leagues (same filter as monitor/signals)
        if _league_excluded(league):
            continue

        # Shorten to first segment before comma/dash for table readability
        league_short = league.split(",")[0].split("-")[0].strip()[:18] or "?"

        if league_short not in league_stats:
            league_stats[league_short] = {
                "q3": {"hit": 0, "miss": 0, "push": 0},
                "q4": {"hit": 0, "miss": 0, "push": 0},
            }

        for quarter in ("q3", "q4"):
            outcome = _resolve_outcome(pred, quarter, row)

            # For BET signals from monitor log, apply confidence > 30% filter.
            # Only filter when confidence is explicitly known (> 0).
            # eval_match_results entries have no confidence, so conf=0.0 → skip filter.
            if outcome not in ("no_bet", "pending"):
                signal = str(pred.get(f"{quarter}_signal") or "").strip().upper().replace("_", " ")
                _NO_BET_SIGS_SET = {"NO BET", "NO_BET"}
                if signal not in _NO_BET_SIGS_SET:
                    conf = float(pred.get(f"{quarter}_confidence") or 0.0)
                    if 0.0 < conf <= 0.30:
                        stats[quarter]["no_bet"] += 1
                        continue
                    # Apply model-specific filter
                    _pick_v = str(pred.get(f"{quarter}_pick") or "")
                    _q_model = MODEL_CONFIG.get(quarter)
                    if _q_model == "v6":
                        _accept, _ = bet_monitor_mod._v6_pick_filter(league, conf, _pick_v)
                        if not _accept:
                            stats[quarter]["no_bet"] += 1
                            continue
                    if _q_model == "v2":
                        _accept, _ = bet_monitor_mod._v2_pick_filter(league, conf, _pick_v)
                        if not _accept:
                            stats[quarter]["no_bet"] += 1
                            continue

            if outcome == "no_bet":
                stats[quarter]["no_bet"] += 1
                # Phantom no-bet tendency check
                won = _nobet_would_win(pred, quarter, row)
                if won is True:
                    nobet_stats[quarter]["hit"] += 1
                elif won is False:
                    nobet_stats[quarter]["miss"] += 1
            elif outcome == "hit":
                stats[quarter]["hit"] += 1
                league_stats[league_short][quarter]["hit"] += 1
            elif outcome == "miss":
                stats[quarter]["miss"] += 1
                league_stats[league_short][quarter]["miss"] += 1
            elif outcome == "push":
                stats[quarter]["push"] += 1
                league_stats[league_short][quarter]["push"] += 1
            else:
                stats[quarter]["pending"] += 1

    # Note: we don't add "missing" matches to NB anymore, since we now filter
    # by league so total_matches no longer matches the filtered pred_map count.

    total_hit = stats["q3"]["hit"] + stats["q4"]["hit"]
    total_miss = stats["q3"]["miss"] + stats["q4"]["miss"]
    total_push = stats["q3"]["push"] + stats["q4"]["push"]
    total_pending = stats["q3"]["pending"] + stats["q4"]["pending"]
    total_no_bet = stats["q3"]["no_bet"] + stats["q4"]["no_bet"]
    resolved_total = total_hit + total_miss + total_push

    # ── Main summary table ───────────────────────────────────────────
    headers = ["Filtro", "W", "L", "NB", "%W", "Ganancia"]
    rows_main: list[list[str]] = []
    for quarter in ("q3", "q4"):
        s = stats[quarter]
        # Push counts as loss
        total_loss = s["miss"] + s["push"]
        resolved = max(s["hit"] + total_loss, 0)
        no_bet_total = s["no_bet"] + s["pending"]
        rows_main.append([
            quarter.upper(),
            str(s["hit"]),
            str(total_loss),
            str(no_bet_total),
            _pct(s["hit"], resolved),
            _profit_text(s["hit"], s["miss"], s["push"]),
        ])
    rows_main.append([
        "GLOBAL",
        str(total_hit),
        str(total_miss + total_push),
        str(total_no_bet + total_pending),
        _pct(total_hit, resolved_total),
        _profit_text(total_hit, total_miss, total_push),
    ])
    out_parts = ["\n".join(_table(headers, rows_main))]

    # ── By-league table (only leagues with at least one bet) ─────────
    active_leagues = [
        lg for lg, qs in league_stats.items()
        if any(qs[q]["hit"] + qs[q]["miss"] + qs[q]["push"] > 0 for q in ("q3", "q4"))
    ]
    if active_leagues:
        # Determine dominant loss quarter globally (Q3 losses vs Q4 losses)
        total_q3_losses = sum(league_stats[lg]["q3"]["miss"] + league_stats[lg]["q3"]["push"] for lg in active_leagues)
        total_q4_losses = sum(league_stats[lg]["q4"]["miss"] + league_stats[lg]["q4"]["push"] for lg in active_leagues)
        loss_quarter = "q3" if total_q3_losses >= total_q4_losses else "q4"
        # Sort: primary = losses in dominant quarter (desc), secondary = total losses (desc)
        active_leagues.sort(
            key=lambda lg: (
                league_stats[lg][loss_quarter]["miss"] + league_stats[lg][loss_quarter]["push"],
                sum(league_stats[lg][q]["miss"] + league_stats[lg][q]["push"] for q in ("q3", "q4")),
            ),
            reverse=True,
        )
        lg_headers = ["Liga", "Q3W", "Q3L", "Q4W", "Q4L", "%W"]
        lg_rows: list[list[str]] = []
        for lg in active_leagues:
            q3s = league_stats[lg]["q3"]
            q4s = league_stats[lg]["q4"]
            total_w = q3s["hit"] + q4s["hit"]
            # Push counts as loss in league table too
            q3_loss = q3s["miss"] + q3s["push"]
            q4_loss = q4s["miss"] + q4s["push"]
            total_r = total_w + q3_loss + q4_loss
            lg_rows.append([
                lg,
                str(q3s["hit"]),
                str(q3_loss),
                str(q4s["hit"]),
                str(q4_loss),
                _pct(total_w, total_r),
            ])
        out_parts.append("\n\nPor liga:\n" + "\n".join(_table(lg_headers, lg_rows)))

    # ── No-bet phantom table ─────────────────────────────────────────
    nb_q3_total = nobet_stats["q3"]["hit"] + nobet_stats["q3"]["miss"]
    nb_q4_total = nobet_stats["q4"]["hit"] + nobet_stats["q4"]["miss"]
    nb_total = nobet_stats["q3"]["hit"] + nobet_stats["q4"]["hit"] + nobet_stats["q3"]["miss"] + nobet_stats["q4"]["miss"]
    if nb_total > 0:
        nb_headers = ["Filtro", "W", "L", "%W", "Ganancia(sim)"]
        nb_rows: list[list[str]] = []
        for quarter in ("q3", "q4"):
            hw = nobet_stats[quarter]["hit"]
            hl = nobet_stats[quarter]["miss"]
            nb_rows.append([
                quarter.upper(),
                str(hw),
                str(hl),
                _pct(hw, hw + hl),
                _profit_text(hw, hl, 0),
            ])
        nb_w = nobet_stats["q3"]["hit"] + nobet_stats["q4"]["hit"]
        nb_l = nobet_stats["q3"]["miss"] + nobet_stats["q4"]["miss"]
        nb_rows.append([
            "GLOBAL",
            str(nb_w),
            str(nb_l),
            _pct(nb_w, nb_w + nb_l),
            _profit_text(nb_w, nb_l, 0),
        ])
        out_parts.append("\n\nNo apostados (si hubiera apostado):\n" + "\n".join(_table(nb_headers, nb_rows)))

    return "".join(out_parts)


def _event_date_title_es(event_date: str, total_matches: int) -> str:
    try:
        dt = datetime.strptime(event_date, "%Y-%m-%d")
    except ValueError:
        return f"Matches del {event_date} [{total_matches}]"

    weekdays = [
        "Lunes",
        "Martes",
        "Miercoles",
        "Jueves",
        "Viernes",
        "Sabado",
        "Domingo",
    ]
    months = [
        "Enero",
        "Febrero",
        "Marzo",
        "Abril",
        "Mayo",
        "Junio",
        "Julio",
        "Agosto",
        "Septiembre",
        "Octubre",
        "Noviembre",
        "Diciembre",
    ]

    wd = weekdays[dt.weekday()]
    month = months[dt.month - 1]
    return f"Matches de {wd} {dt.day:02d} {month} {dt.year} [{total_matches}]"


def _refresh_recent_dates(
    days: int,
    progress_state: dict[str, object] | None = None,
) -> dict[str, object]:
    summaries = _fetch_date_summaries()
    event_dates = [
        str(row.get("event_date", "") or "")
        for row in summaries
        if str(row.get("event_date", "") or "")
    ][:days]

    if progress_state is not None:
        progress_state.update(
            {
                "phase7d": "running",
                "total_dates": len(event_dates),
                "processed_dates": 0,
                "current_date": "",
                "sum_total": 0,
                "sum_updated_ok": 0,
                "sum_still_incomplete": 0,
                "sum_skipped_complete": 0,
                "sum_failed": 0,
            }
        )

    per_date: list[dict[str, object]] = []
    sum_total = 0
    sum_updated_ok = 0
    sum_still_incomplete = 0
    sum_skipped_complete = 0
    sum_failed = 0

    for index, event_date in enumerate(event_dates, start=1):
        if progress_state is not None:
            progress_state.update(
                {
                    "current_date": event_date,
                    "processed_dates": index - 1,
                }
            )

        date_result = _refresh_incomplete_matches_for_date(
            event_date,
            progress_state,
        )
        per_date.append(date_result)

        sum_total += int(date_result.get("total") or 0)
        sum_updated_ok += int(date_result.get("updated_ok") or 0)
        sum_still_incomplete += int(date_result.get("still_incomplete") or 0)
        sum_skipped_complete += int(date_result.get("skipped_complete") or 0)
        sum_failed += int(date_result.get("failed") or 0)

        if progress_state is not None:
            progress_state.update(
                {
                    "processed_dates": index,
                    "sum_total": sum_total,
                    "sum_updated_ok": sum_updated_ok,
                    "sum_still_incomplete": sum_still_incomplete,
                    "sum_skipped_complete": sum_skipped_complete,
                    "sum_failed": sum_failed,
                }
            )

    if progress_state is not None:
        progress_state.update(
            {
                "phase7d": "done",
                "processed_dates": len(event_dates),
                "current_date": "",
            }
        )

    return {
        "days": days,
        "dates": event_dates,
        "per_date": per_date,
        "sum_total": sum_total,
        "sum_updated_ok": sum_updated_ok,
        "sum_still_incomplete": sum_still_incomplete,
        "sum_skipped_complete": sum_skipped_complete,
        "sum_failed": sum_failed,
    }


def _refresh_7d_progress_text(progress_state: dict[str, object]) -> str:
    def _fmt_duration(seconds: float | int) -> str:
        total = max(int(seconds), 0)
        mm, ss = divmod(total, 60)
        hh, mm = divmod(mm, 60)
        if hh:
            return f"{hh}h {mm:02d}m {ss:02d}s"
        return f"{mm:02d}m {ss:02d}s"

    phase = str(progress_state.get("phase7d") or "starting")
    total_dates = int(progress_state.get("total_dates") or 0)
    processed_dates = int(progress_state.get("processed_dates") or 0)
    current_date = str(progress_state.get("current_date") or "")
    sum_total = int(progress_state.get("sum_total") or 0)
    sum_updated_ok = int(progress_state.get("sum_updated_ok") or 0)
    sum_still_incomplete = int(progress_state.get("sum_still_incomplete") or 0)
    sum_skipped_complete = int(progress_state.get("sum_skipped_complete") or 0)
    sum_failed = int(progress_state.get("sum_failed") or 0)
    current_total_matches = int(progress_state.get("total_matches") or 0)
    current_pending_total = int(progress_state.get("pending_total") or 0)
    current_processed = int(progress_state.get("processed") or 0)
    current_updated_ok = int(progress_state.get("updated_ok") or 0)
    current_still_incomplete = int(progress_state.get("still_incomplete") or 0)
    current_failed = int(progress_state.get("failed") or 0)
    current_skipped_complete = int(progress_state.get("skipped_complete") or 0)

    started_monotonic = progress_state.get("started_monotonic")
    elapsed_txt = None
    eta_txt = None
    if isinstance(started_monotonic, (int, float)):
        elapsed = max(0.0, time.monotonic() - float(started_monotonic))
        elapsed_txt = _fmt_duration(elapsed)
        if processed_dates > 0 and total_dates > processed_dates:
            speed = elapsed / float(processed_dates)
            eta = speed * float(total_dates - processed_dates)
            eta_txt = _fmt_duration(eta)

    lines = [
        "[refresh-7d] rango=ultimos 7 dias",
        f"[refresh-7d] progreso_fechas={processed_dates}/{total_dates}",
    ]
    if current_date:
        lines.append(f"[refresh-7d] fecha_actual={current_date}")
        lines.append(
            "[refresh-7d] "
            f"progreso_fecha={current_processed}/{current_pending_total} "
            f"actualizados={current_updated_ok} "
            f"aun_incompletos={current_still_incomplete} "
            f"ya_completos={current_skipped_complete} "
            f"fallidos={current_failed} total_matches={current_total_matches}"
        )
    lines.append(
        "[refresh-7d] "
        f"total={sum_total} actualizados={sum_updated_ok} "
        f"aun_incompletos={sum_still_incomplete} "
        f"ya_completos={sum_skipped_complete} fallidos={sum_failed}"
    )
    if elapsed_txt is not None:
        if eta_txt is not None and phase != "done":
            lines.append(f"[refresh-7d] tiempo elapsed={elapsed_txt} eta={eta_txt}")
        else:
            lines.append(f"[refresh-7d] tiempo elapsed={elapsed_txt}")
    lines.append(
        "[refresh-7d] estado=finalizado" if phase == "done"
        else "[refresh-7d] estado=procesando"
    )
    return "\n".join(lines)


def _pred_outcome_emoji_for_row(pred: dict, quarter: str) -> str:
    available = bool(pred.get(f"{quarter}_available"))
    if not pred:
        return ""
    if not available:
        return "🚫"
    signal = str(pred.get(f"{quarter}_signal") or "").strip().upper().replace("_", " ")
    if signal == "NO BET":
        return "🚫"
    outcome = str(pred.get(f"{quarter}_outcome") or "").lower()
    if outcome == "hit":
        return "✅"
    if outcome == "miss":
        return "❌"
    if outcome == "push":
        return "➖"
    return "⏳"


def _get_match_detail(match_id: str) -> dict | None:
    """Sync version: fetch from DB only (used in backgrounds tasks)"""
    conn = _open_conn()
    data = db_mod.get_match(conn, match_id)
    conn.close()
    return data


async def _get_match_detail_async(match_id: str, update=None) -> dict | None:
    """Async version: fetch from DB, then scrape if not found (for user-facing requests)"""
    conn = _open_conn()
    data = db_mod.get_match(conn, match_id)
    conn.close()
    
    # If not found in DB, try to fetch from scraper in a thread
    if not data:
        download_msg = None
        try:
            logger.info(f"[GET_MATCH_DETAIL] Match {match_id} not in DB, attempting scrape...")
            
            # Send "Downloading..." message to user
            if update:
                download_msg = await update.message.reply_text("Descargando...")
            
            def _fetch_and_save():
                # Create a new connection inside the thread (SQLite thread safety)
                thread_conn = _open_conn()
                fresh = scraper_mod.fetch_match_by_id(match_id)
                if fresh:
                    db_mod.save_match(thread_conn, match_id, fresh)
                    logger.info(f"[GET_MATCH_DETAIL] Successfully scraped and saved match {match_id}")
                thread_conn.close()
                return fresh
            
            data = await asyncio.to_thread(_fetch_and_save)
            
            # Delete "Downloading..." message after scrape completes
            if download_msg:
                try:
                    await download_msg.delete()
                except Exception:
                    pass
                    
        except Exception as exc:
            logger.warning(f"[GET_MATCH_DETAIL] Scrape failed for {match_id}: {exc}")
            if download_msg:
                try:
                    await download_msg.delete()
                except Exception:
                    pass
    
    return data


def _is_ft_complete(data: dict | None) -> bool:
    if not data:
        return False
    status_type = str(data.get("match", {}).get("status_type", "") or "").lower()
    if status_type != "finished":
        return False

    quarters = data.get("score", {}).get("quarters", {})
    required = ("Q1", "Q2", "Q3", "Q4")
    for q in required:
        q_score = quarters.get(q)
        if not isinstance(q_score, dict):
            return False
        if q_score.get("home") is None or q_score.get("away") is None:
            return False

    pbp = data.get("play_by_play", {})
    if not isinstance(pbp, dict):
        return False
    for q in required:
        q_plays = pbp.get(q)
        if not isinstance(q_plays, list) or len(q_plays) == 0:
            return False

    graph_points = data.get("graph_points", [])
    if not isinstance(graph_points, list) or len(graph_points) == 0:
        return False

    max_minute = None
    for point in graph_points:
        try:
            minute = int((point or {}).get("minute"))
        except (TypeError, ValueError, AttributeError):
            continue
        max_minute = minute if max_minute is None else max(max_minute, minute)

    if max_minute is None or max_minute < 48:
        return False

    return True


def _refresh_incomplete_matches_for_date(
    event_date: str,
    progress_state: dict[str, object] | None = None,
) -> dict:
    logger.info(f"[REFRESH_DATE] Starting refresh for date={event_date}")
    rows = _fetch_matches_for_date(event_date)
    total = len(rows)
    updated_ok = 0
    still_incomplete = 0
    failed = 0
    skipped_complete = 0

    if progress_state is not None:
        progress_state.update(
            {
                "phase": "scanning",
                "total_matches": total,
                "pending_total": 0,
                "processed": 0,
                "updated_ok": 0,
                "still_incomplete": 0,
                "failed": 0,
                "skipped_complete": 0,
            }
        )

    conn = _open_conn()
    pending_ids: list[str] = []
    for row in rows:
        match_id = str(row.get("match_id", "") or "")
        if not match_id:
            continue

        current = db_mod.get_match(conn, match_id)
        if _is_ft_complete(current):
            skipped_complete += 1
            continue

        pending_ids.append(match_id)

    pending_total = len(pending_ids)
    if progress_state is not None:
        progress_state.update(
            {
                "phase": "refreshing",
                "pending_total": pending_total,
                "processed": 0,
                "updated_ok": 0,
                "still_incomplete": 0,
                "failed": 0,
                "skipped_complete": skipped_complete,
            }
        )

    for index, match_id in enumerate(pending_ids, start=1):
        try:
            fresh = scraper_mod.fetch_match_by_id(match_id)
            db_mod.save_match(conn, match_id, fresh)
            if _is_ft_complete(fresh):
                updated_ok += 1
            else:
                still_incomplete += 1
        except Exception as exc:
            logger.warning(f"[REFRESH_DATE] Failed match={match_id}: {exc}")
            failed += 1

        if progress_state is not None:
            progress_state.update(
                {
                    "phase": "refreshing",
                    "processed": index,
                    "updated_ok": updated_ok,
                    "still_incomplete": still_incomplete,
                    "failed": failed,
                    "skipped_complete": skipped_complete,
                }
            )

    conn.close()
    if progress_state is not None:
        progress_state.update(
            {
                "phase": "done",
                "processed": pending_total,
                "updated_ok": updated_ok,
                "still_incomplete": still_incomplete,
                "failed": failed,
                "skipped_complete": skipped_complete,
            }
        )
    logger.info(
        "[REFRESH_DATE] Done date=%s total=%s updated_ok=%s still_incomplete=%s skipped_complete=%s failed=%s",
        event_date,
        total,
        updated_ok,
        still_incomplete,
        skipped_complete,
        failed,
    )
    return {
        "date": event_date,
        "total": total,
        "updated_ok": updated_ok,
        "still_incomplete": still_incomplete,
        "failed": failed,
        "skipped_complete": skipped_complete,
    }


def _refresh_date_progress_text(event_date: str, progress_state: dict[str, object]) -> str:
    def _fmt_duration(seconds: float | int) -> str:
        total = max(int(seconds), 0)
        mm, ss = divmod(total, 60)
        hh, mm = divmod(mm, 60)
        if hh:
            return f"{hh}h {mm:02d}m {ss:02d}s"
        return f"{mm:02d}m {ss:02d}s"

    phase = str(progress_state.get("phase") or "starting")
    total_matches = int(progress_state.get("total_matches") or 0)
    pending_total = int(progress_state.get("pending_total") or 0)
    processed = int(progress_state.get("processed") or 0)
    updated_ok = int(progress_state.get("updated_ok") or 0)
    still_incomplete = int(progress_state.get("still_incomplete") or 0)
    failed = int(progress_state.get("failed") or 0)
    skipped_complete = int(progress_state.get("skipped_complete") or 0)

    started_monotonic = progress_state.get("started_monotonic")
    elapsed_txt = None
    eta_txt = None
    if isinstance(started_monotonic, (int, float)):
        elapsed = max(0.0, time.monotonic() - float(started_monotonic))
        elapsed_txt = _fmt_duration(elapsed)
        if processed > 0 and pending_total > processed:
            speed = elapsed / float(processed)
            eta = speed * float(pending_total - processed)
            eta_txt = _fmt_duration(eta)

    lines = [
        f"[refresh-date] date={event_date}",
        f"[refresh-date] total_matches={total_matches}",
    ]
    if phase in ("starting", "scanning"):
        lines.append("[refresh-date] estado=detectando pendientes...")
        lines.append(f"[refresh-date] ya_completos={skipped_complete}")
    else:
        lines.append(f"[refresh-date] pendientes={pending_total}")
        lines.append(
            "[refresh-date] "
            f"progreso={processed}/{pending_total} "
            f"actualizados={updated_ok} "
            f"aun_incompletos={still_incomplete} "
            f"fallidos={failed} "
            f"ya_completos={skipped_complete}"
        )

    if elapsed_txt is not None:
        if eta_txt is not None and phase not in ("done", "cancelled"):
            lines.append(f"[refresh-date] tiempo elapsed={elapsed_txt} eta={eta_txt}")
        else:
            lines.append(f"[refresh-date] tiempo elapsed={elapsed_txt}")

    if phase == "done":
        lines.append("[refresh-date] estado=finalizado")

    return "\n".join(lines)


def _get_missing_dates_suggestions(recent_days: int = 21) -> dict:
    """Return missing dates split into recent and historical.

    Returns a dict with:
      - 'recent':     dates in the last recent_days with 0 matches
      - 'historical': older dates (from min_date in DB to recent cutoff) with 0 matches
      - 'min_date':   earliest event_date in the DB (str or None)
    """
    today = datetime.now().date()
    recent_cutoff = today - timedelta(days=recent_days)

    conn = _open_conn()
    bounds = conn.execute(
        f"""
        SELECT
            MIN(date(datetime(date || ' ' || time, '{UTC_OFFSET_HOURS} hours'))) AS min_date
        FROM matches
        """
    ).fetchone()
    min_date_str = bounds["min_date"] if bounds else None

    existing_rows = conn.execute(
        f"""
        SELECT DISTINCT
            date(datetime(date || ' ' || time, '{UTC_OFFSET_HOURS} hours'))
            AS event_date
        FROM matches
        """
    ).fetchall()
    conn.close()

    existing = {str(r["event_date"]) for r in existing_rows}

    # Recent: last recent_days up to yesterday
    recent_missing = [
        (today - timedelta(days=i)).isoformat()
        for i in range(1, recent_days + 1)
        if (today - timedelta(days=i)).isoformat() not in existing
    ]

    # Historical: from min_date up to recent_cutoff - 1
    historical_missing: list[str] = []
    if min_date_str:
        try:
            min_date = datetime.strptime(min_date_str, "%Y-%m-%d").date()
            cursor = recent_cutoff - timedelta(days=1)
            while cursor >= min_date:
                d = cursor.isoformat()
                if d not in existing:
                    historical_missing.append(d)
                cursor -= timedelta(days=1)
        except ValueError:
            pass

    return {
        "recent": recent_missing,
        "historical": historical_missing,
        "min_date": min_date_str,
    }


def _skip_reason_detail(data: dict | None, *, already_complete: bool = False, two_half: bool = False) -> str:
    """Return a human-readable skip reason for a match that was not ingested."""
    if already_complete:
        return "ya_completo_en_db"
    if two_half:
        return "dos_mitades"
    if data is None:
        return "sin_datos"
    status_type = str(data.get("match", {}).get("status_type", "") or "").lower()
    if status_type != "finished":
        return f"not_finished(status={status_type or 'unknown'})"
    quarters = data.get("score", {}).get("quarters", {})
    required = ("Q1", "Q2", "Q3", "Q4")
    for q in required:
        q_score = quarters.get(q)
        if not isinstance(q_score, dict) or q_score.get("home") is None or q_score.get("away") is None:
            return f"sin_cuartos(falta={q})"
    pbp = data.get("play_by_play", {})
    if not isinstance(pbp, dict):
        return "sin_pbp(no_dict)"
    for q in required:
        q_plays = pbp.get(q)
        if not isinstance(q_plays, list) or len(q_plays) == 0:
            return f"sin_pbp(falta={q})"
    graph_points = data.get("graph_points", [])
    if not isinstance(graph_points, list) or len(graph_points) == 0:
        return "sin_graph(vacio)"
    max_minute = None
    for point in graph_points:
        try:
            minute = int((point or {}).get("minute"))
        except (TypeError, ValueError, AttributeError):
            continue
        max_minute = minute if max_minute is None else max(max_minute, minute)
    if max_minute is None or max_minute < 48:
        return f"sin_graph(max_min={max_minute})"
    return "desconocido"


def _ingest_new_date(event_date: str, limit: int | None = None) -> dict:
    rows_all = scraper_mod.fetch_finished_match_ids_for_date(event_date)
    rows = rows_all[:limit] if limit is not None else rows_all
    conn = _open_conn()
    discovered = db_mod.save_discovered_ft_matches(conn, rows)

    ing_ok = 0
    ing_fail = 0
    skipped_ft = 0

    for row in rows:
        match_id = str(row.get("match_id", "") or "")
        if not match_id:
            continue

        existing = db_mod.get_match(conn, match_id)
        if _is_ft_complete(existing):
            db_mod.mark_discovered_processed(conn, match_id)
            skipped_ft += 1
            continue

        try:
            data = scraper_mod.fetch_match_by_id(match_id)
            if not _is_ft_complete(data):
                db_mod.mark_discovered_processed(conn, match_id)
                skipped_ft += 1
                continue
            if _is_two_half_game(data):
                db_mod.mark_discovered_processed(conn, match_id)
                skipped_ft += 1
                print(f"[date-ingest] skip 2-half match_id={match_id}", flush=True)
                continue
            db_mod.save_match(conn, match_id, data)
            db_mod.mark_discovered_processed(conn, match_id)
            ing_ok += 1
        except Exception as exc:
            db_mod.mark_discovered_error(conn, match_id, str(exc))
            ing_fail += 1

    pending_row = conn.execute(
        """
        SELECT COUNT(*) AS n
        FROM discovered_ft_matches
        WHERE event_date = ? AND processed = 0
        """,
        (event_date,),
    ).fetchone()
    conn.close()

    pending = int(pending_row["n"] if pending_row else 0)
    return {
        "date": event_date,
        "finished_found_total": len(rows_all),
        "finished_selected": len(rows),
        "limit_applied": limit,
        "discovered_rows": discovered,
        "ingested_ok": ing_ok,
        "ingested_fail": ing_fail,
        "skipped_ft": skipped_ft,
        "pending_after": pending,
    }


def _get_discovered_date_state(event_date: str) -> dict:
    conn = _open_conn()
    total_row = conn.execute(
        "SELECT COUNT(*) AS n FROM discovered_ft_matches WHERE event_date = ?",
        (event_date,),
    ).fetchone()
    pending_row = conn.execute(
        "SELECT COUNT(*) AS n FROM discovered_ft_matches WHERE event_date = ? AND processed = 0",
        (event_date,),
    ).fetchone()
    processed_row = conn.execute(
        "SELECT COUNT(*) AS n FROM discovered_ft_matches WHERE event_date = ? AND processed = 1",
        (event_date,),
    ).fetchone()
    conn.close()
    return {
        "discovered_total": int(total_row["n"] if total_row else 0),
        "pending_total": int(pending_row["n"] if pending_row else 0),
        "processed_total": int(processed_row["n"] if processed_row else 0),
    }


def _load_resume_rows(event_date: str, limit: int | None) -> list[dict]:
    conn = _open_conn()
    rows = db_mod.list_pending_discovered_ft(
        conn,
        event_date,
        event_date,
        limit=limit,
    )
    conn.close()
    return rows


def _ingest_new_date_with_progress(
    event_date: str,
    limit: int | None,
    progress_state: dict,
    cancel_event,
) -> dict:
    print(
        f"[date-ingest] start event_date={event_date} limit={limit}",
        flush=True,
    )
    resume_state = _get_discovered_date_state(event_date)
    rows_all: list[dict] | None = None
    if resume_state["pending_total"] > 0 and resume_state["discovered_total"] > 0:
        rows = _load_resume_rows(event_date, limit=limit)
        selected_total = (
            min(resume_state["discovered_total"], limit)
            if limit is not None
            else resume_state["discovered_total"]
        )
        processed_initial = max(0, selected_total - len(rows))
        progress_state["phase"] = "resuming"
        progress_state["finished_found_total"] = resume_state["discovered_total"]
        progress_state["finished_selected"] = selected_total
        progress_state["processed"] = processed_initial
        progress_state["resume_used"] = True
        print(
            (
                "[date-ingest] resume mode "
                f"discovered_total={resume_state['discovered_total']} "
                f"pending_total={resume_state['pending_total']} "
                f"selected_total={selected_total}"
            ),
            flush=True,
        )
    else:
        rows_all = scraper_mod.fetch_finished_match_ids_for_date(event_date)
        rows = rows_all[:limit] if limit is not None else rows_all
        progress_state["phase"] = "ingesting"
        progress_state["finished_found_total"] = len(rows_all)
        progress_state["finished_selected"] = len(rows)
        progress_state["processed"] = 0
        progress_state["resume_used"] = False
        print(
            (
                "[date-ingest] fresh discover mode "
                f"finished_found_total={len(rows_all)} "
                f"finished_selected={len(rows)}"
            ),
            flush=True,
        )
    progress_state["total"] = int(progress_state.get("finished_selected") or len(rows))

    conn = _open_conn()
    if rows_all is not None:
        discovered = db_mod.save_discovered_ft_matches(conn, rows)
    else:
        discovered = len(rows)
    progress_state["discovered_rows"] = discovered

    ing_ok = 0
    ing_fail = 0
    skipped_ft = 0
    skip_reasons: dict[str, int] = {}

    def _record_skip(reason: str) -> None:
        nonlocal skipped_ft
        skipped_ft += 1
        skip_reasons[reason] = skip_reasons.get(reason, 0) + 1

    for index, row in enumerate(rows, start=1):
        if cancel_event.is_set():
            progress_state["phase"] = "cancelled"
            break
        match_id = str(row.get("match_id", "") or "")
        if not match_id:
            progress_state["processed"] = int(progress_state.get("processed") or 0) + 1
            continue

        existing = db_mod.get_match(conn, match_id)
        if _is_ft_complete(existing):
            db_mod.mark_discovered_processed(conn, match_id)
            reason = _skip_reason_detail(existing, already_complete=True)
            _record_skip(reason)
            print(f"[date-ingest] skip match_id={match_id} reason={reason}", flush=True)
            progress_state["processed"] = int(progress_state.get("processed") or 0) + 1
            progress_state["ingested_ok"] = ing_ok
            progress_state["ingested_fail"] = ing_fail
            progress_state["skipped_ft"] = skipped_ft
            progress_state["skip_reasons"] = dict(skip_reasons)
            continue

        try:
            data = scraper_mod.fetch_match_by_id(match_id)
            if not _is_ft_complete(data):
                db_mod.mark_discovered_processed(conn, match_id)
                reason = _skip_reason_detail(data)
                _record_skip(reason)
                print(f"[date-ingest] skip match_id={match_id} reason={reason}", flush=True)
                progress_state["processed"] = int(progress_state.get("processed") or 0) + 1
                progress_state["ingested_ok"] = ing_ok
                progress_state["ingested_fail"] = ing_fail
                progress_state["skipped_ft"] = skipped_ft
                progress_state["skip_reasons"] = dict(skip_reasons)
                if index % DATE_INGEST_PROGRESS_EVERY == 0:
                    progress_state["last_progress_at"] = index
                continue
            if _is_two_half_game(data):
                db_mod.mark_discovered_processed(conn, match_id)
                reason = _skip_reason_detail(data, two_half=True)
                _record_skip(reason)
                print(f"[date-ingest] skip match_id={match_id} reason={reason}", flush=True)
                progress_state["processed"] = int(progress_state.get("processed") or 0) + 1
                progress_state["ingested_ok"] = ing_ok
                progress_state["ingested_fail"] = ing_fail
                progress_state["skipped_ft"] = skipped_ft
                progress_state["skip_reasons"] = dict(skip_reasons)
                if index % DATE_INGEST_PROGRESS_EVERY == 0:
                    progress_state["last_progress_at"] = index
                continue
            db_mod.save_match(conn, match_id, data)
            db_mod.mark_discovered_processed(conn, match_id)
            ing_ok += 1
        except Exception as exc:
            db_mod.mark_discovered_error(conn, match_id, str(exc))
            ing_fail += 1

        progress_state["processed"] = int(progress_state.get("processed") or 0) + 1
        progress_state["ingested_ok"] = ing_ok
        progress_state["ingested_fail"] = ing_fail
        progress_state["skipped_ft"] = skipped_ft
        progress_state["skip_reasons"] = dict(skip_reasons)
        if index % DATE_INGEST_PROGRESS_EVERY == 0:
            progress_state["last_progress_at"] = index

    pending_row = conn.execute(
        """
        SELECT COUNT(*) AS n
        FROM discovered_ft_matches
        WHERE event_date = ? AND processed = 0
        """,
        (event_date,),
    ).fetchone()
    conn.close()

    pending = int(pending_row["n"] if pending_row else 0)
    if cancel_event.is_set():
        progress_state["phase"] = "cancelled"
    else:
        progress_state["phase"] = "done"
    progress_state["pending_after"] = pending
    progress_state["skip_reasons"] = dict(skip_reasons)
    if skip_reasons:
        reasons_txt = "  ".join(f"{k}={v}" for k, v in sorted(skip_reasons.items()))
        print(f"[date-ingest] skip_reasons: {reasons_txt}", flush=True)
    return {
        "date": event_date,
        "finished_found_total": progress_state["finished_found_total"],
        "finished_selected": progress_state["finished_selected"],
        "limit_applied": limit,
        "discovered_rows": discovered,
        "ingested_ok": ing_ok,
        "ingested_fail": ing_fail,
        "skipped_ft": skipped_ft,
        "skip_reasons": dict(skip_reasons),
        "pending_after": pending,
        "cancelled": bool(cancel_event.is_set()),
        "resume_used": bool(progress_state.get("resume_used")),
    }


def _is_message_not_modified_error(exc: Exception) -> bool:
    return isinstance(exc, BadRequest) and "Message is not modified" in str(exc)


async def _safe_edit_message(
    app: Application,
    chat_id: int,
    message_id: int,
    text: str,
    reply_markup: InlineKeyboardMarkup | None = None,
) -> bool:
    try:
        await app.bot.edit_message_text(
            chat_id=chat_id,
            message_id=message_id,
            text=text,
            reply_markup=reply_markup,
        )
        return True
    except Exception as exc:
        if _is_message_not_modified_error(exc):
            print("[date-ingest] telegram edit skipped (message not modified)", flush=True)
            return False
        if isinstance(exc, (TimedOut, NetworkError)):
            print(
                f"[date-ingest] telegram edit timeout/network error: {type(exc).__name__}",
                flush=True,
            )
            return False

        if isinstance(exc, BadRequest):
            text = str(exc)
            if (
                "Message to edit not found" in text
                or "message can't be edited" in text
                or "chat not found" in text
            ):
                print(
                    f"[date-ingest] telegram edit skipped (bad request): {text}",
                    flush=True,
                )
                return False

        print(
            f"[date-ingest] telegram edit unexpected error: {type(exc).__name__}: {exc}",
            flush=True,
        )
        return False


def _date_ingest_progress_text(event_date: str, limit: int | None, progress_state: dict) -> str:
    def _fmt_duration(seconds: float | int) -> str:
        total = max(int(seconds), 0)
        mm, ss = divmod(total, 60)
        hh, mm = divmod(mm, 60)
        if hh:
            return f"{hh}h {mm:02d}m {ss:02d}s"
        return f"{mm:02d}m {ss:02d}s"

    phase = str(progress_state.get("phase") or "starting")
    found_total = int(progress_state.get("finished_found_total") or 0)
    selected = int(progress_state.get("finished_selected") or 0)
    processed = int(progress_state.get("processed") or 0)
    total = int(progress_state.get("total") or 0)
    ing_ok = int(progress_state.get("ingested_ok") or 0)
    ing_fail = int(progress_state.get("ingested_fail") or 0)
    skipped_ft = int(progress_state.get("skipped_ft") or 0)
    resume_used = bool(progress_state.get("resume_used"))
    skip_reasons = progress_state.get("skip_reasons")
    started_monotonic = progress_state.get("started_monotonic")

    elapsed_txt = None
    eta_txt = None
    if isinstance(started_monotonic, (int, float)):
        elapsed = max(0.0, time.monotonic() - float(started_monotonic))
        elapsed_txt = _fmt_duration(elapsed)
        if processed > 0 and total > processed:
            speed = elapsed / float(processed)
            eta = speed * float(total - processed)
            eta_txt = _fmt_duration(eta)

    lines = [f"Traer fecha: {event_date}"]
    if phase == "starting":
        lines.append("Estado: iniciando consulta a SofaScore...")
    else:
        if resume_used:
            lines.append("Modo: reanudando desde descubiertos guardados")
        limit_txt = str(limit) if limit is not None else "sin limite"
        lines.append(
            f"Encontrados: {found_total}  |  Procesando: {selected}  |  Limite: {limit_txt}"
        )
        lines.append(
            f"Progreso: {processed}/{total}  guardados={ing_ok}  errores={ing_fail}"
        )
        if skipped_ft:
            if isinstance(skip_reasons, dict) and skip_reasons:
                reasons_parts = "  ".join(
                    f"{k}={v}" for k, v in sorted(skip_reasons.items())
                )
                lines.append(f"Saltados: {skipped_ft}  ({reasons_parts})")
            else:
                lines.append(f"Saltados: {skipped_ft}")
        if elapsed_txt is not None:
            if eta_txt is not None and phase not in ("done", "cancelled"):
                lines.append(f"Tiempo: {elapsed_txt}  ETA: {eta_txt}")
            else:
                lines.append(f"Tiempo: {elapsed_txt}")
        if phase == "done":
            pending_after = int(progress_state.get("pending_after") or 0)
            if pending_after:
                lines.append(f"Pendientes tras proceso: {pending_after}")
            lines.append("Estado: finalizado ✓")
        elif phase == "cancelled":
            pending_after = int(progress_state.get("pending_after") or 0)
            if pending_after:
                lines.append(f"Pendientes tras proceso: {pending_after}")
            lines.append("Estado: cancelado")
        else:
            lines.append("Estado: procesando...")
    return "\n".join(lines)


def _date_ingest_progress_keyboard(running: bool) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    if running:
        rows.append(
            [InlineKeyboardButton("Cancelar", callback_data="fetchdate:cancel")]
        )
    rows.append([InlineKeyboardButton("Menu principal", callback_data="nav:main")])
    return InlineKeyboardMarkup(rows)


def _refresh_job_progress_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [[InlineKeyboardButton("Menu principal", callback_data="nav:main")]]
    )


def _refresh_job_status_text(job: dict[str, object]) -> str:
    mode = str(job.get("mode") or "date")
    event_date = str(job.get("event_date") or "")
    progress_state = job.get("progress_state")
    if not isinstance(progress_state, dict):
        progress_state = {"phase": "starting"}

    if mode == "7d":
        return _refresh_7d_progress_text(progress_state)
    return _refresh_date_progress_text(event_date, progress_state)


def _follow_status_text(job: dict[str, object]) -> str:
    match_id = str(job.get("match_id") or "-")
    phase = str(job.get("phase") or "running")
    reason = str(job.get("stop_reason") or "")
    cycles = int(job.get("cycles") or 0)
    unchanged = int(job.get("unchanged_cycles") or 0)
    last_tick = str(job.get("last_tick_utc") or "-")
    score = str(job.get("last_score") or "-")
    status = str(job.get("last_status") or "-")
    minute_est = job.get("minute_est")
    minute_txt = f"{minute_est}" if minute_est is not None else "-"

    lines = [
        f"Seguimiento match_id={match_id}",
        f"estado={phase}",
        f"status={status} minuto={minute_txt} score={score}",
        f"ciclos={cycles} sin_cambios={unchanged}",
        f"ultimo_tick_utc={last_tick}",
    ]
    if reason:
        lines.append(f"motivo={reason}")
    return "\n".join(lines)


def _follow_status_keyboard(job: dict[str, object]) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    phase = str(job.get("phase") or "running")
    if phase == "running":
        rows.append([InlineKeyboardButton("Detener seguimiento", callback_data="follow:stop")])

    match_id = str(job.get("match_id") or "")
    event_token = str(job.get("event_token") or "_")
    page = int(job.get("page") or 0)
    if match_id:
        rows.append([InlineKeyboardButton("Abrir match seguido", callback_data=f"follow:open:{match_id}:{event_token}:{page}")])

    rows.append([InlineKeyboardButton("Menu principal", callback_data="nav:main")])
    return InlineKeyboardMarkup(rows)


async def _publish_follow_status(
    app: Application,
    chat_id: int,
    job: dict[str, object],
    *,
    force: bool = False,
) -> None:
    match_id = str(job.get("match_id") or "")
    data = job.get("follow_data")
    pred_row = job.get("follow_pred")
    is_detail = bool(data is not None and isinstance(data, dict))
    
    if is_detail:
        text = _detail_text(match_id, data, pred_row, following=True)
    else:
        text = _follow_status_text(job)

    event_token = str(job.get("event_token") or "_")
    event_date = None if event_token == "_" else event_token
    keyboard = _detail_keyboard(
        match_id,
        event_date=event_date,
        page=int(job.get("page") or 0),
        match_data=data,
        chat_id=chat_id,
    ) if is_detail else _follow_status_keyboard(job)

    last_published = str(job.get("status_last_text") or "")
    if not force and text == last_published:
        return

    if is_detail:
        image_path: Path | None = None
        try:
            image_path = _build_graph_image(match_id, data)
        except Exception:
            image_path = None

        if image_path is not None and image_path.exists():
            prev_message_id = job.get("status_message_id")
            sent = None
            try:
                with image_path.open("rb") as f:
                    sent = await app.bot.send_photo(
                        chat_id=chat_id,
                        photo=f,
                        caption=text,
                        reply_markup=keyboard,
                    )
            except Exception:
                sent = None
            finally:
                try:
                    image_path.unlink()
                except OSError:
                    pass

            if sent is not None:
                if isinstance(prev_message_id, int) and prev_message_id != sent.message_id:
                    try:
                        await app.bot.delete_message(chat_id=chat_id, message_id=prev_message_id)
                    except Exception:
                        pass
                job["status_message_id"] = sent.message_id
                job["status_last_text"] = text
                return

    message_id = job.get("status_message_id")
    if isinstance(message_id, int):
        edited = await _safe_edit_message(
            app,
            chat_id,
            message_id,
            text,
            reply_markup=keyboard,
        )
        if edited:
            job["status_last_text"] = text
            return

    try:
        sent = await app.bot.send_message(
            chat_id=chat_id,
            text=text,
            reply_markup=keyboard,
        )
    except Exception:
        return

    job["status_message_id"] = sent.message_id
    job["status_last_text"] = text


async def _run_follow_match_job(app: Application, chat_id: int) -> None:
    job = FOLLOW_JOBS.get(chat_id)
    if not isinstance(job, dict):
        return

    match_id = str(job.get("match_id") or "")
    interval = int(job.get("interval_seconds") or FOLLOW_REFRESH_SECONDS)
    stale_limit = int(job.get("stale_cycles_limit") or FOLLOW_STALE_CYCLES_LIMIT)

    job["phase"] = "running"
    job["started_utc"] = datetime.utcnow().isoformat(timespec="seconds")
    job["cycles"] = 0
    job["unchanged_cycles"] = 0
    job["stop_reason"] = ""
    last_fp: tuple | None = None
    consecutive_errors = 0

    try:
        while True:
            current = FOLLOW_JOBS.get(chat_id)
            if current is not job:
                return

            if bool(job.get("stop_requested")):
                job["phase"] = "stopped"
                job["stop_reason"] = "manual"
                break

            data = None
            try:
                data = await asyncio.to_thread(_refresh_match_data, match_id)
                if data is None:
                    data = _get_match_detail(match_id)
                if data is None:
                    raise RuntimeError("match no encontrado al refrescar")

                await asyncio.to_thread(_get_or_compute_predictions, match_id, data)
                consecutive_errors = 0
            except Exception as exc:
                consecutive_errors += 1
                job["last_error"] = str(exc)
                if consecutive_errors >= 5:
                    job["phase"] = "stopped"
                    job["stop_reason"] = "errores_consecutivos"
                    break
                await asyncio.sleep(interval)
                continue

            m = data.get("match", {}) if isinstance(data, dict) else {}
            s = data.get("score", {}) if isinstance(data, dict) else {}
            graph_points = data.get("graph_points", []) if isinstance(data, dict) else []
            minute_est = None
            if graph_points:
                try:
                    minute_est = int(graph_points[-1].get("minute", 0))
                except (TypeError, ValueError):
                    minute_est = None

            status_type = str(m.get("status_type", "") or "").lower()
            status_desc = str(m.get("status_description", "") or "")
            home_score = _safe_int(s.get("home"))
            away_score = _safe_int(s.get("away"))
            score_txt = f"{home_score if home_score is not None else '-'}-{away_score if away_score is not None else '-'}"

            fp = (
                status_type,
                status_desc,
                home_score,
                away_score,
                minute_est,
                len(graph_points),
            )
            if fp == last_fp:
                job["unchanged_cycles"] = int(job.get("unchanged_cycles") or 0) + 1
            else:
                job["unchanged_cycles"] = 0
                last_fp = fp
                job["last_change_utc"] = datetime.utcnow().isoformat(timespec="seconds")

            job["cycles"] = int(job.get("cycles") or 0) + 1
            job["last_tick_utc"] = datetime.utcnow().isoformat(timespec="seconds")
            job["last_status"] = status_type or status_desc or "live"
            job["last_score"] = score_txt
            job["minute_est"] = minute_est
            job["follow_data"] = data
            pred_row = _get_or_compute_predictions(match_id, data)
            job["follow_pred"] = pred_row
            await _publish_follow_status(app, chat_id, job)

            if status_type == "finished":
                job["phase"] = "done"
                job["stop_reason"] = "match_finished"
                break

            if int(job.get("unchanged_cycles") or 0) >= stale_limit:
                job["phase"] = "stopped"
                job["stop_reason"] = "sin_cambios"
                break

            await asyncio.sleep(interval)
    finally:
        final_job = FOLLOW_JOBS.get(chat_id)
        if final_job is job:
            await _publish_follow_status(app, chat_id, job, force=True)
            FOLLOW_JOBS.pop(chat_id, None)


async def _run_refresh_date_job(
    app: Application,
    chat_id: int,
    progress_message_id: int,
    event_date: str,
) -> None:
    progress_state: dict[str, object] = {
        "phase": "starting",
        "total_matches": 0,
        "pending_total": 0,
        "processed": 0,
        "updated_ok": 0,
        "still_incomplete": 0,
        "failed": 0,
        "skipped_complete": 0,
        "started_monotonic": time.monotonic(),
    }
    REFRESH_JOBS[chat_id] = {
        "mode": "date",
        "event_date": event_date,
        "message_id": progress_message_id,
        "progress_state": progress_state,
        "suppress_edits": False,
        "force_next_edit": False,
    }

    task = asyncio.create_task(
        asyncio.to_thread(
            _refresh_incomplete_matches_for_date,
            event_date,
            progress_state,
        )
    )

    last_text = ""
    try:
        while not task.done():
            text = _refresh_date_progress_text(event_date, progress_state)
            if text != last_text:
                job = REFRESH_JOBS.get(chat_id, {})
                suppress_edits = bool(job.get("suppress_edits"))
                force_next_edit = bool(job.get("force_next_edit"))
                if force_next_edit and isinstance(job, dict):
                    job["force_next_edit"] = False

                if not suppress_edits or force_next_edit:
                    await _safe_edit_message(
                        app,
                        chat_id=chat_id,
                        message_id=progress_message_id,
                        text=text,
                        reply_markup=_refresh_job_progress_keyboard(),
                    )
                last_text = text

            await asyncio.sleep(REFRESH_DATE_STATUS_INTERVAL_SECONDS)

        result = await task
    except Exception as exc:
        await _safe_edit_message(
            app,
            chat_id=chat_id,
            message_id=progress_message_id,
            text=f"Error refrescando pendientes del dia: {exc}",
            reply_markup=_refresh_job_progress_keyboard(),
        )
        REFRESH_JOBS.pop(chat_id, None)
        return

    rows = _fetch_matches_for_date(event_date)
    pred_map = _fetch_date_pred_outcomes(event_date)
    stats = _pred_stats_text(pred_map, len(rows), match_rows=rows)
    summary = (
        f"Descarga pendientes completada para {event_date}:\n"
        f"Total: {result['total']} | Actualizados: {result['updated_ok']} | "
        f"Aun incompletos: {result['still_incomplete']} | "
        f"Ya completos: {result['skipped_complete']} | Fallidos: {result['failed']}"
    )
    header = f"{summary}\n\nMatches del {event_date}:"
    if stats:
        header += f"\n{stats}"
    await _safe_edit_message(
        app,
        chat_id=chat_id,
        message_id=progress_message_id,
        text=header,
        reply_markup=_matches_keyboard(rows, event_date, 0, pred_map),
    )
    REFRESH_JOBS.pop(chat_id, None)


async def _run_reconcile_job(
    app: Application,
    chat_id: int,
    progress_message_id: int,
) -> None:
    """Background job: find pending BET outcomes, scrape matches, then reconcile."""
    _back_kb = InlineKeyboardMarkup(
        [[InlineKeyboardButton("Menu principal", callback_data="nav:main")]]
    )
    REFRESH_JOBS[chat_id] = {"mode": "reconcile", "message_id": progress_message_id}

    _BET_SIGS = {"BET", "BET HOME", "BET_HOME", "BET AWAY", "BET_AWAY"}

    def _find_pending_match_ids() -> list[str]:
        conn2 = _open_conn()
        try:
            table_cols = [
                row["name"]
                for row in conn2.execute("PRAGMA table_info(eval_match_results)").fetchall()
            ]
            tags = sorted({
                col.split("__", 1)[1]
                for col in table_cols
                if col.startswith("q3_pick__")
            })
            pending_ids: set[str] = set()
            for tag in tags:
                for quarter in ("q3", "q4"):
                    avail_col = f'"{quarter}_available__{tag}"'
                    sig_col   = f'"{quarter}_signal__{tag}"'
                    out_col   = f'"{quarter}_outcome__{tag}"'
                    try:
                        rows2 = conn2.execute(
                            f"SELECT match_id, {sig_col} AS sig FROM eval_match_results"
                            f" WHERE {avail_col} = 1 AND {out_col} = 'pending'"
                        ).fetchall()
                        for r2 in rows2:
                            sig_norm = str(r2["sig"] or "").upper().replace("_", " ")
                            if sig_norm in _BET_SIGS:
                                pending_ids.add(str(r2["match_id"]))
                    except Exception:
                        pass
            return list(pending_ids)
        finally:
            conn2.close()

    try:
        pending_ids = await asyncio.to_thread(_find_pending_match_ids)
    except Exception as exc:
        await _safe_edit_message(app, chat_id, progress_message_id,
                                 f"❌ Error buscando pendientes: {exc}", _back_kb)
        REFRESH_JOBS.pop(chat_id, None)
        return

    if not pending_ids:
        resolved = await asyncio.to_thread(bet_monitor_mod.reconcile_pending_results, DB_PATH)
        await _safe_edit_message(
            app, chat_id, progress_message_id,
            f"✅ Sin apuestas pendientes de marcador.\nbet_monitor_log: {resolved} fila(s) actualizadas.",
            _back_kb,
        )
        REFRESH_JOBS.pop(chat_id, None)
        return

    total = len(pending_ids)
    progress: dict[str, object] = {"done": 0, "ok": 0, "fail": 0, "total": total}

    await _safe_edit_message(
        app, chat_id, progress_message_id,
        f"⏳ Descargando {total} partido(s) pendiente(s)... (0/{total})",
    )

    def _scrape_all() -> None:
        logger.info("[RECONCILE] Iniciando descarga de %d partidos pendientes", total)
        conn3 = _open_conn()
        for mid in pending_ids:
            done_n = int(progress["done"]) + 1
            logger.info("[RECONCILE] (%d/%d) Descargando match_id=%s ...", done_n, total, mid)
            try:
                fresh = scraper_mod.fetch_match_by_id(mid)
                db_mod.save_match(conn3, mid, fresh)
                progress["ok"] = int(progress["ok"]) + 1
                logger.info("[RECONCILE] (%d/%d) match_id=%s OK", done_n, total, mid)
            except Exception as exc2:
                logger.warning("[RECONCILE] (%d/%d) match_id=%s FALLIDO: %s", done_n, total, mid, exc2)
                progress["fail"] = int(progress["fail"]) + 1
            progress["done"] = done_n
        conn3.close()
        logger.info(
            "[RECONCILE] Descarga completa: %d OK, %d fallidos",
            int(progress["ok"]), int(progress["fail"]),
        )

    task = asyncio.create_task(asyncio.to_thread(_scrape_all))
    last_text = ""
    while not task.done():
        done_n = int(progress["done"])
        text = f"⏳ Descargando partido(s) pendiente(s)... ({done_n}/{total})"
        if text != last_text:
            await _safe_edit_message(app, chat_id, progress_message_id, text)
            last_text = text
        await asyncio.sleep(REFRESH_DATE_STATUS_INTERVAL_SECONDS)

    try:
        await task
    except Exception as exc:
        await _safe_edit_message(app, chat_id, progress_message_id,
                                 f"❌ Error durante descarga: {exc}", _back_kb)
        REFRESH_JOBS.pop(chat_id, None)
        return

    resolved = await asyncio.to_thread(bet_monitor_mod.reconcile_pending_results, DB_PATH)
    logger.info("[RECONCILE] Resultados actualizados en DB: %d", resolved)
    lines = [
        "✅ Proceso completado.",
        f"Partidos buscados: {total}",
        f"Descargados: {int(progress['ok'])} ✅  |  Fallidos: {int(progress['fail'])} ❌",
        f"Resultados actualizados: {resolved}",
    ]
    await _safe_edit_message(app, chat_id, progress_message_id, "\n".join(lines), _back_kb)
    REFRESH_JOBS.pop(chat_id, None)


async def _run_refresh_7d_job(
    app: Application,
    chat_id: int,
    progress_message_id: int,
) -> None:
    progress_state: dict[str, object] = {
        "phase7d": "starting",
        "total_dates": 0,
        "processed_dates": 0,
        "current_date": "",
        "sum_total": 0,
        "sum_updated_ok": 0,
        "sum_still_incomplete": 0,
        "sum_skipped_complete": 0,
        "sum_failed": 0,
        "started_monotonic": time.monotonic(),
    }
    REFRESH_JOBS[chat_id] = {
        "mode": "7d",
        "event_date": "",
        "message_id": progress_message_id,
        "progress_state": progress_state,
        "suppress_edits": False,
        "force_next_edit": False,
    }

    task = asyncio.create_task(
        asyncio.to_thread(
            _refresh_recent_dates,
            7,
            progress_state,
        )
    )

    last_text = ""
    try:
        while not task.done():
            text = _refresh_7d_progress_text(progress_state)
            if text != last_text:
                job = REFRESH_JOBS.get(chat_id, {})
                suppress_edits = bool(job.get("suppress_edits"))
                force_next_edit = bool(job.get("force_next_edit"))
                if force_next_edit and isinstance(job, dict):
                    job["force_next_edit"] = False

                if not suppress_edits or force_next_edit:
                    await _safe_edit_message(
                        app,
                        chat_id=chat_id,
                        message_id=progress_message_id,
                        text=text,
                        reply_markup=_refresh_job_progress_keyboard(),
                    )
                last_text = text

            await asyncio.sleep(REFRESH_DATE_STATUS_INTERVAL_SECONDS)

        result7d = await task
    except Exception as exc:
        await _safe_edit_message(
            app,
            chat_id=chat_id,
            message_id=progress_message_id,
            text=f"Error refrescando pendientes de 7 dias: {exc}",
            reply_markup=_refresh_job_progress_keyboard(),
        )
        REFRESH_JOBS.pop(chat_id, None)
        return

    dates = result7d.get("dates") or []
    dates_text = ", ".join(str(d) for d in dates) if dates else "(sin fechas)"
    summary7d = (
        "Descarga pendientes (ultimos 7 dias) completada:\n"
        f"Fechas: {len(dates)}\n"
        f"Total: {int(result7d.get('sum_total') or 0)} | "
        f"Actualizados: {int(result7d.get('sum_updated_ok') or 0)} | "
        f"Aun incompletos: {int(result7d.get('sum_still_incomplete') or 0)} | "
        f"Ya completos: {int(result7d.get('sum_skipped_complete') or 0)} | "
        f"Fallidos: {int(result7d.get('sum_failed') or 0)}\n"
        f"Rango: {dates_text}"
    )
    await _safe_edit_message(
        app,
        chat_id=chat_id,
        message_id=progress_message_id,
        text=summary7d,
        reply_markup=_refresh_job_progress_keyboard(),
    )
    REFRESH_JOBS.pop(chat_id, None)


async def _run_date_ingest_job(
    app: Application,
    chat_id: int,
    progress_message_id: int,
    event_date: str,
    limit: int | None,
) -> None:
    progress_state: dict[str, object] = {
        "phase": "starting",
        "processed": 0,
        "total": 0,
        "ingested_ok": 0,
        "ingested_fail": 0,
        "skipped_ft": 0,
        "discovered_rows": 0,
        "started_monotonic": time.monotonic(),
    }
    cancel_event = threading.Event()
    DATE_INGEST_JOBS[chat_id] = {
        "cancel_event": cancel_event,
        "message_id": progress_message_id,
        "event_date": event_date,
        "limit": limit,
        "progress_state": progress_state,
        "suppress_edits": False,
        "force_next_edit": False,
    }
    task = asyncio.create_task(
        asyncio.to_thread(
            _ingest_new_date_with_progress,
            event_date,
            limit,
            progress_state,
            cancel_event,
        )
    )

    last_text = _date_ingest_progress_text(event_date, limit, progress_state)
    print("[date-ingest] background job started", flush=True)
    try:
        while not task.done():
            text = _date_ingest_progress_text(event_date, limit, progress_state)
            if text != last_text:
                print(text.replace("\n", " | "), flush=True)
                job = DATE_INGEST_JOBS.get(chat_id, {})
                suppress_edits = bool(job.get("suppress_edits"))
                force_next_edit = bool(job.get("force_next_edit"))
                if force_next_edit and isinstance(job, dict):
                    job["force_next_edit"] = False

                if not suppress_edits or force_next_edit:
                    await _safe_edit_message(
                        app,
                        chat_id=chat_id,
                        message_id=progress_message_id,
                        text=text,
                        reply_markup=_date_ingest_progress_keyboard(True),
                    )
                last_text = text
            await asyncio.sleep(DATE_INGEST_STATUS_INTERVAL_SECONDS)

        result = await task
    except Exception as exc:
        print(f"[date-ingest] background job error: {exc}", flush=True)
        await _safe_edit_message(
            app,
            chat_id=chat_id,
            message_id=progress_message_id,
            text=f"No pude traer fecha {event_date}: {exc}",
            reply_markup=_date_ingest_progress_keyboard(False),
        )
        DATE_INGEST_JOBS.pop(chat_id, None)
        return

    progress_state.update(result)
    final_text = _date_ingest_progress_text(event_date, limit, progress_state)
    print(final_text.replace("\n", " | "), flush=True)
    await _safe_edit_message(
        app,
        chat_id=chat_id,
        message_id=progress_message_id,
        text=final_text,
        reply_markup=_date_ingest_progress_keyboard(False),
    )
    print("[date-ingest] background job finished", flush=True)
    DATE_INGEST_JOBS.pop(chat_id, None)


def _main_menu_keyboard(chat_id: int | None = None) -> InlineKeyboardMarkup:
    rows = [
        [InlineKeyboardButton("🏀 Matches en vivo", callback_data="menu:live")],
        [InlineKeyboardButton("📅 Matches por fecha", callback_data="menu:dates")],
        [InlineKeyboardButton("🔍 Match por ID", callback_data="menu:id")],
        [InlineKeyboardButton("🧠 Re entrenar base", callback_data="menu:train")],
        [InlineKeyboardButton("⬇️ Traer fecha nueva", callback_data="menu:fetchdate")],
        [
            InlineKeyboardButton(
                "🔄 Descargar pendientes 7 dias",
                callback_data="menu:refresh7d",
            )
        ],
        [
            InlineKeyboardButton(
                f"🤖 Modelos (Q3={MODEL_CONFIG['q3']} Q4={MODEL_CONFIG['q4']})",
                callback_data="menu:models",
            )
        ],
        [
            InlineKeyboardButton(
                "👁 Monitor Apuestas" + (" 🟢" if (
                    bet_monitor_mod.MONITOR_STATUS.get("running")
                    or (_MONITOR_THREAD is not None and _MONITOR_THREAD.is_alive())
                ) else ""),
                callback_data="menu:monitor",
            )
        ],
        [InlineKeyboardButton("📊 Señales de Hoy", callback_data="monitor:signals_today")],
        [InlineKeyboardButton("📆 Reporte del Mes", callback_data="menu:monthly_report")],
        [InlineKeyboardButton("🔁 Resolver marcadores pendientes", callback_data="menu:reconcile")],
    ]
    if chat_id is not None and chat_id in DATE_INGEST_JOBS:
        rows.append(
            [
                InlineKeyboardButton(
                    "⏳ Status descarga",
                    callback_data="menu:fetchdate:status",
                )
            ]
        )
    if chat_id is not None and chat_id in REFRESH_JOBS:
        rows.append(
            [
                InlineKeyboardButton(
                    "⏳ Status refresh pendientes",
                    callback_data="menu:refresh:status",
                )
            ]
        )
    if chat_id is not None and chat_id in FOLLOW_JOBS:
        rows.append(
            [
                InlineKeyboardButton(
                    "⏳ Status seguimiento",
                    callback_data="menu:follow:status",
                )
            ]
        )
    return InlineKeyboardMarkup(rows)


def _live_minute_text(row: dict) -> str:
    """Show only the most relevant quarter with betting status."""
    def _fmt_mmss(total_seconds: int) -> str:
        minutes, seconds = divmod(max(total_seconds, 0), 60)
        return f"{minutes}:{seconds:02d}"

    def _bet_status(match_id: str, quarter: str) -> tuple[str, str]:
        """Returns (emoji, text) for quarter betting status."""
        if not match_id:
            return ("", "")
        try:
            pred_row = _read_prediction_row(match_id)
            pred = (pred_row or {}).get(quarter, {}) if isinstance(pred_row, dict) else {}
            if not isinstance(pred, dict):
                return ("⏳", "esperando")
            if not pred.get("available"):
                reason = str(pred.get("reason") or pred.get("reasoning") or "").lower()
                if "missing" in reason or "esperando" in reason:
                    return ("⏳", "esperando datos")
                return ("⏸️", "no disponible")
            signal = str(pred.get("signal") or "").strip().upper().replace("_", " ")
            if signal in {"BET", "APUESTA"}:
                return ("🟢", "APOSTAR")
            elif signal in {"LEAN", "NO BET", "NO_BET"}:
                return ("🔴", "NO apostar")
            return ("⏳", "analizando")
        except Exception:
            return ("⏳", "?")
        return ("⏳", "?")

    played_seconds = _safe_int(row.get("played_seconds"))
    status_desc = str(row.get("status_description", "") or "")
    match_id = str(row.get("match_id", "") or "")

    if played_seconds is None:
        return f"⏳ {status_desc or 'live'}"

    # Format current period
    def _period_short(status_desc: str) -> str:
        value = status_desc.strip().lower()
        mapping = {
            "1st quarter": "Q1",
            "2nd quarter": "Q2",
            "3rd quarter": "Q3",
            "4th quarter": "Q4",
            "halftime": "HT",
        }
        return mapping.get(value, status_desc or "live")

    period = _period_short(status_desc)
    minutes, seconds = divmod(played_seconds, 60)
    time_display = f"{minutes}:{seconds:02d}"

    # Determine which quarter to show
    q3_diff = 24 * 60 - played_seconds  # seconds until Q3
    q4_diff = 36 * 60 - played_seconds  # seconds until Q4

    # Decide which quarter is most relevant
    if played_seconds < 24 * 60:  # Before Q3
        # Show time until Q3 with betting status
        q3_emoji, q3_txt = _bet_status(match_id, "q3")
        if q3_diff > 0:
            q3_minutes = q3_diff // 60
            return f"🕐 Q3 en {q3_minutes}min {q3_emoji} | {period} {time_display}"
        else:
            return f"✅ Q3 ahora {q3_emoji} | {period} {time_display}"
    elif played_seconds < 36 * 60:  # Between Q3 and Q4
        # Show time until Q4 with betting status
        q4_emoji, q4_txt = _bet_status(match_id, "q4")
        if q4_diff > 0:
            q4_minutes = q4_diff // 60
            return f"🕐 Q4 en {q4_minutes}min {q4_emoji} | {period} {time_display}"
        else:
            return f"✅ Q4 ahora {q4_emoji} | {period} {time_display}"
    else:  # After Q4 started
        q3_emoji, q3_txt = _bet_status(match_id, "q3")
        q4_emoji, q4_txt = _bet_status(match_id, "q4")
        return f"Q3 {q3_emoji} Q4 {q4_emoji} | {period} {time_display}"


def _live_sort_key(row: dict) -> tuple[int, int, int]:
    played_seconds = _safe_int(row.get("played_seconds"))
    if played_seconds is None:
        return (3, 10**9, 10**9)

    q3_diff = 24 * 60 - played_seconds
    q4_diff = 36 * 60 - played_seconds

    if q3_diff > 0:
        return (0, q3_diff, played_seconds)
    if q4_diff > 0:
        return (1, q4_diff, played_seconds)
    return (2, played_seconds, played_seconds)


def _live_summary_text(result: dict) -> str:
    live_count = len(result.get('live_rows', []))
    if live_count == 0:
        return "🏀 No hay partidos en vivo ahora"

    # Count betting opportunities
    bettable = 0
    waiting = 0
    for row in result.get('live_rows', []):
        match_id = str(row.get("match_id", "") or "")
        played_seconds = _safe_int(row.get("played_seconds"))
        if played_seconds is None:
            continue

        if played_seconds < 24 * 60:
            pred = (_read_prediction_row(match_id) or {}).get("q3", {})
        elif played_seconds < 36 * 60:
            pred = (_read_prediction_row(match_id) or {}).get("q4", {})
        else:
            pred = {}

        if isinstance(pred, dict) and pred.get("available"):
            signal = str(pred.get("signal") or "").upper()
            if signal in {"BET", "APUESTA"}:
                bettable += 1
            else:
                waiting += 1
        else:
            waiting += 1

    parts = [f"🏀 Partidos en vivo: {live_count}"]
    if bettable > 0:
        parts.append(f"🟢 Apostables: {bettable}")
    if waiting > 0:
        parts.append(f"⏳ En espera: {waiting}")

    return "\n".join(parts)


def _live_keyboard(result: dict, page: int) -> InlineKeyboardMarkup:
    live_rows = sorted(
        result.get("live_rows", []),
        key=_live_sort_key,
    )
    total_pages = max((len(live_rows) - 1) // LIVE_PAGE_SIZE + 1, 1)
    page = max(0, min(page, total_pages - 1))

    start = page * LIVE_PAGE_SIZE
    end = start + LIVE_PAGE_SIZE
    page_rows = live_rows[start:end]

    keyboard: list[list[InlineKeyboardButton]] = []
    for row in page_rows:
        match_id = str(row.get("match_id", "") or "")
        if not match_id:
            continue
        league = str(row.get("league", "") or "-")[:15]
        home = _abbr_team(str(row.get("home_team", "")), max_len=12)
        away = _abbr_team(str(row.get("away_team", "")), max_len=12)
        minute = _live_minute_text(row)

        # Add score
        home_score = row.get("home_score")
        away_score = row.get("away_score")
        score_txt = ""
        if home_score is not None and away_score is not None:
            score_txt = f" | {home_score}-{away_score}"

        label = f"{home} vs {away}{score_txt}\n{league} | {minute}"
        keyboard.append(
            [
                InlineKeyboardButton(
                    label,
                    callback_data=f"livematch:{match_id}:{page}",
                )
            ]
        )

    nav_row: list[InlineKeyboardButton] = []
    if page > 0:
        nav_row.append(
            InlineKeyboardButton("<<", callback_data=f"livepage:{page - 1}")
        )
    nav_row.append(
        InlineKeyboardButton(
            f"Pag {page + 1}/{total_pages}",
            callback_data="noop",
        )
    )
    if page < total_pages - 1:
        nav_row.append(
            InlineKeyboardButton(">>", callback_data=f"livepage:{page + 1}")
        )
    keyboard.append(nav_row)
    keyboard.append([InlineKeyboardButton("Refrescar en vivo", callback_data="menu:live")])
    keyboard.append([InlineKeyboardButton("Menu principal", callback_data="nav:main")])
    return InlineKeyboardMarkup(keyboard)


def _ingest_live_matches() -> dict:
    print("[live] consultando lista live de SofaScore...", flush=True)
    live_rows = scraper_mod.fetch_live_match_ids()
    total = len(live_rows)
    print(f"[live] encontrados={total}", flush=True)

    match_ids = [str(row.get("match_id", "") or "") for row in live_rows]
    match_ids = [mid for mid in match_ids if mid]

    conn = _open_conn()

    ing_ok = 0
    ing_fail = 0
    results = scraper_mod.fetch_matches_by_ids(match_ids)
    for index, (match_id, data, error) in enumerate(results, start=1):
        print(
            f"[live] {index}/{total} match_id={match_id}...",
            flush=True,
        )
        if data is None:
            ing_fail += 1
            print(
                f"[live] {index}/{total} fail match_id={match_id} error={error}",
                flush=True,
            )
            continue
        if _is_two_half_game(data):
            print(
                f"[live] {index}/{total} skip 2-half match_id={match_id}",
                flush=True,
            )
            continue
        try:
            db_mod.save_match(conn, match_id, data)
            ing_ok += 1
            print(
                f"[live] {index}/{total} ok match_id={match_id}",
                flush=True,
            )
        except Exception as exc:
            ing_fail += 1
            print(
                f"[live] {index}/{total} fail-save match_id={match_id} error={exc}",
                flush=True,
            )
    conn.close()

    print(
        f"[live] terminado ok={ing_ok} fail={ing_fail}",
        flush=True,
    )

    return {
        "live_rows": live_rows,
        "ingested_ok": ing_ok,
        "ingested_fail": ing_fail,
    }


async def _render_live(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    page: int = 0,
    *,
    refetch: bool = False,
) -> None:
    if refetch or LIVE_RESULT_KEY not in context.user_data:
        await _set_waiting_state(update, text="Espere... trayendo live")
        try:
            result = await asyncio.to_thread(_ingest_live_matches)
        except Exception as exc:
            await _replace_callback_message(
                update,
                text=f"No pude traer matches live: {exc}",
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("Menu principal", callback_data="nav:main")]]
                ),
            )
            return
        context.user_data[LIVE_RESULT_KEY] = result
    else:
        result = context.user_data[LIVE_RESULT_KEY]

    await _replace_callback_message(
        update,
        text=_live_summary_text(result),
        reply_markup=_live_keyboard(result, page),
    )


def _follow_toggle_button(
    chat_id: int | None,
    match_id: str,
    event_token: str,
    page: int,
    match_data: dict | None = None,
) -> InlineKeyboardButton | None:
    status_type = str(
        (match_data or {}).get("match", {}).get("status_type", "") or ""
    ).lower()
    if status_type == "finished":
        return None

    if chat_id is not None:
        job = FOLLOW_JOBS.get(chat_id)
        if isinstance(job, dict):
            followed_match_id = str(job.get("match_id") or "")
            phase = str(job.get("phase") or "")
            stop_requested = bool(job.get("stop_requested"))
            if (
                followed_match_id == str(match_id)
                and phase in {"starting", "running"}
                and not stop_requested
            ):
                return InlineKeyboardButton(
                    "Detener seguimiento",
                    callback_data="follow:stop",
                )

    return InlineKeyboardButton(
        "Iniciar seguimiento",
        callback_data=f"follow:start:{match_id}:{event_token}:{page}",
    )


def _live_detail_keyboard(
    match_id: str,
    page: int,
    match_data: dict | None = None,
    chat_id: int | None = None,
) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    rows.append(
        [InlineKeyboardButton("Ver SofaScore", url=_sofascore_match_url(match_id, match_data))]
    )
    rows.append(
        [
            InlineKeyboardButton(
                "Refresh datos + pred",
                callback_data=f"refreshlive:all:{match_id}:{page}",
            ),
            InlineKeyboardButton(
                "Refresh pred",
                callback_data=f"refreshlive:pred:{match_id}:{page}",
            ),
        ]
    )
    _ftb = _follow_toggle_button(chat_id, match_id, "_", page, match_data)
    if _ftb is not None:
        rows.append([_ftb])
    rows.append(
        [
            InlineKeyboardButton(
                "Borrar match de la base",
                callback_data=f"dellive:confirm:{match_id}:{page}",
            )
        ]
    )
    rows.append([InlineKeyboardButton(
        f"Modelos (Q3={MODEL_CONFIG['q3']} Q4={MODEL_CONFIG['q4']})",
        callback_data=f"matchmodel:open:{match_id}:_:{page}",
    )])
    rows.append([InlineKeyboardButton("Volver a live", callback_data=f"livepage:{page}")])
    rows.append([InlineKeyboardButton("Menu principal", callback_data="nav:main")])
    return InlineKeyboardMarkup(rows)


async def _send_live_detail_message(
    update: Update,
    app: Application,
    match_id: str,
    data: dict,
    pred_row: dict | None,
    page: int,
) -> None:
    chat_id = update.effective_chat.id if update.effective_chat else None
    if chat_id is None:
        return

    detail_text = _detail_text(match_id, data, pred_row)
    keyboard = _live_detail_keyboard(
        match_id,
        page,
        match_data=data,
        chat_id=chat_id,
    )

    image_path: Path | None = None
    try:
        image_path = _build_graph_image(match_id, data)
    except Exception:
        image_path = None

    if image_path is not None and image_path.exists():
        try:
            with image_path.open("rb") as f:
                await app.bot.send_photo(
                    chat_id=chat_id,
                    photo=f,
                    caption=detail_text,
                    reply_markup=keyboard,
                )
        finally:
            try:
                image_path.unlink()
            except OSError:
                pass
        return

    await app.bot.send_message(
        chat_id=chat_id,
        text=detail_text,
        reply_markup=keyboard,
    )


async def _render_live_detail(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    match_id: str,
    page: int,
    *,
    pred_row: dict | None = None,
) -> None:
    await _set_waiting_state(update)
    try:
        data = _get_match_detail(match_id)
        if data is None:
            data = await asyncio.to_thread(_refresh_match_data, match_id)
    except Exception as exc:
        await _replace_callback_message(
            update,
            text=f"No se pudo abrir match live {match_id}: {exc}",
            reply_markup=InlineKeyboardMarkup(
                [[InlineKeyboardButton("Volver a live", callback_data=f"livepage:{page}")]]
            ),
        )
        return

    if pred_row is None:
        pred_row = _get_or_compute_predictions(match_id, data)
    await _send_live_detail_message(
        update,
        context.application,
        match_id,
        data,
        pred_row,
        page,
    )
    if update.callback_query and update.callback_query.message:
        try:
            await update.callback_query.message.delete()
        except Exception:
            pass


async def _refresh_live_detail(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    mode: str,
    match_id: str,
    page: int,
) -> None:
    chat_id = update.effective_chat.id if update.effective_chat else None
    cached = _get_match_detail(match_id)
    await _set_waiting_state(update, text=_refresh_waiting_text(match_id, cached))

    if mode == "all":
        try:
            data = await asyncio.to_thread(_refresh_match_data, match_id)
        except Exception as exc:
            await _replace_callback_message(
                update,
                text=f"No pude refrescar datos de {match_id}: {exc}",
                reply_markup=_live_detail_keyboard(match_id, page, chat_id=chat_id),
            )
            return
        pred_row = _refresh_predictions(match_id, data)
        await _render_live_detail(
            update,
            context,
            match_id,
            page,
            pred_row=pred_row,
        )
        return

    data = _get_match_detail(match_id)
    if not data:
        await _replace_callback_message(
            update,
            text=f"No se encontro el match {match_id}.",
            reply_markup=_live_detail_keyboard(match_id, page, chat_id=chat_id),
        )
        return

    pred_row = _refresh_predictions(match_id, data)
    await _render_live_detail(
        update,
        context,
        match_id,
        page,
        pred_row=pred_row,
    )


def _menu_reply_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [
            [MENU_BUTTON_TEXT, STATS_BUTTON_TEXT],
            [MONTHLY_BUTTON_TEXT, ID_BUTTON_TEXT],
        ],
        resize_keyboard=True,
        is_persistent=True,
    )


def _train_submenu_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "Entrenar (v2 + v4 + calibrar)",
                    callback_data="train:run",
                )
            ],
            [
                InlineKeyboardButton(
                    "Vaciar Resultados",
                    callback_data="train:clear:confirm",
                )
            ],
            [
                InlineKeyboardButton(
                    "Recalcular universo",
                    callback_data="train:recalc:confirm",
                )
            ],
            [InlineKeyboardButton("Stats Modelo", callback_data="train:stats")],
            [InlineKeyboardButton("Estado entrenamiento", callback_data="train:status")],
            [InlineKeyboardButton("Menu principal", callback_data="nav:main")],
        ]
    )


def _monitor_keyboard(running: bool, chat_id: int | None = None) -> InlineKeyboardMarkup:
    rows = []
    if running:
        rows.append([InlineKeyboardButton("⏹ Detener monitor", callback_data="monitor:stop")])
    else:
        rows.append([InlineKeyboardButton("▶️ Iniciar monitor", callback_data="monitor:start")])
    rows.append([InlineKeyboardButton("📋 Itinerario de hoy", callback_data="monitor:schedule")])
    rows.append([InlineKeyboardButton("📊 Señales de hoy", callback_data="monitor:signals_today")])
    rows.append([InlineKeyboardButton("⚙️ Config simulación", callback_data="monitor:betcfg")])
    rows.append([InlineKeyboardButton("📜 Bitácora reciente", callback_data="monitor:log")])
    q3_v = MONITOR_MODEL_CONFIG.get("q3", "v4")
    q4_v = MONITOR_MODEL_CONFIG.get("q4", "v4")
    rows.append([InlineKeyboardButton(
        f"🎯 Monitor modelos: Q3={q3_v}  Q4={q4_v}",
        callback_data="monitor:models",
    )])
    if chat_id is not None:
        pref = _MONITOR_SUBSCRIBERS.get(chat_id)
        if pref is not None:
            signal_type = pref.get("signal_type", "all")
            quarters = pref.get("quarters", ["q3", "q4"])
            bet_mark = "✅" if signal_type == "bet_only" else "☑️"
            all_mark = "✅" if signal_type == "all" else "☑️"
            rows.append([
                InlineKeyboardButton(f"{bet_mark} Solo apuestas", callback_data="monitor:sub_bet_only"),
                InlineKeyboardButton(f"{all_mark} Todo", callback_data="monitor:sub_all"),
            ])
            q3_mark = "✅" if "q3" in quarters else "☑️"
            q4_mark = "✅" if "q4" in quarters else "☑️"
            rows.append([
                InlineKeyboardButton(f"🔔 {q3_mark} Q3 {q4_mark} Q4", callback_data="monitor:quarters_menu"),
            ])
            rows.append([InlineKeyboardButton("🔕 Desuscribirme", callback_data="monitor:unsubscribe")])
        else:
            rows.append([InlineKeyboardButton("🔔 Suscribirme a alertas", callback_data="monitor:subscribe")])
    rows.append([InlineKeyboardButton("Menu principal", callback_data="nav:main")])
    return InlineKeyboardMarkup(rows)


def _monitor_quarters_keyboard(chat_id: int | None = None) -> InlineKeyboardMarkup:
    """Quarters selector (Q3/Q4 checkboxes) that returns to the monitor menu."""
    rows: list[list[InlineKeyboardButton]] = []
    
    if chat_id is not None:
        pref = _MONITOR_SUBSCRIBERS.get(chat_id)
        quarters = pref.get("quarters", ["q3", "q4"]) if pref else ["q3", "q4"]
    else:
        quarters = ["q3", "q4"]
    
    q3_mark = "✅" if "q3" in quarters else "☑️"
    q4_mark = "✅" if "q4" in quarters else "☑️"
    
    rows.append([
        InlineKeyboardButton(f"{q3_mark} Q3", callback_data="monitor:toggle_q3"),
        InlineKeyboardButton(f"{q4_mark} Q4", callback_data="monitor:toggle_q4"),
    ])
    rows.append([InlineKeyboardButton("⬅️ Monitor", callback_data="menu:monitor")])
    return InlineKeyboardMarkup(rows)


def _monitor_model_keyboard(chat_id: int | None = None) -> InlineKeyboardMarkup:
    """Model selector that returns to the monitor menu."""
    q3_v = MONITOR_MODEL_CONFIG.get("q3", "v4")
    q4_v = MONITOR_MODEL_CONFIG.get("q4", "v4")
    rows: list[list[InlineKeyboardButton]] = []
    half = len(AVAILABLE_MODELS) // 2
    rows.append([
        InlineKeyboardButton(
            f"Q3:{'✅' if q3_v == v else ''}{v}",
            callback_data=f"monmodel:set:q3:{_model_version_callback_token(v)}",
        )
        for v in AVAILABLE_MODELS[:half]
    ])
    rows.append([
        InlineKeyboardButton(
            f"Q3:{'✅' if q3_v == v else ''}{v}",
            callback_data=f"monmodel:set:q3:{_model_version_callback_token(v)}",
        )
        for v in AVAILABLE_MODELS[half:]
    ])
    rows.append([
        InlineKeyboardButton(
            f"Q4:{'✅' if q4_v == v else ''}{v}",
            callback_data=f"monmodel:set:q4:{_model_version_callback_token(v)}",
        )
        for v in AVAILABLE_MODELS[:half]
    ])
    rows.append([
        InlineKeyboardButton(
            f"Q4:{'✅' if q4_v == v else ''}{v}",
            callback_data=f"monmodel:set:q4:{_model_version_callback_token(v)}",
        )
        for v in AVAILABLE_MODELS[half:]
    ])
    rows.append([InlineKeyboardButton("⬅️ Monitor", callback_data="menu:monitor")])
    return InlineKeyboardMarkup(rows)


async def _monitor_notify_cb(
    msg: str,
    reply_markup: dict | None = None,
    notify_type: str = "bet",
    quarter: str | None = None,
) -> None:
    """Runs inside the monitor thread's own event loop.

    Sends directly to the Telegram HTTP API via httpx so it never touches
    the bot's asyncio connection and cannot block it.
    """
    import httpx
    if not BOT_TOKEN:
        return
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    if _MONITOR_SUBSCRIBERS:
        targets = list(_MONITOR_SUBSCRIBERS.items())
    else:
        targets = [(cid, "all") for cid in ALLOWED_CHAT_IDS]
    _MAX_RETRIES = 3
    _RETRY_DELAY = 2.0  # seconds between retries

    async with httpx.AsyncClient() as client:
        for cid, pref in targets:
            # Normalize pref: may be a dict or legacy string
            if isinstance(pref, dict):
                signal_type = pref.get("signal_type", "all")
                sub_quarters = [q.lower() for q in pref.get("quarters", ["q3", "q4"])]
            else:
                signal_type = str(pref)
                sub_quarters = ["q3", "q4"]

            # Filter by signal type: "bet_only" subscribers skip NO_BET messages
            if signal_type == "bet_only" and notify_type == "no_bet":
                continue

            # Filter by quarter: skip if subscriber does not have this quarter enabled
            if quarter and quarter.lower() not in sub_quarters:
                continue

            payload: dict = {"chat_id": cid, "text": msg}
            if reply_markup:
                payload["reply_markup"] = reply_markup

            sent = False
            last_exc = None
            for attempt in range(1, _MAX_RETRIES + 1):
                try:
                    resp = await client.post(url, json=payload, timeout=10.0)
                    resp.raise_for_status()
                    sent = True
                    break
                except Exception as exc:
                    last_exc = exc
                    if attempt < _MAX_RETRIES:
                        logger.warning(
                            "[MONITOR_NOTIFY] chat_id=%s intento %d/%d fallido: %s -- reintentando en %.0fs",
                            cid, attempt, _MAX_RETRIES, exc, _RETRY_DELAY,
                        )
                        import asyncio as _aio
                        await _aio.sleep(_RETRY_DELAY)

            if sent:
                logger.info(
                    "[MONITOR_NOTIFY] Enviado a chat_id=%s tipo=%s",
                    cid, notify_type,
                )
            else:
                logger.error(
                    "[MONITOR_NOTIFY] FALLO al enviar a chat_id=%s tipo=%s tras %d intentos: %s",
                    cid, notify_type, _MAX_RETRIES, last_exc,
                )

def _model_submenu_keyboard() -> InlineKeyboardMarkup:
    q3_v = MODEL_CONFIG.get("q3", "v4")
    q4_v = MODEL_CONFIG.get("q4", "v4")
    rows: list[list[InlineKeyboardButton]] = []
    half = len(AVAILABLE_MODELS) // 2

    # Q3 rows (first half and second half)
    rows.append([
        InlineKeyboardButton(
            f"Q3:{'✅' if q3_v == v else ''}{v}",
            callback_data=f"model:set:q3:{_model_version_callback_token(v)}",
        )
        for v in AVAILABLE_MODELS[:half]
    ])
    rows.append([
        InlineKeyboardButton(
            f"Q3:{'✅' if q3_v == v else ''}{v}",
            callback_data=f"model:set:q3:{_model_version_callback_token(v)}",
        )
        for v in AVAILABLE_MODELS[half:]
    ])

    # Q4 rows
    rows.append([
        InlineKeyboardButton(
            f"Q4:{'✅' if q4_v == v else ''}{v}",
            callback_data=f"model:set:q4:{_model_version_callback_token(v)}",
        )
        for v in AVAILABLE_MODELS[:half]
    ])
    rows.append([
        InlineKeyboardButton(
            f"Q4:{'✅' if q4_v == v else ''}{v}",
            callback_data=f"model:set:q4:{_model_version_callback_token(v)}",
        )
        for v in AVAILABLE_MODELS[half:]
    ])

    rows.append([InlineKeyboardButton("Menu principal", callback_data="nav:main")])
    return InlineKeyboardMarkup(rows)


def _match_model_submenu_keyboard(match_id: str, token: str, page: int) -> InlineKeyboardMarkup:
    """Model selector anchored to a specific match detail (for back-navigation)."""
    q3_v = MODEL_CONFIG.get("q3", "v4")
    q4_v = MODEL_CONFIG.get("q4", "v4")
    rows: list[list[InlineKeyboardButton]] = []
    half = len(AVAILABLE_MODELS) // 2

    rows.append([
        InlineKeyboardButton(
            f"Q3:{'✅' if q3_v == v else ''}{v}",
            callback_data=(
                f"matchmodel:set:q3:{_model_version_callback_token(v)}:"
                f"{match_id}:{token}:{page}"
            ),
        )
        for v in AVAILABLE_MODELS[:half]
    ])
    rows.append([
        InlineKeyboardButton(
            f"Q3:{'✅' if q3_v == v else ''}{v}",
            callback_data=(
                f"matchmodel:set:q3:{_model_version_callback_token(v)}:"
                f"{match_id}:{token}:{page}"
            ),
        )
        for v in AVAILABLE_MODELS[half:]
    ])
    rows.append([
        InlineKeyboardButton(
            f"Q4:{'✅' if q4_v == v else ''}{v}",
            callback_data=(
                f"matchmodel:set:q4:{_model_version_callback_token(v)}:"
                f"{match_id}:{token}:{page}"
            ),
        )
        for v in AVAILABLE_MODELS[:half]
    ])
    rows.append([
        InlineKeyboardButton(
            f"Q4:{'✅' if q4_v == v else ''}{v}",
            callback_data=(
                f"matchmodel:set:q4:{_model_version_callback_token(v)}:"
                f"{match_id}:{token}:{page}"
            ),
        )
        for v in AVAILABLE_MODELS[half:]
    ])

    rows.append([InlineKeyboardButton(
        "Regresar (recalcular)",
        callback_data=f"matchmodel:back:{match_id}:{token}:{page}",
    )])
    return InlineKeyboardMarkup(rows)


def _date_model_submenu_keyboard(event_date: str) -> InlineKeyboardMarkup:
    """Model selector anchored to the date view."""
    q3_v = MODEL_CONFIG.get("q3", "v4")
    q4_v = MODEL_CONFIG.get("q4", "v4")
    rows: list[list[InlineKeyboardButton]] = []
    half = len(AVAILABLE_MODELS) // 2

    rows.append([
        InlineKeyboardButton(
            f"Q3:{'✅' if q3_v == v else ''}{v}",
            callback_data=(
                f"datemodel:set:q3:{_model_version_callback_token(v)}:{event_date}"
            ),
        )
        for v in AVAILABLE_MODELS[:half]
    ])
    rows.append([
        InlineKeyboardButton(
            f"Q3:{'✅' if q3_v == v else ''}{v}",
            callback_data=(
                f"datemodel:set:q3:{_model_version_callback_token(v)}:{event_date}"
            ),
        )
        for v in AVAILABLE_MODELS[half:]
    ])
    rows.append([
        InlineKeyboardButton(
            f"Q4:{'✅' if q4_v == v else ''}{v}",
            callback_data=(
                f"datemodel:set:q4:{_model_version_callback_token(v)}:{event_date}"
            ),
        )
        for v in AVAILABLE_MODELS[:half]
    ])
    rows.append([
        InlineKeyboardButton(
            f"Q4:{'✅' if q4_v == v else ''}{v}",
            callback_data=(
                f"datemodel:set:q4:{_model_version_callback_token(v)}:{event_date}"
            ),
        )
        for v in AVAILABLE_MODELS[half:]
    ])

    rows.append([InlineKeyboardButton(
        "Regresar",
        callback_data=f"datemodel:back:{event_date}",
    )])
    return InlineKeyboardMarkup(rows)


def _train_recalc_confirm_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "Confirmar recalcular universo",
                    callback_data="train:recalc:run",
                )
            ],
            [InlineKeyboardButton("Volver", callback_data="menu:train")],
            [InlineKeyboardButton("Menu principal", callback_data="nav:main")],
        ]
    )


def _train_clear_confirm_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "Confirmar vaciar resultados",
                    callback_data="train:clear:run",
                )
            ],
            [InlineKeyboardButton("Volver", callback_data="menu:train")],
            [InlineKeyboardButton("Menu principal", callback_data="nav:main")],
        ]
    )


def _dates_keyboard(
    rows: list[dict],
    page: int,
    pred_stats: dict | None = None,
) -> InlineKeyboardMarkup:
    total_pages = max((len(rows) - 1) // DATE_PAGE_SIZE + 1, 1)
    page = max(0, min(page, total_pages - 1))

    start = page * DATE_PAGE_SIZE
    end = start + DATE_PAGE_SIZE
    page_rows = rows[start:end]

    keyboard: list[list[InlineKeyboardButton]] = []
    for row in page_rows:
        date_str = str(row.get("event_date", ""))
        short = _short_date(date_str)
        ps = (pred_stats or {}).get(date_str, {})
        q3_hit  = ps.get("q3_hit", 0)
        q3_miss = ps.get("q3_miss", 0)
        q3_push = ps.get("q3_push", 0)
        q4_hit  = ps.get("q4_hit", 0)
        q4_miss = ps.get("q4_miss", 0)
        q4_push = ps.get("q4_push", 0)
        if q3_hit or q3_miss or q3_push or q4_hit or q4_miss or q4_push:
            q3_str = f"Q3:{q3_hit}\u2705{q3_miss}\u274c"
            if q3_push:
                q3_str += f"{q3_push}\u2796"
            q4_str = f"Q4:{q4_hit}\u2705{q4_miss}\u274c"
            if q4_push:
                q4_str += f"{q4_push}\u2796"
            text = f"{short} | {q3_str} {q4_str}"
        else:
            total = int(row.get("total_matches", 0) or 0)
            text = f"{short} | M:{total}"
        keyboard.append(
            [
                InlineKeyboardButton(
                    text=text,
                    callback_data=f"date:{date_str}:0",
                )
            ]
        )

    nav_row: list[InlineKeyboardButton] = []
    if page > 0:
        nav_row.append(
            InlineKeyboardButton("<<", callback_data=f"dates:{page - 1}")
        )
    nav_row.append(
        InlineKeyboardButton(
            f"Pag {page + 1}/{total_pages}",
            callback_data="noop",
        )
    )
    if page < total_pages - 1:
        nav_row.append(
            InlineKeyboardButton(">>", callback_data=f"dates:{page + 1}")
        )
    keyboard.append(nav_row)

    keyboard.append([InlineKeyboardButton("Menu principal", callback_data="nav:main")])
    return InlineKeyboardMarkup(keyboard)


def _matches_keyboard(
    rows: list[dict],
    event_date: str,
    page: int,
    pred_map: dict | None = None,
) -> InlineKeyboardMarkup:
    total_pages = max((len(rows) - 1) // MATCH_PAGE_SIZE + 1, 1)
    page = max(0, min(page, total_pages - 1))

    start = page * MATCH_PAGE_SIZE
    end = start + MATCH_PAGE_SIZE
    page_rows = rows[start:end]

    keyboard: list[list[InlineKeyboardButton]] = []
    for row in page_rows:
        match_id = str(row.get("match_id", ""))
        home = _abbr_team(str(row.get("home_team", "")))
        away = _abbr_team(str(row.get("away_team", "")))
        tipoff = str(row.get("display_time", "") or "--:--")
        league = str(row.get("league", "") or "-")
        league_short = league[:10]
        q3 = _winner_short(_safe_int(row.get("q3_home")), _safe_int(row.get("q3_away")))
        q4 = _winner_short(_safe_int(row.get("q4_home")), _safe_int(row.get("q4_away")))
        status_type = str(row.get("status_type", "") or "").strip().lower()
        is_finished = status_type == "finished"
        q4_started = (
            _safe_int(row.get("q4_home")) is not None
            and _safe_int(row.get("q4_away")) is not None
        )
        if not is_finished and not q4_started:
            q3 = "⏳"
        if not is_finished:
            q4 = "⏳"
        pred = (pred_map or {}).get(match_id, {})
        q3e = _pred_outcome_emoji_for_row(pred, "q3")
        q4e = _pred_outcome_emoji_for_row(pred, "q4")
        if not is_finished and not q4_started:
            q3e = ""
        if not is_finished:
            q4e = ""
        # Resolve ⏳ from actual scores when available
        if q3e == "⏳" and pred.get("q3_available") and q4_started:
            _q3h_r = _safe_int(row.get("q3_home"))
            _q3a_r = _safe_int(row.get("q3_away"))
            _q3p_r = str(pred.get("q3_pick") or "").lower()
            if _q3h_r is not None and _q3a_r is not None and _q3p_r in ("home", "away"):
                _q3w_r = _winner_from_scores(_q3h_r, _q3a_r)
                if _q3w_r == "push":
                    q3e = "➖"
                elif _q3w_r in ("home", "away"):
                    q3e = "✅" if _q3w_r == _q3p_r else "❌"
        if q4e == "⏳" and pred.get("q4_available") and is_finished:
            _q4h_r = _safe_int(row.get("q4_home"))
            _q4a_r = _safe_int(row.get("q4_away"))
            _q4p_r = str(pred.get("q4_pick") or "").lower()
            if _q4h_r is not None and _q4a_r is not None and _q4p_r in ("home", "away"):
                _q4w_r = _winner_from_scores(_q4h_r, _q4a_r)
                if _q4w_r == "push":
                    q4e = "➖"
                elif _q4w_r in ("home", "away"):
                    q4e = "✅" if _q4w_r == _q4p_r else "❌"
        
        # Get model prediction emojis
        q3_pred_emoji = ""
        q4_pred_emoji = ""
        q3_pick = str(pred.get("q3_pick") or "").lower()
        q4_pick = str(pred.get("q4_pick") or "").lower()
        
        if pred.get("q3_available") and q3_pick in ("home", "away"):
            q3_pred_emoji = "🏠" if q3_pick == "home" else "✈️"
        if pred.get("q4_available") and q4_pick in ("home", "away"):
            q4_pred_emoji = "🏠" if q4_pick == "home" else "✈️"
        
        label = f"{tipoff} {league_short} | {home} vs {away} | Q3:{q3}{q3e}{q3_pred_emoji} Q4:{q4}{q4e}{q4_pred_emoji}"
        callback = f"match:{match_id}:{event_date}:{page}"
        keyboard.append([InlineKeyboardButton(label, callback_data=callback)])

    nav_row: list[InlineKeyboardButton] = []
    if page > 0:
        nav_row.append(
            InlineKeyboardButton(
                "<<",
                callback_data=f"date:{event_date}:{page - 1}",
            )
        )
    nav_row.append(
        InlineKeyboardButton(
            f"Pag {page + 1}/{total_pages}",
            callback_data="noop",
        )
    )
    if page < total_pages - 1:
        nav_row.append(
            InlineKeyboardButton(
                ">>",
                callback_data=f"date:{event_date}:{page + 1}",
            )
        )
    keyboard.append(nav_row)

    keyboard.append(
        [InlineKeyboardButton("Volver a fechas", callback_data="dates:0")]
    )
    keyboard.append(
        [InlineKeyboardButton("🧠 Recalcular apuestas del dia", callback_data=f"calc:date:{event_date}")]
    )
    keyboard.append(
        [InlineKeyboardButton("⬇️ Descargar pendientes del dia", callback_data=f"refresh:date:{event_date}")]
    )
    keyboard.append(
        [InlineKeyboardButton(
            f"🤖 Cambiar modelo (Q3={MODEL_CONFIG['q3']} Q4={MODEL_CONFIG['q4']})",
            callback_data=f"datemodel:open:{event_date}",
        )]
    )
    keyboard.append([InlineKeyboardButton("Menu principal", callback_data="nav:main")])
    return InlineKeyboardMarkup(keyboard)


def _calc_date_predictions(event_date: str, *, force_recalc: bool = True) -> dict:
    rows = _fetch_matches_for_date(event_date)
    existing = _fetch_date_pred_outcomes(event_date)
    ok = fail = skipped = 0
    for row in rows:
        match_id = str(row.get("match_id", ""))
        if not match_id:
            continue
        if (not force_recalc) and (match_id in existing):
            skipped += 1
            continue
        data = _get_match_detail(match_id)
        if not data:
            fail += 1
            continue
        try:
            result = _compute_and_store_predictions(match_id, data)
            if result is not None:
                ok += 1
            else:
                fail += 1
        except Exception:
            fail += 1
    return {
        "total": len(rows),
        "recalculated": ok + fail,
        "skipped": skipped,
        "ok": ok,
        "fail": fail,
    }


async def _run_calc_date_job(
    app: Application,
    chat_id: int,
    message_id: int,
    event_date: str,
) -> None:
    """Background task: recalculate predictions for all matches on a date,
    editing the message every NOTIFY_EVERY matches to show progress."""
    NOTIFY_EVERY = 10

    rows = await asyncio.to_thread(_fetch_matches_for_date, event_date)
    total = len(rows)
    ok = fail = 0

    async def _edit(text: str, reply_markup=None, parse_mode: str | None = None) -> None:
        try:
            kwargs: dict = {
                "chat_id": chat_id,
                "message_id": message_id,
                "text": text,
            }
            if reply_markup is not None:
                kwargs["reply_markup"] = reply_markup
            if parse_mode is not None:
                kwargs["parse_mode"] = parse_mode
            await app.bot.edit_message_text(**kwargs)
        except Exception:
            pass

    for i, row in enumerate(rows, 1):
        match_id = str(row.get("match_id", ""))
        if not match_id:
            continue
        data = await asyncio.to_thread(_get_match_detail, match_id)
        if not data:
            fail += 1
        else:
            try:
                result = await asyncio.to_thread(
                    _compute_and_store_predictions, match_id, data, dict(MODEL_CONFIG)
                )
                if result is not None:
                    ok += 1
                else:
                    fail += 1
            except Exception:
                fail += 1

        if i % NOTIFY_EVERY == 0 and i < total:
            _rc_label = _event_date_title_es(event_date, total).split("[")[0].strip()
            _rc_model = f"Q3={MODEL_CONFIG['q3']} Q4={MODEL_CONFIG['q4']}"
            await _edit(
                f"Recalculando Partidos {_rc_label}\n"
                f"{i}/{total} Procesados {_rc_model} | OK: {ok} | Fail: {fail}"
            )

    summary = (
        f"Recalculo completado para {event_date}:\n"
        f"Total: {total} | OK: {ok} | Fallidas: {fail}"
    )
    pred_map = await asyncio.to_thread(_fetch_date_pred_outcomes, event_date)
    stats = _pred_stats_text(pred_map, total, match_rows=rows)
    model_line = f"Modelo: Q3={MODEL_CONFIG['q3']} Q4={MODEL_CONFIG['q4']}"
    title = _event_date_title_es(event_date, total)
    header = f"{summary}\n\n{model_line}\n{title}"
    parse_mode: str | None = None
    if stats:
        header += f"\n<pre>{html.escape(stats)}</pre>"
        parse_mode = "HTML"
    keyboard = _matches_keyboard(rows, event_date, 0, pred_map)
    await _edit(header, reply_markup=keyboard, parse_mode=parse_mode)


def _is_two_half_game(data: dict) -> bool:
    """Returns True if match uses 2 halves instead of 4 quarters.

    Detection rules:
    - If Q3 or Q4 already have completed scores it's definitely a 4Q game.
    - Finished game with ≤2 scored quarters → 2-half format.
    - In-progress game at minute ≥36 with no Q3/Q4 data → 4Q game would
      have Q3 data by then, so this must be a 2-half format.
    """
    s = data.get("score", {}) or {}
    m = data.get("match", {}) or {}
    quarters = s.get("quarters", {}) or {}

    def _has_score(qk: str) -> bool:
        q = quarters.get(qk)
        return isinstance(q, dict) and q.get("home") is not None

    if _has_score("Q3") or _has_score("Q4"):
        return False

    status_type = str(m.get("status_type", "") or "").strip().lower()
    if status_type == "finished":
        n_scored = sum(1 for k in ("Q1", "Q2", "Q3", "Q4") if _has_score(k))
        return n_scored <= 2

    # In-progress: at minute >=36 a 4Q game would already have Q3 data
    graph_points = data.get("graph_points", [])
    max_minute = 0
    if graph_points:
        try:
            max_minute = int(graph_points[-1].get("minute", 0))
        except (TypeError, ValueError):
            max_minute = 0
    return max_minute >= 36


def _quarter_line(data: dict, quarter: str) -> str:
    q = data.get("score", {}).get("quarters", {}).get(quarter)
    if not q:
        return f"{quarter}: -"

    def _bold_num(value: int) -> str:
        trans = str.maketrans("0123456789-", "𝟬𝟭𝟮𝟯𝟰𝟱𝟲𝟳𝟴𝟵-")
        return str(value).translate(trans)

    home = _safe_int(q.get("home"))
    away = _safe_int(q.get("away"))
    if home is None or away is None:
        return f"{quarter}: {q.get('home', '-')} - {q.get('away', '-')}"

    home_txt = str(home)
    away_txt = str(away)
    if home > away:
        home_txt = _bold_num(home)
    elif away > home:
        away_txt = _bold_num(away)
    return f"{quarter}: {home_txt} - {away_txt}"


def _graph_summary(graph_points: list[dict]) -> tuple[int, int, int, int, int]:
    if not graph_points:
        return 0, 0, 0, 0, 0

    values: list[int] = []
    for p in graph_points:
        try:
            values.append(int(p.get("value", 0)))
        except (TypeError, ValueError):
            continue

    if not values:
        return 0, 0, 0, 0, 0

    last = values[-1]
    peak_home = max(values)
    peak_away = min(values)
    spread = peak_home - peak_away
    return len(values), peak_home, peak_away, last, spread


def _match_detail_text(match_id: str, data: dict) -> str:
    m = data.get("match", {})
    s = data.get("score", {})
    home_team_name = str(m.get("home_team", "home") or "home")
    away_team_name = str(m.get("away_team", "away") or "away")
    q3 = s.get("quarters", {}).get("Q3")
    q4 = s.get("quarters", {}).get("Q4")
    q2 = s.get("quarters", {}).get("Q2")

    q2_home = _safe_int((q2 or {}).get("home"))
    q2_away = _safe_int((q2 or {}).get("away"))
    q3_home = _safe_int((q3 or {}).get("home"))
    q3_away = _safe_int((q3 or {}).get("away"))
    q4_home = _safe_int((q4 or {}).get("home"))
    q4_away = _safe_int((q4 or {}).get("away"))

    utc_date = str(m.get("date", "") or "")
    utc_time = str(m.get("time", "") or "")
    local_dt = _to_local_datetime(utc_date, utc_time)
    if local_dt is not None:
        fecha_line = f"Fecha: {local_dt.strftime('%Y-%m-%d %H:%M')} UTC{UTC_OFFSET_HOURS:+d}"
    else:
        fecha_line = f"Fecha: {utc_date} {utc_time} UTC"

    status_type = str(m.get("status_type", "") or "")
    status_description = str(m.get("status_description", "") or "")
    minute_est = None
    graph_points = data.get("graph_points", [])
    if graph_points:
        try:
            minute_est = int(graph_points[-1].get("minute", 0))
        except (TypeError, ValueError):
            minute_est = None
    # Show how stale the cached data is for live matches
    _stale_txt = ""
    if status_type not in ("finished", "notstarted"):
        _scraped_at = m.get("scraped_at")
        if _scraped_at:
            try:
                from datetime import timezone as _tz
                _dt_scraped = datetime.fromisoformat(str(_scraped_at).replace("Z", "+00:00"))
                if _dt_scraped.tzinfo is None:
                    _dt_scraped = _dt_scraped.replace(tzinfo=_tz.utc)
                _age_secs = (datetime.now(_tz.utc) - _dt_scraped).total_seconds()
                _age_min = int(_age_secs // 60)
                if _age_min >= 2:
                    _stale_txt = f" (hace {_age_min}min)"
            except Exception:
                pass
    if minute_est is not None:
        minute_suffix = f" | Min: {minute_est}{_stale_txt}"
    elif _stale_txt:
        minute_suffix = f" |{_stale_txt}"
    else:
        minute_suffix = ""

    def _quarter_winner_text(quarter: str, home: int | None, away: int | None) -> str:
        if home is None or away is None:
            return "⏳"
        if quarter == "Q4" and status_type != "finished":
            return "⏳"
        if quarter == "Q3" and status_type != "finished":
            q4 = s.get("quarters", {}).get("Q4")
            if not q4 and minute_est is not None and minute_est < 36 and "Q4" not in status_description:
                return "⏳"
        winner = _winner_from_scores(home, away)
        if winner == "home":
            return f"🏠 {home_team_name}"
        if winner == "away":
            return f"🛫 {away_team_name}"
        if winner == "push":
            return "➖ push"
        return "-"

    def _quarter_line_with_winner(quarter: str, home: int | None, away: int | None) -> str:
        def _bold_num(value: int) -> str:
            trans = str.maketrans("0123456789-", "𝟬𝟭𝟮𝟯𝟰𝟱𝟲𝟳𝟴𝟵-")
            return str(value).translate(trans)

        if home is None or away is None:
            score_txt = "-"
        else:
            home_txt = str(home)
            away_txt = str(away)
            if home > away:
                home_txt = _bold_num(home)
            elif away > home:
                away_txt = _bold_num(away)
            score_txt = f"{home_txt} - {away_txt}"
        winner_txt = _quarter_winner_text(quarter, home, away)
        if winner_txt == "-":
            return f"{quarter}: {score_txt}"
        return f"{quarter}: {score_txt} {winner_txt}"

    def _quarter2_line() -> str:
        if q2_home is None or q2_away is None:
            if status_type != "finished":
                return "Q2: - ⏳"
            return "Q2: -"
        return _quarter_line(data, "Q2")

    two_half = _is_two_half_game(data)
    lines = [
        f"Match ID: {match_id}{minute_suffix}",
        f"{m.get('home_team', '')} vs {m.get('away_team', '')}",
        fecha_line,
        f"Liga: {m.get('league', '')}",
        _quarter_line(data, "Q1"),
        _quarter2_line(),
    ]
    if not two_half:
        lines.append(_quarter_line_with_winner("Q3", q3_home, q3_away))
        lines.append(_quarter_line_with_winner("Q4", q4_home, q4_away))
    return "\n".join(lines)


def _quarter_score(data: dict | None, quarter: str) -> tuple[int | None, int | None]:
    if not data:
        return None, None
    q = data.get("score", {}).get("quarters", {}).get(quarter)
    if not q:
        return None, None
    return _safe_int(q.get("home")), _safe_int(q.get("away"))


def _all_result_tags(conn: sqlite3.Connection) -> list[str]:
    cols = [
        row["name"]
        for row in conn.execute("PRAGMA table_info(eval_match_results)").fetchall()
    ]
    tags: set[str] = set()
    for col in cols:
        if col.startswith("q3_pick__"):
            tags.add(col.split("__", maxsplit=1)[1])
    return sorted(tags)


def _row_value(row: sqlite3.Row, key: str) -> object:
    return row[key] if key in row.keys() else None


def _read_prediction_row(match_id: str) -> dict | None:
    conn = _open_conn()
    tags = _all_result_tags(conn)
    if not tags:
        conn.close()
        return None

    rows = conn.execute(
        """
        SELECT *
        FROM eval_match_results
        WHERE match_id = ?
        ORDER BY updated_at DESC, event_date DESC
        LIMIT 25
        """,
        (match_id,),
    ).fetchall()
    conn.close()

    fallback: dict | None = None
    for row in rows:
        keys = set(row.keys())
        for tag in tags:
            q3_av_key = f"q3_available__{tag}"
            q4_av_key = f"q4_available__{tag}"
            if q3_av_key not in keys or q4_av_key not in keys:
                continue

            q3_av = int(row[q3_av_key] or 0)
            q4_av = int(row[q4_av_key] or 0)

            candidate = {
                "result_tag": tag,
                "event_date": row["event_date"],
                "updated_at": row["updated_at"],
                "q3": {
                    "available": bool(q3_av),
                    "pick": _row_value(row, f"q3_pick__{tag}"),
                    "signal": _row_value(row, f"q3_signal__{tag}"),
                    "outcome": _row_value(row, f"q3_outcome__{tag}"),
                    "confidence": _row_value(row, f"q3_confidence__{tag}"),
                    "threshold_lean": _row_value(row, f"q3_threshold_lean__{tag}"),
                    "threshold_bet": _row_value(row, f"q3_threshold_bet__{tag}"),
                    "reasoning": _row_value(row, f"q3_reasoning__{tag}"),
                    "predicted_home": _row_value(row, f"q3_predicted_home__{tag}"),
                    "predicted_away": _row_value(row, f"q3_predicted_away__{tag}"),
                    "predicted_total": _row_value(row, f"q3_predicted_total__{tag}"),
                    "mae": _row_value(row, f"q3_mae__{tag}"),
                    "mae_home": _row_value(row, f"q3_mae_home__{tag}"),
                    "mae_away": _row_value(row, f"q3_mae_away__{tag}"),
                },
                "q4": {
                    "available": bool(q4_av),
                    "pick": _row_value(row, f"q4_pick__{tag}"),
                    "signal": _row_value(row, f"q4_signal__{tag}"),
                    "outcome": _row_value(row, f"q4_outcome__{tag}"),
                    "confidence": _row_value(row, f"q4_confidence__{tag}"),
                    "threshold_lean": _row_value(row, f"q4_threshold_lean__{tag}"),
                    "threshold_bet": _row_value(row, f"q4_threshold_bet__{tag}"),
                    "reasoning": _row_value(row, f"q4_reasoning__{tag}"),
                    "predicted_home": _row_value(row, f"q4_predicted_home__{tag}"),
                    "predicted_away": _row_value(row, f"q4_predicted_away__{tag}"),
                    "predicted_total": _row_value(row, f"q4_predicted_total__{tag}"),
                    "mae": _row_value(row, f"q4_mae__{tag}"),
                    "mae_home": _row_value(row, f"q4_mae_home__{tag}"),
                    "mae_away": _row_value(row, f"q4_mae_away__{tag}"),
                },
            }

            if q3_av or q4_av:
                return candidate
            if fallback is None:
                fallback = candidate

    return fallback


def _friendly_reason(reason: str) -> str:
    mapping = {
        "match_not_found": "match no encontrado en base/scraper",
        "missing_q3_score": "Q3 aun no disponible",
        "missing_q4_score": "Q4 aun no disponible",
        "missing_graph_points": "sin graph_points suficientes",
        "missing_play_by_play": "sin play-by-play suficiente",
        "no_best_model_entry": "sin modelo seleccionado en comparacion",
        "v3_q3_available_from_minute_24": "Q3 disponible desde minuto 24",
        "v3_q4_available_from_minute_24": "Q4 disponible desde minuto 24",
        "insufficient_graph_or_pbp_coverage": "cobertura de datos insuficiente",
        "match_too_volatile_for_current_signal": "bloqueado por alta volatilidad",
        "confidence_below_minimum_edge": "confianza por debajo del minimo",
        "passed_all_gates": "paso todos los filtros",
        # V12 specific reasons
        "league not bettable": "liga no apostable",
        "LEAGUE NOT BETTABLE": "liga no apostable",
        "Insufficient graph points": "puntos de grafico insuficientes",
        "V12 error": "error en modelo V12",
        "V12 failed": "modelo V12 fallo",
        "No se pudo analizar": "no se pudo analizar",
        "V12 LIVE error": "error en V12 LIVE",
        # V13 specific reasons
        "V13: ": "error en modelo V13",
        "V13 no disponible": "modelo V13 no disponible",
        "Match data not found": "no se encontraron datos del partido",
        "No models found for": "no hay modelo entrenado para este tipo de partido",
        "Low training samples": "modelo con pocos datos de entrenamiento",
        "Volatility too high": "volatilidad del partido demasiado alta",
    }
    # Check for substring matches (for longer English reasons)
    reason_lower = reason.lower()
    if "league not bettable" in reason_lower:
        return "liga no apostable"
    if "insufficient graph points" in reason_lower:
        return "puntos de grafico insuficientes"
    if "insufficient confidence" in reason_lower:
        return "confianza insuficiente para apostar"
    if "too volatile" in reason_lower or "volatility too high" in reason_lower:
        return "partido muy volatil para apostar"
    if "confidence below" in reason_lower:
        return "confianza por debajo del minimo"
    # V13: "Confidence 0.540 < 0.62" pattern
    if "confidence" in reason_lower and "<" in reason_lower:
        return "confianza insuficiente para apostar"
    if "no models found" in reason_lower:
        return "no hay modelo entrenado para este tipo de partido"
    if "match data not found" in reason_lower:
        return "no se encontraron datos del partido"
    if "low training samples" in reason_lower:
        return "modelo con pocos datos de entrenamiento"
    if "missing" in reason_lower and "graph" in reason_lower:
        return "faltan puntos de grafico"
    if "missing" in reason_lower and "play" in reason_lower:
        return "faltan datos de jugadas"
    if "match too early" in reason_lower:
        return "partido muy temprano"
    return mapping.get(reason, reason or "no disponible")


def _league_stats_detail(league: str) -> str | None:
    """Return a one-line stats summary explaining why the league is not bettable."""
    try:
        stats_file = BASE_DIR / "training" / "v12" / "model_outputs" / "league_stats.json"
        if not stats_file.exists():
            return None
        with open(stats_file, "r", encoding="utf-8") as _f:
            _stats = json.load(_f)
        # Exact match first, then base name before comma, then prefix
        s = _stats.get(league, {})
        if not s and "," in league:
            s = _stats.get(league.split(",")[0].strip(), {})
        if not s:
            for key in sorted(_stats.keys(), key=len, reverse=True):
                if league.startswith(key):
                    s = _stats[key]
                    break
        if not s:
            return "📊 Liga sin datos históricos"
        samples = int(s.get("samples", 0))
        home_wr = float(s.get("home_win_rate", 0.5))
        if samples < 30:
            return f"📊 Solo {samples} partidos en histórico (mínimo 30)"
        home_pct = round(home_wr * 100)
        return f"📊 N={samples} partidos | Local gana {home_pct}% — sin ventaja clara"
    except Exception:
        return None


def _enrich_prediction_row_from_infer(
    pred_row: dict | None,
    infer_result: dict,
) -> dict | None:
    if pred_row is None:
        return None

    predictions = infer_result.get("predictions", {}) if isinstance(infer_result, dict) else {}
    minute_estimate = (
        (infer_result.get("match") or {}).get("minute_estimate")
        if isinstance(infer_result, dict)
        else None
    )

    for target in ("q3", "q4"):
        pred = pred_row.get(target)
        src = predictions.get(target)
        if not isinstance(pred, dict) or not isinstance(src, dict):
            continue
        pred["available"] = bool(src.get("available", pred.get("available")))
        pred["pick"] = src.get("predicted_winner") or pred.get("pick")
        pred["signal"] = src.get("final_recommendation") or src.get("bet_signal") or pred.get("signal")
        pred["outcome"] = src.get("result") or pred.get("outcome")
        pred["confidence"] = src.get("confidence", pred.get("confidence"))
        pred["p_home_win"] = src.get("p_home_win", pred.get("p_home_win"))
        pred["p_away_win"] = src.get("p_away_win", pred.get("p_away_win"))
        pred["threshold_lean"] = src.get("threshold_lean", pred.get("threshold_lean"))
        pred["threshold_bet"] = src.get("threshold_bet", pred.get("threshold_bet"))
        pred["reason"] = src.get("reason") or src.get("gate_reason")
        pred["reasoning"] = src.get("reasoning") or pred.get("reason")
        pred["decision_gate"] = src.get("decision_gate")
        pred["gate_reason"] = src.get("gate_reason")
        pred["volatility_index"] = src.get("volatility_index")
        pred["volatility_swings"] = src.get("volatility_swings")
        pred["volatility_lead_changes"] = src.get("volatility_lead_changes")
        pred["gp_count"] = src.get("gate_gp_count")
        pred["pbp_count"] = src.get("gate_pbp_count")
        pred["minute_estimate"] = minute_estimate
        # V12-specific fields
        pred["predicted_home"] = src.get("predicted_home")
        pred["predicted_away"] = src.get("predicted_away")
        pred["predicted_total"] = src.get("predicted_total")
        pred["mae"] = src.get("mae")
        pred["mae_home"] = src.get("mae_home")
        pred["mae_away"] = src.get("mae_away")
        # V13-specific quality metadata
        pred["model_quality"] = src.get("model_quality")
        pred["model_samples"] = src.get("model_samples")
        pred["model_gap"] = src.get("model_gap")
        pred["fallback_used"] = src.get("fallback_used")

    return pred_row


def _needs_unavailable_reason(pred_row: dict | None) -> bool:
    if not isinstance(pred_row, dict):
        return False
    for target in ("q3", "q4"):
        pred = pred_row.get(target, {})
        if isinstance(pred, dict) and (not pred.get("available")) and (not pred.get("reason")):
            return True
    return False


def _missing_prediction_confidence(pred_row: dict | None) -> bool:
    if not isinstance(pred_row, dict):
        return True
    for target in ("q3", "q4"):
        pred = pred_row.get(target, {})
        if not isinstance(pred, dict) or not pred.get("available"):
            continue
        if pred.get("confidence") is None:
            return True
    return False


def _q3_is_locked(data: dict) -> bool:
    """True when Q3 is actively being played — Q2 is done, Q3 not yet done.
    At this point the Q3 prediction window has closed and should be frozen."""
    quarters = (data.get("score", {}) or {}).get("quarters", {}) or {}
    # Q4 started → Q3 is already done, no need to freeze Q3 (it will resolve naturally)
    if "Q4" in quarters:
        return False
    # Q3 score exists (Q3 is in progress or just ended) → freeze
    if "Q3" in quarters:
        return True
    # Use graph point minutes as proxy: past halftime → Q3 in progress
    gp = data.get("graph_points", [])
    if gp:
        try:
            last_min = int(gp[-1].get("minute", 0))
            if last_min >= 24:
                return True
        except (TypeError, ValueError):
            pass
    return False


def _frozen_q3_to_infer_pred(cached_q3: dict) -> dict:
    """Repackage a DB Q3 row back to the inference predictions format so it
    can be passed to save_eval_match_result without change."""
    pick = str(cached_q3.get("pick") or "")
    signal = str(cached_q3.get("signal") or cached_q3.get("final_recommendation") or "NO BET")
    return {
        "available": bool(cached_q3.get("available")),
        "predicted_winner": pick if pick in ("home", "away") else None,
        "p_home_win": cached_q3.get("p_home_win"),
        "p_away_win": cached_q3.get("p_away_win"),
        "confidence": cached_q3.get("confidence"),
        "bet_signal": signal,
        "final_recommendation": signal,
        "suggested_units": 1.0 if signal == "BET" else 0.0,
        "threshold_lean": cached_q3.get("threshold_lean"),
        "threshold_bet": cached_q3.get("threshold_bet"),
        "gate_reason": cached_q3.get("gate_reason") or cached_q3.get("reason"),
        "reasoning": cached_q3.get("reasoning") or cached_q3.get("reason"),
        "decision_gate": cached_q3.get("decision_gate"),
        "gate_gp_count": cached_q3.get("gp_count"),
        "gate_pbp_count": cached_q3.get("pbp_count"),
        "volatility_index": cached_q3.get("volatility_index"),
        "volatility_swings": cached_q3.get("volatility_swings"),
        "volatility_lead_changes": cached_q3.get("volatility_lead_changes"),
        "predicted_home": cached_q3.get("predicted_home"),
        "predicted_away": cached_q3.get("predicted_away"),
        "predicted_total": cached_q3.get("predicted_total"),
        "mae": cached_q3.get("mae"),
        "mae_home": cached_q3.get("mae_home"),
        "mae_away": cached_q3.get("mae_away"),
        "_frozen_q3": True,
    }


def _compute_and_store_predictions(match_id: str, data: dict, config_override: dict | None = None) -> dict | None:
    infer_mod = importlib.import_module("training.infer_match")
    infer_mod.scraper_mod.fetch_event_snapshot = lambda _mid: None

    logger.info(f"[COMPUTE_PRED] Starting predictions for {match_id}")

    # Use override config (match-detail model selector) or fall back to MONITOR config
    _active_config = dict(config_override) if config_override is not None else dict(MONITOR_MODEL_CONFIG)
    logger.info(f"[COMPUTE_PRED] Using config: Q3={_active_config.get('q3')} Q4={_active_config.get('q4')}")

    # Check if any advanced model (v12, v13, v15, v16) is selected for either quarter
    _ADVANCED_MODELS = {"v12", "v13", "v15", "v16", "v17"}
    using_advanced = any(v in _ADVANCED_MODELS for v in _active_config.values())

    if using_advanced:
        # Run advanced-model inference per quarter
        logger.info(f"[COMPUTE_PRED] Using advanced model(s) for {match_id}")
        result = {"ok": True, "predictions": {}}

        for quarter in ["q3", "q4"]:
            qv = _active_config[quarter]
            if qv == "v12":
                qr = _run_v12_inference(match_id, quarter)
            elif qv == "v13":
                qr = _run_v13_inference(match_id, quarter)
            elif qv == "v15":
                qr = _run_v15_inference(match_id, quarter)
            elif qv == "v16":
                qr = _run_v16_inference(match_id, quarter)
            elif qv == "v17":
                qr = _run_v17_inference(match_id, quarter)
            else:
                qr = None  # handled below by regular inference

            if qr is not None:
                if qr.get("ok"):
                    result["predictions"][quarter] = qr["predictions"][quarter]
                else:
                    result["predictions"][quarter] = {
                        "available": False,
                        "reason": qr.get("reason", f"{qv.upper()} falló"),
                    }

        # Run regular inference for any non-advanced quarters
        non_advanced_quarters = [q for q in ("q3", "q4") if _active_config.get(q) not in _ADVANCED_MODELS]
        if non_advanced_quarters:
            logger.info(f"[COMPUTE_PRED] Running regular inference for quarters: {non_advanced_quarters}")
            # Sub in v4 for advanced-model slots so regular inference doesn't break
            reg_config = {q: (v if v not in _ADVANCED_MODELS else "v4") for q, v in _active_config.items()}
            reg_result = infer_mod.run_inference(
                match_id=match_id,
                metric="f1",
                fetch_missing=False,
                force_version=reg_config,
            )
            if reg_result.get("ok"):
                for q in non_advanced_quarters:
                    result["predictions"][q] = reg_result["predictions"].get(
                        q, {"available": False, "reason": "inference_failed"}
                    )
            else:
                for q in non_advanced_quarters:
                    result["predictions"][q] = {"available": False, "reason": "inference_failed"}
    else:
        # Use regular inference
        result = infer_mod.run_inference(
            match_id=match_id,
            metric="f1",
            fetch_missing=False,
            force_version=_active_config,
        )
    _log_raw_inference_result(match_id, source="compute", infer_result=result)
    
    logger.info(f"[COMPUTE_PRED] Inference result ok={result.get('ok')}")
    if result.get("ok"):
        preds = result.get("predictions", {})
        logger.info(
            f"[COMPUTE_PRED] Q3 available={preds.get('q3', {}).get('available')} "
            f"| Q4 available={preds.get('q4', {}).get('available')}"
        )
    
    if not result.get("ok"):
        logger.warning(f"[COMPUTE_PRED] Inference failed for {match_id}")
        return None

    # NOTE: Q3 freeze logic was removed here.
    # infer_match._clip_match_data_for_target now truncates graph_points and
    # play_by_play to minute ≤ 24 (Q3) / ≤ 36 (Q4) before any feature or gate
    # computation, so predictions are stable regardless of when inference runs.
    # To revert the clipping: set infer_match._CLIP_DATA_TO_CUTOFF = False.

    q3h, q3a = _quarter_score(data, "Q3")
    q4h, q4a = _quarter_score(data, "Q4")
    event_date = str(data.get("match", {}).get("date", "") or "")
    if not event_date:
        event_date = datetime.utcnow().date().isoformat()

    conn = _open_conn()
    db_mod.save_eval_match_result(
        conn,
        event_date=event_date,
        match_id=match_id,
        home_team=str(data.get("match", {}).get("home_team", "")),
        away_team=str(data.get("match", {}).get("away_team", "")),
        q3_home_score=q3h,
        q3_away_score=q3a,
        q4_home_score=q4h,
        q4_away_score=q4a,
        result_tag="bot_hybrid_f1",
        predictions=result.get("predictions", {}),
    )
    conn.close()
    logger.info(f"[COMPUTE_PRED] Saved to database for {match_id}")
    base_row = _read_prediction_row(match_id)
    result_row = _enrich_prediction_row_from_infer(base_row, result)
    if result_row is not None:
        result_row["_used_config"] = dict(_active_config)
    return result_row


def _get_or_compute_predictions(match_id: str, data: dict) -> dict | None:
    from_db = _read_prediction_row(match_id)
    if (
        from_db is not None
        and not _needs_unavailable_reason(from_db)
        and not _missing_prediction_confidence(from_db)
    ):
        # If the match is finished and outcomes are still pending, recompute to resolve them
        is_ft = str((data or {}).get("match", {}).get("status_type", "") or "").lower() == "finished"
        if is_ft:
            q3_out = str((from_db.get("q3") or {}).get("outcome") or "").lower()
            q4_out = str((from_db.get("q4") or {}).get("outcome") or "").lower()
            has_pending = q3_out in ("pending", "") or q4_out in ("pending", "")
            if has_pending:
                try:
                    refreshed = _compute_and_store_predictions(match_id, data)
                    if refreshed is not None:
                        return refreshed
                except Exception:
                    pass
        return from_db
    if from_db is not None:
        # When an advanced model (v13, v12, v15, v16) is active, always recompute via
        # _compute_and_store_predictions so the correct engine is called and
        # the raw JSON is logged to console.
        _ADVANCED = {"v12", "v13", "v15", "v16"}
        _using_advanced = any(v in _ADVANCED for v in MONITOR_MODEL_CONFIG.values())
        if _using_advanced:
            try:
                return _compute_and_store_predictions(match_id, data)
            except Exception:
                return from_db
        try:
            infer_mod = importlib.import_module("training.infer_match")
            infer_mod.scraper_mod.fetch_event_snapshot = lambda _mid: None
            infer_result = infer_mod.run_inference(
                match_id=match_id,
                metric="f1",
                fetch_missing=False,
                force_version=dict(MONITOR_MODEL_CONFIG),
            )
            _log_raw_inference_result(match_id, source="enrich-from-db", infer_result=infer_result)
            return _enrich_prediction_row_from_infer(from_db, infer_result)
        except Exception:
            return from_db
    try:
        return _compute_and_store_predictions(match_id, data)
    except Exception:
        return None


def _refresh_predictions(match_id: str, data: dict, config_override: dict | None = None) -> dict | None:
    logger.info(f"[REFRESH_PRED] Starting for match {match_id}")
    try:
        result = _compute_and_store_predictions(match_id, data, config_override=config_override)
        logger.info(f"[REFRESH_PRED] Success: computed predictions for {match_id}")
        return result
    except Exception as exc:
        logger.warning(f"[REFRESH_PRED] Compute failed for {match_id}: {exc}. Falling back to cached row.")
        fallback = _read_prediction_row(match_id)
        logger.info(f"[REFRESH_PRED] Fallback result: available={fallback.get('available') if fallback else None}")
        return fallback


def _refresh_match_data(match_id: str) -> dict | None:
    logger.info(f"[REFRESH_DATA] Starting for match {match_id}")
    try:
        logger.info(f"[REFRESH_DATA] Calling fetch_match_by_id({match_id}) from scraper")
        fresh = scraper_mod.fetch_match_by_id(match_id)
        if not fresh:
            logger.error(f"[REFRESH_DATA] Scraper returned None for {match_id}")
            return None
        
        home = fresh.get("match", {}).get("home_team", "?")
        away = fresh.get("match", {}).get("away_team", "?")
        score_home = fresh.get("score", {}).get("home", "?")
        score_away = fresh.get("score", {}).get("away", "?")
        gp_count = len(fresh.get("graph_points", []))
        pbp_count = sum(len(q) for q in fresh.get("play_by_play", {}).values())
        logger.info(
            f"[REFRESH_DATA] Fetched {home} vs {away} {score_home}-{score_away} | "
            f"graph_points={gp_count} play_by_play={pbp_count}"
        )
        
        logger.info(f"[REFRESH_DATA] Saving to database")
        conn = _open_conn()
        db_mod.save_match(conn, match_id, fresh)
        conn.close()
        logger.info(f"[REFRESH_DATA] Saved match {match_id} to database")
        
        return fresh
    except Exception as exc:
        logger.error(f"[REFRESH_DATA] Error during refresh for {match_id}: {exc}", exc_info=True)
        return None


def _prediction_text(pred_row: dict | None, data: dict | None = None, v10_preds: dict | None = None) -> str:
    if pred_row is None:
        return (
            "Predicciones:\n"
            "No disponibles (no se pudieron calcular)."
        )

    rendered_row = pred_row
    if isinstance(rendered_row, dict):
        rendered_row = {
            **rendered_row,
            "q3": dict(rendered_row.get("q3", {})),
            "q4": dict(rendered_row.get("q4", {})),
        }

    home_team_name = "home"
    away_team_name = "away"
    q4_waiting_preverdict = False
    if isinstance(data, dict):
        match_info_for_name = data.get("match", {})
        home_team_name = str(match_info_for_name.get("home_team") or "home")
        away_team_name = str(match_info_for_name.get("away_team") or "away")

    if isinstance(rendered_row, dict) and isinstance(data, dict):
        match_info = data.get("match", {})
        status_type = str(match_info.get("status_type", "") or "")
        status_description = str(match_info.get("status_description", "") or "")
        minute_est = None
        graph_points = data.get("graph_points", [])
        if graph_points:
            try:
                minute_est = int(graph_points[-1].get("minute", 0))
            except (TypeError, ValueError):
                minute_est = None

        q4_pred = rendered_row.get("q4", {})
        if isinstance(q4_pred, dict) and q4_pred.get("available") and status_type != "finished":
            q4_pred["outcome"] = "pending"

        q3_pred = rendered_row.get("q3", {})
        q4_score = (data.get("score", {}).get("quarters", {}) or {}).get("Q4")
        q4_started = bool(q4_score) or ("Q4" in status_description)
        if minute_est is not None and minute_est >= 36:
            q4_started = True
        # v13 fires Q4 at graph-minute 31 (mid-Q3) — if model already produced a
        # Q4 signal, don't suppress the display with "waiting for Q3"
        q4_has_signal = isinstance(rendered_row.get("q4"), dict) and bool(
            rendered_row["q4"].get("bet_signal") or rendered_row["q4"].get("available")
        )
        if status_type != "finished" and not q4_started and not q4_has_signal:
            q4_waiting_preverdict = True
        if isinstance(q3_pred, dict) and q3_pred.get("available") and status_type != "finished" and not q4_started:
            q3_pred["outcome"] = "pending"

        # Resolve Q3 outcome from actual scores once Q3 is over
        if isinstance(q3_pred, dict) and q3_pred.get("available"):
            q3_done = status_type == "finished" or q4_started
            if q3_done:
                _q3_sc = (data.get("score", {}).get("quarters", {}) or {}).get("Q3", {})
                _q3h = _safe_int((_q3_sc or {}).get("home"))
                _q3a = _safe_int((_q3_sc or {}).get("away"))
                if _q3h is not None and _q3a is not None:
                    _q3_pick = str(q3_pred.get("pick") or "").lower()
                    _q3_winner = _winner_from_scores(_q3h, _q3a)
                    if _q3_winner == "push":
                        q3_pred["outcome"] = "push"
                    elif _q3_winner in ("home", "away") and _q3_pick in ("home", "away"):
                        q3_pred["outcome"] = "hit" if _q3_winner == _q3_pick else "miss"

        # Resolve Q4 outcome from actual scores once game is finished,
        # or when the minute estimate shows the game clock has run out (API lag)
        q4_game_done = status_type == "finished" or (
            minute_est is not None and minute_est >= 48
        )
        if isinstance(q4_pred, dict) and q4_pred.get("available") and q4_game_done:
            _q4_sc = (data.get("score", {}).get("quarters", {}) or {}).get("Q4", {})
            _q4h = _safe_int((_q4_sc or {}).get("home"))
            _q4a = _safe_int((_q4_sc or {}).get("away"))
            if _q4h is not None and _q4a is not None:
                _q4_pick = str(q4_pred.get("pick") or "").lower()
                _q4_winner = _winner_from_scores(_q4h, _q4a)
                if _q4_winner == "push":
                    q4_pred["outcome"] = "push"
                elif _q4_winner in ("home", "away") and _q4_pick in ("home", "away"):
                    q4_pred["outcome"] = "hit" if _q4_winner == _q4_pick else "miss"

    def _emoji(outcome: str) -> str:
        e = _outcome_emoji(outcome)
        return e if e else "🟢"

    def _confidence_pct(pred: dict, pick_norm: str = "") -> str:
        value = pred.get("confidence")
        if value is None:
            return ""
        try:
            pct = round(float(value) * 100)
        except (TypeError, ValueError):
            return ""
        bet_thr = pred.get("threshold_bet")
        try:
            bet_thr_pct = round(float(bet_thr) * 100) if bet_thr is not None else None
        except (TypeError, ValueError):
            bet_thr_pct = None
        p_home = pred.get("p_home_win")
        p_away = pred.get("p_away_win")
        prob_pct = None
        try:
            if pick_norm == "home" and p_home is not None:
                prob_pct = round(float(p_home) * 100)
            elif pick_norm == "away" and p_away is not None:
                prob_pct = round(float(p_away) * 100)
        except (TypeError, ValueError):
            prob_pct = None

        parts = [f"Score {pct}%"]
        if prob_pct is not None:
            parts.append(f"Probabilidad {prob_pct}%")
        return ", ".join(parts)

    def _line(label: str, pred: dict, waiting_preverdict: bool = False) -> str:
        if label.startswith("Q4") and waiting_preverdict:
            return "Q4: Esperando datos Q3"
        if not pred.get("available"):
            reason_code = str(pred.get("reason") or pred.get("gate_reason") or pred.get("outcome") or "unavailable")
            reason_code_norm = reason_code.strip().lower().replace(" ", "_")
            reason_code_simple = reason_code.strip().lower().replace(" ", "_").replace("/", "_")
            if label.startswith("Q4") and reason_code_norm == "missing_q3_score":
                return "Q4: Esperando datos Q3"
            if reason_code_simple == "missing_q1_q2_scores":
                return f"{label}: Esperando datos Q1|Q2"
            reason_txt = _friendly_reason(reason_code)
            minute_est = pred.get("minute_estimate")
            gp_count = pred.get("gp_count")
            pbp_count = pred.get("pbp_count")
            extra = []
            if minute_est is not None:
                extra.append(f"minuto_actual={minute_est}")
            if gp_count is not None and pbp_count is not None:
                extra.append(
                    "datos: "
                    f"grafica={gp_count} eventos, "
                    f"play-by-play={pbp_count} eventos"
                )
            extra_txt = f" | {' '.join(extra)}" if extra else ""

            # Show projection and MAE even when not available (if data exists)
            def _sf(v):
                if v is None:
                    return None
                try:
                    return float(v)
                except (TypeError, ValueError):
                    return None

            pts_home = _sf(pred.get("predicted_home"))
            pts_away = _sf(pred.get("predicted_away"))
            pts_total = _sf(pred.get("predicted_total"))
            mae = _sf(pred.get("mae"))
            mae_home = _sf(pred.get("mae_home"))
            mae_away = _sf(pred.get("mae_away"))
            
            proj_lines = []
            if pts_home is not None and pts_away is not None:
                mae_h_txt = f" (±{mae_home:.1f})" if mae_home is not None else ""
                mae_a_txt = f" (±{mae_away:.1f})" if mae_away is not None else ""
                mae_total_txt = f" (±{mae:.0f})" if mae is not None else ""
                
                overlap_warning = ""
                if mae_home is not None and mae_away is not None:
                    diff = abs(pts_home - pts_away)
                    overlap_threshold = mae_home + mae_away
                    if diff < overlap_threshold:
                        overlap_warning = f" ⚠️ MAE se superpone (dif={diff:.1f} < {overlap_threshold:.1f})"
                
                proj_lines.append(f"Proyeccion: 🏠~{pts_home:.1f}{mae_h_txt} | ✈️~{pts_away:.1f}{mae_a_txt} | Total~{pts_total:.1f}{mae_total_txt}{overlap_warning}")
            
            result_lines = [f"{label}: no disponible ({reason_txt}){extra_txt}"]
            result_lines.extend(proj_lines)
            return "\n".join(result_lines)
        pick = str(pred.get("pick") or "-")
        pick_side_emoji = ""
        pick_norm = pick.strip().lower()
        if pick_norm == "home":
            pick = home_team_name
            pick_side_emoji = "🏠"
        elif pick_norm == "away":
            pick = away_team_name
            pick_side_emoji = "🛫"
        elif pick_norm == "uncertain":
            pick = "incierta"
        signal = pred.get("signal") or "-"
        signal_up = str(signal).upper()
        signal_map = {
            "BET": "APUESTA",
            "BET_HOME": "APUESTA",
            "BET_AWAY": "APUESTA",
            "LEAN": "NO APOSTAR",
            "NO BET": "NO APOSTAR",
            "NO_BET": "NO APOSTAR",
        }
        signal_txt = signal_map.get(signal_up, signal_up)
        outcome = str(pred.get("outcome") or "pending")
        confidence_txt = _confidence_pct(pred, pick_norm)
        pick_txt = f"{pick_side_emoji} {pick}" if pick_side_emoji else pick
        line_emoji = "🚫" if signal_up in {"LEAN", "NO BET", "NO_BET"} else _emoji(outcome)
        if signal_up in {"LEAN", "NO BET", "NO_BET"}:
            first_line = f"{line_emoji}{label}: {signal_txt}"
        else:
            first_line = f"{line_emoji}{label}: {signal_txt} {pick_txt}"
        trend_txt = ""
        if signal_up in {"LEAN", "NO BET", "NO_BET"}:
            if pick_txt != "-":
                trend_txt = f"Tendencia {pick_txt}"
        
        # Show reasoning for NO_BET (translated to Spanish)
        reasoning_txt = ""
        league_detail_txt = ""
        if signal_up in {"LEAN", "NO BET", "NO_BET"}:
            reason = str(pred.get("reasoning") or pred.get("reason") or "")
            if reason:
                reason_translated = _friendly_reason(reason)
                # For V13 confidence reasons, show actual numbers: (35% obtenido < 62% requerido)
                _conf_suffix = ""
                _raw_conf = pred.get("confidence")
                import re as _re
                _conf_match = _re.search(r'[Cc]onfidence\s+([0-9.]+)\s*<\s*([0-9.]+)', reason)
                if _conf_match:
                    _got = float(_conf_match.group(1))
                    _req = float(_conf_match.group(2))
                    _conf_suffix = f" ({_got*100:.0f}% < {_req*100:.0f}%)"
                elif _raw_conf is not None and "confianza" in reason_translated.lower():
                    try:
                        _conf_suffix = f" ({float(_raw_conf)*100:.0f}%)"
                    except (TypeError, ValueError):
                        pass
                reasoning_txt = f"Motivo: {reason_translated.capitalize()}{_conf_suffix}"
                if "league not bettable" in reason.lower():
                    _league = str((data or {}).get("match", {}).get("league") or "")
                    _ld = _league_stats_detail(_league)
                    if _ld:
                        league_detail_txt = _ld
        
        # Show predicted scores and MAE (with safe type conversion from DB)
        def _safe_float(v):
            if v is None:
                return None
            try:
                return float(v)
            except (TypeError, ValueError):
                return None

        pts_home = _safe_float(pred.get("predicted_home"))
        pts_away = _safe_float(pred.get("predicted_away"))
        pts_total = _safe_float(pred.get("predicted_total"))
        mae = _safe_float(pred.get("mae"))
        mae_home = _safe_float(pred.get("mae_home"))
        mae_away = _safe_float(pred.get("mae_away"))
        
        scores_txt = ""
        if pts_home is not None and pts_away is not None:
            mae_h_txt = f" (±{mae_home:.1f})" if mae_home is not None else ""
            mae_a_txt = f" (±{mae_away:.1f})" if mae_away is not None else ""
            mae_total_txt = f" (±{mae:.0f})" if mae is not None else ""

            # Check if prediction intervals overlap (high uncertainty)
            overlap_warning = ""
            if mae_home is not None and mae_away is not None:
                diff = abs(pts_home - pts_away)
                overlap_threshold = mae_home + mae_away
                if diff < overlap_threshold:
                    overlap_warning = f"⚠️ MAE se superpone (dif={diff:.1f} < {overlap_threshold:.1f})"

            scores_txt = f"Proyeccion: 🏠~{pts_home:.1f}{mae_h_txt} | ✈️~{pts_away:.1f}{mae_a_txt} | Total~{pts_total:.1f}{mae_total_txt}"
            if overlap_warning:
                scores_txt = scores_txt + "\n" + overlap_warning
        
        gate_reason_norm = str(pred.get("gate_reason") or pred.get("reason") or "").strip().lower()
        decision_gate_norm = str(pred.get("decision_gate") or "").strip().upper()
        volatility_txt = ""
        if (
            gate_reason_norm == "match_too_volatile_for_current_signal"
            or decision_gate_norm == "BLOCK_HIGH_VOLATILITY"
        ):
            volatility_txt = "⚠️ Alta volatilidad"

        # V13-specific model quality warnings + quality labels in Spanish
        v13_warning_txt = ""
        _pred_disp_cfg = (rendered_row.get("_used_config") if isinstance(rendered_row, dict) else None) or MODEL_CONFIG
        is_v13 = _pred_disp_cfg.get("q3") == "v13" or _pred_disp_cfg.get("q4") == "v13"
        if is_v13:
            mq = pred.get("model_quality")
            ms = pred.get("model_samples")
            mg = pred.get("model_gap")
            fb = pred.get("fallback_used")
            dq = pred.get("data_quality")
            _mq_labels = {"good": "buena", "moderate": "moderada", "low": "baja", "unknown": "desconocida"}
            _dq_labels = {"good": "buena", "moderate": "moderada", "low": "baja", "unknown": "desconocida"}
            _warn_parts = []
            if mq == "low":
                n = ms or 0
                _warn_parts.append(f"🚨 Datos de entrenamiento insuficientes ({n} partidos) — señal no confiable")
            elif mq == "moderate":
                n = ms or 0
                _warn_parts.append(f"⚠️ Modelo con datos reducidos ({n} partidos)")
            if fb:
                _warn_parts.append("ℹ️ Ritmo no clasificado — usando modelo alternativo")
            if mg is not None and float(mg) > 0.20:
                _warn_parts.append(f"⚠️ Sobreajuste detectado (gap={float(mg):.2f})")
            if mq and mq in _mq_labels:
                _warn_parts.append(f"Calidad modelo: {_mq_labels[mq]}")
            if dq and dq in _dq_labels:
                _warn_parts.append(f"Calidad datos: {_dq_labels[dq]}")
            if _warn_parts:
                v13_warning_txt = "\n".join(_warn_parts)

        extra_lines = [line for line in [trend_txt, reasoning_txt, league_detail_txt, scores_txt, volatility_txt, v13_warning_txt, confidence_txt] if line]
        if extra_lines:
            return f"{first_line}\n" + "\n".join(extra_lines)
        return first_line

    _disp_cfg = (rendered_row.get("_used_config") if isinstance(rendered_row, dict) else None) or MODEL_CONFIG
    q3_ver = _disp_cfg["q3"].upper()
    q4_ver = _disp_cfg["q4"].upper()
    if isinstance(data, dict) and _is_two_half_game(data):
        return "Predicciones:\nFormato 2 mitades — sin Q3/Q4"

    def _v10_line(label: str, v10_target_pred: dict, quarter: str) -> str:
        if not v10_target_pred.get("available"):
            reason = str(v10_target_pred.get("reason") or "unavailable")
            if "Missing Q3" in reason or "missing_q3" in reason.lower():
                return f"{label}: Esperando datos Q3"
            if "Missing Q1/Q2" in reason or "missing_q1" in reason.lower():
                return f"{label}: Esperando datos Q1|Q2"
            return f"{label}: sin datos"
        def _sv(v):
            if v is None:
                return None
            try:
                return float(v)
            except (TypeError, ValueError):
                return None
        pts = _sv(v10_target_pred.get("predicted_total")) or 0
        pts_home = _sv(v10_target_pred.get("predicted_home"))
        pts_away = _sv(v10_target_pred.get("predicted_away"))
        mae = _sv(v10_target_pred.get("mae"))
        mae_home = _sv(v10_target_pred.get("mae_home"))
        mae_away = _sv(v10_target_pred.get("mae_away"))
        mae_txt = f" (±{mae:.0f})" if mae is not None else ""
        outcome_suffix = ""
        if isinstance(data, dict):
            q_sc = ((data.get("score", {}).get("quarters", {}) or {}).get(quarter) or {})
            q_home = _safe_int(q_sc.get("home"))
            q_away = _safe_int(q_sc.get("away"))
            if q_home is not None and q_away is not None:
                is_done = (
                    (quarter == "Q3" and (q4_started or status_type == "finished"))
                    or (quarter == "Q4" and (status_type == "finished" or (minute_est is not None and minute_est >= 48)))
                )
                if is_done:
                    actual = q_home + q_away
                    if actual > pts:
                        outcome_suffix = f" ✅ ({actual} pts)"
                    elif actual == pts:
                        outcome_suffix = f" 🔁 Push ({actual} pts)"
                    else:
                        outcome_suffix = f" ❌ ({actual} pts)"
        total_line = f"📈{label}: O ~{pts:.1f}{mae_txt} | APOSTAR{outcome_suffix}"
        # Home / away reference lines
        ref_parts = []
        if pts_home is not None:
            mae_h_txt = f" (±{mae_home:.1f})" if mae_home is not None else ""
            ref_parts.append(f"🏠 ~{pts_home:.1f}{mae_h_txt}")
        if pts_away is not None:
            mae_a_txt = f" (±{mae_away:.1f})" if mae_away is not None else ""
            ref_parts.append(f"🛫 ~{pts_away:.1f}{mae_a_txt}")
        if ref_parts:
            return total_line + "\n" + " | ".join(ref_parts)
        return total_line

    lines = [
        "Predicciones:",
        _line(f"Q3 {q3_ver}", rendered_row.get("q3", {})),
        "",
        _line(f"Q4 {q4_ver}", rendered_row.get("q4", {}), waiting_preverdict=q4_waiting_preverdict),
    ]
    return "\n".join(lines)


def _build_graph_image(match_id: str, data: dict) -> Path:
    tmp_dir = Path(tempfile.gettempdir()) / "pulpa_match_graphs"
    tmp_dir.mkdir(parents=True, exist_ok=True)
    out_path = tmp_dir / f"graph_{match_id}.png"
    ml_mod.plot_graph(data, str(out_path))
    return out_path


async def _send_match_graph(
    app: Application,
    chat_id: int,
    match_id: str,
    data: dict,
) -> None:
    try:
        image_path = _build_graph_image(match_id, data)
    except Exception as exc:
        await app.bot.send_message(
            chat_id=chat_id,
            text=(
                f"No pude generar grafica para {match_id}: {exc}"
            ),
        )
        return

    try:
        with image_path.open("rb") as f:
            await app.bot.send_photo(
                chat_id=chat_id,
                photo=f,
                caption=f"Grafica de presion match {match_id}",
            )
    finally:
        try:
            image_path.unlink()
        except OSError:
            pass


def _get_v10_predictions(match_id: str) -> dict | None:
    """Run v10 Over/Under regression inference synchronously. Returns None on error."""
    try:
        infer_mod = importlib.import_module("training.infer_match")
        result = infer_mod.run_inference_v10(match_id)
        if result.get("ok"):
            return result.get("predictions", {})
        return None
    except Exception:
        return None


_BET_SIGNALS_SET = {"BET", "BET_HOME", "BET_AWAY"}
_NOBET_SIGNALS_SET = {"NO_BET", "NO BET", "LEAN", "UNAVAILABLE", "ERROR", "window_missed"}


def _monitor_full_section(
    match_id: str,
    home_team: str,
    away_team: str,
    current_pred_row: dict | None,
) -> str:
    """Full-detail monitor snapshot block built from inference_debug_log.

    Shows the most-recent monitor evaluation per quarter (BET or NO_BET).
    For NO_BET entries the gate reason (effective threshold) is displayed so
    a raw confidence like 65% with a raised threshold of 70% is clearly
    explained instead of looking inconsistent.
    """
    conn = _open_conn()
    try:
        # Latest entry PER TARGET regardless of signal — we want the last
        # thing the monitor decided for each quarter, BET or NO_BET.
        debug_rows = conn.execute(
            """
            SELECT target, model, scraped_minute, signal, confidence, inference_json
            FROM inference_debug_log
            WHERE match_id = ?
            ORDER BY id DESC
            """,
            (str(match_id),),
        ).fetchall()
        # Fallback: bet_monitor_log for matches where inference_debug_log is empty
        all_rows = conn.execute(
            """
            SELECT target, signal, pick, confidence, scraped_minute, recommendation
            FROM bet_monitor_log
            WHERE match_id = ?
            ORDER BY id DESC
            """,
            (str(match_id),),
        ).fetchall()
    except Exception:
        debug_rows = []
        all_rows = []
    conn.close()

    def _sf(v):
        if v is None:
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    # Build latest-per-target from inference_debug_log
    debug_quarters: dict[str, dict] = {}
    seen: set[str] = set()
    for r in debug_rows:
        t = str(r["target"] or "").lower()
        if t not in ("q3", "q4") or t in seen:
            continue
        seen.add(t)
        try:
            infer = json.loads(r["inference_json"])
            src = (infer.get("predictions") or {}).get(t) or {}
        except Exception:
            src = {}
        sig = str(r["signal"] or src.get("final_recommendation") or "NO_BET").upper().strip()
        debug_quarters[t] = {
            "model": str(r["model"] or "v13").upper(),
            "minute": r["scraped_minute"],
            "signal": sig,
            "confidence": _sf(src.get("confidence")) or _sf(r["confidence"]),
            "pick": str(src.get("predicted_winner") or src.get("winner_pick") or "").lower(),
            # gate reason — explains why NO_BET even when raw conf looks high
            "reasoning": str(src.get("reasoning") or src.get("reason") or "").strip(),
            "pts_home": src.get("predicted_home"),
            "pts_away": src.get("predicted_away"),
            "pts_total": src.get("predicted_total"),
            "mae_home": src.get("mae_home"),
            "mae_away": src.get("mae_away"),
            "mae": src.get("mae"),
            "model_quality": src.get("model_quality"),
            "model_samples": src.get("model_samples"),
            "model_gap": src.get("model_gap"),
            "fallback_used": src.get("fallback_used"),
        }

    # Fallback latest-per-target from bet_monitor_log
    latest_log: dict[str, dict] = {}
    for r in all_rows:
        t = str(r["target"] or "").lower()
        if t in ("q3", "q4") and t not in latest_log:
            latest_log[t] = {
                "signal": str(r["signal"] or "").strip().upper(),
                "pick": str(r["pick"] or "").strip().lower(),
                "confidence": _sf(r["confidence"]),
                "scraped_minute": r["scraped_minute"],
                "recommendation": str(r["recommendation"] or "").strip(),
            }

    if not debug_quarters and not latest_log:
        return ""

    # Prefer debug_quarters; fall back to latest_log when debug_quarters is empty
    if not debug_quarters:
        lines: list[str] = ["📡 Señal monitor (última):"]
        for t in ("q3", "q4"):
            entry = latest_log.get(t)
            if not entry:
                continue
            sig = entry["signal"]
            pick_raw = entry["pick"]
            conf = entry["confidence"]
            minute = entry["scraped_minute"]
            reason = entry.get("recommendation", "")
            pick_name = (
                f"🏠 {home_team}" if pick_raw == "home"
                else f"🛫 {away_team}" if pick_raw == "away"
                else (pick_raw or "-")
            )
            sig_emoji = "🟢" if sig in _BET_SIGNALS_SET else "🔴"
            sig_txt = "APUESTA" if sig in _BET_SIGNALS_SET else "NO APOSTAR"
            conf_txt = f" · Score bruto {round((conf or 0) * 100):.0f}%" if conf is not None else ""
            min_txt = f" · min {minute}" if minute is not None else ""
            lines.append(f"{sig_emoji} {t.upper()}: {sig_txt} {pick_name}{conf_txt}{min_txt}")
            if reason and sig not in _BET_SIGNALS_SET:
                lines.append(f"  ↳ {reason}")
        return "\n".join(lines) if len(lines) > 1 else ""

    # --- Full rendering using inference_debug_log data ---
    minutes_set = {v["minute"] for v in debug_quarters.values() if v.get("minute")}
    min_str = f" (min {', '.join(str(m) for m in sorted(minutes_set))})" if minutes_set else ""
    lines = [f"📡 Señal monitor{min_str}:"]

    for t in ("q3", "q4"):
        q = debug_quarters.get(t)
        if q is None:
            # Quarter not in debug_quarters — try latest_log one-liner
            entry = latest_log.get(t)
            if entry:
                sig = entry["signal"]
                pick_raw = entry["pick"]
                conf = entry["confidence"]
                minute = entry["scraped_minute"]
                reason = entry.get("recommendation", "")
                pick_name = (
                    f"🏠 {home_team}" if pick_raw == "home"
                    else f"🛫 {away_team}" if pick_raw == "away"
                    else (pick_raw or "-")
                )
                sig_emoji = "🟢" if sig in _BET_SIGNALS_SET else "🔴"
                sig_txt = "APUESTA" if sig in _BET_SIGNALS_SET else "NO APOSTAR"
                conf_fmt = f" · Score bruto {round((conf or 0) * 100):.0f}%" if conf is not None else ""
                min_fmt = f" · min {minute}" if minute is not None else ""
                lines.append(f"{sig_emoji} {t.upper()}: {sig_txt} {pick_name}{conf_fmt}{min_fmt}")
                if reason and sig not in _BET_SIGNALS_SET:
                    lines.append(f"  ↳ {reason}")
            continue

        sig = q["signal"]
        is_bet = sig in _BET_SIGNALS_SET
        pick_raw = str(q.get("pick") or "").lower()
        pick_side = "🏠" if pick_raw == "home" else ("🛫" if pick_raw == "away" else "")
        pick_name = (
            home_team if pick_raw == "home"
            else away_team if pick_raw == "away"
            else pick_raw or "-"
        )
        pick_txt = f"{pick_side} {pick_name}".strip() if pick_side else pick_name
        conf = _sf(q.get("confidence"))
        conf_txt = f"Score bruto {round(conf * 100):.0f}%" if conf is not None else ""
        sig_emoji = "🟢" if is_bet else "🔴"
        sig_txt = "APUESTA" if is_bet else "NO APOSTAR"

        lines.append(f"{sig_emoji} {t.upper()} {q['model']}: {sig_txt} {pick_txt}")

        # Gate reason for NO_BET — critical for understanding why conf 65% ≠ BET
        reasoning = str(q.get("reasoning") or "").strip()
        if not is_bet and reasoning:
            lines.append(f"  ↳ {reasoning}")

        # Model quality warnings (relevant for BET and informative for NO_BET)
        mq = q.get("model_quality")
        ms = q.get("model_samples")
        mg = _sf(q.get("model_gap"))
        fb = q.get("fallback_used")
        if mq == "low":
            lines.append(f"🚨 Datos insuficientes ({ms or 0} partidos)")
        elif mq == "moderate":
            lines.append(f"⚠️ Datos reducidos ({ms or 0} partidos)")
        if fb:
            lines.append("ℹ️ Ritmo no clasificado — modelo alternativo")
        if mg is not None and mg > 0.20:
            lines.append(f"⚠️ Sobreajuste (gap={mg:.2f})")

        if is_bet:
            # Score projections (only shown for BET entries)
            ph = _sf(q.get("pts_home"))
            pa = _sf(q.get("pts_away"))
            pt = _sf(q.get("pts_total"))
            mah = _sf(q.get("mae_home"))
            maa = _sf(q.get("mae_away"))
            mae = _sf(q.get("mae"))
            if ph is not None and pa is not None:
                mah_txt = f" (±{mah:.1f})" if mah is not None else ""
                maa_txt = f" (±{maa:.1f})" if maa is not None else ""
                mae_txt = f" (±{mae:.0f})" if mae is not None else ""
                proj = f"Proyeccion: 🏠~{ph:.1f}{mah_txt} | ✈️~{pa:.1f}{maa_txt} | Total~{pt:.1f}{mae_txt}"
                if mah is not None and maa is not None and abs(ph - pa) < (mah + maa):
                    proj += f"\n⚠️ MAE se superpone (dif={abs(ph - pa):.1f} < {mah + maa:.1f})"
                lines.append(proj)

            if conf_txt:
                lines.append(conf_txt)

            # Mismatch warning vs current calculation
            if current_pred_row is not None:
                cur = current_pred_row.get(t)
                if isinstance(cur, dict) and cur.get("available"):
                    cur_sig = str(cur.get("signal") or cur.get("bet_signal") or "").upper()
                    if cur_sig not in _BET_SIGNALS_SET:
                        cur_conf = _sf(cur.get("confidence"))
                        cur_conf_txt = f"{round(cur_conf * 100):.0f}%" if cur_conf is not None else "?"
                        lines.append(f"⚠️ Cálculo actual: NO APOSTAR {cur_conf_txt} — datos del scraper cambiaron desde la alerta")

        lines.append("")  # blank separator between quarters

    # Remove trailing blanks
    while lines and lines[-1] == "":
        lines.pop()

    return "\n".join(lines) if len(lines) > 1 else ""


def _detail_text(match_id: str, data: dict, pred_row: dict | None, following: bool = False) -> str:
    v10_preds = _get_v10_predictions(match_id)
    detail = _match_detail_text(match_id, data) + "\n\n" + _prediction_text(
        pred_row,
        data,
        v10_preds=v10_preds,
    )
    m_info = (data or {}).get("match", {}) if data else {}
    home = str(m_info.get("home_team") or "home")
    away = str(m_info.get("away_team") or "away")
    mon_section = _monitor_full_section(match_id, home, away, pred_row)
    if mon_section:
        detail += "\n\n" + mon_section
    if following:
        detail = "🔴 LIVE SEGUIMIENTO\n\n" + detail
    return detail


def _refresh_waiting_text(match_id: str, data: dict | None) -> str:
    if not data:
        return "Actualizando..."

    m = data.get("match", {})
    utc_date = str(m.get("date", "") or "")
    utc_time = str(m.get("time", "") or "")
    local_dt = _to_local_datetime(utc_date, utc_time)
    if local_dt is not None:
        fecha_line = f"Fecha: {local_dt.strftime('%Y-%m-%d %H:%M')} UTC{UTC_OFFSET_HOURS:+d}"
    else:
        fecha_line = f"Fecha: {utc_date} {utc_time} UTC"

    minute_est = None
    graph_points = data.get("graph_points", [])
    if graph_points:
        try:
            minute_est = int(graph_points[-1].get("minute", 0))
        except (TypeError, ValueError):
            minute_est = None
    minute_suffix = f" | Min: {minute_est}" if minute_est is not None else ""

    text = "\n".join(
        [
            f"Match ID: {match_id}{minute_suffix}",
            f"{m.get('home_team', '')} vs {m.get('away_team', '')}",
            fecha_line,
            f"Liga: {m.get('league', '')}",
            "",
            "Actualizando...",
        ]
    )
    if len(text) > 1000:
        return text[:997] + "..."
    return text


async def _set_waiting_state(update: Update, text: str = "Espere...") -> None:
    query = update.callback_query
    if not query:
        return

    message = query.message
    try:
        if message and getattr(message, "photo", None):
            await query.edit_message_caption(caption=text)
            return

        await query.edit_message_text(text=text)
    except Exception as exc:
        if _is_message_not_modified_error(exc):
            print("[callback] waiting-state edit skipped (message not modified)", flush=True)
            return
        if isinstance(exc, (TimedOut, NetworkError)):
            print(
                f"[callback] waiting-state edit timeout/network error: {type(exc).__name__}",
                flush=True,
            )
            return
        if isinstance(exc, BadRequest):
            text_exc = str(exc)
            if (
                "Message to edit not found" in text_exc
                or "message can't be edited" in text_exc
                or "chat not found" in text_exc
            ):
                print(
                    f"[callback] waiting-state edit skipped (bad request): {text_exc}",
                    flush=True,
                )
                return
        raise


async def _on_error(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    err = context.error
    # BadRequest is a subclass of NetworkError in this version — check it FIRST
    if isinstance(err, BadRequest):
        # Log callback data + full message to help diagnose which handler triggered it
        update_data = ""
        if hasattr(update, "callback_query") and getattr(update, "callback_query", None):
            cb = update.callback_query  # type: ignore[union-attr]
            update_data = f" | callback_data={getattr(cb, 'data', '?')!r}"
        elif hasattr(update, "message") and getattr(update, "message", None):
            msg = update.message  # type: ignore[union-attr]
            update_data = f" | text={getattr(msg, 'text', '?')!r}"
        logger.warning("[telegram] BadRequest%s: %s", update_data, err)
        return
    if isinstance(err, (TimedOut, NetworkError)):
        logger.warning("[telegram] network timeout error: %s", type(err).__name__)
        return
    # Improved error logging with full traceback and update context
    update_info = "unknown"
    try:
        if hasattr(update, "callback_query") and getattr(update, "callback_query", None):
            cb = update.callback_query  # type: ignore[union-attr]
            update_info = f"callback_data={getattr(cb, 'data', '?')!r}"
        elif hasattr(update, "message") and getattr(update, "message", None):
            msg = update.message  # type: ignore[union-attr]
            update_info = f"message_text={getattr(msg, 'text', '?')!r}"
    except Exception:
        pass
    
    logger.error(
        "[telegram] unhandled exception [%s] %s: %s",
        type(err).__name__,
        update_info,
        str(err),
        exc_info=True
    )


async def _replace_callback_message(
    update: Update,
    text: str,
    reply_markup: InlineKeyboardMarkup,
    parse_mode: str | None = None,
) -> None:
    query = update.callback_query
    if not query:
        return

    message = query.message
    # Callbacks sobre mensajes "inaccesibles" (p. ej. antiguos): no se puede editar el texto.
    if message is not None and not isinstance(message, Message):
        chat_id = update.effective_chat.id if update.effective_chat else None
        if chat_id is not None:
            await query.get_bot().send_message(
                chat_id=chat_id,
                text=text,
                reply_markup=reply_markup,
                parse_mode=parse_mode,
            )
        return

    if message and getattr(message, "photo", None):
        chat_id = update.effective_chat.id if update.effective_chat else None
        if chat_id is None:
            return
        await query.get_bot().send_message(
            chat_id=chat_id,
            text=text,
            reply_markup=reply_markup,
            parse_mode=parse_mode,
        )
        try:
            await message.delete()
        except Exception:
            pass
        return

    try:
        await query.edit_message_text(
            text=text,
            reply_markup=reply_markup,
            parse_mode=parse_mode,
        )
    except Exception as _exc:
        _exc_text = str(_exc).lower()
        if "message is not modified" in _exc_text or "message_not_modified" in _exc_text:
            pass  # same content, ignore silently
        elif "message to edit not found" in _exc_text or "chat not found" in _exc_text:
            pass  # message already gone, ignore
        elif "inaccessible message" in _exc_text or "cannot edit" in _exc_text:
            chat_id = update.effective_chat.id if update.effective_chat else None
            if chat_id is not None:
                await query.get_bot().send_message(
                    chat_id=chat_id,
                    text=text,
                    reply_markup=reply_markup,
                    parse_mode=parse_mode,
                )
        else:
            raise


async def _send_detail_message(
    update: Update,
    app: Application,
    match_id: str,
    data: dict,
    pred_row: dict | None,
    event_date: str | None,
    page: int,
) -> None:
    chat_id = update.effective_chat.id if update.effective_chat else None
    if chat_id is None:
        return

    detail_text = _detail_text(match_id, data, pred_row)
    keyboard = _detail_keyboard(
        match_id,
        event_date=event_date,
        page=page,
        match_data=data,
        chat_id=chat_id,
    )

    image_path: Path | None = None
    try:
        image_path = _build_graph_image(match_id, data)
    except Exception:
        image_path = None

    if image_path is not None and image_path.exists():
        try:
            with image_path.open("rb") as f:
                await app.bot.send_photo(
                    chat_id=chat_id,
                    photo=f,
                    caption=detail_text,
                    reply_markup=keyboard,
                )
        finally:
            try:
                image_path.unlink()
            except OSError:
                pass
        return

    await app.bot.send_message(
        chat_id=chat_id,
        text=detail_text,
        reply_markup=keyboard,
    )


def _detail_token(event_date: str | None) -> str:
    return event_date if event_date else "_"


def _normalize_sofascore_slug(value: str | None) -> str:
    text = (value or "").strip().lower()
    if not text:
        return "unknown"
    normalized = unicodedata.normalize("NFKD", text)
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    ascii_text = re.sub(r"[^a-z0-9]+", "-", ascii_text).strip("-")
    return ascii_text or "unknown"


def _sofascore_match_url(match_id: str, match_data: dict | None = None) -> str:
    event_slug = "unknown"
    custom_id = ""
    home_slug = "unknown"
    away_slug = "unknown"

    if match_data and "match" in match_data:
        match_info = match_data["match"]
        event_slug = _normalize_sofascore_slug(match_info.get("event_slug"))
        custom_id = str(match_info.get("custom_id") or "").strip()
        home_slug = _normalize_sofascore_slug(match_info.get("home_slug"))
        away_slug = _normalize_sofascore_slug(match_info.get("away_slug"))
        if home_slug == "unknown":
            home_slug = _normalize_sofascore_slug(match_info.get("home_team"))
        if away_slug == "unknown":
            away_slug = _normalize_sofascore_slug(match_info.get("away_team"))

    if event_slug != "unknown" and custom_id:
        return (
            "https://www.sofascore.com/basketball/match/"
            f"{event_slug}/{custom_id}#id:{match_id}"
        )

    return (
        "https://www.sofascore.com/basketball/match/"
        f"{home_slug}/{away_slug}#id:{match_id}"
    )


def _detail_keyboard(
    match_id: str,
    event_date: str | None = None,
    page: int = 0,
    match_data: dict | None = None,
    chat_id: int | None = None,
) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    token = _detail_token(event_date)
    rows.append(
        [
            InlineKeyboardButton(
                "Ver SofaScore",
                url=_sofascore_match_url(match_id, match_data),
            )
        ]
    )
    rows.append(
        [
            InlineKeyboardButton(
                "🔄 Refresh datos + pred",
                callback_data=f"refresh:all:{match_id}:{token}:{page}",
            ),
            InlineKeyboardButton(
                "🔮 Refresh pred",
                callback_data=f"refresh:pred:{match_id}:{token}:{page}",
            ),
        ]
    )
    _ftb = _follow_toggle_button(chat_id, match_id, token, page, match_data)
    if _ftb is not None:
        rows.append([_ftb])
    rows.append(
        [
            InlineKeyboardButton(
                "🗑 Borrar match de la base",
                callback_data=f"delmatch:confirm:{match_id}:{token}:{page}",
            )
        ]
    )
    rows.append([InlineKeyboardButton(
        f"🤖 Modelos (Q3={MODEL_CONFIG['q3']} Q4={MODEL_CONFIG['q4']})",
        callback_data=f"matchmodel:open:{match_id}:{token}:{page}",
    )])
    rows.append([InlineKeyboardButton(
        "📈 V12 LIVE Bookmaker",
        callback_data=f"v12live:open:{match_id}:{token}:{page}",
    )])
    rows.append([InlineKeyboardButton(
        "📊 Historial inferencias monitor",
        callback_data=f"debuginf:{match_id}:{token}:{page}",
    )])
    if event_date:
        rows.append(
            [
                InlineKeyboardButton(
                    "⬅️ Volver al dia",
                    callback_data=f"date:{event_date}:{page}",
                )
            ]
        )
    rows.append([InlineKeyboardButton("🏠 Menu principal", callback_data="nav:main")])
    return InlineKeyboardMarkup(rows)


async def _render_main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    context.user_data[AWAITING_MATCH_ID_KEY] = False
    context.user_data[AWAITING_FETCH_DATE_KEY] = False
    chat_id = update.effective_chat.id if update.effective_chat else None
    job = DATE_INGEST_JOBS.get(chat_id) if chat_id is not None else None
    text = "Menu principal\n\nSelecciona una opcion:"
    if job:
        text += (
            "\n\nDescarga en curso en segundo plano. "
            "Usa 'Status descarga' para ver el avance."
        )

    if update.callback_query:
        await _replace_callback_message(
            update,
            text=text,
            reply_markup=_main_menu_keyboard(chat_id),
        )
    elif update.message:
        await update.message.reply_text(
            text=text,
            reply_markup=_main_menu_keyboard(chat_id),
        )


async def start_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await _render_main_menu(update, context)


async def _render_dates(update: Update, page: int) -> None:
    rows = await asyncio.to_thread(_fetch_date_summaries)
    if not rows:
        await _replace_callback_message(
            update,
            text="No hay matches en la base.",
            reply_markup=InlineKeyboardMarkup(
                [[InlineKeyboardButton("Menu principal", callback_data="nav:main")]]
            ),
        )
        return

    pred_stats = await asyncio.to_thread(_fetch_dates_pred_stats)
    await _replace_callback_message(
        update,
        text="Fechas disponibles:",
        reply_markup=_dates_keyboard(rows, page, pred_stats),
    )


async def _render_matches_for_date(update: Update, event_date: str, page: int) -> None:
    rows = _fetch_matches_for_date(event_date)
    # Filter excluded leagues (same list as signals/reports)
    rows = [r for r in rows if not any(p in (r.get("league") or "") for p in _MONTHLY_EXCL_PATTERNS)]
    if not rows:
        await _replace_callback_message(
            update,
            text=f"No hay matches para {event_date}.",
            reply_markup=InlineKeyboardMarkup(
                [
                    [InlineKeyboardButton("Volver a fechas", callback_data="dates:0")],
                    [InlineKeyboardButton("Menu principal", callback_data="nav:main")],
                ]
            ),
        )
        return

    pred_map = _fetch_date_pred_outcomes(event_date)
    stats = _pred_stats_text(pred_map, len(rows), match_rows=rows)
    model_line = f"Modelo: Q3={MODEL_CONFIG['q3']} Q4={MODEL_CONFIG['q4']}"
    header = f"{model_line}\n{_event_date_title_es(event_date, len(rows))}"
    if stats:
        _stats_block = f"\n<pre>{html.escape(stats)}</pre>"
        # Telegram message limit is 4096 chars; leave room for the keyboard header
        _MAX = 3900
        if len(header) + len(_stats_block) > _MAX:
            # Truncate stats to fit — keep at least the summary table (first block)
            _first_block_end = stats.find("\n\nPor liga:")
            if _first_block_end != -1:
                _stats_block = f"\n<pre>{html.escape(stats[:_first_block_end])}</pre>"
            # Still too long? Drop stats entirely
            if len(header) + len(_stats_block) > _MAX:
                _stats_block = ""
        header += _stats_block
    await _replace_callback_message(
        update,
        text=header,
        reply_markup=_matches_keyboard(rows, event_date, page, pred_map),
        parse_mode="HTML",
    )


async def _render_inference_debug(
    update: Update,
    match_id: str,
    event_date: str | None,
    page: int,
) -> None:
    """Show inference_debug_log entries for a specific match."""
    import json as _json

    conn = _open_conn()
    try:
        debug_rows = conn.execute(
            "SELECT * FROM inference_debug_log WHERE match_id=? ORDER BY created_at ASC",
            (match_id,),
        ).fetchall()
        # Get team names from bet_monitor_log or matches table
        match_meta = conn.execute(
            "SELECT home_team, away_team, league FROM bet_monitor_log WHERE match_id=? LIMIT 1",
            (match_id,),
        ).fetchone()
        if not match_meta:
            mm = conn.execute(
                "SELECT home_team, away_team FROM matches WHERE match_id=? LIMIT 1",
                (match_id,),
            ).fetchone()
            match_meta = mm
    except Exception:
        debug_rows = []
        match_meta = None
    conn.close()

    token = event_date if event_date else "_"
    back_kb = InlineKeyboardMarkup([[
        InlineKeyboardButton("⬅️ Volver al match", callback_data=f"match:{match_id}:{token}:{page}"),
    ]])

    if not debug_rows:
        await _replace_callback_message(
            update,
            text=(
                f"📊 Historial de inferencias del monitor\n"
                f"Match: {match_id}\n\n"
                "Sin registros en inference_debug_log.\n"
                "El monitor debe haber procesado este partido para generar datos."
            ),
            reply_markup=back_kb,
        )
        return

    home = away = "?"
    league_name = "?"
    if match_meta:
        home = match_meta["home_team"] or "?"
        away = match_meta["away_team"] or "?"
        league_name = (match_meta["league"] if "league" in match_meta.keys() else None) or "?"

    # Find league quality from the most recent inference_json
    _q_icons = {"strong": "✅", "moderate": "🟡", "weak": "❌", "unknown": "⚪"}
    match_lq: str | None = None
    match_lb: bool | None = None
    for _r in reversed(debug_rows):
        try:
            _j = _json.loads(_r["inference_json"] or "{}")
            _tgt_j = ((_j.get("predictions") or {}).get(_r["target"] or "q3") or {})
            if "league_quality" in _tgt_j:
                match_lq = _tgt_j["league_quality"]
                match_lb = bool(_tgt_j.get("league_bettable", False))
                break
        except Exception:
            pass

    # Group by target
    by_target: dict[str, list] = {}
    for r in debug_rows:
        by_target.setdefault(r["target"] or "?", []).append(r)

    lines: list[str] = [
        f"📊 Historial inferencias monitor",
        f"Match: {match_id}",
        f"{home} vs {away}",
        f"Liga: {league_name}",
    ]
    if match_lq:
        _q_icon = _q_icons.get(match_lq, "⚪")
        _bet_note = "  ⚠️ no en entrenamiento" if not match_lb else ""
        lines.append(f"Calidad liga: {match_lq} {_q_icon}{_bet_note}")
    lines.append("")

    def _sig_icon(sig: str) -> str:
        s = (sig or "").upper()
        if s == "BET":
            return "🟢"
        if s == "UNAVAILABLE":
            return "⚪"
        return "🔴"

    for target in ["q3", "q4"]:
        rows_t = by_target.get(target, [])
        if not rows_t:
            continue
        signals = [r["signal"] or "?" for r in rows_t]
        unique_sigs = list(dict.fromkeys(signals))
        flip_tag = "  ⚠️ FLIP" if len(unique_sigs) > 1 else ""
        lines.append(f"── {target.upper()} ({len(rows_t)} checks){flip_tag}")

        for r in rows_t:
            sig = r["signal"] or "?"
            icon = _sig_icon(sig)
            conf = r["confidence"]
            conf_txt = f" {float(conf) * 100:.0f}%" if conf else ""
            gp = r["gp_count"]
            gp_txt = f" gp={gp}" if gp is not None else ""
            mn = r["scraped_minute"]
            mn_txt = f" min={mn}" if mn is not None else ""
            ts = (r["created_at"] or "?")[11:19]

            # Parse debug fields from inference JSON
            reason = ""
            extra_tag = ""
            pred_line = ""
            try:
                pred_json = _json.loads(r["inference_json"] or "{}")
                pred_t = (pred_json.get("predictions") or {}).get(target, {}) or {}
                reason = str(pred_t.get("reasoning") or pred_t.get("reason") or "")[:90]
                dq = str(pred_t.get("data_quality") or "")[:4]
                vx = pred_t.get("volatility_index")
                lq = pred_t.get("league_quality")
                lb = pred_t.get("league_bettable")
                extra_parts = []
                if dq:
                    extra_parts.append(f"cal:{dq}")
                if vx is not None:
                    extra_parts.append(f"vol:{float(vx):.2f}")
                if lq and lb is not None:
                    lq_icon = _q_icons.get(lq, "⚪")
                    extra_parts.append(f"liga:{lq}{lq_icon}" + ("❌" if not lb else ""))
                if extra_parts:
                    extra_tag = "\n    " + " | ".join(extra_parts)
                pt = pred_t.get("predicted_total")
                ph = pred_t.get("predicted_home")
                pa = pred_t.get("predicted_away")
                mae = pred_t.get("mae")
                if pt is not None:
                    pred_line = f"    pred: {float(pt):.1f}"
                    if ph is not None and pa is not None:
                        pred_line += f" ({float(ph):.1f}/{float(pa):.1f})"
                    if mae is not None:
                        pred_line += f"  ±{float(mae):.1f}MAE"
            except Exception:
                pass

            line = f"  [{ts}]{mn_txt}{gp_txt}  {icon} {sig}{conf_txt}"
            if extra_tag:
                line += extra_tag
            lines.append(line)
            if pred_line:
                lines.append(pred_line)
            if reason:
                lines.append(f"    {reason}")

    text = "\n".join(lines)
    # Telegram message limit
    if len(text) > 4000:
        text = text[:3960] + "\n…(truncado)"

    await _replace_callback_message(
        update,
        text=text,
        reply_markup=back_kb,
    )


async def _render_match_detail(
    update: Update,
    app: Application,
    match_id: str,
    event_date: str | None,
    page: int,
    *,
    pred_row: dict | None = None,
    send_graph: bool = True,
) -> None:
    chat_id = update.effective_chat.id if update.effective_chat else None
    await _set_waiting_state(update)
    data = _get_match_detail(match_id)
    if not data:
        await _replace_callback_message(
            update,
            text=f"No se encontro el match {match_id}.",
            reply_markup=_detail_keyboard(
                match_id,
                event_date=event_date,
                page=page,
                chat_id=chat_id,
            ),
        )
        return

    if pred_row is None:
        pred_row = _get_or_compute_predictions(match_id, data)
    await _send_detail_message(
        update,
        app,
        match_id,
        data,
        pred_row,
        event_date,
        page,
    )
    if update.callback_query and update.callback_query.message:
        try:
            await update.callback_query.message.delete()
        except Exception:
            pass


async def _refresh_detail(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    mode: str,
    match_id: str,
    event_date: str | None,
    page: int,
) -> None:
    query = update.callback_query
    if not query:
        return
    chat_id = update.effective_chat.id if update.effective_chat else None

    logger.info(f"[REFRESH_DETAIL] Callback for mode={mode} match={match_id} event_date={event_date}")
    cached = _get_match_detail(match_id)
    await _set_waiting_state(update, text=_refresh_waiting_text(match_id, cached))

    if mode == "all":
        logger.info(f"[REFRESH_DETAIL] Mode='all': refreshing both data and predictions")
        try:
            logger.info(f"[REFRESH_DETAIL] Running _refresh_match_data in thread")
            data = await asyncio.to_thread(_refresh_match_data, match_id)
            if data is None:
                logger.error(f"[REFRESH_DETAIL] Data refresh returned None for {match_id}")
                await _replace_callback_message(
                    update,
                    text=f"No pude refrescar datos de {match_id}: datos nulos del scraper",
                    reply_markup=_detail_keyboard(
                        match_id,
                        event_date=event_date,
                        page=page,
                        chat_id=chat_id,
                    ),
                )
                return
            logger.info(f"[REFRESH_DETAIL] Data refresh succeeded, now computing predictions")
        except Exception as exc:
            logger.error(f"[REFRESH_DETAIL] Exception in _refresh_match_data: {exc}", exc_info=True)
            await _replace_callback_message(
                update,
                text=f"No pude refrescar datos de {match_id}: {exc}",
                reply_markup=_detail_keyboard(
                    match_id,
                    event_date=event_date,
                    page=page,
                    chat_id=chat_id,
                ),
            )
            return
        pred_row = _refresh_predictions(match_id, data)
        logger.info(f"[REFRESH_DETAIL] Rendering match detail after refresh")
        await _render_match_detail(
            update,
            context.application,
            match_id,
            event_date,
            page,
            pred_row=pred_row,
            send_graph=True,
        )
        return

    logger.info(f"[REFRESH_DETAIL] Mode='pred': refreshing predictions only")
    logger.info(f"[REFRESH_DETAIL] Loading from cache for {match_id}")
    data = _get_match_detail(match_id)
    if not data:
        logger.error(f"[REFRESH_DETAIL] No cached data found for {match_id}")
        await _replace_callback_message(
            update,
            text=f"No se encontro el match {match_id}.",
            reply_markup=_detail_keyboard(
                match_id,
                event_date=event_date,
                page=page,
                chat_id=chat_id,
            ),
        )
        return

    logger.info(f"[REFRESH_DETAIL] Computing predictions from cached data")
    pred_row = _refresh_predictions(match_id, data)
    await _render_match_detail(
        update,
        context.application,
        match_id,
        event_date,
        page,
        pred_row=pred_row,
        send_graph=True,
    )


def _tail_lines(text: str, max_lines: int = 16, max_chars: int = 3000) -> str:
    lines = [line for line in text.splitlines() if line.strip()]
    tail = "\n".join(lines[-max_lines:])
    if len(tail) > max_chars:
        tail = tail[-max_chars:]
    return tail or "(sin salida)"


def _safe_pct_from_csv(value: object) -> str:
    try:
        return f"{float(value) * 100:.1f}%"
    except (TypeError, ValueError):
        return "n/a"


def _read_train_v4_effectiveness_summary() -> str:
    lines: list[str] = []
    for target in ("q3", "q4"):
        csv_path = MODEL_OUTPUTS_V4_DIR / f"{target}_metrics.csv"
        if not csv_path.exists():
            lines.append(f"- {target.upper()}: metrics no encontrado")
            continue

        rows: list[dict[str, str]] = []
        with csv_path.open("r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(row)

        if not rows:
            lines.append(f"- {target.upper()}: sin filas de metrics")
            continue

        best = max(rows, key=lambda r: float(r.get("f1") or 0.0))
        ensemble = next(
            (r for r in rows if str(r.get("model") or "") == "ensemble_avg_prob"),
            None,
        )
        best_txt = (
            f"best_f1 model={best.get('model')} "
            f"f1={_safe_pct_from_csv(best.get('f1'))} "
            f"acc={_safe_pct_from_csv(best.get('accuracy'))}"
        )
        if ensemble is not None:
            ens_txt = (
                f"ensemble f1={_safe_pct_from_csv(ensemble.get('f1'))} "
                f"acc={_safe_pct_from_csv(ensemble.get('accuracy'))}"
            )
            lines.append(f"- {target.upper()}: {best_txt} | {ens_txt}")
        else:
            lines.append(f"- {target.upper()}: {best_txt}")

    return "\n".join(lines) if lines else "(sin metrics)"


def _clear_eval_results_table() -> int:
    conn = _open_conn()
    count_row = conn.execute("SELECT COUNT(*) AS n FROM eval_match_results").fetchone()
    before = int(count_row["n"] if count_row else 0)
    conn.execute("DELETE FROM eval_match_results")
    conn.commit()
    conn.close()
    return before


def _delete_match_cascade(match_id: str) -> dict[str, int]:
    conn = _open_conn()

    def _del(sql: str, params: tuple[object, ...]) -> int:
        cur = conn.execute(sql, params)
        affected = int(cur.rowcount or 0)
        return affected if affected >= 0 else 0

    counts = {
        "eval_match_results": _del(
            "DELETE FROM eval_match_results WHERE match_id = ?",
            (match_id,),
        ),
        "discovered_ft_matches": _del(
            "DELETE FROM discovered_ft_matches WHERE match_id = ?",
            (match_id,),
        ),
        "play_by_play": _del(
            "DELETE FROM play_by_play WHERE match_id = ?",
            (match_id,),
        ),
        "graph_points": _del(
            "DELETE FROM graph_points WHERE match_id = ?",
            (match_id,),
        ),
        "quarter_scores": _del(
            "DELETE FROM quarter_scores WHERE match_id = ?",
            (match_id,),
        ),
        "matches": _del(
            "DELETE FROM matches WHERE match_id = ?",
            (match_id,),
        ),
    }
    conn.commit()
    conn.close()
    return counts


def _list_candidate_match_ids_for_universe() -> list[str]:
    conn = _open_conn()
    rows = conn.execute(
        """
        SELECT match_id
        FROM matches
        ORDER BY date DESC, time DESC, match_id DESC
        """
    ).fetchall()
    conn.close()
    return [str(r["match_id"]) for r in rows if r["match_id"] is not None]


def _build_universe_stats_text() -> str:
    conn = _open_conn()
    try:
        total = conn.execute("SELECT COUNT(*) AS n FROM matches").fetchone()["n"]
        finished = conn.execute(
            "SELECT COUNT(*) AS n FROM matches WHERE status_type='finished'"
        ).fetchone()["n"]
        with_q4 = conn.execute(
            """
            SELECT COUNT(DISTINCT match_id) AS n FROM quarter_scores
            WHERE quarter='Q4'
            """
        ).fetchone()["n"]
        full_qs = conn.execute(
            """
            SELECT COUNT(*) AS n FROM (
                SELECT match_id
                FROM quarter_scores
                WHERE quarter IN ('Q1','Q2','Q3','Q4') AND home IS NOT NULL AND away IS NOT NULL
                GROUP BY match_id
                HAVING COUNT(DISTINCT quarter) = 4
            )
            """
        ).fetchone()["n"]
        with_pbp = conn.execute(
            "SELECT COUNT(DISTINCT match_id) AS n FROM play_by_play"
        ).fetchone()["n"]
        with_gp = conn.execute(
            "SELECT COUNT(DISTINCT match_id) AS n FROM graph_points"
        ).fetchone()["n"]
        eval_rows = conn.execute(
            "SELECT COUNT(*) AS n FROM eval_match_results"
        ).fetchone()["n"]
        date_range = conn.execute(
            "SELECT MIN(date) AS d_min, MAX(date) AS d_max FROM matches"
        ).fetchone()
        d_min = date_range["d_min"] or "?"
        d_max = date_range["d_max"] or "?"
    finally:
        conn.close()

    lines = [
        "Universo DB:",
        f"  Total matches: {total}",
        f"  Status finished: {finished}",
        f"  Con Q1-Q4 completo: {full_qs}",
        f"  Con play-by-play: {with_pbp}",
        f"  Con graph_points: {with_gp}",
        f"  Eval calculados: {eval_rows}",
        f"  Rango fechas: {d_min} a {d_max}",
    ]
    return "\n".join(lines)


def _build_model_stats_text() -> str:
    def _table(headers: list[str], rows: list[list[str]]) -> list[str]:
        widths = [len(h) for h in headers]
        for row in rows:
            for i, cell in enumerate(row):
                widths[i] = max(widths[i], len(cell))
        head = " | ".join(h.ljust(widths[i]) for i, h in enumerate(headers))
        sep = "-+-".join("-" * w for w in widths)
        out = [head, sep]
        for row in rows:
            out.append(" | ".join(cell.ljust(widths[i]) for i, cell in enumerate(row)))
        return out

    active_q3 = MODEL_CONFIG.get("q3", "v4")
    active_q4 = MODEL_CONFIG.get("q4", "v4")

    # All version dirs in order; only include those present on disk
    all_version_dirs = [
        ("v2", MODEL_OUTPUTS_V2_DIR),
        ("v4", MODEL_OUTPUTS_V4_DIR),
        ("v6", MODEL_OUTPUTS_V6_DIR),
        ("v6_1", MODEL_OUTPUTS_V6_1_DIR),
        ("v9", MODEL_OUTPUTS_V9_DIR),
        ("v13", MODEL_OUTPUTS_V13_DIR),
        ("v15", MODEL_OUTPUTS_V15_DIR),
        ("v16", MODEL_OUTPUTS_V16_DIR),
        ("v17", MODEL_OUTPUTS_V17_DIR),
    ]
    version_dirs = [
        (v, d) for v, d in all_version_dirs if d.exists()
    ]

    metrics_sections: list[str] = []
    metrics_sections.append(
        f"🏆 Modelos Activos: Q3={active_q3}  Q4={active_q4}"
    )
    metrics_sections.append("")

    headers = ["Filtro", "F1", "ACC", "ROC", "Ntrain", "Ntest", "Ntotal"]
    for version_name, version_dir in version_dirs:
        q3_marker = " ✅" if version_name == active_q3 else ""
        q4_marker = " ✅" if version_name == active_q4 else ""
        metrics_sections.append(f"Modelos {version_name}:")
        for target in ("q3", "q4"):
            marker = q3_marker if target == "q3" else q4_marker
            csv_path = version_dir / f"{target}_metrics.csv"
            if not csv_path.exists():
                metrics_sections.append(f"{target.upper()}: metrics no encontrado")
                continue

            parsed_rows: list[dict[str, str]] = []
            with csv_path.open("r", encoding="utf-8", newline="") as f:
                for row in csv.DictReader(f):
                    if row.get("split") and row.get("split") != "test":
                        continue
                    parsed_rows.append(row)
            if not parsed_rows:
                metrics_sections.append(f"{target.upper()}: sin filas")
                continue

            table_rows: list[list[str]] = []
            for row in parsed_rows:
                table_rows.append(
                    [
                        str(row.get("model", "?")),
                        _safe_pct_from_csv(row.get("f1")),
                        _safe_pct_from_csv(row.get("accuracy")),
                        _safe_pct_from_csv(row.get("roc_auc")),
                        str(row.get("samples_train", "?")),
                        str(row.get("samples_test", "?")),
                        str(row.get("samples_total", "?")),
                    ]
                )

            metrics_sections.append(f"{target.upper()}{marker} (test):")
            metrics_sections.extend(_table(headers, table_rows))
            metrics_sections.append("")

    # Gate config
    gate_lines: list[str] = []
    gate_path = BASE_DIR / "training" / "model_outputs_v2" / "gate_config.json"
    if gate_path.exists():
        try:
            with gate_path.open("r", encoding="utf-8") as f:
                gcfg = json.load(f)
            gen_at = str(gcfg.get("generated_at_utc", "?"))[:19]
            thresholds = gcfg.get("thresholds", {})
            q3_thr = (thresholds.get("q3") or {}).get("default") or {}
            q4_thr = (thresholds.get("q4") or {}).get("36") or {}
            gate_lines.append(f"Gate calibrado: {gen_at}")
            if q3_thr:
                gate_lines.append(
                    f"  Q3: edge>={_safe_pct_from_csv(q3_thr.get('min_edge'))}"
                    f" vol<{q3_thr.get('volatility_block_at')}"
                    f" gp>={q3_thr.get('min_graph_points')}"
                    f" pbp>={q3_thr.get('min_pbp_events')}"
                )
            if q4_thr:
                gate_lines.append(
                    f"  Q4: edge>={_safe_pct_from_csv(q4_thr.get('min_edge'))}"
                    f" vol<{q4_thr.get('volatility_block_at')}"
                    f" gp>={q4_thr.get('min_graph_points')}"
                    f" pbp>={q4_thr.get('min_pbp_events')}"
                )
        except Exception:
            gate_lines.append("Gate config: error leyendo archivo")
    else:
        gate_lines.append("Gate config: no encontrado")

    # ── v12 classification stats (from training_summary.json) ──────────────
    v12_summary = MODEL_OUTPUTS_V12_DIR / "training_summary.json"
    if v12_summary.exists():
        try:
            with v12_summary.open("r", encoding="utf-8") as f:
                s12 = json.load(f)
            q12m = " ✅" if active_q3 == "v12" or active_q4 == "v12" else ""
            metrics_sections.append(f"Modelos v12 (clasificacion){q12m}:")
            clf12 = s12.get("classification", {})
            hdrs12 = ["Modelo", "F1", "ACC", "ROC", "Ntest"]
            for tgt in ("q3", "q4"):
                mk = " ✅" if (tgt == "q3" and active_q3 == "v12") or (tgt == "q4" and active_q4 == "v12") else ""
                entries = clf12.get(tgt, [])
                if not entries:
                    continue
                rows12 = []
                for e in entries:
                    rows12.append([
                        str(e.get("model", "?")),
                        f"{float(e['f1'])*100:.1f}%" if e.get("f1") is not None else "?",
                        f"{float(e['accuracy'])*100:.1f}%" if e.get("accuracy") is not None else "?",
                        f"{float(e['roc_auc'])*100:.1f}%" if e.get("roc_auc") is not None else "?",
                        str(e.get("n_test", "?")),
                    ])
                metrics_sections.append(f"{tgt.upper()}{mk} (val):")
                metrics_sections.extend(_table(hdrs12, rows12))
                metrics_sections.append("")
        except Exception:
            metrics_sections.append("v12: error leyendo training_summary.json")



    # ── v13 stats (from training_summary.json, models_trained) ────────────
    v13_summary = MODEL_OUTPUTS_V13_DIR / "training_summary.json"
    if v13_summary.exists():
        try:
            with v13_summary.open("r", encoding="utf-8") as f:
                s13 = json.load(f)
            q13m = " ✅" if active_q3 == "v13" or active_q4 == "v13" else ""
            trained_at = str(s13.get("trained_at", "?"))[:19]
            metrics_sections.append(f"Modelos v13{q13m} ({trained_at}):")
            hdrs13 = ["Clave", "F1val", "ACCval", "Gap", "Nval"]
            rows13 = []
            for m in s13.get("models_trained", []):
                key = str(m.get("key", "?"))
                tgt = key.split("_")[0] if "_" in key else "?"
                tgt_mk = ""
                if tgt == "q3" and active_q3 == "v13":
                    tgt_mk = "✅"
                elif tgt == "q4" and active_q4 == "v13":
                    tgt_mk = "✅"
                f1v = m.get("val_f1")
                accv = m.get("val_accuracy")
                gap = m.get("train_val_gap")
                rows13.append([
                    f"{tgt_mk}{key}",
                    f"{float(f1v)*100:.1f}%" if f1v is not None else "?",
                    f"{float(accv)*100:.1f}%" if accv is not None else "?",
                    f"{float(gap)*100:.1f}%" if gap is not None else "?",
                    str(m.get("samples_val", "?")),
                ])
            metrics_sections.extend(_table(hdrs13, rows13))
            metrics_sections.append("")
        except Exception:
            metrics_sections.append("v13: error leyendo training_summary.json")
            metrics_sections.append("")

    # ── v15 stats (from training_summary_v15.json, per-league) ────────────
    v15_summary = MODEL_OUTPUTS_V15_DIR / "training_summary_v15.json"
    if v15_summary.exists():
        try:
            with v15_summary.open("r", encoding="utf-8") as f:
                s15 = json.load(f)
            q15m = " ✅" if active_q3 == "v15" or active_q4 == "v15" else ""
            n_trained = s15.get("n_leagues_trained", "?")
            n_skip = s15.get("n_leagues_skipped", "?")
            metrics_sections.append(f"Modelos v15{q15m} ({n_trained} ligas, {n_skip} omitidas):")
            hdrs15 = ["Liga", "Tgt", "F1val", "ACCval", "Ntrain", "Nval"]
            rows15 = []
            for m in s15.get("models", []):
                league = str(m.get("league", "?"))[:20]
                tgt = str(m.get("target", "?"))
                vm = m.get("val_metrics") or {}
                f1v = vm.get("f1")
                accv = vm.get("accuracy")
                tgt_mk = ""
                if tgt == "q3" and active_q3 == "v15":
                    tgt_mk = "✅"
                elif tgt == "q4" and active_q4 == "v15":
                    tgt_mk = "✅"
                rows15.append([
                    league,
                    f"{tgt_mk}{tgt}",
                    f"{float(f1v)*100:.1f}%" if f1v is not None else "?",
                    f"{float(accv)*100:.1f}%" if accv is not None else "?",
                    str(m.get("n_train", "?")),
                    str(m.get("n_val", "?")),
                ])
            metrics_sections.extend(_table(hdrs15, rows15))
            metrics_sections.append("")
        except Exception:
            metrics_sections.append("v15: error leyendo training_summary_v15.json")
            metrics_sections.append("")

    # ── v16 stats (from training_summary_v16.json, per-league) ────────────
    v16_summary = MODEL_OUTPUTS_V16_DIR / "training_summary_v16.json"
    if v16_summary.exists():
        try:
            with v16_summary.open("r", encoding="utf-8") as f:
                s16 = json.load(f)
            q16m = " ✅" if active_q3 == "v16" or active_q4 == "v16" else ""
            n_trained = s16.get("n_leagues_trained", "?")
            n_skip = s16.get("n_leagues_skipped", "?")
            metrics_sections.append(f"Modelos v16{q16m} ({n_trained} ligas, {n_skip} omitidas):")
            hdrs16 = ["Liga", "Tgt", "F1val", "ACCval", "Ntrain", "Nval"]
            rows16 = []
            for m in s16.get("models", []):
                league = str(m.get("league", "?"))[:20]
                tgt = str(m.get("target", "?"))
                vm = m.get("val_metrics") or {}
                f1v = vm.get("f1")
                accv = vm.get("accuracy")
                tgt_mk = ""
                if tgt == "q3" and active_q3 == "v16":
                    tgt_mk = "✅"
                elif tgt == "q4" and active_q4 == "v16":
                    tgt_mk = "✅"
                rows16.append([
                    league,
                    f"{tgt_mk}{tgt}",
                    f"{float(f1v)*100:.1f}%" if f1v is not None else "?",
                    f"{float(accv)*100:.1f}%" if accv is not None else "?",
                    str(m.get("n_train", "?")),
                    str(m.get("n_val", "?")),
                ])
            metrics_sections.extend(_table(hdrs16, rows16))
            metrics_sections.append("")
        except Exception:
            metrics_sections.append("v16: error leyendo training_summary_v16.json")
            metrics_sections.append("")

    # ── v17 stats (from training_summary_v17.json, per-league) ────────────
    v17_summary = MODEL_OUTPUTS_V17_DIR / "training_summary_v17.json"
    if v17_summary.exists():
        try:
            with v17_summary.open("r", encoding="utf-8") as f:
                s17 = json.load(f)
            q17m = " ✅" if active_q3 == "v17" or active_q4 == "v17" else ""
            n_trained = s17.get("n_leagues_trained", "?")
            n_skip = s17.get("n_leagues_skipped", "?")
            metrics_sections.append(f"Modelos v17{q17m} ({n_trained} ligas, {n_skip} omitidas):")
            hdrs17 = ["Liga", "Tgt", "F1val", "ACCval", "Ntrain", "Nval"]
            rows17 = []
            for m in s17.get("models", []):
                league = str(m.get("league", "?"))[:20]
                tgt = str(m.get("target", "?"))
                vm = m.get("val_metrics") or {}
                f1v = vm.get("f1")
                accv = vm.get("accuracy")
                tgt_mk = ""
                if tgt == "q3" and active_q3 == "v17":
                    tgt_mk = "✅"
                elif tgt == "q4" and active_q4 == "v17":
                    tgt_mk = "✅"
                rows17.append([
                    league,
                    f"{tgt_mk}{tgt}",
                    f"{float(f1v)*100:.1f}%" if f1v is not None else "?",
                    f"{float(accv)*100:.1f}%" if accv is not None else "?",
                    str(m.get("n_train", "?")),
                    str(m.get("n_val", "?")),
                ])
            metrics_sections.extend(_table(hdrs17, rows17))
            metrics_sections.append("")
        except Exception:
            metrics_sections.append("v17: error leyendo training_summary_v17.json")
            metrics_sections.append("")

    sections = metrics_sections + gate_lines
    return "\n".join(sections)


# ─── Monthly report ───────────────────────────────────────────────────────────

_MONTHLY_EXCL_PATTERNS = [
    "WNBA", "Women", "women", "Feminina", "Femenina",
    "Playoff", "PLAY OFF", "U21 Espoirs Elite", "Liga Femenina",
    "LF Challenge", "Polish Basketball League",
    "SuperSport Premijer Liga", "Prvenstvo Hrvatske za d",
    "ABA Liga", "Argentina Liga Nacional", "Basketligaen",
    "lite 2", "EYBL", "I B MCKL", "Liga 1 Masculin",
    "Liga Nationala", "NBL1", "PBA Commissioner",
    "Rapid League", "Stoiximan GBL", "Playout",
    "Superleague", "Superliga", "Swedish Basketball Superettan",
    "Swiss Cup", "Финал", "Turkish Basketball Super League",
    "NBA",
    "Big V", "Egyptian Basketball Super League", "Lega A Basket",
    "Liga e Par", "Liga Ouro", "Señal", "LNB",
    "Meridianbet KLS", "MPBL", "Nationale 1", "Poland 2nd Basketball League",
    "Portugal LBP", "Portugal Proliga", "Saku I liiga", "Serie A2",
    "Slovenian Second Basketball", "Super League", "United Cup", "United League",
]
_BET_SIGNALS_MONTHLY = {"BET", "BET HOME", "BET_HOME", "BET AWAY", "BET_AWAY"}


def _build_monthly_excel_bytes(year_month: str, quarters: list[str] | None = None) -> bytes:
    """Build monthly performance Excel from eval_match_results (normal review model).

    Uses the same data source as the per-date stats view (_fetch_date_pred_outcomes),
    so the model version is always the currently configured MODEL_CONFIG, not the
    mixed versions stored in bet_monitor_log.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    if quarters is None:
        quarters = ["q3", "q4"]
    quarters_lower = [q.lower() for q in quarters]

    conn = _open_conn()
    _stake = float(db_mod.get_setting(conn, "sig_bet_size", "100") or 100)
    _odds  = float(db_mod.get_setting(conn, "sig_odds",     "1.4")  or 1.4)
    _bank  = float(db_mod.get_setting(conn, "sig_bank",     "1000") or 1000)
    # Distinct local dates for the month (using UTC_OFFSET_HOURS so date boundaries are correct)
    date_rows = conn.execute(
        f"""
        SELECT DISTINCT date(datetime(date || ' ' || time, '{UTC_OFFSET_HOURS} hours')) AS local_date
        FROM matches
        WHERE local_date LIKE ?
        ORDER BY local_date
        """,
        (f"{year_month}-%",),
    ).fetchall()
    conn.close()
    all_dates = [r["local_date"] for r in date_rows if r["local_date"]]

    def _excl(league: str) -> bool:
        return any(p in league for p in _MONTHLY_EXCL_PATTERNS)

    q3_model = MODEL_CONFIG.get("q3", "v?")
    q4_model = MODEL_CONFIG.get("q4", "v?")

    rows_q3: list[dict] = []
    rows_q4: list[dict] = []
    day_stats: dict[str, dict] = {}

    for date in all_dates:
        match_rows = _fetch_matches_for_date(date)
        pred_map = _fetch_date_pred_outcomes(date)
        match_lookup = {str(r["match_id"]): r for r in match_rows}

        for mid, pred in pred_map.items():
            row = match_lookup.get(mid, {})
            league = str(row.get("league", "") or "").strip()
            if _excl(league):
                continue

            for quarter in quarters_lower:
                if not pred.get(f"{quarter}_available"):
                    continue
                signal = str(pred.get(f"{quarter}_signal") or "").strip().upper().replace("_", " ")
                if signal not in _BET_SIGNALS_MONTHLY:
                    continue
                conf = float(pred.get(f"{quarter}_confidence") or 0.0)
                if 0.0 < conf <= 0.30:
                    continue

                pick = str(pred.get(f"{quarter}_pick") or "")
                model_name = q3_model if quarter == "q3" else q4_model

                # Apply model-specific filter
                _stake_factor = 1.0
                if model_name == "v6":
                    _accept, _stake_factor = bet_monitor_mod._v6_pick_filter(league, conf, pick)
                    if not _accept:
                        continue
                if model_name == "v2":
                    _accept, _stake_factor = bet_monitor_mod._v2_pick_filter(league, conf, pick)
                    if not _accept:
                        continue

                # Resolve outcome from stored field (for past months this is already determined)
                raw_outcome = str(pred.get(f"{quarter}_outcome") or "pending").lower()
                _eff_stake = _stake * _stake_factor
                if raw_outcome == "hit":
                    result_str, profit = "win", _eff_stake * (_odds - 1)
                elif raw_outcome in ("miss", "push"):  # push = loss
                    result_str, profit = "loss", -_eff_stake
                else:
                    result_str, profit = "pending", 0.0

                ts = row.get("scheduled_utc_ts")
                time_local = str(row.get("display_time") or "")

                bet_row = {
                    "match_id": mid,
                    "event_date": date,
                    "time_local": time_local,
                    "home": str(row.get("home_team", "?") or "?"),
                    "away": str(row.get("away_team", "?") or "?"),
                    "league": league,
                    "pick": pick,
                    "confidence": conf if conf > 0 else None,
                    "result": result_str,
                    "profit": profit,
                    "scheduled_utc_ts": ts,
                    "model": model_name,
                    "quarter": quarter,
                    "event_slug": str(row.get("event_slug") or ""),
                    "custom_id": str(row.get("custom_id") or ""),
                    "home_slug": str(row.get("home_slug") or ""),
                    "away_slug": str(row.get("away_slug") or ""),
                }
                if quarter == "q3":
                    rows_q3.append(bet_row)
                else:
                    rows_q4.append(bet_row)

    # ── Build day_stats derived directly from collected rows ────────────────
    day_stats: dict[str, dict] = {}
    for _r in rows_q3 + rows_q4:
        _d = _r["event_date"]
        _q = _r["quarter"]
        if _d not in day_stats:
            day_stats[_d] = {
                "q3": {"w": 0, "l": 0, "p": 0, "profit": 0.0},
                "q4": {"w": 0, "l": 0, "p": 0, "profit": 0.0},
            }
        _ds = day_stats[_d][_q]
        if _r["result"] == "win":
            _ds["w"] += 1
            _ds["profit"] += _r["profit"]
        elif _r["result"] == "loss":
            _ds["l"] += 1
            _ds["profit"] += _r["profit"]
        elif _r["result"] == "pending":
            _ds["p"] += 1

    # ── Excel styles ─────────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    bdr = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    ca = Alignment(horizontal="center", vertical="center")
    la = Alignment(horizontal="left", vertical="center")
    ra = Alignment(horizontal="right")
    green_f = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_f   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gray_f  = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    sum_hdr = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    tot_f   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    BET_HEADERS = [
        "Nro", "Match ID", "Fecha", "Hora (UTC-6)",
        "Home", "Away", "Liga", "Pick", "Modelo", "Confianza",
        "Resultado", "Bank Antes", "Ganancia", "Bank Despues",
    ]

    def _add_bet_sheet(ws, rows: list[dict]):
        for col, h in enumerate(BET_HEADERS, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = h
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = ca
            cell.border = bdr

        current_bank = _bank
        for i, r in enumerate(rows, 1):
            profit = r["profit"]
            if r["result"] == "win":
                outcome_txt, rf = "✅", green_f
            elif r["result"] == "pending":
                outcome_txt, rf = "⏳", gray_f
            else:
                outcome_txt, rf = "❌", red_f

            ts = r.get("scheduled_utc_ts")
            time_local = r.get("time_local", "")
            if time_local:
                date_str, time_str = r["event_date"], time_local
            elif ts:
                try:
                    dt = datetime.fromtimestamp(int(ts), tz=timezone.utc) + timedelta(hours=UTC_OFFSET_HOURS)
                    date_str, time_str = dt.strftime("%Y-%m-%d"), dt.strftime("%H:%M")
                except Exception:
                    date_str, time_str = r["event_date"], ""
            else:
                date_str, time_str = r["event_date"], ""

            pick = r["pick"]
            pick_sym = "🏠" if pick == "home" else ("✈️" if pick == "away" else "")
            bank_before = current_bank
            bank_after  = current_bank + profit
            current_bank = bank_after

            _ev_slug  = r.get("event_slug") or ""
            _cust_id  = r.get("custom_id") or ""
            _h_slug   = r.get("home_slug") or _normalize_sofascore_slug(r.get("home", ""))
            _a_slug   = r.get("away_slug") or _normalize_sofascore_slug(r.get("away", ""))
            _fake_match_data = {"match": {
                "event_slug": _ev_slug,
                "custom_id":  _cust_id,
                "home_slug":  _h_slug,
                "away_slug":  _a_slug,
            }}
            sofascore_url = _sofascore_match_url(r["match_id"], _fake_match_data)
            rn = i + 1
            ws.cell(rn, 1).value = i
            mc = ws.cell(rn, 2)
            mc.value = r["match_id"]
            mc.hyperlink = sofascore_url
            mc.font = Font(color="0563C1", underline="single")
            ws.cell(rn, 3).value = date_str
            ws.cell(rn, 4).value = time_str
            ws.cell(rn, 5).value = r["home"]
            ws.cell(rn, 6).value = r["away"]
            ws.cell(rn, 7).value = r["league"]
            ws.cell(rn, 8).value = f"{pick_sym} {pick}".strip()
            ws.cell(rn, 9).value = r["model"]
            conf = r.get("confidence")
            ws.cell(rn, 10).value = conf
            ws.cell(rn, 11).value = outcome_txt
            ws.cell(rn, 12).value = bank_before
            ws.cell(rn, 13).value = profit
            ws.cell(rn, 14).value = bank_after

            for col in range(1, 15):
                cell = ws.cell(rn, col)
                cell.border = bdr
                if col == 11:
                    cell.fill = rf
                    cell.alignment = ca
                elif col in (1, 3, 4, 8, 9, 10, 11):
                    cell.alignment = ca
                elif col in (5, 6):
                    cell.alignment = la
                if col == 10 and conf is not None:
                    cell.number_format = "0.0%"
                    cell.alignment = ca
                if col in (12, 13, 14):
                    cell.number_format = "$#,##0.00"
                    cell.alignment = ra

        for col_num in range(1, 15):
            cl = get_column_letter(col_num)
            mw = max((len(str(c.value)) for c in ws[cl] if c.value), default=8)
            ws.column_dimensions[cl].width = min(mw + 2, 50)

    def _sort_key(r: dict):
        # Use local time string "HH:MM" for within-day ordering.
        # scheduled_utc_ts is not available in the match_lookup dict, but
        # display_time (stored as time_local) is a reliable "HH:MM" string
        # that sorts lexicographically == chronologically.
        return (r["event_date"], r.get("time_local") or "99:99", r["match_id"])

    rows_q3.sort(key=_sort_key)
    rows_q4.sort(key=_sort_key)

    if rows_q3:
        ws_q3 = wb.create_sheet(f"Q3 ({q3_model})", 0)
        _add_bet_sheet(ws_q3, rows_q3)
    if rows_q4:
        ws_q4 = wb.create_sheet(f"Q4 ({q4_model})", 1 if rows_q3 else 0)
        _add_bet_sheet(ws_q4, rows_q4)

    # ── Resumen sheet ─────────────────────────────────────────────────────────
    ws_res = wb.create_sheet("Resumen", 0)
    _show_q3 = bool(rows_q3)
    _show_q4 = bool(rows_q4)
    _show_both = _show_q3 and _show_q4
    res_hdrs = ["Fecha"]
    if _show_q3:
        res_hdrs += ["Q3 W", "Q3 L", "Q3 ⏳", "Q3 Ef%", "Q3 ROI", "Q3 Ganancia", "Q3 Bank"]
    if _show_q4:
        res_hdrs += ["Q4 W", "Q4 L", "Q4 ⏳", "Q4 Ef%", "Q4 ROI", "Q4 Ganancia", "Q4 Bank"]
    if _show_both:
        res_hdrs += ["Total W", "Total L", "Total ⏳", "Total Ef%", "Total Ganancia", "Bank Final"]
    # Derive format sets dynamically from header names
    _PCT_COLS  = {i for i, h in enumerate(res_hdrs, 1) if "Ef%" in h or "ROI" in h}
    _MON_COLS  = {i for i, h in enumerate(res_hdrs, 1) if "Ganancia" in h or "Bank" in h}
    _GAIN_COLS = {i for i, h in enumerate(res_hdrs, 1) if "Ganancia" in h}

    for c, h in enumerate(res_hdrs, 1):
        cell = ws_res.cell(1, c)
        cell.value = h
        cell.fill = sum_hdr
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = ca
        cell.border = bdr

    running_bank    = _bank
    running_bank_q3 = _bank
    running_bank_q4 = _bank
    tot_w = tot_l = tot_pend = 0
    tot_q3w = tot_q3l = tot_q3pend = 0
    tot_q4w = tot_q4l = tot_q4pend = 0
    tot_profit = tot_q3p = tot_q4p = 0.0

    def _roi_val(w, l):
        played = w + l
        if played == 0:
            return None
        return (w * _stake * (_odds - 1) - l * _stake) / (played * _stake)

    def _ef_val(w, l):
        played = w + l
        if played == 0:
            return None
        return w / played

    for row_i, day in enumerate(sorted(day_stats.keys()), 2):
        ds = day_stats[day]
        q3w, q3l, q3pend, q3p = ds["q3"]["w"], ds["q3"]["l"], ds["q3"]["p"], ds["q3"]["profit"]
        q4w, q4l, q4pend, q4p = ds["q4"]["w"], ds["q4"]["l"], ds["q4"]["p"], ds["q4"]["profit"]
        dw  = q3w  + q4w
        dl  = q3l  + q4l
        dpend = q3pend + q4pend
        dp  = q3p  + q4p

        running_bank_q3 += q3p
        running_bank_q4 += q4p
        running_bank    += dp

        tot_q3w += q3w; tot_q3l += q3l; tot_q3pend += q3pend; tot_q3p += q3p
        tot_q4w += q4w; tot_q4l += q4l; tot_q4pend += q4pend; tot_q4p += q4p
        tot_w   += dw;  tot_l   += dl;  tot_pend   += dpend;  tot_profit += dp

        vals = [day]
        if _show_q3:
            vals += [q3w, q3l, q3pend or None, _ef_val(q3w, q3l), _roi_val(q3w, q3l), q3p, running_bank_q3]
        if _show_q4:
            vals += [q4w, q4l, q4pend or None, _ef_val(q4w, q4l), _roi_val(q4w, q4l), q4p, running_bank_q4]
        if _show_both:
            vals += [dw, dl, dpend or None, _ef_val(dw, dl), dp, running_bank]
        for c, v in enumerate(vals, 1):
            cell = ws_res.cell(row_i, c)
            cell.value = v
            cell.border = bdr
            cell.alignment = ca
            if c in _PCT_COLS and v is not None:
                cell.number_format = "0.0%"
            elif c in _MON_COLS:
                cell.number_format = "$#,##0.00"
                cell.alignment = ra
            if c in _GAIN_COLS:
                cell.fill = green_f if (v or 0) >= 0 else red_f

    # Totals row
    tot_row = len(day_stats) + 2
    tot_vals = ["TOTAL"]
    if _show_q3:
        tot_vals += [tot_q3w, tot_q3l, tot_q3pend or None,
                     _ef_val(tot_q3w, tot_q3l), _roi_val(tot_q3w, tot_q3l), tot_q3p, _bank + tot_q3p]
    if _show_q4:
        tot_vals += [tot_q4w, tot_q4l, tot_q4pend or None,
                     _ef_val(tot_q4w, tot_q4l), _roi_val(tot_q4w, tot_q4l), tot_q4p, _bank + tot_q4p]
    if _show_both:
        tot_vals += [tot_w, tot_l, tot_pend or None,
                     _ef_val(tot_w, tot_l), tot_profit, _bank + tot_profit]
    for c, v in enumerate(tot_vals, 1):
        cell = ws_res.cell(tot_row, c)
        cell.value = v
        cell.fill = tot_f
        cell.font = Font(bold=True)
        cell.border = bdr
        cell.alignment = ca
        if c in _PCT_COLS and v is not None:
            cell.number_format = "0.0%"
        elif c in _MON_COLS:
            cell.number_format = "$#,##0.00"
            cell.alignment = ra

    for col_num in range(1, len(res_hdrs) + 1):
        cl = get_column_letter(col_num)
        mw = max((len(str(c.value)) for c in ws_res[cl] if c.value), default=8)
        ws_res.column_dimensions[cl].width = min(mw + 2, 30)

    if not rows_q3 and not rows_q4:
        ws_res.cell(2, 1).value = "Sin apuestas para el mes"

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


async def _recalculate_month_evals(year_month: str, cid: int, msg, quarters: list[str]) -> None:
    """Re-run inference for every match in the month using current MODEL_CONFIG.

    All heavy work runs in a single background thread so the bot event loop
    stays free to handle other commands while the recalculation is in progress.
    """
    conn = _open_conn()
    match_ids = [
        r["match_id"]
        for r in conn.execute(
            "SELECT match_id FROM matches WHERE date LIKE ? ORDER BY date",
            (f"{year_month}-%",),
        ).fetchall()
    ]
    conn.close()

    total = len(match_ids)
    if total == 0:
        await msg.edit_text(f"❌ Sin partidos encontrados para {year_month}")
        return

    loop = asyncio.get_event_loop()
    _config_snapshot = dict(MODEL_CONFIG)

    def _run_all() -> tuple[int, int]:
        done = errors = 0
        for mid in match_ids:
            try:
                conn2 = _open_conn()
                data = db_mod.get_match(conn2, mid)
                conn2.close()
                if data:
                    _compute_and_store_predictions(mid, data, _config_snapshot)
            except Exception as exc:
                errors += 1
                logger.warning("Recalc %s: %s", mid, exc)
            done += 1
            if done % 25 == 0 or done == total:
                _text = (
                    f"🔄 Recalculando {year_month}...\n"
                    f"{done}/{total} partidos  ({errors} errores)"
                )
                asyncio.run_coroutine_threadsafe(
                    _safe_edit(msg, _text), loop
                )
        return done, errors

    async def _safe_edit(m, text: str) -> None:
        try:
            await m.edit_text(text)
        except Exception:
            pass

    await asyncio.to_thread(_run_all)


async def _send_monthly_report(update: Update, context: ContextTypes.DEFAULT_TYPE, year_month: str) -> None:
    """Show confirmation dialog: use existing data vs recalculate first."""
    cid = update.effective_chat.id if update.effective_chat else None
    q = update.callback_query
    if q:
        try:
            await q.answer()
        except Exception:
            pass

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(
            "📊 Datos actuales",
            callback_data=f"monthly:{year_month}:current",
        )],
        [InlineKeyboardButton(
            "🔄 Recalcular primero (lento)",
            callback_data=f"monthly:{year_month}:recalc",
        )],
        [InlineKeyboardButton("⬅️ Cancelar", callback_data="nav:main")],
    ])
    txt = (
        f"📅 Reporte Mensual: {year_month}\n\n"
        f"Modelo activo: Q3=[{MODEL_CONFIG.get('q3','v?')}]  "
        f"Q4=[{MODEL_CONFIG.get('q4','v?')}]\n\n"
        "• Datos actuales: genera inmediatamente con lo que ya "
        "está en eval_match_results.\n"
        "• Recalcular primero: re-evalúa todos los partidos del "
        "mes con el modelo activo y luego genera el reporte (puede "
        "tardar varios minutos)."
    )
    if q:
        try:
            await q.edit_message_text(txt, reply_markup=kb)
        except Exception:
            await context.bot.send_message(chat_id=cid, text=txt, reply_markup=kb)
    else:
        await context.bot.send_message(chat_id=cid, text=txt, reply_markup=kb)


async def _send_stats_report(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Unificado: genera y envia las estadisticas con manejo de longitud para Telegram."""
    try:
        model_txt = await asyncio.to_thread(_build_model_stats_text)
        universe_txt = await asyncio.to_thread(_build_universe_stats_text)
        stats_txt = model_txt + "\n\n" + universe_txt
    except Exception as exc:
        logger.error("Error building stats: %s", exc, exc_info=True)
        stats_txt = f"Error al leer estadisticas: {exc}"

    MAX_CHUNK = 3800
    escaped = html.escape(stats_txt)
    chunks: list[str] = []
    while escaped:
        chunk = escaped[:MAX_CHUNK]
        if len(escaped) > MAX_CHUNK:
            nl = chunk.rfind("\n")
            if nl > 0:
                chunk = escaped[:nl]
        chunks.append(f"<pre>{chunk}</pre>")
        escaped = escaped[len(chunk):].lstrip("\n")

    message = update.effective_message
    if not message:
        return

    # Si es un callback, actualizamos el mensaje original con la primera parte
    if update.callback_query:
        query = update.callback_query
        try:
            # Edit current message with first part
            await query.edit_message_text(
                text=chunks[0],
                parse_mode="HTML",
                reply_markup=_train_submenu_keyboard()
            )
            # Send following parts as new messages
            for part in chunks[1:]:
                await context.bot.send_message(
                    chat_id=message.chat_id,
                    text=part,
                    parse_mode="HTML"
                )
        except Exception as e:
            if "message is not modified" not in str(e).lower():
                await context.bot.send_message(chat_id=message.chat_id, text=chunks[0], parse_mode="HTML")
    else:
        # Si es un comando o boton de teclado
        for idx, part in enumerate(chunks):
            # Only the LAST chunk gets the reply keyboard to keep the chat usable
            kb = _menu_reply_keyboard() if idx == len(chunks) - 1 else None
            await context.bot.send_message(
                chat_id=message.chat_id,
                text=part,
                parse_mode="HTML",
                reply_markup=kb
            )


async def modelstats_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not _is_allowed(update):
        return
    await _send_stats_report(update, context)




def _train_status_text() -> str:
    running = bool(TRAIN_STATUS.get("running"))
    owner = TRAIN_STATUS.get("owner_chat_id")
    last_exit = TRAIN_STATUS.get("last_exit_code")
    last_finished = TRAIN_STATUS.get("last_finished_utc")
    last_error = str(TRAIN_STATUS.get("last_error_tail") or "")

    if running:
        task = str(TRAIN_STATUS.get("current_task") or "train-v4")
        done = int(TRAIN_STATUS.get("progress_done") or 0)
        total = int(TRAIN_STATUS.get("progress_total") or 0)
        progress_txt = f"\nprogreso={done}/{total}" if total > 0 else ""
        return (
            "Estado entrenamiento: RUNNING\n"
            f"owner_chat_id={owner}\n"
            f"tarea={task}{progress_txt}"
        )

    if last_exit is None:
        return "Estado entrenamiento: IDLE (sin ejecuciones previas)"

    base = (
        "Estado entrenamiento: IDLE\n"
        f"ultimo_exit={last_exit}\n"
        f"ultimo_fin_utc={last_finished}"
    )
    if last_exit != 0 and last_error:
        return base + "\n" + "ultimo_error_tail:\n" + last_error
    return base


def _is_allowed(update: Update) -> bool:
    chat_id = update.effective_chat.id if update.effective_chat else None
    if not ALLOWED_CHAT_IDS:
        return True
    return chat_id in ALLOWED_CHAT_IDS


def _is_train_allowed(chat_id: int | None) -> bool:
    if not ALLOWED_CHAT_IDS:
        return True
    if chat_id is None:
        return False
    return chat_id in ALLOWED_CHAT_IDS


async def _stream_process_output(
    stream: asyncio.StreamReader | None,
    task_label: str,
    is_stderr: bool = False,
) -> str:
    if stream is None:
        return ""

    chunks: list[str] = []
    while True:
        raw = await stream.readline()
        if not raw:
            break
        line = raw.decode("utf-8", errors="replace").rstrip("\r\n")
        if not line:
            continue
        chunks.append(line)
        if is_stderr:
            logger.warning("[TRAIN_PIPE][%s][stderr] %s", task_label, line)
        else:
            logger.info("[TRAIN_PIPE][%s] %s", task_label, line)
    return "\n".join(chunks)


async def _run_train_pipeline(chat_id: int, app: Application) -> None:
    """Run full training pipeline: train-v2 → train-v4 → compare → calibrate."""
    pipeline_steps = [
        ("1/5 train-v2", ["training/model_cli.py", "train-v2"]),
        ("2/5 train-v4", ["training/model_cli.py", "train-v4"]),
        ("3/5 train-v12", ["training/v12/train_v12.py"]),
        ("4/5 compare",  ["training/model_cli.py", "compare"]),
        ("5/5 calibrate", ["training/calibrate_gate.py"]),
    ]
    async with RETRAIN_LOCK:
        TRAIN_STATUS["running"] = True
        TRAIN_STATUS["owner_chat_id"] = chat_id
        TRAIN_STATUS["last_error_tail"] = ""
        TRAIN_STATUS["progress_done"] = 0
        TRAIN_STATUS["progress_total"] = len(pipeline_steps)

        failed_step: tuple[str, int, str] | None = None

        for i, (task_label, script_args) in enumerate(pipeline_steps, start=1):
            TRAIN_STATUS["current_task"] = task_label
            TRAIN_STATUS["progress_done"] = i - 1
            cmd_display = " ".join([sys.executable, "-u", *script_args])
            logger.info("[TRAIN_PIPE] start step=%s cmd=%s", task_label, cmd_display)
            proc = await asyncio.create_subprocess_exec(
                sys.executable,
                "-u",
                *script_args,
                cwd=str(BASE_DIR),
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
            )
            stdout_task = asyncio.create_task(
                _stream_process_output(proc.stdout, task_label, is_stderr=False)
            )
            stderr_task = asyncio.create_task(
                _stream_process_output(proc.stderr, task_label, is_stderr=True)
            )
            stdout_text, stderr_text = await asyncio.gather(stdout_task, stderr_task)
            return_code = await proc.wait()
            logger.info(
                "[TRAIN_PIPE] end step=%s exit=%s",
                task_label,
                return_code,
            )

            if return_code != 0:
                failed_step = (task_label, return_code, stderr_text or stdout_text)
                break

        TRAIN_STATUS["progress_done"] = len(pipeline_steps)

        if failed_step is None:
            metrics_summary = _read_train_v4_effectiveness_summary()
            msg = (
                "Pipeline entrenamiento finalizado OK.\n"
                "Pasos: train-v2, train-v4, compare, calibrate.\n"
                "Efectividad final (test):\n"
                f"{metrics_summary}"
            )
            TRAIN_STATUS["last_error_tail"] = ""
            TRAIN_STATUS["last_exit_code"] = 0
        else:
            label, code, stderr_text = failed_step
            err_tail = _tail_lines(stderr_text)
            msg = (
                f"Pipeline fallo en: {label} (exit={code}).\n"
                "STDERR:\n"
                f"{err_tail}"
            )
            TRAIN_STATUS["last_error_tail"] = err_tail
            TRAIN_STATUS["last_exit_code"] = code

        TRAIN_STATUS["running"] = False
        TRAIN_STATUS["last_finished_utc"] = datetime.utcnow().isoformat(
            timespec="seconds"
        )
        TRAIN_STATUS["current_task"] = ""
        TRAIN_STATUS["progress_done"] = 0
        TRAIN_STATUS["progress_total"] = 0

        await app.bot.send_message(chat_id=chat_id, text=msg)


async def _run_recalc_universe(chat_id: int, app: Application) -> None:
    async with RETRAIN_LOCK:
        TRAIN_STATUS["running"] = True
        TRAIN_STATUS["owner_chat_id"] = chat_id
        TRAIN_STATUS["last_error_tail"] = ""
        TRAIN_STATUS["current_task"] = "recalculo-universo"
        TRAIN_STATUS["progress_done"] = 0
        TRAIN_STATUS["progress_total"] = 0

        try:
            match_ids = await asyncio.to_thread(_list_candidate_match_ids_for_universe)
            total = len(match_ids)
            TRAIN_STATUS["progress_total"] = total

            ok = 0
            fail = 0
            skipped_not_ft = 0

            for index, match_id in enumerate(match_ids, start=1):
                data = _get_match_detail(match_id)
                if not _is_ft_complete(data):
                    skipped_not_ft += 1
                    TRAIN_STATUS["progress_done"] = index
                    continue

                try:
                    result = await asyncio.to_thread(
                        _compute_and_store_predictions,
                        match_id,
                        data,
                    )
                    if result is None:
                        fail += 1
                    else:
                        ok += 1
                except Exception:
                    fail += 1

                TRAIN_STATUS["progress_done"] = index

            msg = (
                "Recalculo universo finalizado.\n"
                f"candidatos={total}\n"
                f"recalculados_ok={ok}\n"
                f"fallidos={fail}\n"
                f"omitidos_no_ft={skipped_not_ft}"
            )
            TRAIN_STATUS["last_error_tail"] = ""
            TRAIN_STATUS["last_exit_code"] = 0
        except Exception as exc:
            err_tail = _tail_lines(str(exc), max_lines=12, max_chars=1500)
            msg = (
                "Recalculo universo fallo.\n"
                f"error={err_tail}"
            )
            TRAIN_STATUS["last_error_tail"] = err_tail
            TRAIN_STATUS["last_exit_code"] = 1
        finally:
            TRAIN_STATUS["running"] = False
            TRAIN_STATUS["last_finished_utc"] = datetime.utcnow().isoformat(
                timespec="seconds"
            )
            TRAIN_STATUS["current_task"] = ""
            TRAIN_STATUS["progress_done"] = 0
            TRAIN_STATUS["progress_total"] = 0

        await app.bot.send_message(chat_id=chat_id, text=msg)


async def myid_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    _ = context
    chat_id = update.effective_chat.id if update.effective_chat else None
    user_id = update.effective_user.id if update.effective_user else None
    await update.effective_message.reply_text(
        "Tus IDs:\n"
        f"chat_id={chat_id}\n"
        f"user_id={user_id}"
    )


async def train_status_cmd(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
) -> None:
    _ = context
    await update.effective_message.reply_text(_train_status_text())


async def dates_cmd(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
) -> None:
    chat_id = update.effective_chat.id if update.effective_chat else None
    rows = await asyncio.to_thread(_fetch_date_summaries)
    if not rows:
        await update.effective_message.reply_text("No hay fechas disponibles.")
        return
    pred_stats = await asyncio.to_thread(_fetch_dates_pred_stats)
    await update.effective_message.reply_text(
        text="Fechas disponibles:",
        reply_markup=_dates_keyboard(rows, 0, pred_stats),
    )


async def monitor_cmd(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
) -> None:
    chat_id = update.effective_chat.id if update.effective_chat else None
    running = bool(
        bet_monitor_mod.MONITOR_STATUS.get("running")
        or (_MONITOR_THREAD is not None and _MONITOR_THREAD.is_alive())
    )
    await update.effective_message.reply_text(
        text="Monitor Apuestas",
        reply_markup=_monitor_keyboard(running, chat_id),
    )


async def signals_cmd(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
) -> None:
    """Show today's signals."""
    if not _is_allowed(update):
        return
    
    today_str = datetime.now().date().isoformat()
    cid = update.effective_chat.id if update.effective_chat else None
    # Reconcile any ⏳ pending bets from quarter_scores already in DB
    await asyncio.to_thread(bet_monitor_mod.reconcile_pending_results, DB_PATH)
    signal_type, quarters = _get_subscriber_pref(cid)
    text = bet_monitor_mod.signals_text_today(DB_PATH, today_str, pref=signal_type, quarters=quarters)
    nav_markup = InlineKeyboardMarkup([
        [InlineKeyboardButton("📋 Reporte Detallado", callback_data="monitor:report_today")],
        [InlineKeyboardButton("Menu principal", callback_data="nav:main")],
    ])
    await update.effective_message.reply_text(text=text, reply_markup=nav_markup)


async def signals_report_cmd(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
) -> None:
    """Download today's detailed report (Excel)."""
    if not _is_allowed(update):
        return
    
    today_str = datetime.now().date().isoformat()
    cid = update.effective_chat.id if update.effective_chat else None
    signal_type, quarters = _get_subscriber_pref(cid)
    
    # Show loading message
    msg = await update.effective_message.reply_text("⏳ Generando Reporte Detallado...")
    
    try:
        # Generate Excel in thread with extended timeout (120 seconds)
        excel_bytes = await asyncio.wait_for(
            asyncio.to_thread(bet_monitor_mod.signals_excel_today, DB_PATH, today_str, quarters),
            timeout=120.0
        )
        filename = f"Reporte_Señales_{today_str}.xlsx"
        from io import BytesIO
        
        # Send document with extended timeout (120 seconds)
        await asyncio.wait_for(
            context.bot.send_document(
                chat_id=cid,
                document=BytesIO(excel_bytes),
                filename=filename,
            ),
            timeout=120.0
        )
        
        # Update loading message to confirmation
        await msg.edit_text("✅ Reporte enviado")
    except asyncio.TimeoutError:
        logger.error("Error generating/sending report: Timeout after 120 seconds")
        await msg.edit_text("❌ Timeout al generar/enviar el reporte (muy lento)")
    except Exception as e:
        logger.error("Error generating report: %s", e)
        await msg.edit_text("❌ Error al generar el reporte")


def _start_monitor_thread() -> None:
    """Spawn the monitor background thread (mirrors monitor:start handler)."""
    global _MONITOR_THREAD, _MONITOR_LOOP
    bet_monitor_mod.set_model_config(dict(MONITOR_MODEL_CONFIG))

    def _run() -> None:
        global _MONITOR_LOOP
        import asyncio as _aio
        loop = _aio.new_event_loop()
        _aio.set_event_loop(loop)
        _MONITOR_LOOP = loop
        stop_ev = _aio.Event()
        _MONITOR_STOP_REF[0] = stop_ev
        bet_monitor_mod.set_notify_callback(_monitor_notify_cb)
        try:
            loop.run_until_complete(bet_monitor_mod.run_monitor(DB_PATH, stop_ev))
        finally:
            loop.close()
            _MONITOR_LOOP = None

    t = threading.Thread(target=_run, daemon=True, name="bet_monitor")
    _MONITOR_THREAD = t
    t.start()


async def _post_init(app: Application) -> None:
    # Ensure DB schema (including settings table) exists
    conn = _open_conn()
    db_mod.init_db(conn)
    # Load persisted model configs from DB
    for quarter in ("q3", "q4"):
        v = db_mod.get_setting(conn, f"bot_model_{quarter}")
        if v and v in AVAILABLE_MODELS:
            MODEL_CONFIG[quarter] = v
        mv = db_mod.get_setting(conn, f"monitor_model_{quarter}")
        if mv and mv in AVAILABLE_MODELS:
            MONITOR_MODEL_CONFIG[quarter] = mv
    conn.close()

    logger.info(
        "[telegram] selector: %d modelos -> %s",
        len(AVAILABLE_MODELS),
        ", ".join(AVAILABLE_MODELS),
    )

    # Restore persisted subscribers
    _load_subscribers()

    # Auto-start the monitor daemon
    _start_monitor_thread()

    await app.bot.set_my_commands(
        [
            BotCommand("start", "Abrir menu principal"),
            BotCommand("myid", "Ver chat_id y user_id"),
            BotCommand("trainstatus", "Estado del ultimo reentreno"),
            BotCommand("dates", "Matches por fecha"),
            BotCommand("signals", "Señales de hoy"),
            BotCommand("report", "Reporte detallado de hoy"),
            BotCommand("monitor", "Menu monitor apuestas"),
            BotCommand("modelstats", "Estadisticas de los modelos"),
        ]
    )
    await app.bot.set_chat_menu_button(menu_button=MenuButtonCommands())


async def _handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    global _MONITOR_THREAD, _MONITOR_LOOP, _MONITOR_STOP_REF
    query = update.callback_query
    if not query:
        return

    await query.answer()
    data = query.data or ""

    if data == "noop":
        return

    if data == "nav:main":
        chat_id = update.effective_chat.id if update.effective_chat else None
        job = DATE_INGEST_JOBS.get(chat_id) if chat_id is not None else None
        if isinstance(job, dict):
            job["suppress_edits"] = True
        refresh_job = REFRESH_JOBS.get(chat_id) if chat_id is not None else None
        if isinstance(refresh_job, dict):
            refresh_job["suppress_edits"] = True
        await _render_main_menu(update, context)
        return

    if data == "menu:fetchdate:status":
        chat_id = update.effective_chat.id if update.effective_chat else None
        job = DATE_INGEST_JOBS.get(chat_id) if chat_id is not None else None
        if not isinstance(job, dict):
            await query.edit_message_text(
                text="No hay una descarga en curso.",
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("Menu principal", callback_data="nav:main")]]
                ),
            )
            return

        event_date = str(job.get("event_date") or "-")
        limit = job.get("limit")
        progress_state = job.get("progress_state")
        if not isinstance(progress_state, dict):
            progress_state = {"phase": "starting"}
            job["progress_state"] = progress_state

        job["suppress_edits"] = False
        job["force_next_edit"] = True
        text = _date_ingest_progress_text(event_date, limit, progress_state)
        await query.edit_message_text(
            text=text,
            reply_markup=_date_ingest_progress_keyboard(True),
        )
        return

    if data == "menu:refresh:status":
        chat_id = update.effective_chat.id if update.effective_chat else None
        job = REFRESH_JOBS.get(chat_id) if chat_id is not None else None
        if not isinstance(job, dict):
            await query.edit_message_text(
                text="No hay una descarga de pendientes en curso.",
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("Menu principal", callback_data="nav:main")]]
                ),
            )
            return

        job["suppress_edits"] = False
        job["force_next_edit"] = True
        text = _refresh_job_status_text(job)
        await query.edit_message_text(
            text=text,
            reply_markup=_refresh_job_progress_keyboard(),
        )
        return

    if data == "menu:follow:status":
        chat_id = update.effective_chat.id if update.effective_chat else None
        job = FOLLOW_JOBS.get(chat_id) if chat_id is not None else None
        if not isinstance(job, dict):
            await query.edit_message_text(
                text="No hay seguimiento activo.",
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("Menu principal", callback_data="nav:main")]]
                ),
            )
            return
        await query.edit_message_text(
            text=_follow_status_text(job),
            reply_markup=_follow_status_keyboard(job),
        )
        if query.message and isinstance(job, dict):
            job["status_message_id"] = query.message.message_id
            job["status_last_text"] = _follow_status_text(job)
        return

    if data == "menu:refresh7d":
        chat_id = update.effective_chat.id if update.effective_chat else None
        if chat_id is None:
            await _replace_callback_message(
                update,
                text="No pude identificar el chat para ejecutar la descarga.",
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("Menu principal", callback_data="nav:main")]]
                ),
            )
            return
        if chat_id in REFRESH_JOBS:
            await query.edit_message_text(
                text="Ya hay una descarga de pendientes en curso.",
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("Ver status", callback_data="menu:refresh:status")]]
                ),
            )
            return

        await _set_waiting_state(
            update,
            text="Descargando pendientes de los ultimos 7 dias... en background",
        )
        message = query.message
        message_id = message.message_id if message else None
        if message_id is None:
            return
        context.application.create_task(
            _run_refresh_7d_job(
                context.application,
                chat_id,
                message_id,
            )
        )
        return

    if data == "fetchdate:cancel":
        chat_id = update.effective_chat.id if update.effective_chat else None
        job = DATE_INGEST_JOBS.get(chat_id) if chat_id is not None else None
        if not job:
            await query.edit_message_text(
                text="No hay una ingesta de fecha en curso.",
                reply_markup=_date_ingest_progress_keyboard(False),
            )
            return
        cancel_event = job.get("cancel_event")
        if cancel_event is not None:
            cancel_event.set()
        await query.edit_message_text(
            text="Cancelacion solicitada. Esperando cierre del match en curso...",
            reply_markup=_date_ingest_progress_keyboard(False),
        )
        return

    if data == "fetchdate:manual":
        context.user_data[AWAITING_FETCH_DATE_KEY] = True
        context.user_data[AWAITING_MATCH_ID_KEY] = False
        await query.edit_message_text(
            text=(
                "Enviar fecha en formato YYYY-MM-DD.\n"
                "Opcional limite: YYYY-MM-DD 100\n"
                "Ejemplo: 2026-03-26 200"
            ),
            reply_markup=InlineKeyboardMarkup(
                [[InlineKeyboardButton("Menu principal", callback_data="nav:main")]]
            ),
        )
        return

    if data.startswith("fetchdate:pick:"):
        chat_id = update.effective_chat.id if update.effective_chat else None
        try:
            event_date = data.split(":", maxsplit=2)[2]
            datetime.strptime(event_date, "%Y-%m-%d")
        except (IndexError, ValueError):
            await query.edit_message_text(
                text="Fecha invalida en el boton.",
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("Menu principal", callback_data="nav:main")]]
                ),
            )
            return
        if chat_id is not None and chat_id in DATE_INGEST_JOBS:
            await query.edit_message_text(
                text="Ya hay una ingesta de fecha en curso. Puedes cancelarla desde su mensaje de progreso.",
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("Menu principal", callback_data="nav:main")]]
                ),
            )
            return
        await query.edit_message_text(
            text=_date_ingest_progress_text(event_date, None, {"phase": "starting"}),
            reply_markup=_date_ingest_progress_keyboard(True),
        )
        if chat_id is not None and query.message is not None:
            context.application.create_task(
                _run_date_ingest_job(
                    context.application,
                    chat_id,
                    query.message.message_id,
                    event_date,
                    None,
                )
            )
        return

        await _render_dates(update, 0)
        return

    if data == "menu:live":
        context.user_data[AWAITING_FETCH_DATE_KEY] = False
        context.user_data[AWAITING_MATCH_ID_KEY] = False
        await _render_live(update, context, 0, refetch=True)
        return

    if data == "menu:dates":
        context.user_data[AWAITING_FETCH_DATE_KEY] = False
        context.user_data[AWAITING_MATCH_ID_KEY] = False
        await query.answer("⏳ Cargando fechas...")
        await _render_dates(update, 0)
        return

    if data.startswith("dates:"):
        try:
            _, page_text = data.split(":", maxsplit=1)
            page = int(page_text)
        except (ValueError, TypeError):
            await _render_dates(update, 0)
            return
        await _render_dates(update, page)
        return

    if data.startswith("livepage:"):
        try:
            _, page_text = data.split(":", maxsplit=1)
            page = int(page_text)
        except (ValueError, TypeError):
            page = 0
        await _render_live(update, context, page, refetch=False)
        return

    if data == "menu:id":
        context.user_data[AWAITING_MATCH_ID_KEY] = True
        context.user_data[AWAITING_FETCH_DATE_KEY] = False
        await query.edit_message_text(
            text=(
                "Enviar Match ID numerico.\n"
                "Ejemplo: 14442355"
            ),
            reply_markup=InlineKeyboardMarkup(
                [[InlineKeyboardButton("Menu principal", callback_data="nav:main")]]
            ),
        )
        return

    if data == "menu:fetchdate":
        context.user_data[AWAITING_FETCH_DATE_KEY] = False
        context.user_data[AWAITING_MATCH_ID_KEY] = False
        missing_data = await asyncio.to_thread(_get_missing_dates_suggestions)
        recent = missing_data["recent"]
        historical = missing_data["historical"]
        min_date = missing_data["min_date"]

        kbd_rows: list[list[InlineKeyboardButton]] = []
        lines: list[str] = []

        # Recent section — up to 5 buttons
        if recent:
            lines.append("Fechas recientes sin datos (ultimos 21 dias):")
            for d in recent[:5]:
                kbd_rows.append(
                    [InlineKeyboardButton(f"📅 {d}", callback_data=f"fetchdate:pick:{d}")]
                )
        else:
            lines.append("Sin fechas faltantes en los ultimos 21 dias.")

        # Historical section — up to 3 buttons (total max 8)
        remaining_slots = 8 - len(kbd_rows)
        if historical and remaining_slots > 0:
            lines.append(f"\nFechas historicas sin datos (desde {min_date}):")
            for d in historical[:remaining_slots]:
                kbd_rows.append(
                    [InlineKeyboardButton(f"🗓 {d}", callback_data=f"fetchdate:pick:{d}")]
                )
            if len(historical) > remaining_slots:
                lines.append(f"(+{len(historical) - remaining_slots} fechas historicas mas — usa ingresar fecha manualmente)")

        hint = "\n".join(lines)
        kbd_rows.append(
            [InlineKeyboardButton("✏️ Ingresar fecha manualmente", callback_data="fetchdate:manual")]
        )
        kbd_rows.append(
            [InlineKeyboardButton("Menu principal", callback_data="nav:main")]
        )
        await query.edit_message_text(
            text=hint,
            reply_markup=InlineKeyboardMarkup(kbd_rows),
        )
        return


    if data == "menu:train":
        chat_id = update.effective_chat.id if update.effective_chat else None
        if not _is_train_allowed(chat_id):
            await query.edit_message_text(
                text=(
                    "No autorizado para reentrenar.\n"
                    f"Tu chat_id es {chat_id}.\n"
                    "Agregalo en TELEGRAM_ALLOWED_CHAT_IDS del .env"
                ),
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("Menu principal", callback_data="nav:main")]]
                ),
            )
            return
        await query.edit_message_text(
            text=(
                "Re-entrenar base.\n"
                "Elige una opcion:"
            ),
            reply_markup=_train_submenu_keyboard(),
        )
        return

    if data == "menu:monitor":
        running = bool(
            bet_monitor_mod.MONITOR_STATUS.get("running", False)
            or (_MONITOR_THREAD is not None and _MONITOR_THREAD.is_alive())
        )
        await query.edit_message_text(
            text=bet_monitor_mod.status_text(),
            reply_markup=_monitor_keyboard(running, chat_id=query.from_user.id if query.from_user else None),
        )
        return

    if data == "menu:monthly_report":
        # Show month picker: current month + last 5 months
        from datetime import datetime as _dt_now
        _now = _dt_now.utcnow()
        _month_buttons = []
        _MONTH_ES = {
            1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
            5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
            9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
        }
        for _i in range(6):
            # go back _i months from current
            _m = _now.month - _i
            _y = _now.year
            while _m <= 0:
                _m += 12
                _y -= 1
            _ym_opt = f"{_y:04d}-{_m:02d}"
            _label = f"{'📅 ' if _i == 0 else ''}{_MONTH_ES[_m]} {_y}"
            _month_buttons.append([InlineKeyboardButton(_label, callback_data=f"monthly_pick:{_ym_opt}")])
        _month_buttons.append([InlineKeyboardButton("⬅️ Cancelar", callback_data="nav:main")])
        _kb_pick = InlineKeyboardMarkup(_month_buttons)
        _cid_pick = update.effective_chat.id if update.effective_chat else None
        _q_pick = update.callback_query
        if _q_pick:
            try:
                await _q_pick.answer()
            except Exception:
                pass
            try:
                await _q_pick.edit_message_text("📆 Selecciona el mes del reporte:", reply_markup=_kb_pick)
            except Exception:
                await context.bot.send_message(chat_id=_cid_pick, text="📆 Selecciona el mes del reporte:", reply_markup=_kb_pick)
        else:
            await context.bot.send_message(chat_id=_cid_pick, text="📆 Selecciona el mes del reporte:", reply_markup=_kb_pick)
        return

    if data.startswith("monthly_pick:"):
        _ym = data.split(":", 1)[1]
        await _send_monthly_report(update, context, _ym)
        return

    if data.startswith("monthly:"):
        # format: monthly:YYYY-MM:current  OR  monthly:YYYY-MM:recalc
        _parts = data.split(":")
        # parts[0]="monthly", parts[1]="YYYY-MM", parts[2]="current"|"recalc"
        # but YYYY-MM has a dash so split(":")  gives 3 or 4 parts
        if len(_parts) >= 3:
            _ym = _parts[1]  # "YYYY-MM"
            _action = _parts[-1]  # "current" or "recalc"
        else:
            _ym = datetime.now().strftime("%Y-%m")
            _action = "current"
        cid = update.effective_chat.id if update.effective_chat else None
        _, quarters = _get_subscriber_pref(cid)
        if update.callback_query:
            try:
                await update.callback_query.answer()
            except Exception:
                pass
        msg = await context.bot.send_message(
            chat_id=cid,
            text=f"⏳ {'Recalculando' if _action == 'recalc' else 'Generando'} reporte {_ym}...",
        )
        try:
            if _action == "recalc":
                await _recalculate_month_evals(_ym, cid, msg, quarters)

            excel_bytes = await asyncio.wait_for(
                asyncio.to_thread(_build_monthly_excel_bytes, _ym, quarters),
                timeout=180.0,
            )
            filename = f"Reporte_Mensual_{_ym}.xlsx"
            from io import BytesIO
            await asyncio.wait_for(
                context.bot.send_document(
                    chat_id=cid,
                    document=BytesIO(excel_bytes),
                    filename=filename,
                ),
                timeout=120.0,
            )
            await msg.edit_text(f"✅ Reporte {_ym} enviado")
        except asyncio.TimeoutError:
            logger.error("Monthly report timeout for %s", _ym)
            await msg.edit_text("❌ Timeout al generar el reporte")
        except Exception as exc:
            logger.error("Monthly report error: %s", exc, exc_info=True)
            await msg.edit_text(f"❌ Error: {exc}")
        return

    if data == "monitor:start":
        _mon_thread = _MONITOR_THREAD
        running = bool(
            bet_monitor_mod.MONITOR_STATUS.get("running", False)
            or (_mon_thread is not None and _mon_thread.is_alive())
        )
        if running:
            await query.edit_message_text(
                text="El monitor ya está activo.\n\n" + bet_monitor_mod.status_text(),
                reply_markup=_monitor_keyboard(True, chat_id=query.from_user.id if query.from_user else None),
            )
            return
        bet_monitor_mod.set_model_config(dict(MONITOR_MODEL_CONFIG))

        _start_monitor_thread()
        await query.edit_message_text(
            text="Monitor iniciado en hilo separado.\n\n" + bet_monitor_mod.status_text(),
            reply_markup=_monitor_keyboard(True, chat_id=query.from_user.id if query.from_user else None),
        )
        return

    if data == "monitor:stop":
        loop = _MONITOR_LOOP
        stop_ev = _MONITOR_STOP_REF[0]
        if loop is not None and stop_ev is not None and not loop.is_closed():
            loop.call_soon_threadsafe(stop_ev.set)
            bet_monitor_mod.MONITOR_STATUS["stop_requested"] = True
        await query.edit_message_text(
            text="Deteniendo monitor...\n\n" + bet_monitor_mod.status_text(),
            reply_markup=_monitor_keyboard(False, chat_id=query.from_user.id if query.from_user else None),
        )
        return

    if data == "monitor:subscribe":
        cid = query.from_user.id if query.from_user else None
        if cid:
            _MONITOR_SUBSCRIBERS.setdefault(cid, {
                "signal_type": "bet_only",
                "quarters": ["q3", "q4"]
            })
            _persist_subscribers()
        running = bool(
            bet_monitor_mod.MONITOR_STATUS.get("running", False)
            or (_MONITOR_THREAD is not None and _MONITOR_THREAD.is_alive())
        )
        await query.edit_message_text(
            text="🔔 Suscrito (solo apuestas). Cambia el filtro abajo.\n\n" + bet_monitor_mod.status_text(),
            reply_markup=_monitor_keyboard(running, chat_id=cid),
        )
        return

    if data == "monitor:unsubscribe":
        cid = query.from_user.id if query.from_user else None
        if cid:
            _MONITOR_SUBSCRIBERS.pop(cid, None)
            _persist_subscribers()
        running = bool(
            bet_monitor_mod.MONITOR_STATUS.get("running", False)
            or (_MONITOR_THREAD is not None and _MONITOR_THREAD.is_alive())
        )
        await query.edit_message_text(
            text="🔕 Desuscrito de alertas del monitor.\n\n" + bet_monitor_mod.status_text(),
            reply_markup=_monitor_keyboard(running, chat_id=cid),
        )
        return

    if data in ("monitor:sub_bet_only", "monitor:sub_all"):
        cid = query.from_user.id if query.from_user else None
        signal_type = "bet_only" if data == "monitor:sub_bet_only" else "all"
        if cid:
            if cid not in _MONITOR_SUBSCRIBERS:
                _MONITOR_SUBSCRIBERS[cid] = {"signal_type": signal_type, "quarters": ["q3", "q4"]}
            else:
                _MONITOR_SUBSCRIBERS[cid]["signal_type"] = signal_type
            _persist_subscribers()
        running = bool(
            bet_monitor_mod.MONITOR_STATUS.get("running", False)
            or (_MONITOR_THREAD is not None and _MONITOR_THREAD.is_alive())
        )
        label = "solo apuestas" if signal_type == "bet_only" else "apuestas + no-apuesta"
        await query.edit_message_text(
            text=f"🔔 Recibirás alertas de {label}.\n\n" + bet_monitor_mod.status_text(),
            reply_markup=_monitor_keyboard(running, chat_id=cid),
        )
        return

    if data == "monitor:quarters_menu":
        cid = query.from_user.id if query.from_user else None
        await query.edit_message_text(
            text="Selecciona qué quarters quieres monitorear:",
            reply_markup=_monitor_quarters_keyboard(chat_id=cid),
        )
        return

    if data in ("monitor:toggle_q3", "monitor:toggle_q4"):
        cid = query.from_user.id if query.from_user else None
        quarter = "q3" if data == "monitor:toggle_q3" else "q4"
        if cid:
            if cid not in _MONITOR_SUBSCRIBERS:
                _MONITOR_SUBSCRIBERS[cid] = {"signal_type": "all", "quarters": ["q3", "q4"]}
            quarters = _MONITOR_SUBSCRIBERS[cid].get("quarters", ["q3", "q4"])
            if quarter in quarters:
                quarters.remove(quarter)
            else:
                quarters.append(quarter)
            _MONITOR_SUBSCRIBERS[cid]["quarters"] = sorted(quarters)
            _persist_subscribers()
        await query.edit_message_text(
            text="Selecciona qué quarters quieres monitorear:",
            reply_markup=_monitor_quarters_keyboard(chat_id=cid),
        )
        return

    if data == "monitor:models":
        q3_v = MODEL_CONFIG.get("q3", "v4")
        q4_v = MODEL_CONFIG.get("q4", "v4")
        await _replace_callback_message(
            update,
            text=(
                f"Modelos del monitor\n"
                f"Activo: Q3={q3_v}  Q4={q4_v}\n\n"
                "Elige el modelo para cada cuarto:"
            ),
            reply_markup=_monitor_model_keyboard(chat_id=query.from_user.id if query.from_user else None),
        )
        return

    if data.startswith("monmodel:set:"):
        # Format: monmodel:set:q3:<idx|slug>  (slug legacy p. ej. v6_1)
        parts = data.split(":")
        if len(parts) == 4:
            _, _, quarter, ver_tok = parts
            version = _model_version_from_callback_token(ver_tok)
            if quarter in ("q3", "q4") and version:
                MONITOR_MODEL_CONFIG[quarter] = version
                conn = _open_conn()
                db_mod.set_setting(conn, f"monitor_model_{quarter}", version)
                conn.close()
                if bet_monitor_mod:
                    bet_monitor_mod.set_model_config({quarter: version})
        q3_v = MONITOR_MODEL_CONFIG.get("q3", "v4")
        q4_v = MONITOR_MODEL_CONFIG.get("q4", "v4")
        await _replace_callback_message(
            update,
            text=(
                f"Modelos del monitor\n"
                f"Activo: Q3={q3_v}  Q4={q4_v}\n\n"
                "Elige el modelo para cada cuarto:"
            ),
            reply_markup=_monitor_model_keyboard(chat_id=query.from_user.id if query.from_user else None),
        )
        return

    if data == "monitor:schedule":
        today_str = datetime.now().date().isoformat()
        text, markup = bet_monitor_mod.schedule_keyboard(DB_PATH, today_str)
        await query.edit_message_text(text=text, reply_markup=markup)
        return

    if data == "monitor:betcfg":
        conn = _open_conn()
        bank  = db_mod.get_setting(conn, "sig_bank",    "1000")
        bet   = db_mod.get_setting(conn, "sig_bet_size", "100")
        odds  = db_mod.get_setting(conn, "sig_odds",     "1.4")
        conn.close()
        text = (
            f"⚙️ Config simulación de banco\n\n"
            f"🏦 Bank inicial: {bank}\n"
            f"💵 Apuesta: {bet}\n"
            f"📈 Momio (odds): {odds}\n\n"
            "Ajusta con los botones:"
        )
        kb = InlineKeyboardMarkup([
            [   InlineKeyboardButton("Bank −500", callback_data="betcfg:bank:-500"),
                InlineKeyboardButton("Bank +500", callback_data="betcfg:bank:+500"),
                InlineKeyboardButton("Bank +1000", callback_data="betcfg:bank:+1000"),
            ],
            [   InlineKeyboardButton("Apuesta −10", callback_data="betcfg:bet:-10"),
                InlineKeyboardButton("Apuesta −50", callback_data="betcfg:bet:-50"),
                InlineKeyboardButton("Apuesta +50", callback_data="betcfg:bet:+50"),
                InlineKeyboardButton("Apuesta +100", callback_data="betcfg:bet:+100"),
            ],
            [   InlineKeyboardButton("Odds −0.1", callback_data="betcfg:odds:-0.1"),
                InlineKeyboardButton("Odds +0.1", callback_data="betcfg:odds:+0.1"),
                InlineKeyboardButton("Odds +0.5", callback_data="betcfg:odds:+0.5"),
            ],
            [InlineKeyboardButton("⬅️ Monitor", callback_data="menu:monitor")],
        ])
        await query.edit_message_text(text=text, reply_markup=kb)
        return

    if data.startswith("betcfg:"):
        # betcfg:bank:+500  /  betcfg:bet:-10  /  betcfg:odds:+0.1
        parts = data.split(":")
        if len(parts) == 3:
            _, field, delta_s = parts
            conn = _open_conn()
            key_map = {"bank": "sig_bank", "bet": "sig_bet_size", "odds": "sig_odds"}
            defaults = {"bank": 1000.0, "bet": 100.0, "odds": 1.4}
            if field in key_map:
                db_key = key_map[field]
                cur = float(db_mod.get_setting(conn, db_key) or str(defaults[field]))
                try:
                    cur = round(cur + float(delta_s), 4)
                    if field == "bank":
                        cur = max(0.0, cur)
                    elif field == "bet":
                        cur = max(1.0, cur)
                    elif field == "odds":
                        cur = max(1.01, round(cur, 2))
                except ValueError:
                    pass
                db_mod.set_setting(conn, db_key, str(cur))
            bank = db_mod.get_setting(conn, "sig_bank",     "1000")
            bet  = db_mod.get_setting(conn, "sig_bet_size",  "100")
            odds = db_mod.get_setting(conn, "sig_odds",       "1.4")
            conn.close()
        text = (
            f"⚙️ Config simulación de banco\n\n"
            f"🏦 Bank inicial: {bank}\n"
            f"💵 Apuesta: {bet}\n"
            f"📈 Momio (odds): {odds}\n\n"
            "Ajusta con los botones:"
        )
        kb = InlineKeyboardMarkup([
            [   InlineKeyboardButton("Bank −500", callback_data="betcfg:bank:-500"),
                InlineKeyboardButton("Bank +500", callback_data="betcfg:bank:+500"),
                InlineKeyboardButton("Bank +1000", callback_data="betcfg:bank:+1000"),
            ],
            [   InlineKeyboardButton("Apuesta −10", callback_data="betcfg:bet:-10"),
                InlineKeyboardButton("Apuesta −50", callback_data="betcfg:bet:-50"),
                InlineKeyboardButton("Apuesta +50", callback_data="betcfg:bet:+50"),
                InlineKeyboardButton("Apuesta +100", callback_data="betcfg:bet:+100"),
            ],
            [   InlineKeyboardButton("Odds −0.1", callback_data="betcfg:odds:-0.1"),
                InlineKeyboardButton("Odds +0.1", callback_data="betcfg:odds:+0.1"),
                InlineKeyboardButton("Odds +0.5", callback_data="betcfg:odds:+0.5"),
            ],
            [InlineKeyboardButton("⬅️ Monitor", callback_data="menu:monitor")],
        ])
        await query.edit_message_text(text=text, reply_markup=kb)
        return

    if data == "menu:reconcile":
        await query.answer()
        chat_id = update.effective_chat.id if update.effective_chat else None
        if chat_id is None:
            return
        if chat_id in REFRESH_JOBS:
            await query.edit_message_text(
                "Ya hay un proceso de reconciliación en curso.",
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("Menu principal", callback_data="nav:main")]]
                ),
            )
            return
        message = query.message
        message_id = message.message_id if message else None
        if message_id is None:
            return
        await query.edit_message_text("⏳ Buscando apuestas pendientes de marcador...")
        context.application.create_task(
            _run_reconcile_job(context.application, chat_id, message_id)
        )
        return

    if data == "monitor:signals_today":
        today_str = datetime.now().date().isoformat()
        cid = query.from_user.id if query.from_user else None
        # Reconcile any ⏳ pending bets from quarter_scores already in DB
        await asyncio.to_thread(bet_monitor_mod.reconcile_pending_results, DB_PATH)
        signal_type, quarters = _get_subscriber_pref(cid)
        text = bet_monitor_mod.signals_text_today(DB_PATH, today_str, pref=signal_type, quarters=quarters)
        nav_markup = InlineKeyboardMarkup([
            [InlineKeyboardButton("📋 Reporte Detallado", callback_data="monitor:report_today")],
            [InlineKeyboardButton("⬅️ Monitor", callback_data="menu:monitor")],
            [InlineKeyboardButton("Menu principal", callback_data="nav:main")],
        ])
        await query.edit_message_text(text=text, reply_markup=nav_markup)
        return

    if data == "monitor:report_today":
        today_str = datetime.now().date().isoformat()
        cid = query.from_user.id if query.from_user else None
        signal_type, quarters = _get_subscriber_pref(cid)
        try:
            # Show loading message
            await query.edit_message_text(text="⏳ Generando Reporte Detallado...", reply_markup=None)
            
            # Generate Excel in thread with extended timeout (120 seconds)
            excel_bytes = await asyncio.wait_for(
                asyncio.to_thread(bet_monitor_mod.signals_excel_today, DB_PATH, today_str, quarters),
                timeout=120.0
            )
            filename = f"Reporte_Señales_{today_str}.xlsx"
            from io import BytesIO
            
            # Send document with extended timeout (120 seconds)
            await asyncio.wait_for(
                context.bot.send_document(
                    chat_id=query.from_user.id if query.from_user else query.message.chat_id,
                    document=BytesIO(excel_bytes),
                    filename=filename,
                ),
                timeout=120.0
            )
            
            # Show confirmation
            await query.edit_message_text(
                text="✅ Reporte enviado",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("⬅️ Volver a Señales", callback_data="monitor:signals_today")
                ]])
            )
        except asyncio.TimeoutError:
            logger.error("Error generating/sending report: Timeout after 120 seconds")
            await query.edit_message_text(
                text="❌ Timeout al generar/enviar el reporte (muy lento)",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("⬅️ Volver a Señales", callback_data="monitor:signals_today")
                ]])
            )
        except Exception as e:
            logger.error("Error generating report: %s", e)
            await query.edit_message_text(
                text="❌ Error al generar el reporte",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("⬅️ Volver a Señales", callback_data="monitor:signals_today")
                ]])
            )
        return

    if data == "monitor:log":
        text, markup = bet_monitor_mod.log_keyboard(DB_PATH)
        await query.edit_message_text(text=text, reply_markup=markup)
        return

    if data == "menu:models":
        q3_v = MODEL_CONFIG.get("q3", "v4")
        q4_v = MODEL_CONFIG.get("q4", "v4")
        await _replace_callback_message(
            update,
            text=(
                f"Selector de modelos\n"
                f"Activo: Q3={q3_v}  Q4={q4_v}\n\n"
                "Elige el modelo para cada cuarto:"
            ),
            reply_markup=_model_submenu_keyboard(),
        )
        return

    if data.startswith("model:set:"):
        # Format: model:set:q3:<idx|slug>
        parts = data.split(":")
        if len(parts) == 4:
            _, _, quarter, ver_tok = parts
            version = _model_version_from_callback_token(ver_tok)
            if quarter in ("q3", "q4") and version:
                MODEL_CONFIG[quarter] = version
                conn = _open_conn()
                db_mod.set_setting(conn, f"bot_model_{quarter}", version)
                conn.close()
        q3_v = MODEL_CONFIG.get("q3", "v4")
        q4_v = MODEL_CONFIG.get("q4", "v4")
        await _replace_callback_message(
            update,
            text=(
                f"Selector de modelos\n"
                f"Activo: Q3={q3_v}  Q4={q4_v}\n\n"
                "Elige el modelo para cada cuarto:"
            ),
            reply_markup=_model_submenu_keyboard(),
        )
        return

    # V12 LIVE Bookmaker handler
    if data.startswith("v12live:open:"):
        try:
            _, _, match_id, token, page_text = data.split(":", maxsplit=4)
            page = int(page_text)
        except (ValueError, TypeError):
            await _render_main_menu(update, context)
            return
        
        await _handle_v12_live_analysis(update, context, query, match_id, page)
        return

    if data.startswith("v12live:refresh:"):
        try:
            parts = data.split(":", maxsplit=4)
            match_id = parts[2]
            token = parts[3]
            page = int(parts[4])
        except (ValueError, TypeError, IndexError):
            await _render_main_menu(update, context)
            return

        await _handle_v12_live_analysis(update, context, query, match_id, page, refresh=True)
        return

    if data.startswith("v12live:track:"):
        try:
            parts = data.split(":", maxsplit=4)
            match_id = parts[2]
            token = parts[3]
            page = int(parts[4])
        except (ValueError, TypeError, IndexError):
            await _render_main_menu(update, context)
            return

        await query.answer("Iniciando seguimiento en vivo...")
        chat_id = query.message.chat_id
        msg_id = query.message.message_id
        job_key = f"{chat_id}:{msg_id}"

        # Cancel existing tracking for this message
        existing_job = V12_LIVE_JOBS.get(job_key)
        if existing_job:
            existing_job["cancelled"] = True

        V12_LIVE_JOBS[job_key] = {
            "chat_id": chat_id,
            "msg_id": msg_id,
            "match_id": match_id,
            "page": page,
            "is_photo": bool(query.message.photo),
            "started_utc": datetime.utcnow().isoformat(),
        }

        context.application.create_task(
            _v12_live_poll(chat_id, msg_id, match_id, page, context.application)
        )
        return

    if data.startswith("v12live:graph:"):
        try:
            parts = data.split(":", maxsplit=4)
            match_id = parts[2]
            token = parts[3]
            page = int(parts[4])
        except (ValueError, TypeError, IndexError):
            await _render_main_menu(update, context)
            return

        await _handle_v12_live_graph(update, context, query, match_id, page)
        return

    if data.startswith("matchmodel:open:"):
        # Format: matchmodel:open:{match_id}:{token}:{page}
        try:
            _, _, match_id, token, page_text = data.split(":", maxsplit=4)
            page = int(page_text)
        except (ValueError, TypeError):
            await _render_main_menu(update, context)
            return
        q3_v = MODEL_CONFIG.get("q3", "v4")
        q4_v = MODEL_CONFIG.get("q4", "v4")
        submenu_text = (
            f"Selector de modelos (match {match_id})\n"
            f"Activo: Q3={q3_v}  Q4={q4_v}\n\n"
            "Elige el modelo para cada cuarto:\n"
            "Al regresar se recalculara la prediccion."
        )
        submenu_kb = _match_model_submenu_keyboard(match_id, token, page)
        msg = query.message
        if msg and getattr(msg, "photo", None):
            chat_id = update.effective_chat.id if update.effective_chat else None
            if chat_id is not None:
                await context.bot.send_message(
                    chat_id=chat_id,
                    text=submenu_text,
                    reply_markup=submenu_kb,
                )
                try:
                    await msg.delete()
                except Exception:
                    pass
        else:
            try:
                await query.edit_message_text(
                    text=submenu_text,
                    reply_markup=submenu_kb,
                )
            except Exception:
                chat_id = update.effective_chat.id if update.effective_chat else None
                if chat_id is not None:
                    await context.bot.send_message(
                        chat_id=chat_id,
                        text=submenu_text,
                        reply_markup=submenu_kb,
                    )
        return

    if data.startswith("matchmodel:set:"):
        # Format: matchmodel:set:{quarter}:{idx|slug}:{match_id}:{token}:{page}
        rest = data[len("matchmodel:set:") :]
        try:
            quarter, ver_tok, match_id, token, page_text = rest.split(":", 4)
            page = int(page_text)
        except (ValueError, TypeError):
            await _render_main_menu(update, context)
            return
        version = _model_version_from_callback_token(ver_tok)
        if not version:
            await _render_main_menu(update, context)
            return
        if quarter in ("q3", "q4") and version in AVAILABLE_MODELS:
            MODEL_CONFIG[quarter] = version
        q3_v = MODEL_CONFIG.get("q3", "v4")
        q4_v = MODEL_CONFIG.get("q4", "v4")
        await _replace_callback_message(
            update,
            text=(
                f"Selector de modelos (match {match_id})\n"
                f"Activo: Q3={q3_v}  Q4={q4_v}\n\n"
                "Elige el modelo para cada cuarto:\n"
                "Al regresar se recalculara la prediccion."
            ),
            reply_markup=_match_model_submenu_keyboard(match_id, token, page),
        )
        return

    if data.startswith("matchmodel:back:"):
        # Format: matchmodel:back:{match_id}:{token}:{page}
        try:
            _, _, match_id, token, page_text = data.split(":", maxsplit=4)
            page = int(page_text)
        except (ValueError, TypeError):
            await _render_main_menu(update, context)
            return
        event_date = None if token == "_" else token
        # Compute with the model the user just selected (MODEL_CONFIG), not MONITOR_MODEL_CONFIG
        _mback_data = _get_match_detail(match_id)
        if _mback_data is None:
            _mback_data = await asyncio.to_thread(_refresh_match_data, match_id)
        await _set_waiting_state(update, text=_refresh_waiting_text(match_id, _mback_data))
        _mback_pred = await asyncio.to_thread(
            _refresh_predictions, match_id, _mback_data, dict(MODEL_CONFIG)
        )
        if event_date is None:
            # Came from live detail
            await _render_live_detail(update, context, match_id, page, pred_row=_mback_pred)
        else:
            await _render_match_detail(
                update, context.application, match_id, event_date, page,
                pred_row=_mback_pred, send_graph=False,
            )
        return

    if data.startswith("datemodel:open:"):
        # Format: datemodel:open:{event_date}
        event_date = data[len("datemodel:open:"):]
        q3_v = MODEL_CONFIG.get("q3", "v4")
        q4_v = MODEL_CONFIG.get("q4", "v4")
        await _replace_callback_message(
            update,
            text=(
                f"Selector de modelos (fecha {event_date})\n"
                f"Activo: Q3={q3_v}  Q4={q4_v}\n\n"
                "Elige el modelo para cada cuarto:"
            ),
            reply_markup=_date_model_submenu_keyboard(event_date),
        )
        return

    if data.startswith("datemodel:set:"):
        # Format: datemodel:set:{quarter}:{idx|slug}:{event_date...} (fecha puede contener ':')
        rest = data[len("datemodel:set:") :]
        try:
            quarter, ver_tok, event_date = rest.split(":", 2)
        except (ValueError, TypeError):
            await _render_main_menu(update, context)
            return
        version = _model_version_from_callback_token(ver_tok)
        if not version:
            await _render_main_menu(update, context)
            return
        if quarter in ("q3", "q4") and version in AVAILABLE_MODELS:
            MODEL_CONFIG[quarter] = version
        q3_v = MODEL_CONFIG.get("q3", "v4")
        q4_v = MODEL_CONFIG.get("q4", "v4")
        await _replace_callback_message(
            update,
            text=(
                f"Selector de modelos (fecha {event_date})\n"
                f"Activo: Q3={q3_v}  Q4={q4_v}\n\n"
                "Elige el modelo para cada cuarto:"
            ),
            reply_markup=_date_model_submenu_keyboard(event_date),
        )
        return

    if data.startswith("datemodel:back:"):
        # Format: datemodel:back:{event_date}
        event_date = data[len("datemodel:back:"):]
        await _render_matches_for_date(update, event_date, 0)
        return

    if data == "train:status":
        await query.edit_message_text(
            text=_train_status_text(),
            reply_markup=_train_submenu_keyboard(),
        )
        return

    if data == "train:stats":
        await _send_stats_report(update, context)
        return

    if data == "train:run":
        chat_id = update.effective_chat.id if update.effective_chat else None
        if not _is_train_allowed(chat_id):
            await query.edit_message_text(
                text=(
                    "No autorizado para reentrenar.\n"
                    f"Tu chat_id es {chat_id}.\n"
                    "Agregalo en TELEGRAM_ALLOWED_CHAT_IDS del .env"
                ),
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("Menu principal", callback_data="nav:main")]]
                ),
            )
            return

        if RETRAIN_LOCK.locked():
            await query.edit_message_text(
                text="Ya hay un reentreno en ejecucion. Espera a que termine.",
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("Menu principal", callback_data="nav:main")]]
                ),
            )
            return

        await query.edit_message_text(
            text=(
                "Pipeline entrenamiento iniciado en background.\n"
                "Pasos: train-v2 → train-v4 → train-v12 → compare → calibrate.\n"
                "Te aviso al terminar."
            ),
            reply_markup=InlineKeyboardMarkup(
                [[InlineKeyboardButton("Menu principal", callback_data="nav:main")]]
            ),
        )
        safe_chat_id = int(chat_id) if chat_id is not None else 0
        context.application.create_task(
            _run_train_pipeline(safe_chat_id, context.application)
        )
        return

    if data == "train:clear:confirm":
        await query.edit_message_text(
            text=(
                "Vas a vaciar la tabla de eval_match_results.\n"
                "Esta accion elimina todos los resultados calculados.\n"
                "No se puede deshacer."
            ),
            reply_markup=_train_clear_confirm_keyboard(),
        )
        return

    if data == "train:clear:run":
        chat_id = update.effective_chat.id if update.effective_chat else None
        if not _is_train_allowed(chat_id):
            await query.edit_message_text(
                text=(
                    "No autorizado.\n"
                    f"Tu chat_id es {chat_id}.\n"
                    "Agregalo en TELEGRAM_ALLOWED_CHAT_IDS del .env"
                ),
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("Menu principal", callback_data="nav:main")]]
                ),
            )
            return
        try:
            deleted_rows = await asyncio.to_thread(_clear_eval_results_table)
            await query.edit_message_text(
                text=f"Tabla vaciada. Filas eliminadas: {deleted_rows}.",
                reply_markup=_train_submenu_keyboard(),
            )
        except Exception as exc:
            await query.edit_message_text(
                text=f"Error al vaciar tabla: {exc}",
                reply_markup=_train_submenu_keyboard(),
            )
        return

    if data == "train:recalc:confirm":
        await query.edit_message_text(
            text=(
                "Vas a recalcular predicciones del universo.\n"
                "El proceso corre en background y puede tardar bastante.\n"
                "Los resultados existentes se sobreescriben si el match ya tiene calculo."
            ),
            reply_markup=_train_recalc_confirm_keyboard(),
        )
        return

    if data == "train:recalc:run":
        chat_id = update.effective_chat.id if update.effective_chat else None
        if not _is_train_allowed(chat_id):
            await query.edit_message_text(
                text=(
                    "No autorizado para reentrenar/recalcular.\n"
                    f"Tu chat_id es {chat_id}.\n"
                    "Agregalo en TELEGRAM_ALLOWED_CHAT_IDS del .env"
                ),
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("Menu principal", callback_data="nav:main")]]
                ),
            )
            return

        if RETRAIN_LOCK.locked():
            await query.edit_message_text(
                text="Ya hay un proceso de entrenamiento/recalculo en ejecucion.",
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("Estado entrenamiento", callback_data="train:status")]]
                ),
            )
            return

        await query.edit_message_text(
            text=(
                "Recalculo de universo iniciado en background.\n"
                "Puedes seguir usando el bot y revisar en Estado entrenamiento."
            ),
            reply_markup=InlineKeyboardMarkup(
                [[InlineKeyboardButton("Menu principal", callback_data="nav:main")]]
            ),
        )
        safe_chat_id = int(chat_id) if chat_id is not None else 0
        context.application.create_task(
            _run_recalc_universe(safe_chat_id, context.application)
        )
        return

    if data.startswith("date:"):
        try:
            _, event_date, page_text = data.split(":", maxsplit=2)
            page = int(page_text)
        except (ValueError, TypeError):
            await _render_dates(update, 0)
            return
        await _render_matches_for_date(update, event_date, page)
        return

    if data.startswith("refresh:date:"):
        event_date = data[len("refresh:date:"):]
        logger.info(f"[REFRESH_DATE] Callback received date={event_date}")
        chat_id = update.effective_chat.id if update.effective_chat else None
        if chat_id is None:
            await _replace_callback_message(
                update,
                text="No pude identificar el chat para ejecutar la descarga.",
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("Volver", callback_data=f"date:{event_date}:0")]]
                ),
            )
            return
        if chat_id in REFRESH_JOBS:
            await query.edit_message_text(
                text="Ya hay una descarga de pendientes en curso.",
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("Ver status", callback_data="menu:refresh:status")]]
                ),
            )
            return

        await _set_waiting_state(
            update,
            text="Descargando pendientes del dia... en background",
        )
        message = query.message
        message_id = message.message_id if message else None
        if message_id is None:
            return
        context.application.create_task(
            _run_refresh_date_job(
                context.application,
                chat_id,
                message_id,
                event_date,
            )
        )
        return

    if data.startswith("refresh:"):
        try:
            _, mode, match_id, event_token, page_text = data.split(
                ":",
                maxsplit=4,
            )
            page = int(page_text)
        except (ValueError, TypeError):
            await _render_dates(update, 0)
            return
        event_date = None if event_token == "_" else event_token
        await _refresh_detail(
            update,
            context,
            mode,
            match_id,
            event_date,
            page,
        )
        return

    if data.startswith("follow:start:"):
        try:
            _, _, match_id, event_token, page_text = data.split(":", maxsplit=4)
            page = int(page_text)
        except (ValueError, TypeError):
            await _render_main_menu(update, context)
            return

        chat_id = update.effective_chat.id if update.effective_chat else None
        if chat_id is None:
            await _replace_callback_message(
                update,
                text="No pude identificar el chat para iniciar seguimiento.",
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("Menu principal", callback_data="nav:main")]]
                ),
            )
            return

        existing = FOLLOW_JOBS.get(chat_id)
        if isinstance(existing, dict) and str(existing.get("phase") or "") == "running":
            existing_mid = str(existing.get("match_id") or "")
            if existing_mid != match_id:
                await _replace_callback_message(
                    update,
                    text=(
                        f"Ya hay seguimiento activo para {existing_mid}.\n"
                        "Detenlo antes de iniciar otro."
                    ),
                    reply_markup=_follow_status_keyboard(existing),
                )
                return

        data_row = _get_match_detail(match_id)
        status_type = str((data_row or {}).get("match", {}).get("status_type", "") or "").lower()
        if status_type == "finished":
            await _replace_callback_message(
                update,
                text=f"El match {match_id} ya esta finalizado. No requiere seguimiento.",
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("Menu principal", callback_data="nav:main")]]
                ),
            )
            return

        event_date = None if event_token == "_" else event_token
        FOLLOW_JOBS[chat_id] = {
            "match_id": match_id,
            "event_token": event_token,
            "page": page,
            "phase": "starting",
            "stop_requested": False,
            "interval_seconds": FOLLOW_REFRESH_SECONDS,
            "stale_cycles_limit": FOLLOW_STALE_CYCLES_LIMIT,
        }
        if query.message:
            FOLLOW_JOBS[chat_id]["status_message_id"] = query.message.message_id
        
        try:
            pred_row = _get_or_compute_predictions(match_id, data_row)
            FOLLOW_JOBS[chat_id]["follow_data"] = data_row
            FOLLOW_JOBS[chat_id]["follow_pred"] = pred_row
        except Exception:
            pred_row = None

        await _publish_follow_status(
            context.application,
            chat_id,
            FOLLOW_JOBS[chat_id],
            force=True,
        )
        context.application.create_task(_run_follow_match_job(context.application, chat_id))

        job = FOLLOW_JOBS.get(chat_id)
        if not isinstance(job, dict):
            await _render_main_menu(update, context)
            return

        return

    if data == "follow:stop":
        chat_id = update.effective_chat.id if update.effective_chat else None
        job = FOLLOW_JOBS.get(chat_id) if chat_id is not None else None
        if not isinstance(job, dict):
            await _replace_callback_message(
                update,
                text="No hay seguimiento activo para detener.",
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("Menu principal", callback_data="nav:main")]]
                ),
            )
            return

        job["stop_requested"] = True
        await _replace_callback_message(
            update,
            text="Deteniendo seguimiento...",
            reply_markup=_follow_status_keyboard(job),
        )
        return

    if data.startswith("follow:open:"):
        try:
            _, _, match_id, event_token, page_text = data.split(":", maxsplit=4)
            page = int(page_text)
        except (ValueError, TypeError):
            await _render_main_menu(update, context)
            return

        event_date = None if event_token == "_" else event_token
        if event_date is None:
            await _render_live_detail(update, context, match_id, page)
            return
        await _render_match_detail(
            update,
            context.application,
            match_id,
            event_date,
            page,
        )
        return

    if data.startswith("delmatch:confirm:"):
        try:
            _, _, match_id, event_token, page_text = data.split(":", maxsplit=4)
            page = int(page_text)
        except (ValueError, TypeError):
            await _render_dates(update, 0)
            return
        event_date = None if event_token == "_" else event_token
        await _replace_callback_message(
            update,
            text=(
                f"Vas a borrar el match {match_id} de la base y sus dependencias.\n"
                "Incluye: quarter_scores, play_by_play, graph_points, eval_match_results.\n"
                "Esta accion no se puede deshacer."
            ),
            reply_markup=InlineKeyboardMarkup(
                [
                    [
                        InlineKeyboardButton(
                            "Confirmar borrado",
                            callback_data=f"delmatch:run:{match_id}:{event_token}:{page}",
                        )
                    ],
                    [
                        InlineKeyboardButton(
                            "Cancelar",
                            callback_data=f"match:{match_id}:{event_token}:{page}",
                        )
                    ],
                ]
            ),
        )
        return

    if data.startswith("delmatch:run:"):
        try:
            _, _, match_id, event_token, page_text = data.split(":", maxsplit=4)
            page = int(page_text)
        except (ValueError, TypeError):
            await _render_dates(update, 0)
            return
        chat_id = update.effective_chat.id if update.effective_chat else None
        if not _is_train_allowed(chat_id):
            await _replace_callback_message(
                update,
                text="No autorizado para borrar matches.",
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("Menu principal", callback_data="nav:main")]]
                ),
            )
            return
        counts = await asyncio.to_thread(_delete_match_cascade, match_id)
        event_date = None if event_token == "_" else event_token
        if event_date:
            rows_after = _fetch_matches_for_date(event_date)
            if rows_after:
                await _render_matches_for_date(update, event_date, page)
            else:
                summary = (
                    f"Match {match_id} borrado.\n"
                    f"matches={counts['matches']} quarter_scores={counts['quarter_scores']}\n"
                    f"play_by_play={counts['play_by_play']} graph_points={counts['graph_points']}\n"
                    f"eval={counts['eval_match_results']} discovered={counts['discovered_ft_matches']}\n\n"
                    f"Ya no quedan matches en {event_date}."
                )
                await _replace_callback_message(
                    update,
                    text=summary,
                    reply_markup=InlineKeyboardMarkup(
                        [
                            [InlineKeyboardButton("Volver a fechas", callback_data="dates:0")],
                            [InlineKeyboardButton("Menu principal", callback_data="nav:main")],
                        ]
                    ),
                )
        else:
            summary = (
                f"Match {match_id} borrado.\n"
                f"matches={counts['matches']} quarter_scores={counts['quarter_scores']}\n"
                f"play_by_play={counts['play_by_play']} graph_points={counts['graph_points']}\n"
                f"eval={counts['eval_match_results']} discovered={counts['discovered_ft_matches']}"
            )
            await _replace_callback_message(
                update,
                text=summary,
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("Menu principal", callback_data="nav:main")]]
                ),
            )
        return

    if data.startswith("calc:date:"):
        event_date = data[len("calc:date:"):]
        chat_id = update.effective_chat.id if update.effective_chat else None
        message = query.message if query else None
        message_id = message.message_id if message else None
        if chat_id is None or message_id is None:
            return
        rows_preview = _fetch_matches_for_date(event_date)
        total_preview = len(rows_preview)
        _rc_label = _event_date_title_es(event_date, total_preview).split("[")[0].strip()
        _rc_model = f"Q3={MODEL_CONFIG['q3']} Q4={MODEL_CONFIG['q4']}"
        await _set_waiting_state(
            update,
            text=f"Recalculando Partidos {_rc_label}\n0/{total_preview} Procesados {_rc_model}",
        )
        context.application.create_task(
            _run_calc_date_job(context.application, chat_id, message_id, event_date)
        )
        return

    if data.startswith("refreshlive:"):
        try:
            _, mode, match_id, page_text = data.split(":", maxsplit=3)
            page = int(page_text)
        except (ValueError, TypeError):
            await _render_live(update, context, 0, refetch=False)
            return
        await _refresh_live_detail(update, context, mode, match_id, page)
        return

    if data.startswith("dellive:confirm:"):
        try:
            _, _, match_id, page_text = data.split(":", maxsplit=3)
            page = int(page_text)
        except (ValueError, TypeError):
            await _render_live(update, context, 0, refetch=False)
            return
        await _replace_callback_message(
            update,
            text=(
                f"Vas a borrar el match {match_id} de la base y sus dependencias.\n"
                "Incluye: quarter_scores, play_by_play, graph_points, eval_match_results.\n"
                "Esta accion no se puede deshacer."
            ),
            reply_markup=InlineKeyboardMarkup(
                [
                    [
                        InlineKeyboardButton(
                            "Confirmar borrado",
                            callback_data=f"dellive:run:{match_id}:{page}",
                        )
                    ],
                    [InlineKeyboardButton("Cancelar", callback_data=f"livematch:{match_id}:{page}")],
                ]
            ),
        )
        return

    if data.startswith("dellive:run:"):
        try:
            _, _, match_id, page_text = data.split(":", maxsplit=3)
            page = int(page_text)
        except (ValueError, TypeError):
            await _render_live(update, context, 0, refetch=False)
            return
        chat_id = update.effective_chat.id if update.effective_chat else None
        if not _is_train_allowed(chat_id):
            await _replace_callback_message(
                update,
                text="No autorizado para borrar matches.",
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("Menu principal", callback_data="nav:main")]]
                ),
            )
            return
        await asyncio.to_thread(_delete_match_cascade, match_id)
        await _render_live(update, context, page, refetch=True)
        return

    if data.startswith("livematch:"):
        try:
            _, match_id, page_text = data.split(":", maxsplit=2)
            page = int(page_text)
        except (ValueError, TypeError):
            await _render_live(update, context, 0, refetch=False)
            return
        await _render_live_detail(update, context, match_id, page)
        return

    if data.startswith("notifmatch:"):
        # "Ver match" from a BET/NO_BET notification — open detail as a NEW
        # message so the original notification stays visible for reference.
        try:
            _, match_id, event_date, page_text = data.split(":", maxsplit=3)
            page = int(page_text)
        except (ValueError, TypeError):
            await query.answer()
            return
        event_date_val = None if event_date == "_" else event_date
        chat_id = update.effective_chat.id if update.effective_chat else None
        if chat_id is None:
            await query.answer()
            return
        # Dismiss the spinner on the notification without touching the message
        await query.answer()
        # Fetch data + predictions then send as a fresh message
        data_row = _get_match_detail(match_id)
        if not data_row:
            data_row = await asyncio.to_thread(_refresh_match_data, match_id)
        if data_row is None:
            await context.application.bot.send_message(
                chat_id=chat_id,
                text=f"No se pudo obtener datos del match {match_id}.",
            )
            return
        pred_row = await asyncio.to_thread(_get_or_compute_predictions, match_id, data_row)
        detail_text = _detail_text(match_id, data_row, pred_row)
        keyboard = _detail_keyboard(
            match_id,
            event_date=event_date_val,
            page=page,
            match_data=data_row,
            chat_id=chat_id,
        )
        image_path: Path | None = None
        try:
            image_path = _build_graph_image(match_id, data_row)
        except Exception:
            image_path = None
        if image_path is not None and image_path.exists():
            try:
                with image_path.open("rb") as _f:
                    await context.application.bot.send_photo(
                        chat_id=chat_id,
                        photo=_f,
                        caption=detail_text,
                        reply_markup=keyboard,
                    )
            finally:
                try:
                    image_path.unlink()
                except OSError:
                    pass
        else:
            await context.application.bot.send_message(
                chat_id=chat_id,
                text=detail_text,
                reply_markup=keyboard,
            )
        return

    if data.startswith("match:"):
        try:
            _, match_id, event_date, page_text = data.split(":", maxsplit=3)
            page = int(page_text)
        except (ValueError, TypeError):
            await _render_dates(update, 0)
            return
        event_date = None if event_date == "_" else event_date

        # Cancel any V12 LIVE tracking for this match
        if update.callback_query and update.callback_query.message:
            chat_id = update.callback_query.message.chat_id
            msg_id = update.callback_query.message.message_id
            # Cancel all V12 live jobs for this chat/message
            for job_key in list(V12_LIVE_JOBS.keys()):
                job = V12_LIVE_JOBS[job_key]
                if job.get("chat_id") == chat_id:
                    job["cancelled"] = True

        await _render_match_detail(
            update,
            context.application,
            match_id,
            event_date,
            page,
        )
        return

    if data.startswith("debuginf:"):
        try:
            _, match_id, event_date_tok, page_text = data.split(":", maxsplit=3)
            page = int(page_text)
        except (ValueError, TypeError):
            await _render_dates(update, 0)
            return
        event_date_val = None if event_date_tok == "_" else event_date_tok
        await _render_inference_debug(update, match_id, event_date_val, page)
        return


async def _handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message:
        return

    raw = (update.message.text or "").strip()
    if raw.lower() == MENU_BUTTON_TEXT.lower():
        await _render_main_menu(update, context)
        return

    if raw.lower() == STATS_BUTTON_TEXT.lower():
        await _send_stats_report(update, context)
        return

    if raw.lower() == MONTHLY_BUTTON_TEXT.lower():
        from datetime import datetime as _dt_now
        _now = _dt_now.utcnow()
        _MONTH_ES = {
            1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
            5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
            9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
        }
        _month_buttons = []
        for _i in range(6):
            _m = _now.month - _i
            _y = _now.year
            while _m <= 0:
                _m += 12
                _y -= 1
            _ym_opt = f"{_y:04d}-{_m:02d}"
            _label = f"{'📅 ' if _i == 0 else ''}{_MONTH_ES[_m]} {_y}"
            _month_buttons.append([InlineKeyboardButton(_label, callback_data=f"monthly_pick:{_ym_opt}")])
        _month_buttons.append([InlineKeyboardButton("⬅️ Cancelar", callback_data="nav:main")])
        await update.message.reply_text(
            "📆 Selecciona el mes del reporte:",
            reply_markup=InlineKeyboardMarkup(_month_buttons),
        )
        return

    if raw.lower() == ID_BUTTON_TEXT.lower():
        context.user_data[AWAITING_MATCH_ID_KEY] = True
        await update.message.reply_text(
            "Ingresa el Match ID (número):",
            reply_markup=_menu_reply_keyboard(),
        )
        return

    # Auto-detect: a bare numeric string that looks like a match ID
    if re.fullmatch(r"\d{6,}", raw) and not context.user_data.get(AWAITING_FETCH_DATE_KEY):
        context.user_data[AWAITING_MATCH_ID_KEY] = False
        data = await _get_match_detail_async(raw, update=update)
        if not data:
            await update.message.reply_text(
                f"No se encontró el match {raw}.",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("Buscar otro ID", callback_data="menu:id")],
                    [InlineKeyboardButton("Menu principal", callback_data="nav:main")],
                ]),
            )
            return
        wait_msg = await update.message.reply_text("Espere...", reply_markup=_menu_reply_keyboard())
        await _send_detail_message(
            update, context.application, raw, data,
            _get_or_compute_predictions(raw, data), None, 0,
        )
        try:
            await wait_msg.delete()
        except Exception:
            pass
        return

    if context.user_data.get(AWAITING_FETCH_DATE_KEY):
        parts = [p for p in raw.replace(",", " ").split() if p]
        if not parts:
            await update.message.reply_text(
                "Fecha invalida. Usa YYYY-MM-DD o YYYY-MM-DD 100"
            )
            return

        try:
            date_obj = datetime.strptime(parts[0], "%Y-%m-%d")
            event_date = date_obj.date().isoformat()
        except ValueError:
            await update.message.reply_text(
                "Fecha invalida. Usa formato YYYY-MM-DD."
            )
            return

        limit: int | None = None
        if len(parts) >= 2:
            try:
                limit = int(parts[1])
            except ValueError:
                await update.message.reply_text(
                    "Limite invalido. Usa entero positivo."
                )
                return
            if limit <= 0:
                await update.message.reply_text(
                    "Limite invalido. Usa entero positivo."
                )
                return

        context.user_data[AWAITING_FETCH_DATE_KEY] = False
        chat_id = update.effective_chat.id if update.effective_chat else None
        if chat_id is not None and chat_id in DATE_INGEST_JOBS:
            await update.message.reply_text(
                "Ya hay una ingesta de fecha en curso en este chat. Puedes cancelarla desde su mensaje de progreso.",
                reply_markup=_menu_reply_keyboard(),
            )
            return
        status_msg = await update.message.reply_text(
            _date_ingest_progress_text(
                event_date,
                limit,
                {"phase": "starting"},
            ),
            reply_markup=_date_ingest_progress_keyboard(True),
        )
        if chat_id is not None:
            context.application.create_task(
                _run_date_ingest_job(
                    context.application,
                    chat_id,
                    status_msg.message_id,
                    event_date,
                    limit,
                )
            )
        await update.message.reply_text(
            "Proceso iniciado. Te voy avisando el progreso aqui.",
            reply_markup=_menu_reply_keyboard(),
        )
        return

    if not context.user_data.get(AWAITING_MATCH_ID_KEY):
        return

    match_id = raw
    if not re.fullmatch(r"\d{6,}", match_id):
        await update.message.reply_text(
            "ID invalido. Debe ser numerico. Ejemplo: 14442355"
        )
        return

    context.user_data[AWAITING_MATCH_ID_KEY] = False

    data = await _get_match_detail_async(match_id, update=update)
    if not data:
        await update.message.reply_text(
            text=f"No se encontro el match {match_id}.",
            reply_markup=InlineKeyboardMarkup(
                [
                    [InlineKeyboardButton("Buscar otro ID", callback_data="menu:id")],
                    [InlineKeyboardButton("Menu principal", callback_data="nav:main")],
                ]
            ),
        )
        return

    wait_msg = await update.message.reply_text(
        text="Espere...",
        reply_markup=_menu_reply_keyboard(),
    )
    await _send_detail_message(
        update,
        context.application,
        match_id,
        data,
        _get_or_compute_predictions(match_id, data),
        None,
        0,
    )
    try:
        await wait_msg.delete()
    except Exception:
        pass


# ── /reporte command ──────────────────────────────────────────────────────────

REPORT_JOBS: dict[int, dict] = {}


def _report_progress_text(ps: dict) -> str:
    total = int(ps.get("total") or 0)
    processed = int(ps.get("processed") or 0)
    phase = str(ps.get("phase") or "...")
    current = str(ps.get("current") or "")
    pct = round(100 * processed / total) if total else 0
    bar_full = 10
    bar_done = int(bar_full * pct / 100)
    bar = "█" * bar_done + "░" * (bar_full - bar_done)
    phase_map = {
        "loading": "Cargando datos...",
        "processing": "Procesando partidos...",
        "writing": "Escribiendo Excel...",
        "done": "Listo ✓",
        "starting": "Iniciando...",
    }
    phase_text = phase_map.get(phase, phase)
    lines = [
        f"📊 Generando reporte de modelos",
        f"[{bar}] {pct}% ({processed}/{total})",
        f"Estado: {phase_text}",
    ]
    if current and phase == "processing":
        lines.append(f"📌 {current[:60]}")
    return "\n".join(lines)


async def _run_report_job(
    app: Application,
    chat_id: int,
    message_id: int,
    month: str,
) -> None:
    """Non-blocking report generation task. Polls progress and sends xlsx when done."""
    progress_state: dict = {
        "total": 0, "processed": 0,
        "phase": "starting", "current": "", "output_path": "",
    }
    REPORT_JOBS[chat_id] = {"message_id": message_id, "progress_state": progress_state}

    task = asyncio.create_task(
        asyncio.to_thread(
            _generate_report_sync,
            month,
            progress_state,
        )
    )

    last_text = ""
    try:
        while not task.done():
            text = _report_progress_text(progress_state)
            if text != last_text:
                await _safe_edit_message(
                    app,
                    chat_id=chat_id,
                    message_id=message_id,
                    text=text,
                )
                last_text = text
            await asyncio.sleep(3)

        result = await task
    except Exception as exc:
        await _safe_edit_message(
            app,
            chat_id=chat_id,
            message_id=message_id,
            text=f"❌ Error generando reporte: {exc}",
        )
        REPORT_JOBS.pop(chat_id, None)
        return

    if result.get("error"):
        await _safe_edit_message(
            app,
            chat_id=chat_id,
            message_id=message_id,
            text=f"❌ Error: {result['error']}",
        )
        REPORT_JOBS.pop(chat_id, None)
        return

    out_path = result.get("output_path")
    total = int(progress_state.get("total") or 0)
    await _safe_edit_message(
        app,
        chat_id=chat_id,
        message_id=message_id,
        text=f"📊 Reporte {month} listo — {total} partidos analizados.\nEnviando archivo...",
    )

    try:
        with open(out_path, "rb") as f:
            await app.bot.send_document(
                chat_id=chat_id,
                document=f,
                filename=Path(out_path).name,
                caption=f"📊 Comparación de modelos — {month}",
            )
    except Exception as exc:
        await app.bot.send_message(
            chat_id=chat_id,
            text=f"⚠️ No se pudo enviar el archivo: {exc}\nGuardado en: {out_path}",
        )

    REPORT_JOBS.pop(chat_id, None)


def _generate_report_sync(month: str, progress_state: dict) -> dict:
    """Blocking wrapper for generate_report, safe to run in a thread."""
    try:
        report_mod = importlib.import_module("training.report_model_comparison")
        out = report_mod.generate_report(month, progress_state=progress_state)
        return {"output_path": str(out)}
    except Exception as exc:
        return {"error": str(exc)}


async def reporte_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle /reporte [YYYY-MM] command."""
    if not _is_allowed(update):
        return

    chat_id = update.effective_chat.id if update.effective_chat else None
    if chat_id is None:
        return

    if chat_id in REPORT_JOBS:
        await update.message.reply_text("Ya hay un reporte en curso para este chat. Espera que termine.")
        return

    args = context.args or []
    month = args[0].strip() if args else ""

    if not month:
        from datetime import datetime as _dt
        month = _dt.utcnow().strftime("%Y-%m")
        await update.message.reply_text(
            f"No se especificó mes. Generando reporte de {month}.\n"
            f"Uso: /reporte YYYY-MM"
        )

    try:
        from datetime import datetime as _dt
        _dt.strptime(month, "%Y-%m")
    except ValueError:
        await update.message.reply_text(
            f"Formato de mes inválido: '{month}'. Usa YYYY-MM (ej: 2026-03)."
        )
        return

    msg = await update.message.reply_text(_report_progress_text({"total": 0, "processed": 0, "phase": "starting", "current": ""}))
    context.application.create_task(
        _run_report_job(context.application, chat_id, msg.message_id, month)
    )


def main() -> None:
    if not BOT_TOKEN:
        raise SystemExit(
            "Falta TELEGRAM_BOT_TOKEN. Define la variable en .env"
        )

    # Create custom request with extended timeouts
    request = CustomHTTPXRequest()
    app = Application.builder().token(BOT_TOKEN).request(request).post_init(_post_init).build()

    app.add_handler(CommandHandler("start", start_cmd))
    app.add_handler(CommandHandler("myid", myid_cmd))
    app.add_handler(CommandHandler("trainstatus", train_status_cmd))
    app.add_handler(CommandHandler("dates", dates_cmd))
    app.add_handler(CommandHandler("signals", signals_cmd))
    app.add_handler(CommandHandler("report", signals_report_cmd))
    app.add_handler(CommandHandler("monitor", monitor_cmd))
    app.add_handler(CommandHandler("reporte", reporte_cmd))
    app.add_handler(CommandHandler("modelstats", modelstats_cmd))
    app.add_handler(CallbackQueryHandler(_handle_callback))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, _handle_text))
    app.add_error_handler(_on_error)

    print("[telegram-bot] iniciado")
    if ALLOWED_CHAT_IDS:
        print(f"[telegram-bot] allowed_chat_ids={sorted(ALLOWED_CHAT_IDS)}")
    else:
        print("[telegram-bot] allowed_chat_ids=ALL (sin restriccion)")
    
    # Start polling with retry logic for bootstrap network errors
    max_retries = 3
    retry_delay = 5  # seconds
    for attempt in range(1, max_retries + 1):
        try:
            app.run_polling(allowed_updates=Update.ALL_TYPES)
            break  # If successful, exit the retry loop
        except KeyboardInterrupt:
            print("\n[telegram-bot] Interrupción del usuario. Saliendo...")
            break
        except (NetworkError, TimedOut) as e:
            if attempt < max_retries:
                logger.error(
                    f"[bootstrap] NetworkError en intento {attempt}/{max_retries}: {e}"
                )
                logger.info(f"[bootstrap] Reintentando en {retry_delay} segundos...")
                time.sleep(retry_delay)
            else:
                logger.error(
                    f"[bootstrap] Error definitivo después de {max_retries} intentos: {e}"
                )
                raise


if __name__ == "__main__":
    main()

