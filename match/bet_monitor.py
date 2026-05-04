"""
bet_monitor.py — Betting window monitor daemon.

Smart alarm system that:
  1. Fetches today's / tomorrow's basketball schedule and stores it in
     `bet_monitor_schedule`.
  2. For each match, wakes up near the Q3 / Q4 window (estimated from live
     graph_points), fetches live data, runs inference, and fires a Telegram
     notification when the model says BET.
  3. Exposes MONITOR_STATUS + helper functions for the Telegram bot UI.
  4. Recovers on restart: reads the schedule table and resumes any pending
     matches.

Usage (inside telegam_bot.py):
    import bet_monitor
    bet_monitor.set_notify_callback(my_async_send_fn)
    bet_monitor.set_model_config({"q3": "v4", "q4": "v4"})
    stop_ev = asyncio.Event()
    task = asyncio.create_task(bet_monitor.run_monitor(DB_PATH, stop_ev))
"""

from __future__ import annotations

import asyncio
import importlib
import json
import logging
import os
import re
import sqlite3
import sys
import time
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Awaitable, Callable, TextIO

BASE_DIR = Path(__file__).resolve().parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

logger = logging.getLogger(__name__)

# ── Timing constants ──────────────────────────────────────────────────────────
UTC_OFFSET_HOURS = -6        # must match telegram_bot.py

Q3_MINUTE = 24               # game-minute where Q3 data is required
Q4_MINUTE = 36               # game-minute where Q4 data is required
WAKE_BEFORE_MINUTES = 4      # start polling this many game-minutes before cutoff
POLL_NEAR_SECS = 100         # poll interval when within WAKE_BEFORE window
IDLE_POLL_SECS = 300         # poll interval when far from next window
NO_BET_CONFIRM_TICKS = 2     # extra ticks before sending uncertain NO BET notification
Q4_WAITING_MAX_TICKS = 6     # max re-eval ticks waiting for valid Q3 progression
Q4_STALE_MAX_TICKS = 4       # max ticks with frozen graph_points before aborting Q4
UNAVAILABLE_LOG_EVERY = 1    # log cadence for transient UNAVAILABLE retry ticks
SECS_PER_GAME_MIN = 170      # initial estimate: ~2.8 real-min per game-minute
MIN_GP_Q3 = 16               # graph_points with minute ≤ 24 required for Q3 check
MIN_GP_Q4 = 26               # graph_points with minute ≤ 36 required for Q4 check
SKIP_Q3 = True               # temporarily disable Q3 monitoring to reduce CPU load
MAX_FETCH_ERRORS = 4         # consecutive errors before discarding a match
SCHEDULE_REFRESH_HOURS = 8   # re-fetch schedule this often
PENDING_RECHECK_SECS = 15 * 60  # periodically backfill pending outcomes
PENDING_RECHECK_MAX_MATCHES = 40  # cap scrape workload per cycle
PENDING_SCHEDULE_RECHECK_SECS = 15 * 60  # periodically revisit pending schedule rows
PENDING_SCHEDULE_MAX_MATCHES = 20  # cap pending-schedule scrape workload
PENDING_SCHEDULE_MIN_AGE_SECS = 45 * 60  # only recheck matches older than this
PENDING_RECHECK_MAX_FETCHES_IN_FLIGHT = 2  # skip periodic rechecks only if fetch load is high
NO_GRAPH_REAL_SECS = 55 * 60 # discard if no graph_points 55 real-min after start
MAX_CONCURRENT_FETCHES = 6   # max simultaneous Playwright fetches across all watchers
FINAL_FETCH_EXTRA_SECS = 300 # extra real-seconds after estimated end before final save
FINAL_FETCH_MIN_GP = 8       # require at least this many graph_points to attempt save
MONITOR_LOG_DIR = BASE_DIR / "logs"  # daily monitor logs written here

# ── Global state (read by telegram_bot.py) ────────────────────────────────────
MONITOR_STATUS: dict = {
    "running": False,
    "started_at": None,
    "today_total": 0,
    "tomorrow_total": 0,
    "checked_q3": 0,
    "checked_q4": 0,
    "bets_sent": 0,
    "no_bet": 0,
    "discarded": 0,
    "active_matches": [],
    "last_event": "",
}

# Injected by telegram_bot.py before starting the monitor
_notify_cb: Callable[..., Awaitable[None]] | None = None
_model_config: dict[str, str] = {"q3": "v4", "q4": "v4"}

# Persistent engine cache: loaded once, reused across inference calls
_ENGINE_CACHE: dict[str, object] = {}
_daily_log_handler: logging.Handler | None = None
_ANSI_RE = re.compile(r"\x1b\[[0-9;]*m")


def _is_bet_signal(signal: str) -> bool:
    """True for BET, BET_HOME, BET_AWAY — but not NO_BET."""
    s = str(signal).upper()
    return "BET" in s and "NO_BET" not in s and "NO BET" not in s


def _q3_timing() -> tuple[int, int, int]:
    """Return (cutoff_minute, min_gp, wake_before) for Q3.

    V13/V15/V16/V17 were trained with graph_points up to minute 22, so they
    can fire inference at minute ~22 — near halftime for 10-min quarter
    leagues (Q2 ends at minute 20) and 2 min pre-Q3 for 12-min leagues.
    A wake_before of 6 means polling starts at minute 16.
    Older models need data up to minute 24 and wake at minute 20.
    """
    model = _model_config.get("q3", "v4")
    if model in ("v13", "v15", "v16", "v17"):
        # Cutoff at minute 22; start polling at minute 16.
        return 22, 14, 6
    # default (v2, v3, v4, v6, v9, v12, ...)
    return Q3_MINUTE, MIN_GP_Q3, WAKE_BEFORE_MINUTES


def _q4_timing() -> tuple[int, int, bool, int]:
    """Return (cutoff_minute, min_gp, requires_q3_score, wake_before) for Q4.

    V13/V15/V16/V17 fire Q4 inference at minute ~31 (pre-Q4 or start of Q4
    in 10-min quarter leagues). A large wake_before ensures the monitor
    starts polling long before Q4 so the user has time to place the bet.
    Older models need minute-36 data with Q3 complete.
    """
    model = _model_config.get("q4", "v4")
    if model in ("v13", "v15", "v16", "v17"):
        # Cutoff at minute 31 (late Q3 / pre-Q4).
        # requires_q3_score=True ensures Q3 final score exists before Q4 inference.
        # wake_before=6: start polling at minute ~25 (Q3 underway), never in Q2.
        return 31, 16, True, 6
    # default (v4, v6, v9, v12, ...)
    return Q4_MINUTE, MIN_GP_Q4, True, WAKE_BEFORE_MINUTES


def set_notify_callback(cb: Callable[..., Awaitable[None]]) -> None:
    global _notify_cb
    _notify_cb = cb


def set_model_config(config: dict[str, str]) -> None:
    global _model_config
    _model_config = {**_model_config, **config}


# Lazy semaphore — created inside the monitor thread's event loop
_fetch_sem: asyncio.Semaphore | None = None


def _get_fetch_sem() -> asyncio.Semaphore:
    global _fetch_sem
    if _fetch_sem is None:
        _fetch_sem = asyncio.Semaphore(MAX_CONCURRENT_FETCHES)
    return _fetch_sem


def _fetches_in_flight() -> int:
    """Best-effort count of active monitor fetches using semaphore usage."""
    sem = _fetch_sem
    if sem is None:
        return 0
    cur_value = getattr(sem, "_value", MAX_CONCURRENT_FETCHES)
    try:
        in_flight = MAX_CONCURRENT_FETCHES - int(cur_value)
    except (TypeError, ValueError):
        return 0
    return max(0, in_flight)


# ── Internal helpers ──────────────────────────────────────────────────────────

def _decorate_quarter_tokens(msg: str) -> str:
    """Decorate Q3/Q4 tokens and match IDs for easier terminal scanning."""
    out = str(msg or "")
    out = re.sub(r"\bQ3\b(?!\s*🔵)", "Q3 🔵", out)
    out = re.sub(r"\bQ4\b(?!\s*🟠)", "Q4 🟠", out)

    # Highlight common match-id patterns in terminal logs.
    out = re.sub(
        r"(match_id\s*=\s*)(\d{6,12})",
        lambda m: f"{m.group(1)}{_color_text(m.group(2), '95')}",
        out,
        flags=re.IGNORECASE,
    )
    out = re.sub(
        r"\((\d{6,12})\)",
        lambda m: f"({_color_text(m.group(1), '95')})",
        out,
    )
    return out


def _monitor_local_now() -> datetime:
    """Monitor local time using configured UTC offset."""
    return datetime.now(timezone.utc) + timedelta(hours=UTC_OFFSET_HOURS)


def _monitor_local_today_str() -> str:
    return _monitor_local_now().date().isoformat()


def _monitor_local_tomorrow_str() -> str:
    return (_monitor_local_now().date() + timedelta(days=1)).isoformat()


def _supports_ansi_color() -> bool:
    if os.getenv("NO_COLOR"):
        return False
    try:
        return bool(sys.stdout.isatty())
    except Exception:
        return False


def _color_text(text: str, code: str = "93") -> str:
    """Wrap text with ANSI color if terminal supports it."""
    if not _supports_ansi_color():
        return text
    return f"\x1b[{code}m{text}\x1b[0m"


def _format_wait_eta(seconds: float) -> str:
    total_min = max(0, int(round(seconds / 60.0)))
    days, rem_min = divmod(total_min, 24 * 60)
    hours, mins = divmod(rem_min, 60)
    parts: list[str] = []
    if days > 0:
        parts.append(f"{days}d")
    if hours > 0 or days > 0:
        parts.append(f"{hours}h")
    parts.append(f"{mins}m")
    return " ".join(parts)


def _format_sched_local_label(scheduled_ts: int | None) -> str:
    """Return local scheduled time label for logs."""
    if not scheduled_ts:
        return "????-??-?? ??:??"
    try:
        dt_local = datetime.fromtimestamp(int(scheduled_ts), tz=timezone.utc) + timedelta(hours=UTC_OFFSET_HOURS)
    except (TypeError, ValueError, OSError, OverflowError):
        return "????-??-?? ??:??"
    return dt_local.strftime('%Y-%m-%d %H:%M')


class _DailyMonitorFileHandler(logging.Handler):
    """Write monitor logs to logs/monitor-YYYY-MM-DD.log and rotate daily."""

    def __init__(self, log_dir: Path):
        super().__init__()
        self.log_dir = log_dir
        self.current_date = ""
        self._fh: TextIO | None = None

    def _target_path(self, date_str: str) -> Path:
        return self.log_dir / f"monitor-{date_str}.log"

    def _ensure_file_for_date(self, date_str: str) -> None:
        if self._fh is not None and self.current_date == date_str:
            return
        if self._fh is not None:
            try:
                self._fh.close()
            except Exception:
                pass
        self.log_dir.mkdir(parents=True, exist_ok=True)
        path = self._target_path(date_str)
        self._fh = path.open("a", encoding="utf-8")
        self.current_date = date_str

    def emit(self, record: logging.LogRecord) -> None:
        try:
            date_str = _monitor_local_now().date().isoformat()
            self._ensure_file_for_date(date_str)
            if self._fh is None:
                return
            msg = self.format(record)
            msg = _ANSI_RE.sub("", msg)
            self._fh.write(msg + "\n")
            self._fh.flush()
        except Exception:
            self.handleError(record)

    def close(self) -> None:
        try:
            if self._fh is not None:
                self._fh.close()
                self._fh = None
        finally:
            super().close()


def _ensure_daily_file_logging() -> None:
    global _daily_log_handler
    if _daily_log_handler is not None:
        return
    handler = _DailyMonitorFileHandler(MONITOR_LOG_DIR)
    handler.setLevel(logging.INFO)
    handler.setFormatter(
        logging.Formatter(
            "%(asctime)s [%(levelname)s] %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
    )
    logger.addHandler(handler)
    _daily_log_handler = handler

def _log(msg: str, level: str = "info") -> None:
    dec_msg = _decorate_quarter_tokens(msg)
    MONITOR_STATUS["last_event"] = dec_msg
    getattr(logger, level)("[MONITOR] %s", dec_msg)


async def _notify(
    msg: str,
    reply_markup: dict | None = None,
    notify_type: str = "bet",
    quarter: str | None = None,
) -> None:
    """notify_type: 'bet' | 'no_bet' | 'result' | 'filtered_bet'"""
    if _notify_cb:
        try:
            await _notify_cb(msg, reply_markup, notify_type, quarter)
        except Exception as exc:
            logger.error("[MONITOR] notify error: %s", exc)
    else:
        logger.info("[MONITOR][NOTIFY] %s", msg)


def _should_notify_filtered_bets(db_path: str) -> bool:
    """Read persisted toggle for informational notifications on filtered BETs."""
    try:
        conn = _open_db(db_path)
        row = conn.execute(
            "SELECT value FROM settings WHERE key='monitor_notify_filtered_bet'"
        ).fetchone()
        conn.close()
        if row is None:
            return True
        val = str(row[0] if not isinstance(row, sqlite3.Row) else row["value"]).strip().lower()
        return val in {"1", "true", "yes", "on"}
    except Exception:
        return True


def _format_filtered_bet_notification(
    *,
    match_id: str,
    data: dict,
    quarter_label: str,
    model_used: str,
    home: str,
    away: str,
    league: str,
    event_date: str,
    current_minute: int | None,
    pick: str,
    confidence: float,
    p_pick: float | None,
    reason: str,
) -> str:
    pick_sym = "🏠" if pick == "home" else ("✈️" if pick == "away" else "❔")
    pick_name = home if pick == "home" else (away if pick == "away" else "Sin pick")

    qs = ((data.get("score") or {}).get("quarters") or {})
    q3 = qs.get("Q3") if isinstance(qs.get("Q3"), dict) else None
    if q3 is None:
        q3_txt = "-"
    else:
        q3h = int(q3.get("home", 0) or 0)
        q3a = int(q3.get("away", 0) or 0)
        q3_txt = f"{q3h} - {q3a}"

    p_pick_txt = "-"
    if p_pick is not None:
        p_pick_txt = f"{p_pick * 100:.1f}%"

    min_txt = "-" if current_minute is None else str(current_minute)
    return (
        f"🟡 BET FILTRADA {quarter_label} [{model_used}]\n\n"
        f"Match ID: {match_id} | Min: {min_txt}\n"
        f"{home} vs {away}\n"
        f"Fecha: {event_date} UTC{UTC_OFFSET_HOURS:+d}\n"
        f"Liga: {league}\n"
        f"Q3 actual: {q3_txt}\n\n"
        f"Pick sugerido: {pick_sym} {pick_name}\n"
        f"Confianza: {confidence * 100:.1f}%\n"
        f"p_pick: {p_pick_txt}\n"
        f"Motivo filtro: {reason}"
    )


# ── DB ────────────────────────────────────────────────────────────────────────

def init_tables(db_path: str) -> None:
    conn = _open_db(db_path)
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS bet_monitor_schedule (
            match_id          TEXT PRIMARY KEY,
            event_date        TEXT NOT NULL,
            home_team         TEXT,
            away_team         TEXT,
            league            TEXT,
            scheduled_utc_ts  INTEGER,
            scheduled_utc     TEXT,
            status            TEXT DEFAULT 'pending',
            q3_checked        INTEGER DEFAULT 0,
            q4_checked        INTEGER DEFAULT 0,
            q3_signal         TEXT,
            q4_signal         TEXT,
            q3_pick           TEXT,
            q4_pick           TEXT,
            q3_confidence     REAL,
            q4_confidence     REAL,
            q3_model          TEXT,
            q4_model          TEXT,
            q3_notified       INTEGER DEFAULT 0,
            q4_notified       INTEGER DEFAULT 0,
            skip_reason       TEXT,
            created_at        TEXT DEFAULT (datetime('now')),
            updated_at        TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS bet_monitor_log (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            match_id       TEXT NOT NULL,
            event_date     TEXT,
            home_team      TEXT,
            away_team      TEXT,
            league         TEXT,
            target         TEXT,
            model          TEXT,
            signal         TEXT,
            recommendation TEXT,
            pick           TEXT,
            confidence     REAL,
            scraped_minute INTEGER,
            notified_at    TEXT,
            result         TEXT DEFAULT 'pending',
            result_checked INTEGER DEFAULT 0,
            created_at     TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS inference_debug_log (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            match_id       TEXT NOT NULL,
            target         TEXT NOT NULL,
            model          TEXT,
            scraped_minute INTEGER,
            signal         TEXT,
            confidence     REAL,
            gp_count       INTEGER,
            inference_json TEXT,
            created_at     TEXT DEFAULT (datetime('now'))
        );

        -- Add columns if they don't exist yet (safe to run on existing DBs)
        CREATE TABLE IF NOT EXISTS _dummy_q3model (x);
    """)
    # ALTER TABLE is not transactional in SQLite executescript; do it separately
    for col_def in [
        ("bet_monitor_schedule", "q3_model",       "TEXT"),
        ("bet_monitor_schedule", "q4_model",       "TEXT"),
        ("bet_monitor_schedule", "final_fetched",  "INTEGER DEFAULT 0"),
        ("bet_monitor_schedule", "final_fetch_at", "TEXT"),
        ("bet_monitor_log",      "model",           "TEXT"),
        ("bet_monitor_log",      "result_checked",  "INTEGER DEFAULT 0"),
    ]:
        try:
            conn.execute(
                f"ALTER TABLE {col_def[0]} ADD COLUMN {col_def[1]} {col_def[2]}"
            )
        except Exception:
            pass  # column already exists
    conn.commit()
    conn.close()


def _open_db(db_path: str) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA busy_timeout=30000")
    return conn


def _upsert_schedule_row(conn: sqlite3.Connection, row: dict) -> None:
    conn.execute(
        """
        INSERT INTO bet_monitor_schedule
            (match_id, event_date, home_team, away_team, league,
             scheduled_utc_ts, scheduled_utc, updated_at)
        VALUES (:match_id, :event_date, :home_team, :away_team, :league,
                :scheduled_utc_ts, :scheduled_utc, datetime('now'))
        ON CONFLICT(match_id) DO UPDATE SET
            event_date        = excluded.event_date,
            home_team         = excluded.home_team,
            away_team         = excluded.away_team,
            league            = excluded.league,
            scheduled_utc_ts  = excluded.scheduled_utc_ts,
            scheduled_utc     = excluded.scheduled_utc,
            updated_at        = excluded.updated_at
        WHERE bet_monitor_schedule.status = 'pending'
        """,
        row,
    )


def _log_daily_summary(conn: sqlite3.Connection, local_date: str) -> None:
    """Log a breakdown of today's matches: total in DB / pending / done / filtered by league,
    then list the next 30 matches (by scheduled_utc_ts) after the current moment."""
    total = conn.execute(
        "SELECT COUNT(*) FROM bet_monitor_schedule WHERE event_date = ?", (local_date,)
    ).fetchone()[0]
    if total == 0:
        return
    done_count = conn.execute(
        "SELECT COUNT(*) FROM bet_monitor_schedule WHERE event_date = ? AND status IN ('done', 'discarded')",
        (local_date,),
    ).fetchone()[0]
    # active = not done/discarded
    active = conn.execute(
        "SELECT COUNT(*) FROM bet_monitor_schedule WHERE event_date = ? AND status NOT IN ('done', 'discarded')",
        (local_date,),
    ).fetchone()[0]
    pending = len(_get_pending_rows(conn, local_date))
    filtered_by_league = active - pending
    _log(
        f"Resumen {local_date}: total={total}  "
        f"pendientes={pending}  "
        f"filtrados_liga={filtered_by_league}  "
        f"terminados/descartados={done_count}"
    )
    # List next 30 matches of the day after current time (any status, any league)
    now_ts = int(time.time())
    upcoming = conn.execute(
        """
        SELECT home_team, away_team, match_id, scheduled_utc_ts, league, status
        FROM bet_monitor_schedule
        WHERE event_date = ?
          AND scheduled_utc_ts > ?
        ORDER BY scheduled_utc_ts ASC
        LIMIT 30
        """,
        (local_date, now_ts),
    ).fetchall()
    if upcoming:
        _log(f"Próximos {len(upcoming)} partidos del día ({local_date}) después de ahora:")
        for r in upcoming:
            sched_label = _format_sched_local_label(int(r["scheduled_utc_ts"] or 0))
            _log(
                f"  {sched_label}  {r['home_team']} vs {r['away_team']} ({r['match_id']})"
                f"  [{r['league']}]  status={r['status']}"
            )


def _get_pending_rows(conn: sqlite3.Connection, local_date: str) -> list[dict]:
    rows = conn.execute(
        """
        SELECT * FROM bet_monitor_schedule
        WHERE event_date = ?
          AND status NOT IN ('done', 'discarded')
          AND NOT (league LIKE '%WNBA%' OR league LIKE '%Women%' OR league LIKE '%women%' OR league LIKE '%Feminina%' OR league LIKE '%Femenina%')
          AND NOT (league LIKE '%Playoff%' OR league LIKE '%PLAY OFF%')
          AND NOT league LIKE '%U21 Espoirs Elite%'
          AND NOT league LIKE '%Liga Femenina%'
          AND NOT league LIKE '%LF Challenge%'
          AND NOT league LIKE '%Polish Basketball League%'
          AND NOT league LIKE '%SuperSport Premijer Liga%'
          AND NOT league LIKE '%Prvenstvo Hrvatske za d%'
          AND NOT league LIKE '%ABA Liga%'
          AND NOT league LIKE '%Argentina Liga Nacional%'
          AND NOT league LIKE '%Basketligaen%'
          AND NOT league LIKE '%lite 2%'
          AND NOT league LIKE '%EYBL%'
          AND NOT league LIKE '%I B MCKL%'
          AND NOT league LIKE '%Liga 1 Masculin%'
          AND NOT league LIKE '%Liga Nationala%'
          AND NOT league LIKE '%NBL1%'
          AND NOT league LIKE '%PBA Commissioner%'
          AND NOT league LIKE '%Rapid League%'
          AND NOT league LIKE '%Stoiximan GBL%'
          AND NOT league LIKE '%Playout%'
          AND NOT league LIKE '%Superleague%'
          AND NOT league LIKE '%Superliga%'
          AND NOT league LIKE '%Swedish Basketball Superettan%'
          AND NOT league LIKE '%Swiss Cup%'
          AND NOT league LIKE '%Финал%'
          AND NOT league LIKE '%Turkish Basketball Super League%'
          AND NOT league LIKE '%NBA%'
          AND NOT league LIKE '%Big V%'
          AND NOT league LIKE '%Egyptian Basketball Super League%'
          AND NOT league LIKE '%Lega A Basket%'
          AND NOT league LIKE '%Liga e Par%'
          AND NOT league LIKE '%Liga Ouro%'
          AND NOT league LIKE '%Señal%'
          AND NOT league LIKE '%LNB%'
          AND NOT league LIKE '%Meridianbet KLS%'
          AND NOT league LIKE '%MPBL%'
          AND NOT league LIKE '%Nationale 1%'
          AND NOT league LIKE '%Poland 2nd Basketball League%'
          AND NOT league LIKE '%Portugal LBP%'
          AND NOT league LIKE '%Portugal Proliga%'
          AND NOT league LIKE '%Saku I liiga%'
          AND NOT league LIKE '%Serie A2%'
          AND NOT league LIKE '%Slovenian Second Basketball%'
          AND NOT league LIKE '%Super League%'
          AND NOT league LIKE '%United Cup%'
          AND NOT league LIKE '%United League%'
        ORDER BY scheduled_utc_ts ASC
        """,
        (local_date,),
    ).fetchall()
    return [dict(r) for r in rows]


def _update_row(conn: sqlite3.Connection, match_id: str, **kwargs) -> None:
    if not kwargs:
        return
    kwargs["_mid"] = match_id
    kwargs["_ts"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
    sets = ", ".join(f"{k} = :{k}" for k in kwargs if not k.startswith("_"))
    sets += ", updated_at = :_ts"
    conn.execute(
        f"UPDATE bet_monitor_schedule SET {sets} WHERE match_id = :_mid",
        kwargs,
    )
    conn.commit()


def _insert_log(conn: sqlite3.Connection, **kwargs) -> None:
    kwargs.setdefault("notified_at", datetime.now(timezone.utc).isoformat(timespec="seconds"))
    cols = ", ".join(kwargs.keys())
    ph = ", ".join(f":{k}" for k in kwargs)
    conn.execute(f"INSERT INTO bet_monitor_log ({cols}) VALUES ({ph})", kwargs)
    conn.commit()


def reconcile_pending_results(db_path: str) -> int:
    """Update pending bet_monitor_log rows by cross-referencing quarter_scores.

    Runs synchronously (DB-only, no scraping).  Safe to call at any time.
    Returns the number of rows resolved.
    """
    conn = _open_db(db_path)
    pending = conn.execute(
        """
        SELECT id, match_id, target, pick
        FROM bet_monitor_log
        WHERE result = 'pending'
          AND signal NOT IN (
              'UNAVAILABLE', 'ERROR', 'NO_BET', 'NO BET',
              'window_missed', 'no_data', 'no_graph'
          )
        """,
    ).fetchall()

    resolved = 0
    for row in pending:
        q_key = str(row["target"] or "").upper()
        pick = str(row["pick"] or "")
        qs = conn.execute(
            "SELECT home, away FROM quarter_scores"
            " WHERE match_id = ? AND quarter = ?",
            (str(row["match_id"]), q_key),
        ).fetchone()
        if not qs:
            continue
        h, a = qs["home"], qs["away"]
        if h is None or a is None:
            continue
        try:
            h, a = int(h), int(a)
        except (TypeError, ValueError):
            continue
        if h == a:
            result_key = "push"
        elif pick == "home":
            result_key = "win" if h > a else "loss"
        else:
            result_key = "win" if a > h else "loss"
        conn.execute(
            "UPDATE bet_monitor_log SET result = ?, result_checked = 1"
            " WHERE id = ?",
            (result_key, row["id"]),
        )
        resolved += 1

    if resolved:
        conn.commit()
        _log(f"reconcile bet_monitor_log: {resolved} resultado(s) actualizado(s)")

    # ── Phase 2: reconcile eval_match_results ─────────────────────────────────
    # Find all model-tag columns (q3_pick__TAG → TAG) so we can reconcile
    # q3_outcome__TAG / q4_outcome__TAG = 'pending' using quarter_scores.
    eval_resolved = 0
    try:
        table_cols = [
            row[1]
            for row in conn.execute("PRAGMA table_info(eval_match_results)").fetchall()
        ]
        tags = sorted({
            col.split("__", 1)[1]
            for col in table_cols
            if col.startswith("q3_pick__")
        })

        _BET_SIGS = {"BET", "BET HOME", "BET_HOME", "BET AWAY", "BET_AWAY"}

        for tag in tags:
            pick_q3 = f'"q3_pick__{tag}"'
            pick_q4 = f'"q4_pick__{tag}"'
            sig_q3  = f'"q3_signal__{tag}"'
            sig_q4  = f'"q4_signal__{tag}"'
            out_q3  = f'"q3_outcome__{tag}"'
            out_q4  = f'"q4_outcome__{tag}"'
            avail_q3 = f'"q3_available__{tag}"'
            avail_q4 = f'"q4_available__{tag}"'

            pending_eval = conn.execute(
                f"""
                SELECT match_id,
                       {pick_q3} AS pick_q3, {sig_q3} AS sig_q3,
                       {pick_q4} AS pick_q4, {sig_q4} AS sig_q4,
                       {out_q3}  AS out_q3,  {out_q4}  AS out_q4,
                       {avail_q3} AS avail_q3, {avail_q4} AS avail_q4
                FROM eval_match_results
                WHERE ({avail_q3} = 1 AND {out_q3} = 'pending')
                   OR ({avail_q4} = 1 AND {out_q4} = 'pending')
                """
            ).fetchall()

            for erow in pending_eval:
                mid = str(erow["match_id"])
                updates: list[tuple[str, str]] = []

                for quarter, out_col, pick_val, sig_val, avail_val in (
                    ("Q3", f"q3_outcome__{tag}", erow["pick_q3"], erow["sig_q3"], erow["avail_q3"]),
                    ("Q4", f"q4_outcome__{tag}", erow["pick_q4"], erow["sig_q4"], erow["avail_q4"]),
                ):
                    if not avail_val:
                        continue
                    sig_norm = str(sig_val or "").upper().replace("_", " ")
                    if sig_norm not in _BET_SIGS:
                        continue
                    if str(erow[f"out_{quarter.lower()}"] or "") != "pending":
                        continue

                    qs = conn.execute(
                        "SELECT home, away FROM quarter_scores"
                        " WHERE match_id = ? AND quarter = ?",
                        (mid, quarter),
                    ).fetchone()
                    if not qs:
                        continue
                    h, a = qs["home"], qs["away"]
                    if h is None or a is None:
                        continue
                    try:
                        h, a = int(h), int(a)
                    except (TypeError, ValueError):
                        continue

                    pick = str(pick_val or "").lower()
                    if h == a:
                        outcome = "push"
                    elif pick == "home":
                        outcome = "hit" if h > a else "miss"
                    else:
                        outcome = "hit" if a > h else "miss"

                    updates.append((f'"{out_col}"', outcome))

                for col_quoted, outcome in updates:
                    conn.execute(
                        f"UPDATE eval_match_results SET {col_quoted} = ?, updated_at = datetime('now')"
                        " WHERE match_id = ?",
                        (outcome, mid),
                    )
                    eval_resolved += 1

        if eval_resolved:
            conn.commit()
            _log(f"reconcile eval_match_results: {eval_resolved} outcome(s) actualizado(s)")
    except Exception as exc:
        _log(f"reconcile eval_match_results error: {exc}")

    conn.close()
    total = resolved + eval_resolved
    return total


def _find_pending_result_match_ids(conn: sqlite3.Connection) -> list[str]:
    """Collect match_ids that still have pending outcomes to be reconciled."""
    pending_ids: set[str] = set()

    # 1) Pending rows in bet_monitor_log with real bet signals.
    log_rows = conn.execute(
        """
        SELECT DISTINCT match_id
        FROM bet_monitor_log
        WHERE result = 'pending'
          AND signal IN ('BET', 'BET_HOME', 'BET_AWAY')
        """
    ).fetchall()
    for row in log_rows:
        pending_ids.add(str(row["match_id"]))

    # 2) Pending outcomes in eval_match_results across discovered model tags.
    try:
        table_cols = [
            row[1]
            for row in conn.execute("PRAGMA table_info(eval_match_results)").fetchall()
        ]
        tags = sorted({
            col.split("__", 1)[1]
            for col in table_cols
            if col.startswith("q3_pick__")
        })

        _BET_SIGS = {"BET", "BET HOME", "BET_HOME", "BET AWAY", "BET_AWAY"}
        for tag in tags:
            for quarter in ("q3", "q4"):
                avail_col = f'"{quarter}_available__{tag}"'
                sig_col = f'"{quarter}_signal__{tag}"'
                out_col = f'"{quarter}_outcome__{tag}"'
                rows = conn.execute(
                    f"SELECT match_id, {sig_col} AS sig FROM eval_match_results "
                    f"WHERE {avail_col} = 1 AND {out_col} = 'pending'"
                ).fetchall()
                for row in rows:
                    sig_norm = str(row["sig"] or "").upper().replace("_", " ")
                    if sig_norm in _BET_SIGS:
                        pending_ids.add(str(row["match_id"]))
    except Exception:
        # Table/columns may not exist in early setups.
        pass

    return sorted(pending_ids)


async def _recheck_pending_outcomes_once(db_path: str) -> dict[str, int]:
    """Scrape pending matches and reconcile outcomes in one background cycle."""
    conn = _open_db(db_path)
    try:
        pending_ids = _find_pending_result_match_ids(conn)
    finally:
        conn.close()

    if not pending_ids:
        resolved = await asyncio.to_thread(reconcile_pending_results, db_path)
        return {
            "found": 0,
            "checked": 0,
            "scraped_ok": 0,
            "scraped_fail": 0,
            "resolved": int(resolved),
        }

    ids_to_scrape = pending_ids[:PENDING_RECHECK_MAX_MATCHES]

    def _scrape_pending_ids() -> tuple[int, int]:
        import db as db_mod
        import scraper as scraper_mod

        ok = 0
        fail = 0
        save_conn = db_mod.get_conn(db_path)
        db_mod.init_db(save_conn)
        try:
            for mid in ids_to_scrape:
                try:
                    fresh = scraper_mod.fetch_match_by_id(mid)
                    if fresh:
                        db_mod.save_match(save_conn, mid, fresh)
                        ok += 1
                    else:
                        fail += 1
                except Exception:
                    fail += 1
        finally:
            save_conn.close()
        return ok, fail

    scraped_ok, scraped_fail = await asyncio.to_thread(_scrape_pending_ids)
    resolved = await asyncio.to_thread(reconcile_pending_results, db_path)
    return {
        "found": len(pending_ids),
        "checked": len(ids_to_scrape),
        "scraped_ok": int(scraped_ok),
        "scraped_fail": int(scraped_fail),
        "resolved": int(resolved),
    }


async def _recheck_pending_finished_schedule_once(db_path: str) -> dict[str, int]:
    """Scrape overdue pending schedule rows and finalize those already FT."""

    def _sync_job() -> dict[str, int]:
        import db as db_mod
        import scraper as scraper_mod

        sched_conn = _open_db(db_path)
        now_ts = int(time.time())
        max_sched_ts = now_ts - PENDING_SCHEDULE_MIN_AGE_SECS
        rows = sched_conn.execute(
            """
            SELECT match_id, home_team, away_team, scheduled_utc_ts
            FROM bet_monitor_schedule
            WHERE status = 'pending'
              AND final_fetched = 0
              AND scheduled_utc_ts > 0
              AND scheduled_utc_ts <= ?
            ORDER BY scheduled_utc_ts ASC
            LIMIT ?
            """,
            (max_sched_ts, PENDING_SCHEDULE_MAX_MATCHES),
        ).fetchall()

        if not rows:
            sched_conn.close()
            return {
                "found": 0,
                "checked": 0,
                "scraped_ok": 0,
                "scraped_fail": 0,
                "finished_saved": 0,
            }

        save_conn = db_mod.get_conn(db_path)
        db_mod.init_db(save_conn)
        checked = 0
        scraped_ok = 0
        scraped_fail = 0
        finished_saved = 0

        try:
            for row in rows:
                mid = str(row["match_id"])
                checked += 1
                try:
                    fresh = scraper_mod.fetch_match_by_id(mid)
                except Exception:
                    scraped_fail += 1
                    continue

                if not fresh:
                    scraped_fail += 1
                    continue

                scraped_ok += 1
                st = str((fresh.get("match", {}) or {}).get("status_type", "") or "").lower()
                if st != "finished":
                    continue

                try:
                    db_mod.save_match(save_conn, mid, fresh)
                    _update_row(
                        sched_conn,
                        mid,
                        status="done",
                        final_fetched=1,
                        final_fetch_at=datetime.now(timezone.utc).isoformat(timespec="seconds"),
                        skip_reason="backfill_finished",
                    )
                    finished_saved += 1
                except Exception:
                    # Save/update error for this row should not stop whole batch
                    continue
        finally:
            save_conn.close()
            sched_conn.close()

        return {
            "found": len(rows),
            "checked": checked,
            "scraped_ok": scraped_ok,
            "scraped_fail": scraped_fail,
            "finished_saved": finished_saved,
        }

    return await asyncio.to_thread(_sync_job)


# ── Schedule fetcher ──────────────────────────────────────────────────────────

def _fetch_all_events_for_date_sync(local_date: str) -> list[dict]:
    """Return all basketball events whose LOCAL date equals local_date.

    Queries two UTC dates (local_date and local_date+1) to catch overnight
    games, then filters by startTimestamp converted to local time.
    """
    from playwright.sync_api import sync_playwright

    UA = (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    )
    hdrs = {
        "Referer": "https://www.sofascore.com/",
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "en-US,en;q=0.9",
    }

    local_dt = datetime.strptime(local_date, "%Y-%m-%d").date()
    utc_dates = [local_dt.isoformat(), (local_dt + timedelta(days=1)).isoformat()]

    all_events: list = []
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx = browser.new_context(user_agent=UA)
        page = ctx.new_page()
        try:
            page.goto(
                "https://www.sofascore.com/basketball",
                wait_until="domcontentloaded",
                timeout=45_000,
            )
        except Exception:
            pass
        for utc_date in utc_dates:
            url = (
                "https://api.sofascore.com/api/v1/"
                f"sport/basketball/scheduled-events/{utc_date}"
            )
            try:
                resp = ctx.request.get(url, headers=hdrs, timeout=30_000)
                if resp.ok:
                    body = resp.json() or {}
                    all_events.extend(body.get("events", []))
            except Exception as exc:
                logger.warning("[MONITOR] schedule fetch %s: %s", utc_date, str(exc).split("\n")[0][:160])
        browser.close()

    out: list[dict] = []
    seen: set[str] = set()
    for ev in all_events:
        mid = str(ev.get("id", "") or "")
        if not mid or mid in seen:
            continue
        ts = int(ev.get("startTimestamp", 0) or 0)
        if ts == 0:
            continue
        dt_utc = datetime.fromtimestamp(ts, tz=timezone.utc)
        dt_local = dt_utc + timedelta(hours=UTC_OFFSET_HOURS)
        if dt_local.date().isoformat() != local_date:
            continue
        seen.add(mid)
        out.append({
            "match_id": mid,
            "event_date": local_date,
            "home_team": (ev.get("homeTeam") or {}).get("name", ""),
            "away_team": (ev.get("awayTeam") or {}).get("name", ""),
            "league": ((ev.get("tournament") or {}).get("name", "")),
            "scheduled_utc_ts": ts,
            "scheduled_utc": dt_utc.strftime("%Y-%m-%d %H:%M:%S"),
            "status_type": str((ev.get("status") or {}).get("type", "") or ""),
        })

    return sorted(out, key=lambda x: x["scheduled_utc_ts"])


# ── Live data helpers ─────────────────────────────────────────────────────────

def _get_minute(data: dict) -> int | None:
    gp = data.get("graph_points", [])
    if not gp:
        return None
    try:
        return int(gp[-1].get("minute", 0))
    except (TypeError, ValueError):
        return None


def _count_gp_up_to(data: dict, max_min: int) -> int:
    return sum(
        1 for p in data.get("graph_points", [])
        if int(p.get("minute", 0) or 0) <= max_min
    )


def _has_scores(data: dict, *quarters: str) -> bool:
    qs = (data.get("score", {}) or {}).get("quarters", {}) or {}
    return all(
        isinstance(qs.get(q), dict)
        and qs[q].get("home") is not None
        and qs[q].get("away") is not None
        for q in quarters
    )


def _q3_has_real_progress_for_q4(data: dict) -> bool:
    """True when Q3 is truly usable for Q4, not just placeholder 0-0."""
    if not _has_scores(data, "Q1", "Q2", "Q3"):
        return False
    qs = (data.get("score", {}) or {}).get("quarters", {}) or {}
    q3 = qs.get("Q3") or {}
    try:
        q3_total = int(q3.get("home", 0) or 0) + int(q3.get("away", 0) or 0)
    except (TypeError, ValueError):
        q3_total = 0
    if q3_total > 0:
        return True
    # If Q4 already has a score row, Q3 is implicitly complete.
    return _has_scores(data, "Q4")


def _is_two_half(data: dict) -> bool:
    s = data.get("score", {}) or {}
    qs = s.get("quarters", {}) or {}
    if qs.get("Q3") or qs.get("Q4"):
        return False
    st = str((data.get("match", {}) or {}).get("status_type", "") or "").lower()
    if st == "finished":
        scored = sum(
            1 for k in ("Q1", "Q2", "Q3", "Q4")
            if isinstance(qs.get(k), dict) and qs[k].get("home") is not None
        )
        return scored <= 2
    gp = data.get("graph_points", [])
    if gp:
        try:
            return int(gp[-1].get("minute", 0)) >= 36
        except (TypeError, ValueError):
            pass
    return False


async def _final_fetch_and_save(
    match_id: str,
    db_path: str,
    conn_sched: sqlite3.Connection,
    home: str,
    away: str,
    current_minute: int | None,
    secs_per_gmin: float,
    stop_event: asyncio.Event,
    sleep_fn,
) -> None:
    """Wait for the estimated end of the match, then scrape and persist
    the final result to the matches DB.

    Only attempts the save if the live data already has graph_points
    (FINAL_FETCH_MIN_GP threshold). Matches without graph data are skipped
    to avoid polluting the DB with empty records.
    """
    import scraper as scraper_mod

    # Estimate real seconds until minute 48 (end of Q4)
    GAME_END_MINUTE = 48
    mins_remaining = max(0, GAME_END_MINUTE - (current_minute or GAME_END_MINUTE))
    wait_secs = mins_remaining * secs_per_gmin + FINAL_FETCH_EXTRA_SECS
    _log(
        f"{home} vs {away}: esperando fin estimado "
        f"(~{wait_secs / 60:.0f} min) antes de guardar resultado final"
    )
    if await sleep_fn(wait_secs):
        return  # stop_event fired

    # Poll every 90s until status=finished, then save.  Give up after 40 min.
    MAX_FINISH_WAIT = 40 * 60
    POLL_INTERVAL = 90
    waited_finish = 0
    last_data: dict | None = None

    while waited_finish <= MAX_FINISH_WAIT:
        if stop_event.is_set():
            return
        try:
            async with _get_fetch_sem():
                last_data = await asyncio.to_thread(
                    scraper_mod.fetch_match_by_id, match_id
                )
        except Exception as exc:
            logger.warning(
                "[MONITOR] final_fetch %s error: %s",
                match_id, str(exc).split("\n")[0][:120],
            )
            if await sleep_fn(POLL_INTERVAL):
                return
            waited_finish += POLL_INTERVAL
            continue

        if last_data:
            st = str(
                (last_data.get("match", {}) or {}).get("status_type", "") or ""
            ).lower()
            if st == "finished":
                break
            _log(
                f"{home} vs {away}: final_fetch esperando fin "
                f"(status={st}, espera={waited_finish}s)"
            )

        if await sleep_fn(POLL_INTERVAL):
            return
        waited_finish += POLL_INTERVAL

    data = last_data
    if not data:
        _log(f"{home} vs {away}: final_fetch sin datos al finalizar")
        return

    gp_total = len(data.get("graph_points") or [])
    if gp_total < FINAL_FETCH_MIN_GP:
        _log(
            f"{home} vs {away}: final_fetch omitido "
            f"(sólo {gp_total} graph_points, mínimo {FINAL_FETCH_MIN_GP})"
        )
        return

    # Save to matches DB
    try:
        db_conn = __import__("db").get_conn(db_path)
        __import__("db").init_db(db_conn)
        __import__("db").save_match(db_conn, match_id, data)
        db_conn.close()
        _update_row(
            conn_sched, match_id,
            final_fetched=1,
            final_fetch_at=datetime.now(timezone.utc).isoformat(timespec="seconds"),
        )
        _log(f"{home} vs {away}: resultado final guardado (gp={gp_total})")
    except Exception as exc:
        logger.warning(
            "[MONITOR] final_fetch save error %s: %s", match_id, exc
        )


# ── Inference runner ──────────────────────────────────────────────────────────

def _get_v12_mae(target: str, metric_type: str) -> float | None:
    """Load V12 MAE for target/type from all_metrics.csv."""
    try:
        import csv as _csv
        from pathlib import Path as _Path
        csv_path = _Path(__file__).parent / "training" / "model_outputs_v12" / "all_metrics.csv"
        if not csv_path.exists():
            raise FileNotFoundError
        metrics: dict[str, float] = {}
        with csv_path.open("r", encoding="utf-8", newline="") as f:
            for row in _csv.DictReader(f):
                key = str(row.get("model") or "")
                metrics[key] = float(row.get("mae") or 0)
        for suffix in ("men_or_open", "women", ""):
            k = f"{target}_{metric_type}{'_' + suffix if suffix else ''}"
            if k in metrics:
                return metrics[k]
    except Exception:
        pass
    fallbacks = {
        "q3_home": 3.63, "q3_away": 3.60, "q3_total": 5.33,
        "q4_home": 3.63, "q4_away": 3.60, "q4_total": 5.33,
    }
    return fallbacks.get(f"{target}_{metric_type}")


# ── V15/V16 inference helpers ──────────────────────────────────────────────────

def _load_match_data_for_engine(match_id: str) -> dict | None:
    """Load match data from DB for V15/V16/V17 engine (uses v15 dataset DB path)."""
    try:
        v15_ds = importlib.import_module("training.v15.dataset")
        db_mod = importlib.import_module("db")
        conn = v15_ds.get_db_connection()
        data = db_mod.get_match(conn, match_id)
        conn.close()
        return data
    except Exception as exc:
        logger.warning("[MONITOR] load_match_data_for_engine %s: %s", match_id, exc)
        return None


def _extract_engine_data(data: dict, match_id: str, target: str):
    """Extract (league, quarter_scores, graph_points, pbp_events) from data dict."""
    m = data.get("match", {})
    s = data.get("score", {})
    quarters = s.get("quarters", {})
    gp: list[dict] = data.get("graph_points") or []
    # Flatten quarter-keyed PBP → flat list with 'quarter' key injected
    pbp: list[dict] = []
    for q_code, evts in (data.get("play_by_play") or {}).items():
        for e in (evts or []):
            pbp.append({**e, "quarter": q_code})
    league = str(m.get("league") or "Unknown")

    def _qsc(q_code: str, side: str) -> int:
        return int((quarters.get(q_code) or {}).get(side) or 0)

    quarter_scores: dict[str, int] = {
        "q1_home": _qsc("Q1", "home"), "q1_away": _qsc("Q1", "away"),
        "q2_home": _qsc("Q2", "home"), "q2_away": _qsc("Q2", "away"),
    }
    if target == "q4":
        quarter_scores["q3_home"] = _qsc("Q3", "home")
        quarter_scores["q3_away"] = _qsc("Q3", "away")
    return league, quarter_scores, gp, pbp


def _v15v16_pred_to_mon_dict(pred: object, target: str) -> dict:
    """Convert V15/V16/V17 Prediction to the standard predictions[target] dict."""
    sig = getattr(pred, "signal", "NO_BET")
    d = getattr(pred, "debug", None)
    winner_pick = "home" if sig == "BET_HOME" else ("away" if sig == "BET_AWAY" else None)
    return {
        "available": True,
        "predicted_winner": winner_pick,
        "confidence": getattr(pred, "confidence", None),
        "bet_signal": sig,
        "final_recommendation": sig,
        "predicted_total": getattr(d, "pred_total", None) if d else None,
        "predicted_home": getattr(d, "pred_home", None) if d else None,
        "predicted_away": getattr(d, "pred_away", None) if d else None,
        "reasoning": getattr(pred, "reason", ""),
        "mae": getattr(d, "reg_mae_total", None) if d else None,
        "mae_home": None,
        "mae_away": None,
        "league_quality": None,
        "league_bettable": sig != "NO_BET",
        "volatility_index": None,
        "data_quality": None,
        # V15/V16/V17 extras for terminal log
        "model_found": getattr(d, "model_found", None) if d else None,
        "gp_count": getattr(d, "gp_count", None) if d else None,
        "pbp_count": getattr(d, "pbp_count", None) if d else None,
        "threshold": getattr(pred, "threshold", None),
        "probability": getattr(pred, "probability", None),
        "gates": [
            {"name": g.name, "passed": g.passed, "reason": g.reason}
            for g in (getattr(d, "gates", None) or [])
        ] if d else [],
    }


def _run_inference_sync(match_id: str, target: str) -> dict:
    """Run inference synchronously (called via asyncio.to_thread)."""
    version = _model_config.get(target, "v4")

    # V12/V13 have their own inference engines — not handled by infer_match.py
    if version == "v12":
        try:
            v12_mod = importlib.import_module("training.v12.infer_match_v12")
            pred = v12_mod.run_inference(
                match_id=match_id,
                target=target,
                fetch_missing=False,
            )
            if isinstance(pred, dict) and not pred.get("ok", True):
                return {"ok": False, "reason": pred.get("reason", "V12 failed")}

            def _attr(obj, name, default=None):
                if isinstance(obj, dict):
                    return obj.get(name, default)
                return getattr(obj, name, default)

            return {
                "ok": True,
                "predictions": {
                    target: {
                        "available": True,
                        "predicted_winner": _attr(pred, "winner_pick"),
                        "confidence": _attr(pred, "winner_confidence"),
                        "bet_signal": _attr(pred, "winner_signal"),
                        "final_recommendation": _attr(pred, "final_signal"),
                        "predicted_total": _attr(pred, "predicted_total"),
                        "predicted_home": _attr(pred, "predicted_home"),
                        "predicted_away": _attr(pred, "predicted_away"),
                        "reasoning": _attr(pred, "reasoning"),
                        "mae": _get_v12_mae(target, "total"),
                        "mae_home": _get_v12_mae(target, "home"),
                        "mae_away": _get_v12_mae(target, "away"),
                        "league_quality": _attr(pred, "league_quality"),
                        "league_bettable": _attr(pred, "league_bettable"),
                        "volatility_index": _attr(pred, "volatility_index"),
                        "data_quality": _attr(pred, "data_quality"),
                    }
                },
            }
        except Exception as exc:
            return {"ok": False, "reason": f"V12 error: {exc}"}

    if version == "v13":
        try:
            v13_mod = importlib.import_module("training.v13.infer_match_v13")
            v13_result = v13_mod.run_inference(match_id=match_id, target=target)
            if not v13_result.get("ok", False):
                return {"ok": False, "reason": v13_result.get("reason", "V13 failed")}
            pred = v13_result.get("prediction")

            def _attr(obj, name, default=None):
                if isinstance(obj, dict):
                    return obj.get(name, default)
                return getattr(obj, name, default)

            return {
                "ok": True,
                "predictions": {
                    target: {
                        "available": True,
                        "predicted_winner": _attr(pred, "winner_pick"),
                        "confidence": _attr(pred, "winner_confidence"),
                        "bet_signal": _attr(pred, "winner_signal"),
                        "final_recommendation": _attr(pred, "final_signal"),
                        "predicted_total": _attr(pred, "predicted_total"),
                        "predicted_home": _attr(pred, "predicted_home"),
                        "predicted_away": _attr(pred, "predicted_away"),
                        "reasoning": _attr(pred, "reasoning"),
                        "mae": _attr(pred, "mae"),
                        "mae_home": _attr(pred, "mae_home"),
                        "mae_away": _attr(pred, "mae_away"),
                        "league_quality": _attr(pred, "league_quality"),
                        "league_bettable": _attr(pred, "league_bettable"),
                        "volatility_index": _attr(pred, "volatility_index"),
                        "data_quality": _attr(pred, "data_quality"),
                        "model_quality": _attr(pred, "model_quality"),
                        "model_samples": _attr(pred, "model_samples"),
                        "model_gap": _attr(pred, "model_gap"),
                        "model_f1": _attr(pred, "model_f1"),
                        "fallback_used": _attr(pred, "fallback_used", False),
                    }
                },
            }
        except Exception as exc:
            return {"ok": False, "reason": f"V13 error: {exc}"}

    if version == "v15":
        try:
            data = _load_match_data_for_engine(match_id)
            if data is None:
                return {"ok": False, "reason": "V15: match data not found in DB"}
            league, quarter_scores, gp, pbp = _extract_engine_data(data, match_id, target)
            v15_mod = importlib.import_module("training.v15.inference")
            engine = _ENGINE_CACHE.get("v15")
            if engine is None:
                engine = v15_mod.V15Engine.load()
                _ENGINE_CACHE["v15"] = engine
            pred = engine.predict(
                match_id=match_id, target=target, league=league,
                quarter_scores=quarter_scores, graph_points=gp, pbp_events=pbp,
            )
            return {"ok": True, "predictions": {target: _v15v16_pred_to_mon_dict(pred, target)}}
        except Exception as exc:
            return {"ok": False, "reason": f"V15 error: {exc}"}

    if version == "v16":
        try:
            data = _load_match_data_for_engine(match_id)
            if data is None:
                return {"ok": False, "reason": "V16: match data not found in DB"}
            league, quarter_scores, gp, pbp = _extract_engine_data(data, match_id, target)
            # V16 inference module also names its engine class V15Engine
            v16_mod = importlib.import_module("training.v16.inference")
            engine = _ENGINE_CACHE.get("v16")
            if engine is None:
                engine = v16_mod.V15Engine.load()
                _ENGINE_CACHE["v16"] = engine
            pred = engine.predict(
                match_id=match_id, target=target, league=league,
                quarter_scores=quarter_scores, graph_points=gp, pbp_events=pbp,
            )
            return {"ok": True, "predictions": {target: _v15v16_pred_to_mon_dict(pred, target)}}
        except Exception as exc:
            return {"ok": False, "reason": f"V16 error: {exc}"}

    if version == "v17":
        try:
            data = _load_match_data_for_engine(match_id)
            if data is None:
                return {"ok": False, "reason": "V17: match data not found in DB"}
            league, quarter_scores, gp, pbp = _extract_engine_data(data, match_id, target)
            v17_mod = importlib.import_module("training.v17.inference")
            engine = _ENGINE_CACHE.get("v17")
            if engine is None:
                engine = v17_mod.V15Engine.load()
                _ENGINE_CACHE["v17"] = engine
            pred = engine.predict(
                match_id=match_id, target=target, league=league,
                quarter_scores=quarter_scores, graph_points=gp, pbp_events=pbp,
            )
            return {"ok": True, "predictions": {target: _v15v16_pred_to_mon_dict(pred, target)}}
        except Exception as exc:
            return {"ok": False, "reason": f"V17 error: {exc}"}

    infer_mod = importlib.import_module("training.infer_match")
    # Suppress live snapshot fetch — we already have the data
    infer_mod.scraper_mod.fetch_event_snapshot = lambda _mid: None
    return infer_mod.run_inference(
        match_id=match_id,
        metric="f1",
        fetch_missing=False,
        force_version=version,
        target_only=target,
    )


def _friendly_reason(reason: str) -> str:
    """Translate gate reason codes to Spanish labels."""
    reason_lower = str(reason).lower()
    if "league not bettable" in reason_lower:
        return "liga no apostable"
    if "insufficient graph points" in reason_lower or "missing" in reason_lower and "graph" in reason_lower:
        return "puntos de grafico insuficientes"
    if "insufficient confidence" in reason_lower:
        return "confianza insuficiente para apostar"
    if "too volatile" in reason_lower:
        return "partido muy volatil"
    if "confidence below" in reason_lower:
        return "confianza por debajo del minimo"
    if "missing" in reason_lower and "play" in reason_lower:
        return "faltan datos de jugadas"
    if "match too early" in reason_lower:
        return "partido muy temprano"
    mapping = {
        "league not bettable": "liga no apostable",
        "insufficient_graph_or_pbp_coverage": "cobertura de datos insuficiente",
        "match_too_volatile_for_current_signal": "partido muy volatil",
        "confidence_below_minimum_edge": "confianza por debajo del minimo",
        "missing_q3_score": "Q3 aun no disponible",
        "missing_q1_q2_scores": "sin scores Q1/Q2",
    }
    return mapping.get(reason, reason or "no disponible")


def _league_stats_detail(league: str) -> str | None:
    """Return a one-line stats summary explaining why the league is not bettable."""
    try:
        stats_file = Path(__file__).resolve().parent / "training" / "v12" / "model_outputs" / "league_stats.json"
        if not stats_file.exists():
            return None
        with open(stats_file, "r", encoding="utf-8") as _f:
            _stats = json.load(_f)
        # Exact match first, then base name before comma, then prefix
        s = _stats.get(league, {})
        if not s and "," in league:
            s = _stats.get(league.split(",")[0].strip(), {})
        if not s:
            for key in sorted(_stats.keys(), key=len, reverse=True):
                if league.startswith(key):
                    s = _stats[key]
                    break
        if not s:
            return "📊 Liga sin datos históricos"
        samples = int(s.get("samples", 0))
        home_wr = float(s.get("home_win_rate", 0.5))
        if samples < 30:
            return f"📊 Solo {samples} partidos en histórico (mínimo 30)"
        home_pct = round(home_wr * 100)
        return f"📊 N={samples} partidos | Local gana {home_pct}% — sin ventaja clara"
    except Exception:
        return None


def _is_definitive_no_bet(pred: dict) -> bool:
    """True when the NO BET reason is categorical and won't change with more data.
    These are notified immediately; all other reasons wait for confirmation ticks."""
    reasoning = str(
        pred.get("reasoning") or pred.get("gate_reason") or pred.get("reason") or ""
    ).lower()
    return "league not bettable" in reasoning


def _insert_inference_debug(
    db_path: str,
    match_id: str,
    target: str,
    model: str,
    scraped_minute: int | None,
    signal: str,
    confidence: float,
    gp_count: int,
    inference_json: str,
) -> None:
    """Persist raw inference result to inference_debug_log for later debugging."""
    try:
        conn = _open_db(db_path)
        conn.execute(
            """
            INSERT INTO inference_debug_log
                (match_id, target, model, scraped_minute, signal, confidence,
                 gp_count, inference_json)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (match_id, target, model, scraped_minute, signal, confidence,
             gp_count, inference_json),
        )
        conn.commit()
        conn.close()
    except Exception as exc:
        logger.warning("[MONITOR] inference_debug insert error: %s", exc)


def _bold_num(value) -> str:
    trans = str.maketrans("0123456789", "𝟬𝟭𝟮𝟯𝟰𝟱𝟲𝟳𝟴𝟵")
    return str(value).translate(trans)


def _league_bet_history(db_path: str, league: str, target: str, model: str, pick: str, confidence: float) -> str:
    """Return a compact historical stats block for the league/pick/model/confidence.

    Pulls from two sources:
    - bet_monitor_log: live-monitored bets (result = 'win'/'loss'/'push')
    - eval_match_results: batch-evaluated historical data (outcome = 'hit'/'miss'/'push')
    Results are deduplicated by match_id; bet_monitor_log takes priority.
    """
    try:
        conn = _open_db(db_path)
        league_l = league.lower()
        target_l = target.lower()   # e.g. "q4"
        model_l  = model.lower()    # e.g. "v6"

        # ── Source 1: live bet_monitor_log ────────────────────────────────────
        live_rows = conn.execute(
            """
            SELECT match_id, pick, confidence, result
            FROM bet_monitor_log
            WHERE LOWER(league) = ?
              AND LOWER(target) = ?
              AND LOWER(model)  = ?
              AND signal IN ('BET','BET_HOME','BET_AWAY')
              AND result IN ('win','loss','push')
            ORDER BY id
            """,
            (league_l, target_l, model_l),
        ).fetchall()
        seen_match_ids = {str(r["match_id"]) for r in live_rows}

        # ── Source 2: eval_match_results (batch evaluation / post-training) ──
        # Columns follow the pattern q{N}_{field}__{model}, e.g. q4_pick__v6
        qt = target_l  # "q3" or "q4"
        pick_col   = f"{qt}_pick__{model_l}"
        signal_col = f"{qt}_signal__{model_l}"
        outcome_col = f"{qt}_outcome__{model_l}"
        conf_col   = f"{qt}_confidence__{model_l}"
        eval_extra: list[dict] = []
        try:
            eval_rows = conn.execute(
                f"""
                SELECT e.match_id,
                       e."{pick_col}"    AS pick,
                       e."{conf_col}"    AS confidence,
                       e."{outcome_col}" AS outcome
                FROM eval_match_results e
                INNER JOIN matches m ON e.match_id = m.match_id
                WHERE LOWER(m.league) = ?
                  AND e."{signal_col}" = 'BET'
                  AND e."{outcome_col}" IN ('hit','miss','push')
                ORDER BY e.event_date
                """,
                (league_l,),
            ).fetchall()
            for er in eval_rows:
                mid = str(er["match_id"])
                if mid in seen_match_ids:
                    continue  # already counted from bet_monitor_log
                outcome_map = {"hit": "win", "miss": "loss", "push": "push"}
                result = outcome_map.get(str(er["outcome"] or "").lower())
                if result is None:
                    continue
                eval_extra.append({
                    "pick": er["pick"],
                    "confidence": er["confidence"],
                    "result": result,
                })
        except Exception:
            pass  # column may not exist for this model version

        conn.close()
    except Exception:
        return ""

    # Merge: live first, then eval extras (as plain dicts for uniform access)
    rows = [{"pick": r["pick"], "confidence": r["confidence"], "result": r["result"]} for r in live_rows]
    rows.extend(eval_extra)

    if not rows:
        return "📊 Historial liga\nSin registros previos para esta liga"

    total = len(rows)
    wins_total = sum(1 for r in rows if r["result"] == "win")

    # Confidence buckets: (label, lo_inclusive, hi_exclusive)
    _BUCKETS = [
        ("0-35%",  0.00, 0.36),
        ("36-40%", 0.36, 0.41),
        ("41-45%", 0.41, 0.46),
        ("46-50%", 0.46, 0.51),
        ("51-55%", 0.51, 0.56),
        ("56-60%", 0.56, 0.61),
        ("61-65%", 0.61, 0.66),
        ("66-70%", 0.66, 0.71),
        ("71-75%", 0.71, 0.76),
        ("76-80%", 0.76, 0.81),
        ("81%+",   0.81, 1.01),
    ]

    def _bucket_lines(lbl: str, direction: str) -> list[str]:
        dir_rows = [r for r in rows if str(r["pick"] or "").lower() == direction.lower()]
        lines = []
        for bucket_lbl, lo, hi in _BUCKETS:
            b = [r for r in dir_rows if lo <= float(r["confidence"] or 0) < hi]
            if not b:
                continue
            w = sum(1 for r in b if r["result"] == "win")
            lines.append(f"{lbl}: {bucket_lbl}  {w}/{len(b)} ({w/len(b)*100:.0f}%)")
        return lines

    opp_pick = "away" if pick == "home" else "home"
    pick_lbl = "🏠 home" if pick == "home" else "✈️ away"
    opp_lbl  = "✈️ away" if pick == "home" else "🏠 home"

    lines = ["📊 Historial liga"]
    lines.append(f"Liga total: {wins_total}/{total} ({wins_total/total*100:.0f}%)")
    pick_lines = _bucket_lines(pick_lbl, pick)
    if pick_lines:
        lines.extend(pick_lines)
    opp_lines = _bucket_lines(opp_lbl, opp_pick)
    if opp_lines:
        lines.extend(opp_lines)

    return "\n".join(lines)


def _format_bet_notification(
    match_id: str,
    data: dict,
    quarter_label: str,   # "Q3" or "Q4"
    home: str,
    away: str,
    league: str,
    event_date: str,
    current_minute: int | None,
    pick_sym: str,
    pick_name: str,
    confidence: float,
    model: str,
    is_bet: bool = True,
    pred: dict | None = None,
    db_path: str | None = None,
    pick: str | None = None,
) -> str:
    """Build a match-detail-style BET/NO BET notification message."""
    m_info = data.get("match") or {}
    status_type = str(m_info.get("status_type") or "")
    quarters = (data.get("score") or {}).get("quarters", {})
    # Show all completed/in-progress quarters up to and including the current one.
    # Including Q4 in Q4 alerts: shows partial Q4 score and lets Q3 auto-mark
    # as final when Q4 data exists (fixes ⏳ on Q3 for short-quarter leagues).
    if quarter_label == "Q3":
        q_order = ["Q1", "Q2"]
    else:
        q_order = ["Q1", "Q2", "Q3", "Q4"]
    q_end_minute = {"Q1": 12, "Q2": 24, "Q3": 36, "Q4": 48}

    def _q_line(q: str) -> str:
        qd = quarters.get(q)
        if not qd or qd.get("home") is None:
            return f"{q}: - ⏳"
        h, a = int(qd["home"]), int(qd["away"])
        # Score text with bold winner
        if h > a:
            score_txt = f"{_bold_num(h)} - {a}"
        elif a > h:
            score_txt = f"{h} - {_bold_num(a)}"
        else:
            score_txt = f"{h} - {a}"
        # Determine if quarter is final
        is_final = (
            status_type == "finished"
            or (current_minute is not None and current_minute >= q_end_minute.get(q, 99))
            # Quarter is final if a later quarter already has scores
            or any(
                quarters.get(later_q, {}).get("home") is not None
                for later_q in q_order[q_order.index(q) + 1:]
            )
        )
        if not is_final:
            return f"{q}: {score_txt} ⏳"
        if h > a:
            return f"{q}: {score_txt} 🏠 {home}"
        elif a > h:
            return f"{q}: {score_txt} ✈️ {away}"
        return f"{q}: {score_txt}"

    score_lines = "\n".join(_q_line(q) for q in q_order)

    match_time = str(m_info.get("time") or "")
    match_date_str = str(m_info.get("date") or event_date)
    try:
        from datetime import datetime as _dt
        if match_date_str and match_time:
            _utc_dt = _dt.strptime(f"{match_date_str} {match_time}", "%Y-%m-%d %H:%M")
            _local_dt = _utc_dt + timedelta(hours=UTC_OFFSET_HOURS)
            fecha_txt = f"{_local_dt.strftime('%Y-%m-%d %H:%M')} UTC{UTC_OFFSET_HOURS:+d}"
        else:
            fecha_txt = match_date_str
    except Exception:
        fecha_txt = f"{match_date_str} {match_time}" if match_time else match_date_str

    if is_bet:
        header = f"🟢 APUESTA {quarter_label} [{model}]"
        _p = pred or {}
        _footer_parts = [f"Pick: {pick_sym} {pick_name}"]

        # Projection (V12 and other models that expose predicted_home/away)
        def _sf_bet(v):
            try:
                return float(v) if v is not None else None
            except (TypeError, ValueError):
                return None
        _ph = _sf_bet(_p.get("predicted_home"))
        _pa = _sf_bet(_p.get("predicted_away"))
        _pt = _sf_bet(_p.get("predicted_total"))
        _mae = _sf_bet(_p.get("mae"))
        _mae_h = _sf_bet(_p.get("mae_home"))
        _mae_a = _sf_bet(_p.get("mae_away"))
        if _ph is not None and _pa is not None:
            _mh_txt = f" (±{_mae_h:.1f})" if _mae_h is not None else ""
            _ma_txt = f" (±{_mae_a:.1f})" if _mae_a is not None else ""
            _mt_txt = f" (±{_mae:.0f})" if _mae is not None else ""
            _overlap = ""
            if _mae_h is not None and _mae_a is not None:
                _diff = abs(_ph - _pa)
                _thresh = _mae_h + _mae_a
                if _diff < _thresh:
                    _overlap = f" ⚠️ MAE se superpone (dif={_diff:.1f} < {_thresh:.1f})"
            _pt_val = _pt if _pt is not None else (_ph + _pa)
            _footer_parts.append(
                f"Proyeccion: 🏠~{_ph:.1f}{_mh_txt} | ✈️~{_pa:.1f}{_ma_txt}"
                f" | Total~{_pt_val:.1f}{_mt_txt}{_overlap}"
            )

        _footer_parts.append(f"Confianza: {confidence * 100:.1f}%")

        # Historical league stats
        if db_path and pick:
            _hist = _league_bet_history(db_path, league, quarter_label, model, pick, confidence)
            _footer_parts.append(_hist)

        pick_line = "\n".join(_footer_parts)
    else:
        header = f"🔴 NO APOSTAR {quarter_label} [{model}]"
        _p = pred or {}
        _footer_parts = ["Señal: NO BET"]

        # Tendency
        if pick_sym and pick_name:
            _footer_parts.append(f"Tendencia {pick_sym} {pick_name}")

        # Reason
        _reason_raw = str(_p.get("gate_reason") or _p.get("reasoning") or _p.get("reason") or "").strip()
        if _reason_raw:
            _footer_parts.append(f"Motivo: {_friendly_reason(_reason_raw).capitalize()}")
        if "league not bettable" in _reason_raw.lower():
            _ld = _league_stats_detail(league)
            if _ld:
                _footer_parts.append(_ld)

        # Projection (V12 or any model that sets predicted_home/away)
        def _sf(v):
            try:
                return float(v) if v is not None else None
            except (TypeError, ValueError):
                return None
        _ph = _sf(_p.get("predicted_home"))
        _pa = _sf(_p.get("predicted_away"))
        _pt = _sf(_p.get("predicted_total"))
        _mae = _sf(_p.get("mae"))
        _mae_h = _sf(_p.get("mae_home"))
        _mae_a = _sf(_p.get("mae_away"))
        if _ph is not None and _pa is not None:
            _mh_txt = f" (±{_mae_h:.1f})" if _mae_h is not None else ""
            _ma_txt = f" (±{_mae_a:.1f})" if _mae_a is not None else ""
            _mt_txt = f" (±{_mae:.0f})" if _mae is not None else ""
            _overlap = ""
            if _mae_h is not None and _mae_a is not None:
                _diff = abs(_ph - _pa)
                _thresh = _mae_h + _mae_a
                if _diff < _thresh:
                    _overlap = f" ⚠️ MAE se superpone (dif={_diff:.1f} < {_thresh:.1f})"
            _pt_val = _pt if _pt is not None else (_ph + _pa)
            _footer_parts.append(
                f"Proyeccion: 🏠~{_ph:.1f}{_mh_txt} | ✈️~{_pa:.1f}{_ma_txt}"
                f" | Total~{_pt_val:.1f}{_mt_txt}{_overlap}"
            )

        # Score (confidence as percentage)
        if confidence > 0:
            _footer_parts.append(f"Score {confidence * 100:.0f}%")

        pick_line = "\n".join(_footer_parts)

    return (
        f"{header}\n\n"
        f"Match ID: {match_id} | Min: {current_minute or '?'}\n"
        f"{home} vs {away}\n"
        f"Fecha: {fecha_txt}\n"
        f"Liga: {league}\n"
        f"{score_lines}\n\n"
        f"{pick_line}"
    )


# ── v2 pick filter (Q3) ───────────────────────────────────────────────────────

def _v2_dynamic_stake(confidence: float) -> float:
    """Map confidence to stake factor per rule 8."""
    if confidence < 0.45:
        return 0.5   # small
    if confidence < 0.75:
        return 1.0   # normal / high (65–75 treated as normal cap)
    if confidence < 0.80:
        return 0.5   # small (75–80)
    return 0.3        # micro (80%+)


def _v2_pick_filter(league: str, confidence: float, pick: str) -> tuple[bool, float]:
    """Apply v2 model (Q3) filtering rules.

    Returns (accept, stake_factor).
    stake_factor 1.0 = normal, 0.5 = half, 0.3 = micro, 0.0 = reject.
    """
    ll = league.lower()
    pk = pick.lower()

    # ── 1. Hard ban: youth phases (rule 9) ───────────────────────────────
    for _y in ("u14", "u18", "u19", "youth", "kadetska"):
        if _y in ll:
            return False, 0.0

    # ── 2. Phase modifier: playoff / final / knockout → stake ×0.5 ───────
    _PHASE_KEYS = (
        "play-in", "play in", "playoff", "play off", "semi",
        "final", "knockout", "superfinal", "classification",
    )
    phase_reduced = any(pp in ll for pp in _PHASE_KEYS)

    # ── 3. B1 League (rule 2) ─────────────────────────────────────────────
    if "b1 league" in ll:
        if confidence > 0.80:
            return False, 0.0
        if confidence < 0.33:
            return False, 0.0
        sf = 0.5 if 0.75 <= confidence <= 0.80 else _v2_dynamic_stake(confidence)
        return (True, sf * 0.5) if phase_reduced else (True, sf)

    # ── 4. B2 League (rule 3) ─────────────────────────────────────────────
    if "b2 league" in ll:
        if confidence > 0.90 or confidence < 0.30:
            return False, 0.0
        sf = _v2_dynamic_stake(confidence)
        return (True, sf * 0.5) if phase_reduced else (True, sf)

    # ── 5. Priority leagues: accept from 30% (rule 10) ───────────────────
    _PRIORITY = (
        "euroleague", "china cba", "germany bbl", "liga acb",
        "poland 1st", "bulgaria nbl", "colombia lpb", "lmb apertura",
    )
    if any(p in ll for p in _PRIORITY):
        if confidence < 0.30:
            return False, 0.0
        sf = _v2_dynamic_stake(confidence)
        return (True, sf * 0.5) if phase_reduced else (True, sf)

    # ── 6. Away underdogs in strong leagues (rule 6) ─────────────────────
    _STRONG = ("euroleague", "b1 league", "b2 league", "china cba", "germany bbl", "bnxt")
    if pk == "away" and 0.30 <= confidence <= 0.45 and any(s in ll for s in _STRONG):
        return (True, 0.25) if phase_reduced else (True, 0.5)

    # ── 7. Volatile leagues: higher threshold (rules 5 & 7) ──────────────
    _VOLATILE = (
        "puerto rico bsn", "cibacopa", "brazil nbb", "israeli national",
        "primera feb", "indonesian bl", "new zealand nbl", "slb",
    )
    _HOME_PENALTY = ("brazil nbb", "puerto rico bsn", "china cba", "new zealand nbl")
    is_volatile = any(v in ll for v in _VOLATILE)
    min_conf = 0.55 if is_volatile else 0.30
    if any(h in ll for h in _HOME_PENALTY) and pk == "home":
        min_conf = max(min_conf, 0.30) + 0.05  # +5%

    if confidence < min_conf:
        return False, 0.0

    # Volatile + overconfidence (≥80%): reject (rule 1 + volatile interaction)
    if is_volatile and confidence >= 0.80:
        return False, 0.0

    # ── 8. Global overconfidence ≥80%: micro stake (rule 1) ──────────────
    if confidence >= 0.80:
        return (True, 0.15) if phase_reduced else (True, 0.3)

    # ── Global stake dynamic (rule 8) ────────────────────────────────────
    sf = _v2_dynamic_stake(confidence)
    return (True, sf * 0.5) if phase_reduced else (True, sf)


# ── v6.2 pick filter ────────────────────────────────────────────────────────────

def _load_v62_exclusion_rules() -> list[tuple[str, str, str]]:
    import json
    from pathlib import Path
    try:
        p = Path('match/training/v6_2_league_name_exclusions.json')
        if not p.exists():
            return []
        cfg = json.loads(p.read_text(encoding='utf-8'))
        pats = []
        for cat in cfg.get('categories', []):
            for p_str in cat.get('patterns', []):
                p_str = str(p_str).strip()
                if p_str:
                    pats.append((cat.get('name', 'uncategorized'), p_str, p_str.lower()))
        return pats
    except Exception:
        return []

_V62_EXCLUSIONS = None

def _v6_2_pick_filter_explain(
    league: str,
    confidence: float,
    pick: str,
    p_pick: float | None = None,
) -> tuple[bool, float, str]:
    """Apply v6.2 model filtering rules.

    Returns (accept, stake_factor, reason).
    stake_factor: 1.0, 0.75, 0.50, 0.25 based on kelly.
    """
    global _V62_EXCLUSIONS
    if _V62_EXCLUSIONS is None:
        _V62_EXCLUSIONS = _load_v62_exclusion_rules()
        
    ll = league.lower()
    for cname, raw, low in _V62_EXCLUSIONS:
        if low in ll:
            return False, 0.0, f"league_excluded:{raw}"

    # Report_v62 uses p_pick probability, not confidence score.
    # If p_pick is not provided, derive it from confidence score in [0,1]:
    # p_pick = 0.5 + confidence/2.
    if p_pick is None:
        c = max(0.0, min(float(confidence or 0.0), 1.0))
        p_pick = 0.5 + (c / 2.0)
    p_pick = max(0.0, min(float(p_pick), 1.0))

    min_conf_prob = 0.58
    if p_pick < min_conf_prob:
        return False, 0.0, f"p_pick_below_min:{p_pick:.3f}<{min_conf_prob:.3f}"

    odds = 1.4
    break_even = 1.0 / odds
    edge = p_pick - break_even
    if edge <= 0:
        return False, 0.0, f"negative_edge:p_pick={p_pick:.3f}<=be={break_even:.3f}@odds={odds:.2f}"

    b = odds - 1.0
    q = 1.0 - p_pick
    k_raw = ((b * p_pick) - q) / b if b > 0 else 0.0
    
    if k_raw <= 0:
        return False, 0.0, f"kelly_non_positive:k_raw={k_raw:.5f}"
        
    kelly_mult = 0.25
    kelly_cap = 0.05
    k_used = min(k_raw * kelly_mult, kelly_cap)

    k_strength = max(0.0, min(k_used / kelly_cap, 1.0))
    
    if p_pick >= 0.97 and edge >= 0.24 and k_strength >= 0.995:
        stake_val = 100.0
    elif p_pick >= 0.90 and edge >= 0.18 and k_strength >= 0.85:
        stake_val = 75.0
    elif p_pick >= 0.80 and edge >= 0.09 and k_strength >= 0.55:
        stake_val = 50.0
    elif k_strength < 0.15:
        stake_val = 0.0
    else:
        stake_val = 25.0

    if stake_val < 25.0:
        return False, 0.0, f"stake_below_min:stake={stake_val:.0f}"

    stake_factor = stake_val / 100.0
    return True, stake_factor, f"accepted:stake={stake_val:.0f}"


def _v6_2_pick_filter(
    league: str,
    confidence: float,
    pick: str,
    p_pick: float | None = None,
) -> tuple[bool, float]:
    """Compatibility wrapper for existing callers."""
    accept, stake_factor, _reason = _v6_2_pick_filter_explain(
        league=league,
        confidence=confidence,
        pick=pick,
        p_pick=p_pick,
    )
    return accept, stake_factor


# ── v6 pick filter ────────────────────────────────────────────────────────────

def _v6_pick_filter(league: str, confidence: float, pick: str) -> tuple[bool, float]:
    """Apply v6 model filtering rules.

    Returns (accept, stake_factor).
    stake_factor: 1.0 = normal stake, 0.5 = reduced, 0.25 = minimal.
    accept=False means reject completely (don't show / don't notify).
    """
    ll = league.lower()

    # ── Hard ban: youth / fringe tournament phase keywords ────────────────
    _BAN_PHASE = [
        "u14", "u16", "u18", "u19", "youth", "superfinal",
        "bronzekamp", "3rd place", "5-8", "9-12", "cadets", "kadetska",
    ]
    for bp in _BAN_PHASE:
        if bp in ll:
            return False, 0.0

    # ── Hard ban: 100% confidence (unrealistic) ────────────────────────
    if confidence >= 1.0:
        return False, 0.0

    # ── Hard ban: specific league/phase keywords ──────────────────────────
    _BANNED_LEAGUES = [
        "1st division",
        "belgian basketball 2nd division",
        "belize elite basketball",
        "brazil nbb",
        "germany bbl",
        "golden square",
        "slb",
        "israeli national",
        "uruguay lub championship round",
        "knock-out",
        "relegation round",
        "ketvirtfinalia",
        "championship round",
        "primera vuelta",
        "segunda vuelta",
    ]
    for b in _BANNED_LEAGUES:
        if b in ll:
            return False, 0.0

    # ── Phase modifier: raise threshold +5% for knockout phases ──────────
    # Exception: Euroleague Play-in stays at normal threshold.
    _PHASE_PLUS5 = [
        "playoffs", "play offs", "play-in", "knockout",
        "quarterfinal", "quarter-finals", "semifinal", "semi-finals", "final",
    ]
    is_euroleague = "euroleague" in ll
    phase_extra = 0.0
    for pp in _PHASE_PLUS5:
        if pp in ll:
            if is_euroleague and pp == "play-in":
                break  # Euroleague Play-in: no penalty
            phase_extra = 0.05
            break

    # ── Adjusted leagues: per-league rules ───────────────────────────────
    if "b1 league" in ll:
        # Underdogs <45% or strong favorites >78% are OK; avoid 50–70%
        if confidence < 0.45:
            return True, 1.0
        if confidence > 0.78:
            return True, 1.0
        if 0.50 <= confidence <= 0.70:
            return False, 0.0
        return True, 1.0  # 45–50% or 70–78%: borderline, accept

    if "b2 league" in ll:
        # >80% → treat as 65% (still accept); stake reduced; avoid <30%
        if confidence < 0.30 + phase_extra:
            return False, 0.0
        return True, 0.5

    if "primera feb" in ll:
        if confidence < 0.60 + phase_extra:
            return False, 0.0
        return True, 1.0

    if "lmb" in ll or "lmb apertura" in ll:
        if confidence < 0.60 + phase_extra:
            return False, 0.0
        return True, 0.5  # Stake lower

    if "colombia lpb" in ll:
        min_conf = 0.50 + phase_extra if pick == "home" else 0.35 + phase_extra
        if confidence < min_conf:
            return False, 0.0
        return True, 1.0

    if "korean basketball league" in ll:
        min_conf = 0.50 + phase_extra if pick == "home" else 0.35 + phase_extra
        if confidence < min_conf:
            return False, 0.0
        return True, 1.0

    # ── Trusted leagues: accept from 35% ─────────────────────────────────
    _TRUSTED_35 = [
        "betsafe-lkl", "china cba", "euroleague", "france pro a",
        "nbl men", "serie b group", "liga acb", "puerto rico bsn",
    ]
    for t in _TRUSTED_35:
        if t in ll:
            if confidence < 0.35 + phase_extra:
                return False, 0.0
            return True, 1.0  # Trusted: accept full range (including 70%+)

    # ── Trusted leagues: accept from 38% ─────────────────────────────────
    _TRUSTED_38 = ["bulgaria nbl"]
    for t in _TRUSTED_38:
        if t in ll:
            if confidence < 0.38 + phase_extra:
                return False, 0.0
            return True, 1.0

    # ── Leagues requiring strictly >40% confidence ────────────────────────
    _MIN40 = ["segunda feb", "1. a skl"]
    for t in _MIN40:
        if t in ll:
            if confidence <= 0.40 + phase_extra:
                return False, 0.0
            return True, 1.0

    # ── Global rules (all other leagues) ─────────────────────────────────
    if confidence < 0.30 + phase_extra:
        return False, 0.0

    # 70–80%: only trusted leagues may bet this range → reject for others
    if 0.70 <= confidence <= 0.80:
        return False, 0.0

    # 80%+: possible overconfidence; accept but reduced stake
    if confidence > 0.80:
        return True, 0.5

    return True, 1.0


async def _check_quarter(
    match_id: str,
    data: dict,
    target: str,          # "q3" | "q4"
    db_path: str,
    conn_sched: sqlite3.Connection,
    home: str,
    away: str,
    league: str,
    event_date: str,
    current_minute: int | None,
    suppress_no_bet_notify: bool = False,
) -> tuple[str, bool, dict]:
    """Run inference for one quarter.  Returns (signal, notified, pred)."""
    import db as db_mod

    quarter_label = target.upper()
    _log(f"{home} vs {away}: analizando {quarter_label} (min {current_minute})")

    # Persist scraped data so infer_match can load from DB
    try:
        tmp = _open_db(db_path)
        db_mod.save_match(tmp, match_id, data)
        tmp.close()
    except Exception as exc:
        logger.warning("[MONITOR] save_match error %s: %s", match_id, exc)

    # Run inference in thread
    try:
        result = await asyncio.to_thread(_run_inference_sync, match_id, target)
    except Exception as exc:
        logger.error("[MONITOR] inference error %s %s: %s", match_id, target, exc)
        return "ERROR", False, {}

    pred: dict = {}
    if result.get("ok"):
        pred = result.get("predictions", {}).get(target, {}) or {}

    available = bool(pred.get("available", False))
    signal = str(
        pred.get("final_recommendation") or pred.get("bet_signal") or "NO_BET"
    ).upper().strip()
    pick = str(pred.get("predicted_winner") or "")
    try:
        confidence = float(pred.get("confidence") or 0.0)
    except (TypeError, ValueError):
        confidence = 0.0

    # Update counters
    if target == "q3":
        MONITOR_STATUS["checked_q3"] = MONITOR_STATUS.get("checked_q3", 0) + 1
    else:
        MONITOR_STATUS["checked_q4"] = MONITOR_STATUS.get("checked_q4", 0) + 1

    # Persist raw inference JSON for later debugging
    try:
        _gp_count = len(data.get("graph_points") or [])
        _insert_inference_debug(
            db_path=db_path,
            match_id=match_id,
            target=target,
            model=_model_config.get(target, "-"),
            scraped_minute=current_minute,
            signal=signal if available else "UNAVAILABLE",
            confidence=round(confidence, 4) if available else 0.0,
            gp_count=_gp_count,
            inference_json=json.dumps(result, default=str),
        )
    except Exception as _exc:
        logger.warning("[MONITOR] debug log error: %s", _exc)

    # ── Terminal debug dump ────────────────────────────────────────────────
    _model_used = _model_config.get(target, "-")
    # Use model-aware cutoff for GP count display
    _cutoff = (22 if target == "q3" else 31) if _model_used in ("v13", "v15", "v16", "v17") else (24 if target == "q3" else 36)
    _gp_all = data.get("graph_points") or []
    _gp_used = [p for p in _gp_all if int(p.get("minute", 0)) <= _cutoff]

    # Build V15/V16-specific extra lines
    _extra_lines = ""
    if _model_used in ("v15", "v16", "v17") and result.get("ok"):
        _p = result.get("predictions", {}).get(target, {})
        _gates_info = _p.get("gates", [])
        _failed_gates = [g for g in _gates_info if not g.get("passed")]
        _extra_lines = (
            f"  Model found: {_p.get('model_found')} | GP≤{_cutoff}={_p.get('gp_count')} | PBP={_p.get('pbp_count')}\n"
            f"  Probability: {(_p.get('probability') or 0)*100:.1f}% | Threshold: {_p.get('threshold')}\n"
        )
        if _failed_gates:
            _extra_lines += f"  Gates failed: {', '.join(g['name'] + '(' + g['reason'] + ')' for g in _failed_gates)}\n"

    _quarter_dbg = _decorate_quarter_tokens(quarter_label)

    print(
        f"\n{'='*60}\n"
        f"[MONITOR DEBUG] {home} vs {away} — {_quarter_dbg} | match_id={match_id}\n"
        f"  Model    : {_model_used}  |  Q3={_model_config.get('q3', '-')}  Q4={_model_config.get('q4', '-')}\n"
        f"  Signal   : {signal if available else 'UNAVAILABLE'}\n"
        f"  Confidence: {confidence*100:.1f}%\n"
        f"  Min scraped: {current_minute}\n"
        f"  GP total={len(_gp_all)} | GP used (≤min{_cutoff})={len(_gp_used)}\n"
        f"{_extra_lines}"
        f"  Inference JSON:\n{json.dumps(result, ensure_ascii=False, default=str, indent=2)}\n"
        f"{'='*60}",
        flush=True,
    )
    # ── End terminal debug dump ───────────────────────────────────────────

    # Persist log entry
    log_conn = _open_db(db_path)
    _insert_log(
        log_conn,
        match_id=match_id,
        event_date=event_date,
        home_team=home,
        away_team=away,
        league=league,
        target=target,
        model=_model_config.get(target, "-"),
        signal=signal if available else "UNAVAILABLE",
        recommendation=str(pred.get("reason") or signal),
        pick=pick,
        confidence=round(confidence, 4),
        scraped_minute=current_minute,
        result="pending",
    )
    log_conn.close()

    if not available:
        reason = str(pred.get("reason") or "unavailable")
        _log(f"{home} vs {away}: {quarter_label} no disponible ({reason})")
        return "UNAVAILABLE", False, pred

    notified = False
    if _is_bet_signal(signal):
        def _filtered_markup() -> dict:
            return {
                "inline_keyboard": [[
                    {"text": "🔍 Ver match", "callback_data": f"notifmatch:{match_id}:{event_date}:0"},
                    {"text": "📱 Sofascore", "url": (
                        f"https://www.sofascore.com/{data.get('match', {}).get('event_slug') or ''}/{data.get('match', {}).get('custom_id') or ''}#id:{match_id}"
                        if (data.get("match", {}).get("event_slug") and data.get("match", {}).get("custom_id"))
                        else f"https://www.sofascore.com/basketball/event/{match_id}"
                    )},
                ]]
            }

        if confidence <= 0.30:
            _log(f"🔕 BET {quarter_label} [{_model_config.get(target, '-')}]: {home} vs {away} → confianza {confidence*100:.0f}% ≤ 30%, ignorando")
            if _should_notify_filtered_bets(db_path) and not suppress_no_bet_notify:
                await _notify(
                    _format_filtered_bet_notification(
                        match_id=match_id,
                        data=data,
                        quarter_label=quarter_label,
                        model_used=_model_config.get(target, "-"),
                        home=home,
                        away=away,
                        league=league,
                        event_date=event_date,
                        current_minute=current_minute,
                        pick=pick,
                        confidence=confidence,
                        p_pick=None,
                        reason="confidence_below_min:<=30%",
                    ),
                    reply_markup=_filtered_markup(),
                    notify_type="filtered_bet",
                    quarter=target,
                )
                notified = True
            return signal, notified, pred
        try:
            p_home = float(pred.get("p_home_win") or 0.0)
        except (TypeError, ValueError):
            p_home = 0.0
        if pick == "home":
            p_pick = p_home
        elif pick == "away":
            p_pick = 1.0 - p_home
        else:
            p_pick = 0.5 + (max(0.0, min(confidence, 1.0)) / 2.0)
        # Apply model-specific filter rules
        model_used = _model_config.get(target, "-")
        _v6_stake = 1.0
        if model_used == "v6":
            _v6_accept, _v6_stake = _v6_pick_filter(league, confidence, pick)
            if not _v6_accept:
                _log(f"🔕 BET {quarter_label} [v6]: {home} vs {away} → filtrado (conf={confidence*100:.0f}%, liga='{league}')")                
                if _should_notify_filtered_bets(db_path) and not suppress_no_bet_notify:
                    await _notify(
                        _format_filtered_bet_notification(
                            match_id=match_id,
                            data=data,
                            quarter_label=quarter_label,
                            model_used=model_used,
                            home=home,
                            away=away,
                            league=league,
                            event_date=event_date,
                            current_minute=current_minute,
                            pick=pick,
                            confidence=confidence,
                            p_pick=p_pick,
                            reason="v6_filter_reject",
                        ),
                        reply_markup=_filtered_markup(),
                        notify_type="filtered_bet",
                        quarter=target,
                    )
                    notified = True
                return signal, notified, pred
        elif model_used == "v6_2":
            _v6_2_accept, _v6_stake, _v6_2_reason = _v6_2_pick_filter_explain(
                league, confidence, pick, p_pick=p_pick
            )
            if not _v6_2_accept:
                _log(
                    f"🔕 BET {quarter_label} [v6_2]: {home} vs {away} → filtrado "
                    f"(conf={confidence*100:.0f}%, p_pick={p_pick*100:.1f}%, liga='{league}', reason={_v6_2_reason})"
                )
                if _should_notify_filtered_bets(db_path) and not suppress_no_bet_notify:
                    await _notify(
                        _format_filtered_bet_notification(
                            match_id=match_id,
                            data=data,
                            quarter_label=quarter_label,
                            model_used=model_used,
                            home=home,
                            away=away,
                            league=league,
                            event_date=event_date,
                            current_minute=current_minute,
                            pick=pick,
                            confidence=confidence,
                            p_pick=p_pick,
                            reason=_v6_2_reason,
                        ),
                        reply_markup=_filtered_markup(),
                        notify_type="filtered_bet",
                        quarter=target,
                    )
                    notified = True
                return signal, notified, pred
        elif model_used == "v2":
            _v2_accept, _v6_stake = _v2_pick_filter(league, confidence, pick)
            if not _v2_accept:
                _log(f"🔕 BET {quarter_label} [v2]: {home} vs {away} → filtrado (conf={confidence*100:.0f}%, liga='{league}')")
                if _should_notify_filtered_bets(db_path) and not suppress_no_bet_notify:
                    await _notify(
                        _format_filtered_bet_notification(
                            match_id=match_id,
                            data=data,
                            quarter_label=quarter_label,
                            model_used=model_used,
                            home=home,
                            away=away,
                            league=league,
                            event_date=event_date,
                            current_minute=current_minute,
                            pick=pick,
                            confidence=confidence,
                            p_pick=p_pick,
                            reason="v2_filter_reject",
                        ),
                        reply_markup=_filtered_markup(),
                        notify_type="filtered_bet",
                        quarter=target,
                    )
                    notified = True
                return signal, notified, pred

        pick_sym = "🏠" if pick == "home" else ("✈️" if pick == "away" else "?")
        pick_name = home if pick == "home" else (away if pick == "away" else pick)
        msg = _format_bet_notification(
            match_id=match_id,
            data=data,
            quarter_label=quarter_label,
            home=home,
            away=away,
            league=league,
            event_date=event_date,
            current_minute=current_minute,
            pick_sym=pick_sym,
            pick_name=pick_name,
            confidence=confidence,
            model=model_used,
            pred=pred,
            db_path=db_path,
            pick=pick,
        )
        if model_used == "v6_2":
            msg += f"\n💰 Stake sugerido (Kelly): ${int(_v6_stake * 100)}"
        elif _v6_stake < 1.0:
            msg += f"\n⚠️ Stake sugerido: {int(_v6_stake * 100)}% del stake normal"
        markup = {
            "inline_keyboard": [[
                {"text": "🔍 Ver match", "callback_data": f"notifmatch:{match_id}:{event_date}:0"},
                {"text": "📱 Sofascore", "url": (
                    f"https://www.sofascore.com/{data.get('match', {}).get('event_slug') or ''}/{data.get('match', {}).get('custom_id') or ''}#id:{match_id}"
                    if (data.get("match", {}).get("event_slug") and data.get("match", {}).get("custom_id"))
                    else f"https://www.sofascore.com/basketball/event/{match_id}"
                )},
            ]]
        }
        await _notify(msg, reply_markup=markup, notify_type="bet", quarter=target)
        notified = True
        _log(f"🟢 BET {quarter_label} [{model_used}]: {home} vs {away} → {pick} ({confidence * 100:.0f}%)")
        # Schedule result check after match finishes
        asyncio.ensure_future(
            _resolve_bet_result(
                match_id, target, pick, pick_name,
                home, away, league, event_date, db_path, model_used,
            )
        )
    else:
        model_used = _model_config.get(target, "-")
        # Definitive reasons (e.g. league_not_bettable) fire immediately;
        # uncertain reasons wait for confirmation ticks (suppress_no_bet_notify).
        _send_now = not suppress_no_bet_notify or _is_definitive_no_bet(pred)
        _log(
            f"🔴 NO BET {quarter_label} [{model_used}]: {home} vs {away} | signal={signal}"
            + ("" if _send_now else " [esperando confirmacion]")
        )
        if _send_now:
            MONITOR_STATUS["no_bet"] = MONITOR_STATUS.get("no_bet", 0) + 1
            _nb_pick_sym = "🏠" if pick == "home" else ("✈️" if pick == "away" else "")
            _nb_pick_name = home if pick == "home" else (away if pick == "away" else "")
            no_bet_msg = _format_bet_notification(
                match_id=match_id,
                data=data,
                quarter_label=quarter_label,
                home=home,
                away=away,
                league=league,
                event_date=event_date,
                current_minute=current_minute,
                pick_sym=_nb_pick_sym,
                pick_name=_nb_pick_name,
                confidence=confidence,
                model=model_used,
                is_bet=False,
                pred=pred,
            )
            markup = {
                "inline_keyboard": [[
                    {"text": "🔍 Ver match", "callback_data": f"notifmatch:{match_id}:{event_date}:0"},
                    {"text": "📱 Sofascore", "url": (
                        f"https://www.sofascore.com/{data.get('match', {}).get('event_slug') or ''}/{data.get('match', {}).get('custom_id') or ''}#id:{match_id}"
                        if (data.get("match", {}).get("event_slug") and data.get("match", {}).get("custom_id"))
                        else f"https://www.sofascore.com/basketball/event/{match_id}"
                    )},
                ]]
            }
            await _notify(no_bet_msg, reply_markup=markup, notify_type="no_bet", quarter=target)
            notified = True

    return signal, notified, pred


async def _resolve_bet_result(
    match_id: str,
    target: str,
    pick: str,
    pick_name: str,
    home: str,
    away: str,
    league: str,
    event_date: str,
    db_path: str,
    model_used: str = "-",
) -> None:
    """Wait for the match to finish then notify win/loss for the bet quarter."""
    import scraper as scraper_mod

    quarter_label = target.upper()
    q_key = {"q3": "Q3", "q4": "Q4"}.get(target, target.upper())
    _log(f"Revision resultado {quarter_label}: {home} vs {away}")

    MAX_WAIT_SECS = 4 * 3600   # give up after 4 h
    POLL_SECS = 120
    waited = 0

    data: dict | None = None
    while waited < MAX_WAIT_SECS:
        await asyncio.sleep(POLL_SECS)
        waited += POLL_SECS
        try:
            async with _get_fetch_sem():
                data = await asyncio.to_thread(scraper_mod.fetch_match_by_id, match_id)
        except Exception as exc:
            logger.warning("[MONITOR] result-poll %s: %s", match_id, str(exc).split("\n")[0][:160])
            continue
        if not data:
            continue
        st = str((data.get("match", {}) or {}).get("status_type", "") or "").lower()
        if st == "finished":
            break
    else:
        _log(f"Revision resultado {quarter_label} {home} vs {away}: timeout sin finalizar")
        return

    # Read the actual quarter score
    qs = ((data or {}).get("score", {}) or {}).get("quarters", {}) or {}
    qd = qs.get(q_key)
    if not isinstance(qd, dict):
        _log(f"Revision resultado {quarter_label}: sin data de cuarto {q_key}")
        return

    q_home = qd.get("home")
    q_away = qd.get("away")
    if q_home is None or q_away is None:
        _log(f"Revision resultado {quarter_label}: scores nulos para {q_key}")
        return

    try:
        q_home = int(q_home)
        q_away = int(q_away)
    except (TypeError, ValueError):
        return

    if q_home == q_away:
        outcome = "PERDIDA ❌ (empate)"
        result_key = "push"
    elif pick == "home":
        outcome = "GANADA ✅" if q_home > q_away else "PERDIDA ❌"
        result_key = "win" if q_home > q_away else "loss"
    else:
        outcome = "GANADA ✅" if q_away > q_home else "PERDIDA ❌"
        result_key = "win" if q_away > q_home else "loss"

    # Format the new message with model and time info
    pick_emoji = "🏠" if pick == "home" else ("✈️" if pick == "away" else "")
    
    # Extract event time from match data
    event_time_str = event_date
    _match_info = (data.get("match") or {}) if data else {}
    start_ts = _match_info.get("startTimestamp", 0)
    if start_ts:
        try:
            dt_utc = datetime.fromtimestamp(start_ts, tz=timezone.utc)
            dt_local = dt_utc + timedelta(hours=UTC_OFFSET_HOURS)
            # Map month numbers to Spanish names
            months_es = {
                1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
                7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
            }
            days_es = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
            day_name = days_es[dt_local.weekday()]
            month_name = months_es[dt_local.month].lower()
            event_time_str = f"{day_name} {dt_local.day} {month_name} {dt_local.year}, {dt_local.hour:02d}:{dt_local.minute:02d}"
        except Exception:
            pass

    msg = (
        f"{'✅' if result_key == 'win' else '❌'} "
        f"{q_key} [{model_used}] — {outcome} {pick_emoji} {pick_name}\n"
        f"{home} vs {away} {q_key}: {q_home}-{q_away}\n"
        f"{league}\n"
        f"{event_time_str}"
    )
    _m = (data.get("match") or {}) if data else {}
    _ev_slug = _m.get("event_slug") or ""
    _custom_id = _m.get("custom_id") or ""
    if _ev_slug and _custom_id:
        _sf_url = f"https://www.sofascore.com/{_ev_slug}/{_custom_id}#id:{match_id}"
    else:
        _sf_url = f"https://www.sofascore.com/basketball/event/{match_id}"
    _result_markup = {"inline_keyboard": [
        [
            {"text": "🔍 Ver Match", "callback_data": f"match:{match_id}:_:0"},
            {"text": "📱 Sofascore", "url": _sf_url},
        ],
    ]}
    await _notify(msg, reply_markup=_result_markup, notify_type="result", quarter=target)
    _log(f"Resultado {quarter_label} {home} vs {away}: {outcome} ({q_home}-{q_away})")

    # Update the log row
    try:
        conn = _open_db(db_path)
        row_id = conn.execute(
            """
            SELECT id FROM bet_monitor_log
            WHERE match_id = ? AND target = ? AND result = 'pending'
            ORDER BY id DESC LIMIT 1
            """,
            (match_id, target),
        ).fetchone()
        if row_id:
            conn.execute(
                "UPDATE bet_monitor_log SET result = ?, result_checked = 1 WHERE id = ?",
                (result_key, row_id["id"]),
            )
            conn.commit()
        conn.close()
    except Exception as exc:
        logger.warning("[MONITOR] result update error %s: %s", match_id, exc)


# ── Match watcher coroutine ───────────────────────────────────────────────────

async def _watch_match(
    match_id: str,
    row: dict,
    db_path: str,
    stop_event: asyncio.Event,
) -> None:
    """Watch one match from start to finish, checking Q3 and Q4.

    Smart-sleep strategy:
      - Before Q3 window: sleep estimate = (min_to_wake * secs_per_game_min)
      - Inside Q3/Q4 window: poll every POLL_NEAR_SECS until data is ready
      - After Q3+6 min without data: skip Q3 and proceed to Q4
      - If no graph_points 55 min after scheduled start: discard entirely
    """
    import scraper as scraper_mod

    home = str(row.get("home_team") or "?")
    away = str(row.get("away_team") or "?")
    league = str(row.get("league") or "")
    event_date = str(row.get("event_date") or "")
    scheduled_ts = int(row.get("scheduled_utc_ts") or 0)
    sched_label = _format_sched_local_label(scheduled_ts)

    _log(f"Vigilando: {home} vs {away} ({match_id}) | {sched_label}")
    MONITOR_STATUS["active_matches"] = list(
        set(MONITOR_STATUS.get("active_matches", []) + [match_id])
    )

    q3_done = bool(row.get("q3_checked")) or SKIP_Q3
    q4_done = bool(row.get("q4_checked"))
    if SKIP_Q3 and not bool(row.get("q3_checked")):
        _log(f"{home} vs {away}: [SKIP_Q3] Q3 deshabilitado — directo a Q4")
    q3_no_bet_ticks = 0   # confirmation ticks for uncertain Q3 NO BET
    q4_no_bet_ticks = 0   # confirmation ticks for uncertain Q4 NO BET
    q4_waiting_score_ticks = 0  # ticks while waiting for Q3 score to appear
    q3_unavailable_ticks = 0  # retries while Q3 inference is temporarily unavailable
    q4_unavailable_ticks = 0  # retries while Q4 inference is temporarily unavailable
    q4_stale_ticks = 0  # ticks where graph_points stop moving while waiting Q4
    q4_last_gp_total = -1
    errors = 0
    secs_per_gmin = float(SECS_PER_GAME_MIN)
    last_gmin: int | None = None
    last_gmin_wall: float = 0.0

    conn = _open_db(db_path)

    async def _sleep(seconds: float) -> bool:
        """Interruptible sleep.  Returns True if stop_event fired."""
        end = time.monotonic() + max(1.0, seconds)
        while time.monotonic() < end:
            if stop_event.is_set():
                return True
            await asyncio.sleep(min(10.0, end - time.monotonic()))
        return False

    try:
        # Wait until close to match start (2-min buffer)
        if scheduled_ts > 0:
            wait = scheduled_ts - time.time() - 120
            if wait > 120:
                eta = _format_wait_eta(wait)
                eta_color = _color_text(eta, "96")
                _log(
                    f"{home} vs {away}: inicio en {eta_color} "
                    f"({wait / 60:.0f} min), esperando"
                )
                if await _sleep(wait):
                    return

        # Main watch loop
        while not stop_event.is_set():
            if q3_done and q4_done:
                _update_row(conn, match_id, status="done")
                # Schedule final data save after estimated match end
                gp_total = len(data.get("graph_points") or []) if data else 0
                if gp_total >= FINAL_FETCH_MIN_GP:
                    await _final_fetch_and_save(
                        match_id=match_id,
                        db_path=db_path,
                        conn_sched=conn,
                        home=home,
                        away=away,
                        current_minute=minute,
                        secs_per_gmin=secs_per_gmin,
                        stop_event=stop_event,
                        sleep_fn=_sleep,
                    )
                break

            # Fetch live data
            try:
                async with _get_fetch_sem():
                    data = await asyncio.to_thread(
                        scraper_mod.fetch_match_by_id, match_id
                    )
                errors = 0
            except Exception as exc:
                errors += 1
                # Truncate Playwright call-log to first meaningful line
                exc_short = str(exc).split("\n")[0][:160]
                logger.warning(
                    "[MONITOR] %s fetch error #%d: %s", match_id, errors, exc_short
                )
                if errors >= MAX_FETCH_ERRORS:
                    _update_row(
                        conn, match_id,
                        status="discarded", skip_reason="fetch_errors",
                    )
                    MONITOR_STATUS["discarded"] = (
                        MONITOR_STATUS.get("discarded", 0) + 1
                    )
                    _log(f"{home} vs {away}: descartado ({errors} errores)")
                    break
                if await _sleep(POLL_NEAR_SECS):
                    break
                continue

            if not data:
                if await _sleep(POLL_NEAR_SECS):
                    break
                continue

            # 2-half format → discard (not Q1/Q2/Q3/Q4)
            if _is_two_half(data):
                _update_row(
                    conn, match_id,
                    status="discarded", skip_reason="dos_mitades",
                )
                MONITOR_STATUS["discarded"] = (
                    MONITOR_STATUS.get("discarded", 0) + 1
                )
                _log(f"{home} vs {away}: formato 2 mitades, descartado")
                break

            status_type = str(
                (data.get("match", {}) or {}).get("status_type", "") or ""
            ).lower()
            minute = _get_minute(data)

            # Calibrate secs per game-minute from observed changes
            if minute is not None:
                now_wall = time.monotonic()
                if last_gmin is not None and minute > last_gmin:
                    rate = (now_wall - last_gmin_wall) / (minute - last_gmin)
                    old_rate = secs_per_gmin
                    secs_per_gmin = 0.7 * secs_per_gmin + 0.3 * rate
                    secs_per_gmin = max(60.0, min(360.0, secs_per_gmin))
                    if abs(secs_per_gmin - old_rate) > 15:
                        logger.debug(
                            "[MONITOR] %s vs %s: ritmo actualizado %.0fs/gmin → %.0fs/gmin "
                            "(crudo=%.0fs, min=%d)",
                            home, away, old_rate, secs_per_gmin, rate, minute,
                        )
                last_gmin = minute
                last_gmin_wall = now_wall

            # Match finished — do final checks and exit
            if status_type == "finished":
                if not q3_done:
                    q3_cut, q3_mgp, _q3wb = _q3_timing()
                    gp3 = _count_gp_up_to(data, q3_cut)
                    if _has_scores(data, "Q1", "Q2") and gp3 >= q3_mgp:
                        sig, notified, _ = await _check_quarter(
                            match_id, data, "q3", db_path, conn,
                            home, away, league, event_date, minute,
                        )
                    else:
                        sig, notified = "no_data", False
                    _update_row(
                        conn, match_id,
                        q3_checked=1, q3_signal=sig, q3_notified=int(notified),
                        q3_model=_model_config.get("q3", "-"),
                    )
                    q3_done = True

                if not q4_done:
                    # Game already finished — Q4 window is closed, never notify
                    _log(f"{home} vs {away}: Q4 no notificado (partido ya terminado)")
                    _update_row(
                        conn, match_id,
                        q4_checked=1, q4_signal="too_late",
                        q4_model=_model_config.get("q4", "-"),
                    )
                    q4_done = True

                _update_row(conn, match_id, status="done")

                # Save final data immediately (match is already finished)
                gp_total = len(data.get("graph_points") or [])
                if gp_total >= FINAL_FETCH_MIN_GP:
                    try:
                        db_conn = __import__("db").get_conn(db_path)
                        __import__("db").init_db(db_conn)
                        __import__("db").save_match(db_conn, match_id, data)
                        db_conn.close()
                        _update_row(
                            conn, match_id,
                            final_fetched=1,
                            final_fetch_at=datetime.now(timezone.utc).isoformat(timespec="seconds"),
                        )
                        _log(f"{home} vs {away}: resultado final guardado al detectar fin (gp={gp_total})")
                    except Exception as _exc:
                        logger.warning("[MONITOR] final_fetch save (finished branch) %s: %s", match_id, _exc)
                break

            # ── Q3 logic ─────────────────────────────────────────────────────
            if not q3_done:
                if minute is None:
                    # No graph data yet
                    elapsed_real = time.time() - scheduled_ts if scheduled_ts > 0 else 0
                    if elapsed_real > NO_GRAPH_REAL_SECS:
                        _update_row(
                            conn, match_id,
                            q3_checked=1, q3_signal="no_graph",
                            q4_checked=1, q4_signal="no_graph",
                            status="discarded", skip_reason="no_graph_by_ht",
                        )
                        MONITOR_STATUS["discarded"] = (
                            MONITOR_STATUS.get("discarded", 0) + 1
                        )
                        _log(f"{home} vs {away}: sin gráfica tras {elapsed_real/60:.0f}min reales, descartado")
                        break
                    if await _sleep(POLL_NEAR_SECS):
                        break
                    continue

                q3_cut, q3_mgp, q3_wake_before = _q3_timing()
                gp3 = _count_gp_up_to(data, q3_cut)

                if minute > q3_cut + 6:
                    # Q3 window has passed
                    _update_row(conn, match_id, q3_checked=1, q3_signal="window_missed")
                    q3_done = True
                    _log(f"{home} vs {away}: ventana Q3 pasada (min {minute})")

                elif minute >= q3_cut - q3_wake_before:
                    # In Q3 window — check if data is ready
                    if _has_scores(data, "Q1", "Q2") and gp3 >= q3_mgp:
                        _is_last_q3 = q3_no_bet_ticks >= NO_BET_CONFIRM_TICKS
                        sig, notified, _ = await _check_quarter(
                            match_id, data, "q3", db_path, conn,
                            home, away, league, event_date, minute,
                            suppress_no_bet_notify=not _is_last_q3,
                        )
                        if notified or _is_last_q3 or sig == "ERROR":
                            _update_row(
                                conn, match_id,
                                q3_checked=1, q3_signal=sig, q3_notified=int(notified),
                                q3_model=_model_config.get("q3", "-"),
                            )
                            q3_done = True
                        elif sig == "UNAVAILABLE":
                            q3_unavailable_ticks += 1
                            if q3_unavailable_ticks % UNAVAILABLE_LOG_EVERY == 0:
                                _log(
                                    f"{home} vs {away}: Q3 UNAVAILABLE "
                                    f"(tick {q3_unavailable_ticks}), "
                                    f"re-evaluando en {POLL_NEAR_SECS}s"
                                )
                            q3_no_bet_ticks = 0
                            if await _sleep(POLL_NEAR_SECS):
                                break
                            continue
                        else:
                            q3_unavailable_ticks = 0
                            q3_no_bet_ticks += 1
                            _log(
                                f"{home} vs {away}: Q3 NO BET incierto "
                                f"(tick {q3_no_bet_ticks}/{NO_BET_CONFIRM_TICKS}), "
                                f"re-evaluando en {POLL_NEAR_SECS}s"
                            )
                            if await _sleep(POLL_NEAR_SECS):
                                break
                            continue
                    else:
                        _log(
                            f"{home} vs {away}: Q3 ventana abierta pero datos insuficientes "
                            f"(min={minute} gp={gp3} Q1Q2={_has_scores(data,'Q1','Q2')})"
                        )
                        if await _sleep(POLL_NEAR_SECS):
                            break
                        continue

                else:
                    # Still before Q3 window — smart sleep
                    mins_to_wake = (q3_cut - q3_wake_before) - minute
                    sleep_secs = max(30.0, min(mins_to_wake * secs_per_gmin, IDLE_POLL_SECS))
                    _log(
                        f"{home} vs {away}: Q3 en ~{mins_to_wake:.0f} game-min, "
                        f"durmiendo {sleep_secs:.0f}s"
                    )
                    if await _sleep(sleep_secs):
                        break
                    continue

            # ── Q4 logic ─────────────────────────────────
            if not q4_done:
                if minute is None:
                    if await _sleep(POLL_NEAR_SECS):
                        break
                    continue

                q4_cut, q4_mgp, q4_need_q3, q4_wake_before = _q4_timing()
                gp4 = _count_gp_up_to(data, q4_cut)
                # Wake at minute 27 — well into Q3 — instead of the old
                # minute 32 window.  For 10-min quarter leagues Q3 ends at
                # minute ~30, so waking at 32 was already inside Q4.
                # Polling every POLL_NEAR_SECS from minute 27 lets us detect
                # Q3 completion as soon as the score appears and fire Q4
                # inference before Q4 gets underway.
                Q4_EARLIEST_MINUTE = 27

                Q4_LAST_MINUTE = 38  # don't bet if Q4 is in its final 2 minutes
                if minute > q4_cut + 6 or minute >= Q4_LAST_MINUTE:
                    _update_row(conn, match_id, q4_checked=1, q4_signal="window_missed")
                    q4_done = True
                    _log(f"{home} vs {away}: ventana Q4 pasada (min {minute}, cutoff={q4_cut})")

                elif minute >= Q4_EARLIEST_MINUTE:
                    # Q3 is well underway — poll until Q3 score exists then fire.
                    score_ok = (_q3_has_real_progress_for_q4(data) if q4_need_q3
                                else _has_scores(data, "Q1", "Q2"))
                    # Budget: real seconds left before Q4_LAST_MINUTE cutoff
                    mins_left = Q4_LAST_MINUTE - minute
                    budget_secs = max(30.0, mins_left * secs_per_gmin)
                    # Adaptive poll: tighter as deadline approaches
                    if mins_left <= 2:
                        poll_secs = 30.0
                    elif mins_left <= 4:
                        poll_secs = 50.0
                    else:
                        poll_secs = POLL_NEAR_SECS

                    if score_ok and gp4 >= q4_mgp:
                        q4_waiting_score_ticks = 0
                        q4_stale_ticks = 0
                        q4_last_gp_total = len(data.get("graph_points") or [])
                        _is_last_q4 = q4_no_bet_ticks >= NO_BET_CONFIRM_TICKS
                        sig, notified, _ = await _check_quarter(
                            match_id, data, "q4", db_path, conn,
                            home, away, league, event_date, minute,
                            suppress_no_bet_notify=not _is_last_q4,
                        )
                        if notified or _is_last_q4 or sig == "ERROR":
                            _update_row(
                                conn, match_id,
                                q4_checked=1, q4_signal=sig, q4_notified=int(notified),
                                q4_model=_model_config.get("q4", "-"),
                            )
                            q4_done = True
                        elif sig == "UNAVAILABLE":
                            q4_unavailable_ticks += 1
                            _log(
                                f"🟡 {home} vs {away}: Q4 UNAVAILABLE "
                                f"(tick {q4_unavailable_ticks}, presupuesto~{budget_secs:.0f}s), "
                                f"re-evaluando en {poll_secs:.0f}s"
                            )
                            q4_no_bet_ticks = 0
                            if await _sleep(poll_secs):
                                break
                            continue
                        else:
                            q4_unavailable_ticks = 0
                            q4_no_bet_ticks += 1
                            _log(
                                f"🟡 {home} vs {away}: Q4 NO BET incierto "
                                f"(tick {q4_no_bet_ticks}/{NO_BET_CONFIRM_TICKS}, "
                                f"min={minute}, presupuesto~{budget_secs:.0f}s), "
                                f"re-evaluando en {poll_secs:.0f}s"
                            )
                            if await _sleep(poll_secs):
                                break
                            continue
                    else:
                        q4_waiting_score_ticks += 1
                        gp_total_now = len(data.get("graph_points") or [])
                        if gp_total_now <= q4_last_gp_total:
                            q4_stale_ticks += 1
                        else:
                            q4_stale_ticks = 0
                            q4_last_gp_total = gp_total_now

                        if (
                            q4_waiting_score_ticks >= Q4_WAITING_MAX_TICKS
                            or q4_stale_ticks >= Q4_STALE_MAX_TICKS
                        ):
                            q4_abort_reason = (
                                "graph_stale_timeout"
                                if q4_stale_ticks >= Q4_STALE_MAX_TICKS
                                else "q3_placeholder_timeout"
                            )
                            _update_row(
                                conn, match_id,
                                q4_checked=1,
                                q4_signal=q4_abort_reason,
                                q4_model=_model_config.get("q4", "-"),
                            )
                            q4_done = True
                            _log(
                                f"🟡 {home} vs {away}: Q4 abortado ({q4_abort_reason}) "
                                f"(tick={q4_waiting_score_ticks}, stale={q4_stale_ticks}, "
                                f"min={minute}, gp={gp4}/{q4_mgp})"
                            )
                            continue

                        need_q3_str = f" Q3={_has_scores(data,'Q3')}" if q4_need_q3 else ""
                        q3_prog_str = (
                            f" Q3_prog={_q3_has_real_progress_for_q4(data)}"
                            if q4_need_q3 else ""
                        )
                        _log(
                            f"🟡 {home} vs {away}: Q4 esperando fin Q3 "
                            f"(min={minute} gp={gp4}/{q4_mgp}{need_q3_str}{q3_prog_str} "
                            f"tick={q4_waiting_score_ticks} stale={q4_stale_ticks} "
                            f"presupuesto~{budget_secs:.0f}s), "
                            f"re-eval en {poll_secs:.0f}s"
                        )
                        if await _sleep(poll_secs):
                            break
                        continue

                else:
                    # Tiered sleep before Q4_EARLIEST_MINUTE — adapts to calibrated pace
                    mins_to_wake = Q4_EARLIEST_MINUTE - minute
                    if mins_to_wake > 10:
                        # Far away: sleep at most 3 min so we can recalibrate pace
                        sleep_secs = min(mins_to_wake * secs_per_gmin * 0.5, 180.0)
                    elif mins_to_wake > 5:
                        sleep_secs = min(mins_to_wake * secs_per_gmin * 0.6, 120.0)
                    elif mins_to_wake > 2:
                        sleep_secs = min(mins_to_wake * secs_per_gmin * 0.7, 60.0)
                    else:
                        sleep_secs = max(20.0, mins_to_wake * secs_per_gmin * 0.8)
                    sleep_secs = max(20.0, sleep_secs)
                    _log(
                        f"🟡 {home} vs {away}: Q4 en ~{mins_to_wake:.0f} game-min "
                        f"(min={minute}, ritmo={secs_per_gmin:.0f}s/gmin, "
                        f"esperando min {Q4_EARLIEST_MINUTE}), durmiendo {sleep_secs:.0f}s"
                    )
                    if await _sleep(sleep_secs):
                        break
                    continue

    except asyncio.CancelledError:
        _log(f"{home} vs {away}: watcher cancelado")
    except Exception as exc:
        logger.error(
            "[MONITOR] watcher exception %s: %s", match_id, exc, exc_info=True
        )
    finally:
        conn.close()
        MONITOR_STATUS["active_matches"] = [
            m for m in MONITOR_STATUS.get("active_matches", []) if m != match_id
        ]
        _log(f"Fin watcher: {home} vs {away}")


# ── Main monitor loop ─────────────────────────────────────────────────────────

async def run_monitor(db_path: str, stop_event: asyncio.Event) -> None:
    """Main monitor coroutine.  Run as asyncio.create_task(run_monitor(...))."""
    _ensure_daily_file_logging()

    async def _refresh_schedule_dates(today_str: str, tomorrow_str: str) -> None:
        for d in (today_str, tomorrow_str):
            try:
                rows = await asyncio.to_thread(
                    _fetch_all_events_for_date_sync, d
                )
                conn = _open_db(db_path)
                for row in rows:
                    rdata = {k: v for k, v in row.items() if k != "status_type"}
                    _upsert_schedule_row(conn, rdata)
                conn.commit()
                conn.close()
                count = len(rows)
                if d == today_str:
                    MONITOR_STATUS["today_total"] = count
                else:
                    MONITOR_STATUS["tomorrow_total"] = count
                _log(f"Itinerario {d}: {count} partidos")
            except Exception as exc:
                logger.error("[MONITOR] schedule error %s: %s", d, exc)

    MONITOR_STATUS.update({
        "running": True,
        "started_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "stop_requested": False,
        "checked_q3": 0,
        "checked_q4": 0,
        "bets_sent": 0,
        "no_bet": 0,
        "discarded": 0,
        "active_matches": [],
    })

    init_tables(db_path)
    reconcile_pending_results(db_path)
    _log(f"Log diario habilitado en: {MONITOR_LOG_DIR}")
    _log("Monitor iniciado")

    active_tasks: dict[str, asyncio.Task] = {}
    last_refresh_wall: float = 0.0
    last_pending_recheck_wall: float = 0.0
    last_pending_schedule_recheck_wall: float = 0.0
    last_date: str = ""

    _today_str = _monitor_local_today_str()
    _tomorrow_str = _monitor_local_tomorrow_str()
    await _refresh_schedule_dates(_today_str, _tomorrow_str)
    last_refresh_wall = time.monotonic()
    last_date = _today_str

    # ── Startup summary: log today's match counts by status/filter
    _summary_conn = _open_db(db_path)
    _log_daily_summary(_summary_conn, _today_str)
    _summary_conn.close()

    # ── Resume: launch watchers for rows already in DB before fetching schedule
    _resume_conn = _open_db(db_path)
    _existing = _get_pending_rows(_resume_conn, _today_str) + _get_pending_rows(_resume_conn, _tomorrow_str)
    _resume_conn.close()
    if _existing:
        _log(f"Retomando {len(_existing)} partido(s) pendientes de la base")
        for _row in _existing:
            _mid = _row["match_id"]
            _sched_ts = int(_row.get("scheduled_utc_ts") or 0)
            if _sched_ts > 0 and (_sched_ts - time.time()) > 20 * 3600:
                continue
            _task = asyncio.create_task(
                _watch_match(_mid, _row, db_path, stop_event),
                name=f"watch_{_mid}",
            )
            active_tasks[_mid] = _task
            _log(
                f"Retomado: {_row.get('home_team')} vs {_row.get('away_team')} ({_mid}) | "
                f"{_format_sched_local_label(_sched_ts)}"
            )

    try:
        while not stop_event.is_set():
            today_str = _monitor_local_today_str()
            tomorrow_str = _monitor_local_tomorrow_str()
            now_wall = time.monotonic()

            # Refresh schedule if stale or new day
            if (
                today_str != last_date
                or now_wall - last_refresh_wall > SCHEDULE_REFRESH_HOURS * 3600
            ):
                await _refresh_schedule_dates(today_str, tomorrow_str)
                last_refresh_wall = now_wall
                last_date = today_str

            # Launch watcher tasks for pending matches
            conn = _open_db(db_path)
            pending_today = _get_pending_rows(conn, today_str)
            pending_tomorrow = _get_pending_rows(conn, tomorrow_str)
            conn.close()

            for row in pending_today + pending_tomorrow:
                mid = row["match_id"]

                # Already being watched and still running
                existing = active_tasks.get(mid)
                if existing and not existing.done():
                    continue
                # Clean up finished task slot
                if existing and existing.done():
                    del active_tasks[mid]

                # Don't start tomorrow's matches more than 20h from now
                sched_ts = int(row.get("scheduled_utc_ts") or 0)
                if sched_ts > 0 and (sched_ts - time.time()) > 20 * 3600:
                    continue

                task = asyncio.create_task(
                    _watch_match(mid, row, db_path, stop_event),
                    name=f"watch_{mid}",
                )
                active_tasks[mid] = task
                _log(
                    f"Tarea lanzada: {row.get('home_team')} vs "
                    f"{row.get('away_team')} ({mid}) | "
                    f"{_format_sched_local_label(sched_ts)}"
                )

            # Clean up done tasks
            for k in [k for k, t in active_tasks.items() if t.done()]:
                del active_tasks[k]

            # Periodically backfill unfinished outcomes created while monitor was off.
            if now_wall - last_pending_recheck_wall >= PENDING_RECHECK_SECS:
                fetches_in_flight = _fetches_in_flight()
                if fetches_in_flight > PENDING_RECHECK_MAX_FETCHES_IN_FLIGHT:
                    _log(
                        "recheck pendientes omitido por carga: "
                        f"fetches={fetches_in_flight}/>{PENDING_RECHECK_MAX_FETCHES_IN_FLIGHT}"
                    )
                else:
                    try:
                        summary = await _recheck_pending_outcomes_once(db_path)
                        if summary["found"] > 0 or summary["resolved"] > 0:
                            _log(
                                "recheck pendientes: "
                                f"found={summary['found']} "
                                f"scan={summary['checked']} "
                                f"ok={summary['scraped_ok']} "
                                f"fail={summary['scraped_fail']} "
                                f"resolved={summary['resolved']}"
                            )
                    except Exception as exc:
                        logger.warning("[MONITOR] pending recheck error: %s", exc)
                last_pending_recheck_wall = now_wall

            # Periodically revisit pending schedule rows and persist FT matches.
            if now_wall - last_pending_schedule_recheck_wall >= PENDING_SCHEDULE_RECHECK_SECS:
                fetches_in_flight = _fetches_in_flight()
                if fetches_in_flight > PENDING_RECHECK_MAX_FETCHES_IN_FLIGHT:
                    _log(
                        "recheck schedule FT omitido por carga: "
                        f"fetches={fetches_in_flight}/>{PENDING_RECHECK_MAX_FETCHES_IN_FLIGHT}"
                    )
                else:
                    try:
                        sched_summary = await _recheck_pending_finished_schedule_once(db_path)
                        if (
                            sched_summary["found"] > 0
                            or
                            sched_summary["checked"] > 0
                            or sched_summary["finished_saved"] > 0
                        ):
                            _log(
                                "recheck schedule FT: "
                                f"found={sched_summary['found']} "
                                f"scan={sched_summary['checked']} "
                                f"ok={sched_summary['scraped_ok']} "
                                f"fail={sched_summary['scraped_fail']} "
                                f"saved={sched_summary['finished_saved']}"
                            )
                    except Exception as exc:
                        logger.warning("[MONITOR] pending schedule recheck error: %s", exc)
                last_pending_schedule_recheck_wall = now_wall

            await asyncio.sleep(60)

    except asyncio.CancelledError:
        _log("Monitor cancelado")
    except Exception as exc:
        logger.error("[MONITOR] main loop exception: %s", exc, exc_info=True)
    finally:
        for task in active_tasks.values():
            if not task.done():
                task.cancel()
        if active_tasks:
            await asyncio.gather(*active_tasks.values(), return_exceptions=True)
        MONITOR_STATUS["running"] = False
        MONITOR_STATUS["active_matches"] = []
        _log("Monitor detenido")


# ── Status / display helpers (called by telegram_bot.py) ─────────────────────

def status_text() -> str:
    """One-shot status string for the Telegram bot menu."""
    running = MONITOR_STATUS.get("running", False)
    estado = "🟢 ACTIVO" if running else "🔴 DETENIDO"
    started = MONITOR_STATUS.get("started_at") or "-"
    today = MONITOR_STATUS.get("today_total", 0)
    tmrw = MONITOR_STATUS.get("tomorrow_total", 0)
    q3 = MONITOR_STATUS.get("checked_q3", 0)
    q4 = MONITOR_STATUS.get("checked_q4", 0)
    bets = MONITOR_STATUS.get("bets_sent", 0)
    no_bet = MONITOR_STATUS.get("no_bet", 0)
    discarded = MONITOR_STATUS.get("discarded", 0)
    active = MONITOR_STATUS.get("active_matches", [])
    last = MONITOR_STATUS.get("last_event") or "-"

    lines = [
        f"Monitor Apuestas: {estado}",
        f"Inicio UTC: {started}",
        f"Hoy: {today} partidos  |  Manana: {tmrw}",
        f"Chequeados: Q3={q3}  Q4={q4}",
        f"Apuestas: {bets}  Sin apuesta: {no_bet}  Descartados: {discarded}",
    ]
    if active:
        lines.append(f"Vigilando ahora: {len(active)} partido(s)")
    lines.append(f"Ultima actividad: {last}")
    return "\n".join(lines)


def schedule_text(db_path: str, local_date: str) -> str:
    """Return a formatted schedule for local_date, capped at ~3800 chars.

    Rows with a BET signal are shown first; the rest are shown after,
    trimmed if the total would exceed Telegram's 4096-char message limit.
    """
    conn = _open_db(db_path)
    rows = conn.execute(
        """
        SELECT * FROM bet_monitor_schedule
        WHERE event_date = ?
        ORDER BY scheduled_utc_ts ASC
        """,
        (local_date,),
    ).fetchall()
    conn.close()

    if not rows:
        return f"Sin itinerario guardado para {local_date}."

    def _fmt_row(row: sqlite3.Row) -> str:
        ts = int(row["scheduled_utc_ts"] or 0)
        if ts:
            dt_l = (
                datetime.fromtimestamp(ts, tz=timezone.utc)
                + timedelta(hours=UTC_OFFSET_HOURS)
            )
            hora = dt_l.strftime("%H:%M")
        else:
            hora = "--:--"
        status_short = {"pending": "⏳", "done": "✔", "discarded": "✗"}.get(
            str(row["status"] or "pending"), "?"
        )
        q3s = str(row["q3_signal"] or "-")[:6]
        q4s = str(row["q4_signal"] or "-")[:6]
        n3 = "✉" if int(row["q3_notified"] or 0) else ""
        n4 = "✉" if int(row["q4_notified"] or 0) else ""
        home_s = str(row["home_team"] or "?")[:11]
        away_s = str(row["away_team"] or "?")[:11]
        return (
            f"{hora} {status_short} {home_s} vs {away_s} | "
            f"Q3:{q3s}{n3} Q4:{q4s}{n4}"
        )

    # Rows with a BET come first so they are never truncated
    bet_rows = [
        r for r in rows
        if _is_bet_signal(str(r["q3_signal"] or "")) or _is_bet_signal(str(r["q4_signal"] or ""))
    ]
    other_rows = [
        r for r in rows
        if r not in bet_rows
    ]

    header = f"Itinerario {local_date} ({len(rows)} partidos):"
    parts = [header]
    if bet_rows:
        parts.append("— Apuestas —")
        for r in bet_rows:
            parts.append(_fmt_row(r))
        parts.append("— Resto —")

    LIMIT = 3800
    used = sum(len(p) + 1 for p in parts)
    omitted = 0
    for r in other_rows:
        line = _fmt_row(r)
        if used + len(line) + 1 > LIMIT:
            omitted += 1
        else:
            parts.append(line)
            used += len(line) + 1

    if omitted:
        parts.append(f"... y {omitted} partido(s) mas")

    return "\n".join(parts)


def log_text(db_path: str, limit: int = 20) -> str:
    """Return recent bet_monitor_log entries, capped at Telegram's limit."""
    conn = _open_db(db_path)
    rows = conn.execute(
        "SELECT * FROM bet_monitor_log ORDER BY created_at DESC LIMIT ?",
        (limit,),
    ).fetchall()
    conn.close()

    if not rows:
        return "Sin registros en la bitacora del monitor."

    lines = [f"Bitacora monitor (ultimos {len(rows)}):"]
    for row in rows:
        ts = str(row["created_at"] or "-")[:16]
        tgt = str(row["target"] or "?").upper()
        sig = str(row["signal"] or "-")
        pick = str(row["pick"] or "-")
        conf = row["confidence"]
        conf_txt = f" {float(conf) * 100:.0f}%" if conf else ""
        result = str(row["result"] or "pending")
        res_sym = {"win": "✅", "loss": "❌", "push": "➖", "pending": "⏳"}.get(result, "")
        home_s = str(row["home_team"] or "?")[:10]
        away_s = str(row["away_team"] or "?")[:10]
        lines.append(
            f"{ts} | {tgt} {sig}{conf_txt} {res_sym} | {home_s} vs {away_s} → {pick}"
        )
    text = "\n".join(lines)
    if len(text) > 3800:
        text = text[:3750] + "\n... (truncado)"
    return text


def schedule_keyboard(db_path: str, local_date: str):
    """Return (header_text, InlineKeyboardMarkup) — one button per match."""
    from telegram import InlineKeyboardButton, InlineKeyboardMarkup

    conn = _open_db(db_path)
    rows = conn.execute(
        "SELECT * FROM bet_monitor_schedule WHERE event_date = ?"
        " ORDER BY scheduled_utc_ts ASC",
        (local_date,),
    ).fetchall()
    conn.close()

    nav = [
        [InlineKeyboardButton("⬅️ Monitor", callback_data="menu:monitor")],
        [InlineKeyboardButton("Menu principal", callback_data="nav:main")],
    ]

    if not rows:
        return (f"Sin itinerario para {local_date}.", InlineKeyboardMarkup(nav))

    def _sig_emoji(row) -> str:
        q3 = str(row["q3_signal"] or "")
        q4 = str(row["q4_signal"] or "")
        if _is_bet_signal(q3) or _is_bet_signal(q4):
            return "🟢"
        if q3 in ("NO_BET", "NO BET") or q4 in ("NO_BET", "NO BET"):
            return "🔴"
        if str(row["status"] or "") in ("done", "discarded"):
            return "✗"
        return "⏳"

    bet_rows = [
        r for r in rows
        if _is_bet_signal(str(r["q3_signal"] or "")) or _is_bet_signal(str(r["q4_signal"] or ""))
    ]
    other_rows = [r for r in rows if r not in bet_rows]

    buttons = []
    for r in (bet_rows + other_rows)[:80]:
        emoji = _sig_emoji(r)
        ts = int(r["scheduled_utc_ts"] or 0)
        hora = (
            (datetime.fromtimestamp(ts, tz=timezone.utc) + timedelta(hours=UTC_OFFSET_HOURS)).strftime("%H:%M")
            if ts else "--:--"
        )
        home_s = str(r["home_team"] or "?")[:12]
        away_s = str(r["away_team"] or "?")[:12]
        q3s = str(r["q3_signal"] or "-")[:6]
        q4s = str(r["q4_signal"] or "-")[:6]
        q3m = str(r["q3_model"] or "")
        q4m = str(r["q4_model"] or "")
        qmodels = f"{q3m}/{q4m}" if (q3m or q4m) else "-"
        label = (
            f"{emoji} {hora} {home_s} vs {away_s} "
            f"Q3:{q3s} Q4:{q4s} [{qmodels}]"
        )
        buttons.append([
            InlineKeyboardButton(label, callback_data=f"monitor:match:{r['match_id']}")
        ])

    buttons.extend(nav)
    header = f"Itinerario {local_date} — {len(rows)} partidos ({len(bet_rows)} apuestas):"
    return header, InlineKeyboardMarkup(buttons)


def log_keyboard(db_path: str, limit: int = 20):
    """Return (header_text, InlineKeyboardMarkup) — one button per log row."""
    from telegram import InlineKeyboardButton, InlineKeyboardMarkup
    conn = _open_db(db_path)
    rows = conn.execute(
        "SELECT * FROM bet_monitor_log ORDER BY created_at DESC LIMIT ?",
        (limit,),
    ).fetchall()
    conn.close()

    nav = [
        [InlineKeyboardButton("⬅️ Monitor", callback_data="menu:monitor")],
        [InlineKeyboardButton("Menu principal", callback_data="nav:main")],
    ]

    if not rows:
        return ("Sin registros en la bitacora.", InlineKeyboardMarkup(nav))

    buttons = []
    for r in rows:
        result = str(r["result"] or "pending")
        res_emoji = {"win": "✅", "loss": "❌", "push": "➖", "pending": "⏳"}.get(result, "⏳")
        sig = str(r["signal"] or "-")
        sig_emoji = "🟢" if _is_bet_signal(sig) else ("🔴" if "NO" in sig.upper() else "⚪")
        tgt = str(r["target"] or "?").upper()
        model = str(r["model"] or "-")
        conf = r["confidence"]
        conf_txt = f" {float(conf) * 100:.0f}%" if conf else ""
        home_s = str(r["home_team"] or "?")[:10]
        away_s = str(r["away_team"] or "?")[:10]
        label = f"{res_emoji}{sig_emoji} {tgt}{conf_txt} [{model}] {home_s} vs {away_s}"[:64]
        ev_date = str(r["event_date"] or "_")
        cd = f"match:{r['match_id']}:{ev_date}:0"
        buttons.append([InlineKeyboardButton(label, callback_data=cd)])

    buttons.extend(nav)
    header = f"Bitácora monitor (últimos {len(rows)} registros):"
    return header, InlineKeyboardMarkup(buttons)


def signals_text_today(
    db_path: str,
    local_date: str,
    pref: str = "all",
    quarters: list[str] | None = None,
) -> str:
    """Return a formatted list of today's signals (BET and/or NO_BET).

    pref='bet_only' hides NO_BET-only matches and NO_BET lines inside
    matches that have at least one BET signal.
    pref='all' (default) shows everything.
    
    quarters: list of quarters to include (e.g., ["q3", "q4"]). If None, includes all.
    """
    if quarters is None:
        quarters = ["q3", "q4"]
    quarters_lower = [q.lower() for q in quarters]
    conn = _open_db(db_path)

    # Latest log entry per (match_id, target) for today, ordered by schedule time
    rows = conn.execute(
        """
        SELECT l.match_id, l.target, l.signal, l.pick, l.confidence,
               l.recommendation, l.result, l.created_at,
               l.home_team, l.away_team, l.league, l.model,
               s.scheduled_utc_ts
        FROM bet_monitor_log l
        INNER JOIN (
            SELECT match_id, target, MAX(id) AS max_id
            FROM bet_monitor_log
            WHERE event_date = ?
            GROUP BY match_id, target
        ) latest ON l.match_id = latest.match_id
                 AND l.target  = latest.target
                 AND l.id      = latest.max_id
        LEFT JOIN bet_monitor_schedule s ON l.match_id = s.match_id
        WHERE NOT (l.league LIKE '%WNBA%' OR l.league LIKE '%Women%' OR l.league LIKE '%women%' OR l.league LIKE '%Feminina%' OR l.league LIKE '%Femenina%')
          AND NOT (l.league LIKE '%Playoff%' OR l.league LIKE '%PLAY OFF%')
          AND NOT l.league LIKE '%U21 Espoirs Elite%'
          AND NOT l.league LIKE '%Liga Femenina%'
          AND NOT l.league LIKE '%LF Challenge%'
          AND NOT l.league LIKE '%Polish Basketball League%'
          AND NOT l.league LIKE '%SuperSport Premijer Liga%'
          AND NOT l.league LIKE '%Prvenstvo Hrvatske za d%'
          AND NOT l.league LIKE '%ABA Liga%'
          AND NOT l.league LIKE '%Argentina Liga Nacional%'
          AND NOT l.league LIKE '%Basketligaen%'
          AND NOT l.league LIKE '%lite 2%'
          AND NOT l.league LIKE '%EYBL%'
          AND NOT l.league LIKE '%I B MCKL%'
          AND NOT l.league LIKE '%Liga 1 Masculin%'
          AND NOT l.league LIKE '%Liga Nationala%'
          AND NOT l.league LIKE '%NBL1%'
          AND NOT l.league LIKE '%PBA Commissioner%'
          AND NOT l.league LIKE '%Rapid League%'
          AND NOT l.league LIKE '%Stoiximan GBL%'
          AND NOT l.league LIKE '%Playout%'
          AND NOT l.league LIKE '%Superleague%'
          AND NOT l.league LIKE '%Superliga%'
          AND NOT l.league LIKE '%Swedish Basketball Superettan%'
          AND NOT l.league LIKE '%Swiss Cup%'
          AND NOT l.league LIKE '%Финал%'
          AND NOT l.league LIKE '%Turkish Basketball Super League%'
          AND NOT l.league LIKE '%NBA%'
          AND NOT l.league LIKE '%Big V%'
          AND NOT l.league LIKE '%Egyptian Basketball Super League%'
          AND NOT l.league LIKE '%Lega A Basket%'
          AND NOT l.league LIKE '%Liga e Par%'
          AND NOT l.league LIKE '%Liga Ouro%'
          AND NOT l.league LIKE '%Señal%'
          AND NOT l.league LIKE '%LNB%'
          AND NOT l.league LIKE '%Meridianbet KLS%'
          AND NOT l.league LIKE '%MPBL%'
          AND NOT l.league LIKE '%Nationale 1%'
          AND NOT l.league LIKE '%Poland 2nd Basketball League%'
          AND NOT l.league LIKE '%Portugal LBP%'
          AND NOT l.league LIKE '%Portugal Proliga%'
          AND NOT l.league LIKE '%Saku I liiga%'
          AND NOT l.league LIKE '%Serie A2%'
          AND NOT l.league LIKE '%Slovenian Second Basketball%'
          AND NOT l.league LIKE '%Super League%'
          AND NOT l.league LIKE '%United Cup%'
          AND NOT l.league LIKE '%United League%'
          AND (l.signal NOT IN ('BET', 'BET_HOME', 'BET_AWAY') OR l.confidence > 0.30)
        ORDER BY COALESCE(s.scheduled_utc_ts, 0) ASC, l.created_at ASC
        """,
        (local_date,),
    ).fetchall()

    # Scheduled matches for today (to show pending ones)
    sched_rows = conn.execute(
        """
        SELECT match_id, home_team, away_team, league,
               scheduled_utc_ts, status
        FROM bet_monitor_schedule
        WHERE event_date = ?
          AND NOT (league LIKE '%WNBA%' OR league LIKE '%Women%' OR league LIKE '%women%' OR league LIKE '%Feminina%' OR league LIKE '%Femenina%')
          AND NOT league LIKE '%Playoff%'
          AND NOT league LIKE '%U21 Espoirs Elite%'
          AND NOT league LIKE '%Liga Femenina%'
          AND NOT league LIKE '%LF Challenge%'
          AND NOT league LIKE '%Polish Basketball League%'
          AND NOT league LIKE '%SuperSport Premijer Liga%'
          AND NOT league LIKE '%Prvenstvo Hrvatske za d%'
          AND NOT league LIKE '%ABA Liga%'
          AND NOT league LIKE '%Argentina Liga Nacional%'
          AND NOT league LIKE '%Basketligaen%'
          AND NOT league LIKE '%lite 2%'
          AND NOT league LIKE '%EYBL%'
          AND NOT league LIKE '%I B MCKL%'
          AND NOT league LIKE '%Liga 1 Masculin%'
          AND NOT league LIKE '%Liga Nationala%'
          AND NOT league LIKE '%NBL1%'
          AND NOT league LIKE '%PBA Commissioner%'
          AND NOT league LIKE '%Rapid League%'
          AND NOT league LIKE '%Stoiximan GBL%'
          AND NOT league LIKE '%Playout%'
          AND NOT league LIKE '%Superleague%'
          AND NOT league LIKE '%Superliga%'
          AND NOT league LIKE '%Swedish Basketball Superettan%'
          AND NOT league LIKE '%Swiss Cup%'
          AND NOT league LIKE '%Финал%'
          AND NOT league LIKE '%Turkish Basketball Super League%'
          AND NOT league LIKE '%NBA%'
          AND NOT league LIKE '%Big V%'
          AND NOT league LIKE '%Egyptian Basketball Super League%'
          AND NOT league LIKE '%Lega A Basket%'
          AND NOT league LIKE '%Liga e Par%'
          AND NOT league LIKE '%Liga Ouro%'
          AND NOT league LIKE '%Señal%'
          AND NOT league LIKE '%LNB%'
          AND NOT league LIKE '%Meridianbet KLS%'
          AND NOT league LIKE '%MPBL%'
          AND NOT league LIKE '%Nationale 1%'
          AND NOT league LIKE '%Poland 2nd Basketball League%'
          AND NOT league LIKE '%Portugal LBP%'
          AND NOT league LIKE '%Portugal Proliga%'
          AND NOT league LIKE '%Saku I liiga%'
          AND NOT league LIKE '%Serie A2%'
          AND NOT league LIKE '%Slovenian Second Basketball%'
          AND NOT league LIKE '%Super League%'
          AND NOT league LIKE '%United Cup%'
          AND NOT league LIKE '%United League%'
        ORDER BY scheduled_utc_ts ASC
        """,
        (local_date,),
    ).fetchall()

    # Quarter scores stored after match finishes
    _qs_rows = conn.execute(
        """
        SELECT qs.match_id, qs.quarter, qs.home, qs.away
        FROM quarter_scores qs
        INNER JOIN bet_monitor_schedule s ON qs.match_id = s.match_id
        WHERE s.event_date = ?
        """,
        (local_date,),
    ).fetchall()
    conn.close()

    # Build: quarter_actual[(match_id, "Q3")] = "home" | "away" | "push"
    quarter_actual: dict[tuple, str] = {}
    for qr in _qs_rows:
        h = qr["home"]
        a = qr["away"]
        if h is None or a is None:
            continue
        try:
            h, a = int(h), int(a)
        except (TypeError, ValueError):
            continue
        if h > a:
            winner = "home"
        elif a > h:
            winner = "away"
        else:
            winner = "push"
        quarter_actual[(str(qr["match_id"]), str(qr["quarter"]))] = winner

    # Spanish date header
    _DAY_ES = [
        "Lunes", "Martes", "Miércoles", "Jueves",
        "Viernes", "Sábado", "Domingo",
    ]
    _MON_ES = [
        "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
    ]
    try:
        _d = datetime.strptime(local_date, "%Y-%m-%d")
        _date_hdr = (
            f"{_DAY_ES[_d.weekday()]} {_d.day}"
            f" {_MON_ES[_d.month - 1]} {_d.year}"
        )
    except ValueError:
        _date_hdr = local_date

    def _hora(ts) -> str:
        ts_int = int(ts or 0)
        if not ts_int:
            return "--:--"
        return (
            datetime.fromtimestamp(ts_int, tz=timezone.utc)
            + timedelta(hours=UTC_OFFSET_HOURS)
        ).strftime("%H:%M")

    def _pick_sym(pick: str) -> str:
        if pick == "home":
            return "🏠"
        if pick == "away":
            return "✈️"
        return pick

    def _sig_line(
        tgt: str,
        s: dict,
        show_no_bet: bool = True,
        mid: str = "",
    ) -> str | None:
        sig    = str(s.get("signal") or "-").upper()
        pick   = str(s.get("pick") or "")
        conf   = s.get("confidence") or 0.0
        result = str(s.get("result") or "pending")
        model  = str(s.get("model") or "")
        conf_txt  = f" {float(conf) * 100:.0f}%" if conf else ""
        res_sym   = {"win": "✅", "loss": "❌", "push": "➖", "pending": "⏳"}.get(result, "⏳")
        model_txt = f" [{model}]" if model else ""
        pick_sym  = _pick_sym(pick)
        if _is_bet_signal(sig):
            return (
                f"  🟢 {tgt.upper()}{model_txt}"
                f" → {pick_sym}{conf_txt} {res_sym}"
            ).rstrip()
        # NO_BET line
        if not show_no_bet:
            return None
        tendency = f" {pick_sym}{conf_txt}" if pick_sym else ""
        # Real quarter result (if available)
        _q_key = tgt.upper()   # "Q3" or "Q4"
        actual = quarter_actual.get((mid, _q_key))
        if actual == "home":
            real_txt = " (🏠)"
        elif actual == "away":
            real_txt = " (✈️)"
        elif actual == "push":
            real_txt = " (=)"
        else:
            real_txt = ""
        return f"  🔴 {tgt.upper()}{model_txt}{tendency}{real_txt}"

    # Group signals by match_id, preserving insertion order
    match_signals: dict[str, dict] = {}
    match_meta: dict[str, dict] = {}
    for row in rows:
        mid = str(row["match_id"])
        tgt = str(row["target"] or "").lower()
        # Apply v6 filter: suppress BET signals rejected by the filter
        _row_model  = str(row["model"]      or "") if "model"      in row.keys() else ""
        _row_signal = str(row["signal"]     or "") if "signal"     in row.keys() else ""
        _row_league = str(row["league"]     or "") if "league"     in row.keys() else ""
        _row_conf   = float(row["confidence"] or 0.0) if "confidence" in row.keys() else 0.0
        _row_pick   = str(row["pick"]       or "") if "pick"       in row.keys() else ""
        if _is_bet_signal(_row_signal.upper()):
            if _row_model == "v6":
                _accept, _ = _v6_pick_filter(_row_league, _row_conf, _row_pick)
                if not _accept:
                    continue
            if _row_model == "v6_2":
                _accept, _ = _v6_2_pick_filter(_row_league, _row_conf, _row_pick)
                if not _accept:
                    continue
            if _row_model == "v2":
                _accept, _ = _v2_pick_filter(_row_league, _row_conf, _row_pick)
                if not _accept:
                    continue
        if mid not in match_signals:
            match_signals[mid] = {}
            match_meta[mid] = {
                "home":   str(row["home_team"] or "?"),
                "away":   str(row["away_team"] or "?"),
                "league": str(row["league"] or "?"),
                "ts":     row["scheduled_utc_ts"],
            }
        match_signals[mid][tgt] = dict(row)

    # Scheduled matches that have no signals at all yet
    signaled_ids = set(match_signals.keys())
    pending_sched = [
        dict(r) for r in sched_rows
        if r["match_id"] not in signaled_ids
    ]

    # Filter by quarters: remove quarters not in quarters_lower
    for mid in match_signals:
        quarters_to_remove = [q for q in match_signals[mid] if q not in quarters_lower]
        for q in quarters_to_remove:
            del match_signals[mid][q]

    show_no_bet = pref == "all"

    bet_mids = [
        m for m in match_signals
        if any(
            _is_bet_signal(str(match_signals[m].get(t, {}).get("signal", "")))
            for t in ("q3", "q4")
        )
    ]
    nobet_mids = [m for m in match_signals if m not in bet_mids]

    def _render_match(mid: str, force_show_no_bet: bool) -> list[str]:
        meta = match_meta[mid]
        hora = _hora(meta["ts"])
        lines_m = [f"{hora}  {meta['home']} vs {meta['away']} ({meta['league']})"]
        for tgt in ("q3", "q4"):
            if tgt in match_signals[mid]:
                line = _sig_line(
                    tgt,
                    match_signals[mid][tgt],
                    show_no_bet=force_show_no_bet,
                    mid=mid,
                )
                if line is not None:
                    lines_m.append(line)
        # Skip match block if no lines were added
        if len(lines_m) == 1:
            return []
        return lines_m

    # ── Stats + simulated bank (FIRST, before signals) ────────────────────
    # Read sim config from settings
    cfg_conn = _open_db(db_path)
    try:
        _bank0  = float(cfg_conn.execute(
            "SELECT value FROM settings WHERE key='sig_bank'"
        ).fetchone()["value"] or 1000)
    except Exception:
        _bank0 = 1000.0
    try:
        _bet_sz = float(cfg_conn.execute(
            "SELECT value FROM settings WHERE key='sig_bet_size'"
        ).fetchone()["value"] or 100)
    except Exception:
        _bet_sz = 100.0
    try:
        _odds   = float(cfg_conn.execute(
            "SELECT value FROM settings WHERE key='sig_odds'"
        ).fetchone()["value"] or 1.4)
    except Exception:
        _odds = 1.4
    cfg_conn.close()

    def _calc_stats(mids: list, target: str = None) -> dict:
        """Count wins/losses/push/pending for BET signals in mids.
        
        If target is provided (e.g., "q3" or "q4"), filter to that quarter only.
        """
        w = l = p = pending = 0
        for mid in mids:
            targets = [target] if target else ("q3", "q4")
            for tgt in targets:
                s = match_signals[mid].get(tgt)
                if not s:
                    continue
                if not _is_bet_signal(str(s.get("signal") or "")):
                    continue
                r = str(s.get("result") or "pending")
                if r == "win":
                    w += 1
                elif r in ("loss", "push"):
                    l += 1
                elif r == "pending":
                    pending += 1
        return {"w": w, "l": l, "pending": pending}

    def _calc_nobet_stats(mids: list) -> dict:
        """For NO_BET signals: what would outcome have been?"""
        w = l = p = pending = 0
        for mid in mids:
            for tgt in ("q3", "q4"):
                s = match_signals[mid].get(tgt)
                if not s:
                    continue
                if _is_bet_signal(str(s.get("signal") or "")):
                    continue
                pick = str(s.get("pick") or "")
                actual = quarter_actual.get((mid, tgt.upper()))
                if actual is None:
                    pending += 1
                    continue
                if actual == "push":
                    l += 1
                elif actual == pick:
                    w += 1
                else:
                    l += 1
        return {"w": w, "l": l, "pending": pending}

    def _roi_line(st: dict, label: str, bank: float, bet: float, odds: float) -> str:
        played = st["w"] + st["l"]
        if played == 0 and st["pending"] == 0:
            return ""
        profit = st["w"] * bet * (odds - 1) - st["l"] * bet
        bank_end = bank + profit
        roi = (profit / (played * bet) * 100) if played > 0 else 0.0
        hit = (st["w"] / played * 100) if played > 0 else 0.0
        pend_txt = f"+{st['pending']}⏳\n" if st["pending"] else ""
        sign = "+" if profit >= 0 else ""
        return (
            f"{label}\n"
            f"{st['w']}✅{hit:.0f}%\n"
            f"{st['l']}❌\n"
            f"{pend_txt}"
            f"ROI {sign}{roi:.1f}%\n"
            f"Bank ${bank:.0f}→${bank_end:.0f}"
        )

    def _format_table_stats(st_q3: dict, st_q4: dict, model_q3: str, model_q4: str,
                            bank: float, bet: float, odds: float,
                            active_quarters: list[str] | None = None) -> str:
        """Format Q3 and Q4 stats side-by-side in a table.
        active_quarters: list of lowercase quarter names to show (e.g. ['q3','q4']).
        Columns for inactive quarters are omitted.
        """
        if active_quarters is None:
            active_quarters = ["q3", "q4"]
        show_q3 = "q3" in active_quarters
        show_q4 = "q4" in active_quarters
        lines = []

        def _col(text: str) -> str:
            return text.ljust(16)

        # Header
        header = ""
        if show_q3:
            header += _col(f"Q3 [{model_q3}]") + " "
        if show_q4:
            header += _col(f"Q4 [{model_q4}]")
        lines.append(f"  {header.rstrip()}")

        # Wins & hit rate
        played_q3 = st_q3["w"] + st_q3["l"]
        hit_q3 = (st_q3["w"] / played_q3 * 100) if played_q3 > 0 else 0.0
        played_q4 = st_q4["w"] + st_q4["l"]
        hit_q4 = (st_q4["w"] / played_q4 * 100) if played_q4 > 0 else 0.0
        wins_row = ""
        if show_q3:
            wins_row += _col(f"{st_q3['w']}✅{hit_q3:.0f}%") + " "
        if show_q4:
            wins_row += _col(f"{st_q4['w']}✅{hit_q4:.0f}%")
        lines.append(f"  {wins_row.rstrip()}")

        # Losses
        loss_row = ""
        if show_q3:
            loss_row += _col(f"{st_q3['l']}❌") + " "
        if show_q4:
            loss_row += _col(f"{st_q4['l']}❌")
        lines.append(f"  {loss_row.rstrip()}")

        # Pending
        has_pending = (show_q3 and st_q3["pending"]) or (show_q4 and st_q4["pending"])
        if has_pending:
            pend_row = ""
            if show_q3:
                pend_row += _col(f"+{st_q3['pending']}⏳" if st_q3["pending"] else "") + " "
            if show_q4:
                pend_row += _col(f"+{st_q4['pending']}⏳" if st_q4["pending"] else "")
            lines.append(f"  {pend_row.rstrip()}")

        # ROI
        profit_q3 = st_q3["w"] * bet * (odds - 1) - st_q3["l"] * bet
        roi_q3 = (profit_q3 / (played_q3 * bet) * 100) if played_q3 > 0 else 0.0
        sign_q3 = "+" if profit_q3 >= 0 else ""
        profit_q4 = st_q4["w"] * bet * (odds - 1) - st_q4["l"] * bet
        roi_q4 = (profit_q4 / (played_q4 * bet) * 100) if played_q4 > 0 else 0.0
        sign_q4 = "+" if profit_q4 >= 0 else ""
        roi_row = ""
        if show_q3:
            roi_row += _col(f"ROI {sign_q3}{roi_q3:.1f}%") + " "
        if show_q4:
            roi_row += _col(f"ROI {sign_q4}{roi_q4:.1f}%")
        lines.append(f"  {roi_row.rstrip()}")

        # Bank
        bank_row = ""
        if show_q3:
            bank_row += _col(f"${bank:.0f}→${bank + profit_q3:.0f}") + " "
        if show_q4:
            bank_row += _col(f"${bank:.0f}→${bank + profit_q4:.0f}")
        lines.append(f"  {bank_row.rstrip()}")

        return "\n".join(lines)

    bet_st_q3 = _calc_stats(bet_mids, "q3")
    bet_st_q4 = _calc_stats(bet_mids, "q4")
    nobet_st  = _calc_nobet_stats(nobet_mids) if show_no_bet else None
    # Also compute nobet lines inside bet matches
    nobet_in_bet = _calc_nobet_stats(bet_mids) if show_no_bet else None

    # Build stats section early with separate Q3 and Q4 simulations
    stat_lines = []
    
    # Read active models from in-memory config (set_model_config keeps this current)
    q3_model = _model_config.get("q3", "v4")
    q4_model = _model_config.get("q4", "v4")
    
    # Build table with Q3 and Q4 side by side
    played_q3 = bet_st_q3["w"] + bet_st_q3["l"]
    played_q4 = bet_st_q4["w"] + bet_st_q4["l"]
    has_bets = played_q3 > 0 or played_q4 > 0 or bet_st_q3["pending"] or bet_st_q4["pending"]
    
    if has_bets:
        table_txt = _format_table_stats(bet_st_q3, bet_st_q4, q3_model, q4_model, _bank0, _bet_sz, _odds,
                                        active_quarters=quarters_lower)
        stat_lines.append(table_txt)
    if show_no_bet:
        # Merge nobet_mids + nobet quarters inside bet matches
        nb_all = {
            "w": (nobet_st["w"] if nobet_st else 0)
                 + (nobet_in_bet["w"] if nobet_in_bet else 0),
            "l": (nobet_st["l"] if nobet_st else 0)
                 + (nobet_in_bet["l"] if nobet_in_bet else 0),
            "pending": (nobet_st["pending"] if nobet_st else 0)
                       + (nobet_in_bet["pending"] if nobet_in_bet else 0),
        }
        nb_line = _roi_line(
            nb_all, "📉 No apostado",
            _bank0, _bet_sz, _odds,
        )
        if nb_line:
            stat_lines.append(nb_line)

    # Build message with stats at TOP (after title)
    lines: list[str] = [f"📊 Señales {_date_hdr}:"]
    
    # Add stats section immediately after title (so it's never truncated)
    if stat_lines:
        cfg_txt = (
            f"(odds {_odds} · apuesta ${int(_bet_sz)}"
            f" · Bank ${int(_bank0)})"
        )
        lines.append(f"── Resumen {cfg_txt} ──")
        lines.extend(stat_lines)
        lines.append("")

    # Then add signal sections
    if bet_mids:
        lines.append("— 🟢 Apuestas —")
        for mid in bet_mids:
            # In bet_only mode, suppress NO_BET lines within the match
            lines.extend(_render_match(mid, force_show_no_bet=show_no_bet))

    if show_no_bet and nobet_mids:
        lines.append("— 🔴 Sin apuesta —")
        for mid in nobet_mids:
            lines.extend(_render_match(mid, force_show_no_bet=True))

    if pending_sched:
        _active_sched = [r for r in pending_sched if str(r.get("status") or "pending") != "discarded"]
        if _active_sched:
            lines.append(f"— ⏳ Sin evaluar ({len(_active_sched)}) —")
            for r in _active_sched[:20]:
                hora = _hora(r.get("scheduled_utc_ts"))
                home_s = str(r.get("home_team") or "?")[:14]
                away_s = str(r.get("away_team") or "?")[:14]
                lines.append(f"  {hora} {home_s} vs {away_s}")
            if len(_active_sched) > 20:
                lines.append(f"  ... y {len(_active_sched) - 20} más")

    if not match_signals and not pending_sched:
        lines.append("Sin datos para hoy. ¿El monitor está corriendo?")

    text = "\n".join(lines)
    if len(text) > 3800:
        text = text[:3750] + "\n... (truncado)"
    return text


def signals_report_today(
    db_path: str,
    local_date: str,
) -> str:
    """Return a detailed match-by-match report showing bankroll progression."""
    conn = _open_db(db_path)

    # Get all BET signals for today, ordered by schedule time
    rows = conn.execute(
        """
        SELECT l.match_id, l.target, l.signal, l.pick, l.confidence,
               l.result, l.created_at,
               l.home_team, l.away_team, l.league, l.model,
               s.scheduled_utc_ts
        FROM bet_monitor_log l
        INNER JOIN (
            SELECT match_id, target, MAX(id) AS max_id
            FROM bet_monitor_log
            WHERE event_date = ? AND signal IN ('BET', 'BET_HOME', 'BET_AWAY')
            GROUP BY match_id, target
        ) latest ON l.match_id = latest.match_id
                 AND l.target  = latest.target
                 AND l.id      = latest.max_id
        LEFT JOIN bet_monitor_schedule s ON l.match_id = s.match_id
        ORDER BY COALESCE(s.scheduled_utc_ts, 0) ASC, l.created_at ASC
        """,
        (local_date,),
    ).fetchall()

    # Quarter scores
    _qs_rows = conn.execute(
        """
        SELECT qs.match_id, qs.quarter, qs.home, qs.away
        FROM quarter_scores qs
        INNER JOIN bet_monitor_schedule s ON qs.match_id = s.match_id
        WHERE s.event_date = ?
        """,
        (local_date,),
    ).fetchall()
    
    # Read config
    try:
        _bank0  = float(conn.execute(
            "SELECT value FROM settings WHERE key='sig_bank'"
        ).fetchone()["value"] or 1000)
    except Exception:
        _bank0 = 1000.0
    try:
        _bet_sz = float(conn.execute(
            "SELECT value FROM settings WHERE key='sig_bet_size'"
        ).fetchone()["value"] or 100)
    except Exception:
        _bet_sz = 100.0
    try:
        _odds   = float(conn.execute(
            "SELECT value FROM settings WHERE key='sig_odds'"
        ).fetchone()["value"] or 1.4)
    except Exception:
        _odds = 1.4
    
    conn.close()

    # Build quarter_actual map
    quarter_actual: dict[tuple, str] = {}
    for qr in _qs_rows:
        h = qr["home"]
        a = qr["away"]
        if h is None or a is None:
            continue
        try:
            h, a = int(h), int(a)
        except (TypeError, ValueError):
            continue
        if h > a:
            winner = "home"
        elif a > h:
            winner = "away"
        else:
            winner = "push"
        quarter_actual[(str(qr["match_id"]), str(qr["quarter"]))] = winner

    # Build report
    lines = []
    current_bank = _bank0
    
    for i, row in enumerate(rows, 1):
        mid = str(row["match_id"])
        tgt = str(row["target"] or "").upper()
        home = str(row["home_team"] or "?")
        away = str(row["away_team"] or "?")
        pick = str(row["pick"] or "")
        result = str(row["result"] or "pending")
        model = str(row["model"] or "-")
        
        # Determine actual result
        actual = quarter_actual.get((mid, tgt))
        
        # Calculate outcome
        if result == "win":
            outcome = "✅ GANADA"
            profit = _bet_sz * (_odds - 1)
        elif result == "loss":
            outcome = "❌ PERDIDA"
            profit = -_bet_sz
        elif result == "push":
            outcome = "➖ EMPATE"
            profit = 0
        else:  # pending
            outcome = "⏳ PENDIENTE"
            profit = 0
        
        pick_sym = "🏠" if pick == "home" else ("✈️" if pick == "away" else "")
        pick_name = home if pick == "home" else (away if pick == "away" else pick)
        
        bank_before = current_bank
        bank_after = current_bank + profit
        current_bank = bank_after
        
        # Format line
        line = (
            f"{i}. {home} vs {away} ({tgt} [{model}])\n"
            f"   {pick_sym} {pick_name} → {outcome}\n"
            f"   Banco: ${bank_before:.0f} → ${bank_after:.0f} ({profit:+.0f})"
        )
        lines.append(line)
    
    # Header
    _DAY_ES = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    _MON_ES = [
        "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
    ]
    try:
        _d = datetime.strptime(local_date, "%Y-%m-%d")
        _date_hdr = (
            f"{_DAY_ES[_d.weekday()]} {_d.day} {_MON_ES[_d.month - 1]} {_d.year}"
        )
    except ValueError:
        _date_hdr = local_date

    text = f"📋 Reporte {_date_hdr}\n\n"
    if lines:
        text += "\n\n".join(lines)
        text += f"\n\nBanco Final: ${current_bank:.0f}"
    else:
        text += "Sin apuestas para hoy."
    
    if len(text) > 3800:
        text = text[:3750] + "\n... (truncado)"
    return text


def _normalize_sofascore_slug(value: str | None) -> str:
    """Normalize team/event names for SofaScore URL."""
    import unicodedata
    import re
    text = (value or "").strip().lower()
    if not text:
        return "unknown"
    normalized = unicodedata.normalize("NFKD", text)
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    ascii_text = re.sub(r"[^a-z0-9]+", "-", ascii_text).strip("-")
    return ascii_text or "unknown"


def _sofascore_match_url(match_id: str, match_data: dict | None = None, home_team: str | None = None, away_team: str | None = None) -> str:
    """Generate SofaScore URL for a match."""
    event_slug = "unknown"
    custom_id = ""
    home_slug = _normalize_sofascore_slug(home_team) if home_team else "unknown"
    away_slug = _normalize_sofascore_slug(away_team) if away_team else "unknown"

    if match_data and "match" in match_data:
        match_info = match_data["match"]
        event_slug = _normalize_sofascore_slug(match_info.get("event_slug"))
        custom_id = str(match_info.get("custom_id") or "").strip()
        home_slug = _normalize_sofascore_slug(match_info.get("home_slug"))
        away_slug = _normalize_sofascore_slug(match_info.get("away_slug"))
        if home_slug == "unknown":
            home_slug = _normalize_sofascore_slug(match_info.get("home_team"))
        if away_slug == "unknown":
            away_slug = _normalize_sofascore_slug(match_info.get("away_team"))

    if event_slug != "unknown" and custom_id:
        return (
            "https://www.sofascore.com/basketball/match/"
            f"{event_slug}/{custom_id}#id:{match_id}"
        )

    return (
        "https://www.sofascore.com/basketball/match/"
        f"{home_slug}/{away_slug}#id:{match_id}"
    )


def signals_excel_today(
    db_path: str,
    local_date: str,
    quarters: list[str] | None = None,
) -> bytes:
    """Generate an Excel file with Q3 and Q4 sheets showing bankroll progression.
    
    quarters: list of quarters to include (e.g., ["q3", "q4"]). If None, includes all.
    
    Filters:
    - Excluye ligas de mujeres (WNBA, Women's, etc.)
    """
    if quarters is None:
        quarters = ["q3", "q4"]
    quarters_lower = [q.lower() for q in quarters]
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import scraper as scraper_mod

    conn = _open_db(db_path)

    # Get all BET signals for today, grouped by target
    rows_q3 = [] if "q3" not in quarters_lower else conn.execute(
        """
        SELECT l.match_id, l.target, l.signal, l.pick, l.confidence,
               l.result, l.created_at,
               l.home_team, l.away_team, l.league, l.model,
               s.scheduled_utc_ts
        FROM bet_monitor_log l
        INNER JOIN (
            SELECT match_id, target, MAX(id) AS max_id
            FROM bet_monitor_log
            WHERE event_date = ? AND signal IN ('BET', 'BET_HOME', 'BET_AWAY') AND target = 'q3'
            GROUP BY match_id, target
        ) latest ON l.match_id = latest.match_id
                 AND l.target  = latest.target
                 AND l.id      = latest.max_id
        LEFT JOIN bet_monitor_schedule s ON l.match_id = s.match_id
        WHERE l.confidence > 0.30
          AND NOT (l.league LIKE '%WNBA%' OR l.league LIKE '%Women%' OR l.league LIKE '%women%' OR l.league LIKE '%Feminina%' OR l.league LIKE '%Femenina%')
          AND NOT (l.league LIKE '%Playoff%' OR l.league LIKE '%PLAY OFF%')
          AND NOT l.league LIKE '%U21 Espoirs Elite%'
          AND NOT l.league LIKE '%Liga Femenina%'
          AND NOT l.league LIKE '%LF Challenge%'
          AND NOT l.league LIKE '%Polish Basketball League%'
          AND NOT l.league LIKE '%SuperSport Premijer Liga%'
          AND NOT l.league LIKE '%Prvenstvo Hrvatske za d%'
          AND NOT l.league LIKE '%ABA Liga%'
          AND NOT l.league LIKE '%Argentina Liga Nacional%'
          AND NOT l.league LIKE '%Basketligaen%'
          AND NOT l.league LIKE '%lite 2%'
          AND NOT l.league LIKE '%EYBL%'
          AND NOT l.league LIKE '%I B MCKL%'
          AND NOT l.league LIKE '%Liga 1 Masculin%'
          AND NOT l.league LIKE '%Liga Nationala%'
          AND NOT l.league LIKE '%NBL1%'
          AND NOT l.league LIKE '%PBA Commissioner%'
          AND NOT l.league LIKE '%Rapid League%'
          AND NOT l.league LIKE '%Stoiximan GBL%'
          AND NOT l.league LIKE '%Playout%'
          AND NOT l.league LIKE '%Superleague%'
          AND NOT l.league LIKE '%Superliga%'
          AND NOT l.league LIKE '%Swedish Basketball Superettan%'
          AND NOT l.league LIKE '%Swiss Cup%'
          AND NOT l.league LIKE '%Финал%'
          AND NOT l.league LIKE '%Turkish Basketball Super League%'
          AND NOT l.league LIKE '%NBA%'
          AND NOT l.league LIKE '%Big V%'
          AND NOT l.league LIKE '%Egyptian Basketball Super League%'
          AND NOT l.league LIKE '%Lega A Basket%'
          AND NOT l.league LIKE '%Liga e Par%'
          AND NOT l.league LIKE '%Liga Ouro%'
          AND NOT l.league LIKE '%Señal%'
          AND NOT l.league LIKE '%LNB%'
          AND NOT l.league LIKE '%Meridianbet KLS%'
          AND NOT l.league LIKE '%MPBL%'
          AND NOT l.league LIKE '%Nationale 1%'
          AND NOT l.league LIKE '%Poland 2nd Basketball League%'
          AND NOT l.league LIKE '%Portugal LBP%'
          AND NOT l.league LIKE '%Portugal Proliga%'
          AND NOT l.league LIKE '%Saku I liiga%'
          AND NOT l.league LIKE '%Serie A2%'
          AND NOT l.league LIKE '%Slovenian Second Basketball%'
          AND NOT l.league LIKE '%Super League%'
          AND NOT l.league LIKE '%United Cup%'
          AND NOT l.league LIKE '%United League%'
        ORDER BY COALESCE(s.scheduled_utc_ts, 0) ASC, l.created_at ASC
        """,
        (local_date,),
    ).fetchall()

    rows_q4 = [] if "q4" not in quarters_lower else conn.execute(
        """
        SELECT l.match_id, l.target, l.signal, l.pick, l.confidence,
               l.result, l.created_at,
               l.home_team, l.away_team, l.league, l.model,
               s.scheduled_utc_ts
        FROM bet_monitor_log l
        INNER JOIN (
            SELECT match_id, target, MAX(id) AS max_id
            FROM bet_monitor_log
            WHERE event_date = ? AND signal IN ('BET', 'BET_HOME', 'BET_AWAY') AND target = 'q4'
            GROUP BY match_id, target
        ) latest ON l.match_id = latest.match_id
                 AND l.target  = latest.target
                 AND l.id      = latest.max_id
        LEFT JOIN bet_monitor_schedule s ON l.match_id = s.match_id
        WHERE l.confidence > 0.30
          AND NOT (l.league LIKE '%WNBA%' OR l.league LIKE '%Women%' OR l.league LIKE '%women%' OR l.league LIKE '%Feminina%' OR l.league LIKE '%Femenina%')
          AND NOT (l.league LIKE '%Playoff%' OR l.league LIKE '%PLAY OFF%')
          AND NOT l.league LIKE '%U21 Espoirs Elite%'
          AND NOT l.league LIKE '%Liga Femenina%'
          AND NOT l.league LIKE '%LF Challenge%'
          AND NOT l.league LIKE '%Polish Basketball League%'
          AND NOT l.league LIKE '%SuperSport Premijer Liga%'
          AND NOT l.league LIKE '%Prvenstvo Hrvatske za d%'
          AND NOT l.league LIKE '%ABA Liga%'
          AND NOT l.league LIKE '%Argentina Liga Nacional%'
          AND NOT l.league LIKE '%Basketligaen%'
          AND NOT l.league LIKE '%lite 2%'
          AND NOT l.league LIKE '%EYBL%'
          AND NOT l.league LIKE '%I B MCKL%'
          AND NOT l.league LIKE '%Liga 1 Masculin%'
          AND NOT l.league LIKE '%Liga Nationala%'
          AND NOT l.league LIKE '%NBL1%'
          AND NOT l.league LIKE '%PBA Commissioner%'
          AND NOT l.league LIKE '%Rapid League%'
          AND NOT l.league LIKE '%Stoiximan GBL%'
          AND NOT l.league LIKE '%Playout%'
          AND NOT l.league LIKE '%Superleague%'
          AND NOT l.league LIKE '%Superliga%'
          AND NOT l.league LIKE '%Swedish Basketball Superettan%'
          AND NOT l.league LIKE '%Swiss Cup%'
          AND NOT l.league LIKE '%Финал%'
          AND NOT l.league LIKE '%Turkish Basketball Super League%'
          AND NOT l.league LIKE '%NBA%'
          AND NOT l.league LIKE '%Big V%'
          AND NOT l.league LIKE '%Egyptian Basketball Super League%'
          AND NOT l.league LIKE '%Lega A Basket%'
          AND NOT l.league LIKE '%Liga e Par%'
          AND NOT l.league LIKE '%Liga Ouro%'
          AND NOT l.league LIKE '%Señal%'
          AND NOT l.league LIKE '%LNB%'
          AND NOT l.league LIKE '%Meridianbet KLS%'
          AND NOT l.league LIKE '%MPBL%'
          AND NOT l.league LIKE '%Nationale 1%'
          AND NOT l.league LIKE '%Poland 2nd Basketball League%'
          AND NOT l.league LIKE '%Portugal LBP%'
          AND NOT l.league LIKE '%Portugal Proliga%'
          AND NOT l.league LIKE '%Saku I liiga%'
          AND NOT l.league LIKE '%Serie A2%'
          AND NOT l.league LIKE '%Slovenian Second Basketball%'
          AND NOT l.league LIKE '%Super League%'
          AND NOT l.league LIKE '%United Cup%'
          AND NOT l.league LIKE '%United League%'
        ORDER BY COALESCE(s.scheduled_utc_ts, 0) ASC, l.created_at ASC
        """,
        (local_date,),
    ).fetchall()

    # Quarter scores
    _qs_rows = conn.execute(
        """
        SELECT qs.match_id, qs.quarter, qs.home, qs.away
        FROM quarter_scores qs
        INNER JOIN bet_monitor_schedule s ON qs.match_id = s.match_id
        WHERE s.event_date = ?
        """,
        (local_date,),
    ).fetchall()
    
    # Read config
    try:
        _bank0  = float(conn.execute(
            "SELECT value FROM settings WHERE key='sig_bank'"
        ).fetchone()["value"] or 1000)
    except Exception:
        _bank0 = 1000.0
    try:
        _bet_sz = float(conn.execute(
            "SELECT value FROM settings WHERE key='sig_bet_size'"
        ).fetchone()["value"] or 100)
    except Exception:
        _bet_sz = 100.0
    try:
        _odds   = float(conn.execute(
            "SELECT value FROM settings WHERE key='sig_odds'"
        ).fetchone()["value"] or 1.4)
    except Exception:
        _odds = 1.4
    
    conn.close()

    # Apply v6/v6_2 filter to rows
    def _apply_v6_filter_rows(rows):
        out = []
        for row in rows:
            if str(row.get("model") or "") == "v6":
                _accept, _ = _v6_pick_filter(
                    str(row.get("league") or ""),
                    float(row.get("confidence") or 0),
                    str(row.get("pick") or ""),
                )
                if not _accept:
                    continue
            elif str(row.get("model") or "") == "v6_2":
                _accept, _ = _v6_2_pick_filter(
                    str(row.get("league") or ""),
                    float(row.get("confidence") or 0),
                    str(row.get("pick") or ""),
                )
                if not _accept:
                    continue
            out.append(row)
        return out

    rows_q3 = _apply_v6_filter_rows(rows_q3)
    rows_q4 = _apply_v6_filter_rows(rows_q4)

    # Fetch match_data for all match_ids to get event_slug and custom_id for SofaScore URLs
    all_rows = (rows_q3 or []) + (rows_q4 or [])
    unique_match_ids = list(set(str(row["match_id"]) for row in all_rows))
    
    match_data_map: dict[str, dict | None] = {}
    if unique_match_ids:
        try:
            # Fetch match data from scraper (with timeout)
            results = scraper_mod.fetch_matches_by_ids(unique_match_ids)
            for mid, data, error in results:
                match_data_map[str(mid)] = data if data else None
        except Exception as e:
            # If scraper fails, fall back to using None (URLs will use home/away fallback)
            logger.warning(f"Failed to fetch match data for Excel: {e}")
            for mid in unique_match_ids:
                match_data_map[str(mid)] = None

    # Build quarter_actual map: (match_id, quarter) -> ("home"|"away"|"push"|None)
    quarter_actual: dict[tuple, str | None] = {}
    for qr in _qs_rows:
        h = qr["home"]
        a = qr["away"]
        mid = str(qr["match_id"])
        q = str(qr["quarter"]).upper()
        if h is None or a is None:
            quarter_actual[(mid, q)] = None
            continue
        try:
            h, a = int(h), int(a)
        except (TypeError, ValueError):
            quarter_actual[(mid, q)] = None
            continue
        if h > a:
            winner = "home"
        elif a > h:
            winner = "away"
        else:
            winner = "push"
        quarter_actual[(mid, q)] = winner

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    # Colors for results
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    def _add_sheet(ws, rows, target_quarter):
        """Add a sheet with match data."""
        # Headers
        headers = ["Nro", "Match ID", "Fecha", "Hora (UTC-6)", "Home", "Away", "Liga",
                   "Pick", "Modelo", "Confianza", "Resultado", "Q Result", "Bank Antes", "Ganancia", "Bank Despues"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        current_bank = _bank0
        row_num = 2
        
        for i, row in enumerate(rows, 1):
            mid = str(row["match_id"])
            tgt = str(row["target"] or "").upper()
            home = str(row["home_team"] or "?")
            away = str(row["away_team"] or "?")
            league = str(row["league"] or "")
            pick = str(row["pick"] or "")
            model = str(row["model"] or "")
            confidence = row["confidence"]
            result = str(row["result"] or "pending")
            ts = row["scheduled_utc_ts"]
            
            # Calculate outcome (push = loss)
            if result == "win":
                outcome = "✅"
                profit = _bet_sz * (_odds - 1)
                result_fill = green_fill
            else:  # loss, push, pending -> all treated as loss for display
                if result == "pending":
                    outcome = "⏳"
                    profit = 0
                    result_fill = gray_fill
                else:  # loss or push
                    outcome = "❌"
                    profit = -_bet_sz
                    result_fill = red_fill
            
            # Format date and time
            if ts:
                try:
                    dt = datetime.fromtimestamp(int(ts), tz=timezone.utc)
                    # Apply UTC-6 offset
                    dt_local = dt + timedelta(hours=UTC_OFFSET_HOURS)
                    date_str = dt_local.strftime("%Y-%m-%d")
                    time_str = dt_local.strftime("%H:%M")
                except Exception:
                    date_str = ""
                    time_str = ""
            else:
                date_str = ""
                time_str = ""
            
            # Get quarter result
            q_key = (mid, tgt)
            quarter_result = quarter_actual.get(q_key)
            if quarter_result == "home":
                q_result_text = f"🏠 {home}"
            elif quarter_result == "away":
                q_result_text = f"✈️ {away}"
            elif quarter_result == "push":
                q_result_text = "➖ Push"
            else:
                q_result_text = "⏳"
            
            pick_sym = "🏠" if pick == "home" else ("✈️" if pick == "away" else "")
            
            bank_before = current_bank
            bank_after = current_bank + profit
            current_bank = bank_after
            
            # Build SofaScore URL with match_data if available
            sofascore_url = _sofascore_match_url(mid, match_data=match_data_map.get(mid), home_team=home, away_team=away)
            
            # Row data
            ws.cell(row=row_num, column=1).value = i
            # Match ID with hyperlink
            match_cell = ws.cell(row=row_num, column=2)
            match_cell.value = mid
            match_cell.hyperlink = sofascore_url
            match_cell.font = Font(color="0563C1", underline="single")
            
            ws.cell(row=row_num, column=3).value = date_str
            ws.cell(row=row_num, column=4).value = time_str
            ws.cell(row=row_num, column=5).value = home
            ws.cell(row=row_num, column=6).value = away
            ws.cell(row=row_num, column=7).value = league
            ws.cell(row=row_num, column=8).value = f"{pick_sym} {pick}".strip()
            ws.cell(row=row_num, column=9).value = model
            ws.cell(row=row_num, column=10).value = confidence
            ws.cell(row=row_num, column=11).value = outcome
            ws.cell(row=row_num, column=12).value = q_result_text
            ws.cell(row=row_num, column=13).value = bank_before
            ws.cell(row=row_num, column=14).value = profit
            ws.cell(row=row_num, column=15).value = bank_after
            
            # Format cells
            for col in range(1, 16):
                cell = ws.cell(row=row_num, column=col)
                cell.border = border
                if col == 11:  # Resultado
                    cell.fill = result_fill
                    cell.alignment = center_align
                elif col in (1, 3, 4, 7, 8, 9, 10, 11, 12):
                    cell.alignment = center_align
                elif col in (5, 6):
                    cell.alignment = left_align
                if col == 10:  # Confianza - percentage format
                    if confidence is not None:
                        cell.number_format = '0.0%'
                    cell.alignment = center_align
                if col in (13, 14, 15):
                    cell.number_format = '$#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
            
            row_num += 1
        
        # Auto-fit column widths based on content
        from openpyxl.utils import get_column_letter
        for col_num in range(1, 16):
            col_letter = get_column_letter(col_num)
            max_length = 0
            column_cells = ws[f'{col_letter}:{col_letter}']
            for cell in column_cells:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            # Add padding and set width (max 50)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
    
    # Add Q3 sheet
    if rows_q3:
        ws_q3 = wb.create_sheet("Q3", 0)
        _add_sheet(ws_q3, rows_q3, "Q3")
    
    # Add Q4 sheet
    if rows_q4:
        ws_q4 = wb.create_sheet("Q4", 1 if rows_q3 else 0)
        _add_sheet(ws_q4, rows_q4, "Q4")
    
    # If no sheets, add empty summary
    if not rows_q3 and not rows_q4:
        ws = wb.create_sheet("Info", 0)
        ws.cell(row=1, column=1).value = "Sin apuestas para hoy"
    
    # Save to bytes
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def signals_excel_monthly(
    db_path: str,
    year_month: str,
    quarters: list[str] | None = None,
) -> bytes:
    """Generate an Excel report for a full month (year_month='YYYY-MM').

    Applies identical filters as signals_excel_today:
    - League exclusions
    - BET confidence > 0.30
    - Push = loss in bankroll simulation

    Sheets: Q3, Q4 (bet rows by date), Resumen (per-day summary + totals).
    """
    if quarters is None:
        quarters = ["q3", "q4"]
    quarters_lower = [q.lower() for q in quarters]

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    conn = _open_db(db_path)

    _BET_SIGNALS_SQL = "('BET', 'BET_HOME', 'BET_AWAY')"
    _LEAGUE_FILTER_SQL = """
        AND NOT (l.league LIKE '%WNBA%' OR l.league LIKE '%Women%'
                 OR l.league LIKE '%women%' OR l.league LIKE '%Feminina%' OR l.league LIKE '%Femenina%')
        AND NOT (l.league LIKE '%Playoff%' OR l.league LIKE '%PLAY OFF%')
        AND NOT l.league LIKE '%U21 Espoirs Elite%'
        AND NOT l.league LIKE '%Liga Femenina%'
        AND NOT l.league LIKE '%LF Challenge%'
        AND NOT l.league LIKE '%Polish Basketball League%'
        AND NOT l.league LIKE '%SuperSport Premijer Liga%'
        AND NOT l.league LIKE '%Prvenstvo Hrvatske za d%'
        AND NOT l.league LIKE '%ABA Liga%'
        AND NOT l.league LIKE '%Argentina Liga Nacional%'
        AND NOT l.league LIKE '%Basketligaen%'
        AND NOT l.league LIKE '%lite 2%'
        AND NOT l.league LIKE '%EYBL%'
        AND NOT l.league LIKE '%I B MCKL%'
        AND NOT l.league LIKE '%Liga 1 Masculin%'
        AND NOT l.league LIKE '%Liga Nationala%'
        AND NOT l.league LIKE '%NBL1%'
        AND NOT l.league LIKE '%PBA Commissioner%'
        AND NOT l.league LIKE '%Rapid League%'
        AND NOT l.league LIKE '%Stoiximan GBL%'
        AND NOT l.league LIKE '%Playout%'
        AND NOT l.league LIKE '%Superleague%'
        AND NOT l.league LIKE '%Superliga%'
        AND NOT l.league LIKE '%Swedish Basketball Superettan%'
        AND NOT l.league LIKE '%Swiss Cup%'
        AND NOT l.league LIKE '%Финал%'
        AND NOT l.league LIKE '%Turkish Basketball Super League%'
        AND NOT l.league LIKE '%NBA%'
        AND NOT l.league LIKE '%Big V%'
        AND NOT l.league LIKE '%Egyptian Basketball Super League%'
        AND NOT l.league LIKE '%Lega A Basket%'
        AND NOT l.league LIKE '%Liga e Par%'
        AND NOT l.league LIKE '%Liga Ouro%'
        AND NOT l.league LIKE '%Señal%'
        AND NOT l.league LIKE '%LNB%'
        AND NOT l.league LIKE '%Meridianbet KLS%'
        AND NOT l.league LIKE '%MPBL%'
        AND NOT l.league LIKE '%Nationale 1%'
        AND NOT l.league LIKE '%Poland 2nd Basketball League%'
        AND NOT l.league LIKE '%Portugal LBP%'
        AND NOT l.league LIKE '%Portugal Proliga%'
        AND NOT l.league LIKE '%Saku I liiga%'
        AND NOT l.league LIKE '%Serie A2%'
        AND NOT l.league LIKE '%Slovenian Second Basketball%'
        AND NOT l.league LIKE '%Super League%'
        AND NOT l.league LIKE '%United Cup%'
        AND NOT l.league LIKE '%United League%'
    """

    def _fetch_quarter_rows(target: str) -> list:
        return conn.execute(
            f"""
            SELECT l.match_id, l.target, l.signal, l.pick, l.confidence,
                   l.result, l.created_at,
                   l.home_team, l.away_team, l.league, l.model,
                   l.event_date,
                   s.scheduled_utc_ts
            FROM bet_monitor_log l
            INNER JOIN (
                SELECT match_id, target, MAX(id) AS max_id
                FROM bet_monitor_log
                WHERE event_date LIKE ? AND signal IN {_BET_SIGNALS_SQL}
                  AND target = ?
                GROUP BY match_id, target
            ) latest ON l.match_id = latest.match_id
                     AND l.target  = latest.target
                     AND l.id      = latest.max_id
            LEFT JOIN bet_monitor_schedule s ON l.match_id = s.match_id
            WHERE l.confidence > 0.30
            {_LEAGUE_FILTER_SQL}
            ORDER BY l.event_date ASC, COALESCE(s.scheduled_utc_ts, 0) ASC, l.created_at ASC
            """,
            (f"{year_month}-%", target),
        ).fetchall()

    rows_q3 = _fetch_quarter_rows("q3") if "q3" in quarters_lower else []
    rows_q4 = _fetch_quarter_rows("q4") if "q4" in quarters_lower else []

    # Quarter scores for the month (all match_ids present in log)
    all_match_ids = list({str(r["match_id"]) for r in (rows_q3 + rows_q4)})
    quarter_actual: dict[tuple, str | None] = {}
    if all_match_ids:
        placeholders = ",".join("?" * len(all_match_ids))
        qs_rows = conn.execute(
            f"""
            SELECT match_id, quarter, home, away
            FROM quarter_scores
            WHERE match_id IN ({placeholders})
            """,
            all_match_ids,
        ).fetchall()
        for qr in qs_rows:
            mid = str(qr["match_id"])
            q = str(qr["quarter"]).upper()
            h, a = qr["home"], qr["away"]
            if h is None or a is None:
                quarter_actual[(mid, q)] = None
                continue
            try:
                h, a = int(h), int(a)
            except (TypeError, ValueError):
                quarter_actual[(mid, q)] = None
                continue
            quarter_actual[(mid, q)] = "home" if h > a else ("away" if a > h else "push")

    try:
        _bank0 = float(
            conn.execute("SELECT value FROM settings WHERE key='sig_bank'").fetchone()["value"]
            or 1000
        )
    except Exception:
        _bank0 = 1000.0
    try:
        _bet_sz = float(
            conn.execute("SELECT value FROM settings WHERE key='sig_bet_size'").fetchone()["value"]
            or 100
        )
    except Exception:
        _bet_sz = 100.0
    try:
        _odds = float(
            conn.execute("SELECT value FROM settings WHERE key='sig_odds'").fetchone()["value"]
            or 1.4
        )
    except Exception:
        _odds = 1.4

    conn.close()

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    summary_header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    total_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    # Collect per-day stats for summary sheet
    # day -> {q3: {w,l,nb,profit}, q4: {w,l,nb,profit}}
    day_stats: dict[str, dict] = {}

    def _add_bet_sheet(ws, rows, tgt_label: str):
        nonlocal day_stats
        headers = [
            "Nro", "Match ID", "Fecha", "Hora (UTC-6)",
            "Home", "Away", "Liga",
            "Pick", "Modelo", "Confianza",
            "Resultado", "Q Result",
            "Bank Antes", "Ganancia", "Bank Despues",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = h
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        current_bank = _bank0
        tgt_key = tgt_label.lower()

        for i, row in enumerate(rows, 1):
            mid = str(row["match_id"])
            home = str(row["home_team"] or "?")
            away = str(row["away_team"] or "?")
            league = str(row["league"] or "")
            pick = str(row["pick"] or "")
            model = str(row["model"] or "")
            confidence = row["confidence"]
            result = str(row["result"] or "pending")
            event_date = str(row["event_date"] or "")
            ts = row["scheduled_utc_ts"]

            if result == "win":
                outcome, profit, result_fill = "✅", _bet_sz * (_odds - 1), green_fill
            elif result == "pending":
                outcome, profit, result_fill = "⏳", 0.0, gray_fill
            else:  # loss or push → loss
                outcome, profit, result_fill = "❌", -_bet_sz, red_fill

            # Date/time
            if ts:
                try:
                    dt = datetime.fromtimestamp(int(ts), tz=timezone.utc)
                    dt_local = dt + timedelta(hours=UTC_OFFSET_HOURS)
                    date_str = dt_local.strftime("%Y-%m-%d")
                    time_str = dt_local.strftime("%H:%M")
                except Exception:
                    date_str = event_date
                    time_str = ""
            else:
                date_str = event_date
                time_str = ""

            q_key = (mid, tgt_label.upper())
            qr = quarter_actual.get(q_key)
            q_result_text = (
                f"🏠 {home}" if qr == "home"
                else (f"✈️ {away}" if qr == "away"
                      else ("➖ Push" if qr == "push" else "⏳"))
            )

            pick_sym = "🏠" if pick == "home" else ("✈️" if pick == "away" else "")
            bank_before = current_bank
            bank_after = current_bank + profit
            current_bank = bank_after

            # Accumulate day stats
            if event_date not in day_stats:
                day_stats[event_date] = {
                    "q3": {"w": 0, "l": 0, "profit": 0.0},
                    "q4": {"w": 0, "l": 0, "profit": 0.0},
                }
            ds = day_stats[event_date][tgt_key]
            if result == "win":
                ds["w"] += 1
                ds["profit"] += profit
            elif result != "pending":
                ds["l"] += 1
                ds["profit"] += profit

            sofascore_url = _sofascore_match_url(mid, home_team=home, away_team=away)
            row_num = i + 1

            ws.cell(row=row_num, column=1).value = i
            match_cell = ws.cell(row=row_num, column=2)
            match_cell.value = mid
            match_cell.hyperlink = sofascore_url
            match_cell.font = Font(color="0563C1", underline="single")
            ws.cell(row=row_num, column=3).value = date_str
            ws.cell(row=row_num, column=4).value = time_str
            ws.cell(row=row_num, column=5).value = home
            ws.cell(row=row_num, column=6).value = away
            ws.cell(row=row_num, column=7).value = league
            ws.cell(row=row_num, column=8).value = f"{pick_sym} {pick}".strip()
            ws.cell(row=row_num, column=9).value = model
            ws.cell(row=row_num, column=10).value = confidence
            ws.cell(row=row_num, column=11).value = outcome
            ws.cell(row=row_num, column=12).value = q_result_text
            ws.cell(row=row_num, column=13).value = bank_before
            ws.cell(row=row_num, column=14).value = profit
            ws.cell(row=row_num, column=15).value = bank_after

            for col in range(1, 16):
                cell = ws.cell(row=row_num, column=col)
                cell.border = border
                if col == 11:
                    cell.fill = result_fill
                    cell.alignment = center_align
                elif col in (1, 3, 4, 7, 8, 9, 10, 11, 12):
                    cell.alignment = center_align
                elif col in (5, 6):
                    cell.alignment = left_align
                if col == 10 and confidence is not None:
                    cell.number_format = "0.0%"
                    cell.alignment = center_align
                if col in (13, 14, 15):
                    cell.number_format = "$#,##0.00"
                    cell.alignment = Alignment(horizontal="right")

        # Auto-fit
        from openpyxl.utils import get_column_letter as gcl
        for col_num in range(1, 16):
            col_letter = gcl(col_num)
            max_len = max(
                (len(str(c.value)) for c in ws[col_letter] if c.value),
                default=8,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    if rows_q3:
        ws_q3 = wb.create_sheet("Q3", 0)
        _add_bet_sheet(ws_q3, rows_q3, "Q3")
    if rows_q4:
        ws_q4 = wb.create_sheet("Q4", 1 if rows_q3 else 0)
        _add_bet_sheet(ws_q4, rows_q4, "Q4")

    # ── Resumen sheet ───────────────────────────────────────────────────────
    ws_res = wb.create_sheet("Resumen", 0)

    def _res_header(ws, col, text):
        cell = ws.cell(row=1, column=col)
        cell.value = text
        cell.fill = summary_header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = center_align
        cell.border = border

    res_headers = [
        "Fecha", "Q3 W", "Q3 L", "Q3 ROI", "Q3 Ganancia",
        "Q4 W", "Q4 L", "Q4 ROI", "Q4 Ganancia",
        "Total W", "Total L", "Total Ganancia", "Bank Final",
    ]
    for c, h in enumerate(res_headers, 1):
        _res_header(ws_res, c, h)

    running_bank = _bank0
    tot_w = tot_l = 0
    tot_profit = 0.0

    for row_i, day in enumerate(sorted(day_stats.keys()), 2):
        ds = day_stats[day]
        q3w = ds["q3"]["w"]; q3l = ds["q3"]["l"]; q3p = ds["q3"]["profit"]
        q4w = ds["q4"]["w"]; q4l = ds["q4"]["l"]; q4p = ds["q4"]["profit"]
        dw = q3w + q4w; dl = q3l + q4l; dp = q3p + q4p
        tot_w += dw; tot_l += dl; tot_profit += dp
        running_bank += dp

        def _roi_val(w, l):
            played = w + l
            if played == 0:
                return None
            return (w * _bet_sz * (_odds - 1) - l * _bet_sz) / (played * _bet_sz)

        vals = [
            day,
            q3w, q3l, _roi_val(q3w, q3l), q3p,
            q4w, q4l, _roi_val(q4w, q4l), q4p,
            dw, dl, dp, running_bank,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws_res.cell(row=row_i, column=c)
            cell.value = v
            cell.border = border
            if c == 1:
                cell.alignment = center_align
            elif c in (2, 3, 6, 7, 10, 11):
                cell.alignment = center_align
            elif c in (4, 8):
                if v is not None:
                    cell.number_format = "0.0%"
                cell.alignment = center_align
            elif c in (5, 9, 12, 13):
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            # colour profit cells
            if c == 12:
                cell.fill = green_fill if (v or 0) >= 0 else red_fill

    # Totals row
    total_row = len(day_stats) + 2
    total_roi = (
        (tot_w * _bet_sz * (_odds - 1) - tot_l * _bet_sz) / ((tot_w + tot_l) * _bet_sz)
        if (tot_w + tot_l) > 0
        else None
    )
    total_vals = [
        "TOTAL", "", "", "", "",
        "", "", "", "",
        tot_w, tot_l, tot_profit, _bank0 + tot_profit,
    ]
    for c, v in enumerate(total_vals, 1):
        cell = ws_res.cell(row=total_row, column=c)
        cell.value = v
        cell.fill = total_fill
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = center_align
        if c in (12, 13):
            cell.number_format = "$#,##0.00"
            cell.alignment = Alignment(horizontal="right")

    # ROI in totals row (col 4 as overall)
    roi_cell = ws_res.cell(row=total_row, column=4)
    roi_cell.value = total_roi
    roi_cell.fill = total_fill
    roi_cell.font = Font(bold=True)
    roi_cell.border = border
    roi_cell.number_format = "0.0%"
    roi_cell.alignment = center_align

    # Auto-fit resumen
    from openpyxl.utils import get_column_letter as gcl
    for col_num in range(1, len(res_headers) + 1):
        col_letter = gcl(col_num)
        max_len = max(
            (len(str(c.value)) for c in ws_res[col_letter] if c.value),
            default=8,
        )
        ws_res.column_dimensions[col_letter].width = min(max_len + 2, 30)

    if not rows_q3 and not rows_q4:
        ws_res.cell(row=2, column=1).value = "Sin apuestas para el mes"

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
