"""
Reporte comparativo V12 vs V13 — out-of-sample (Abril 2026)

Re-corre inferencia de V12 y V13 sobre matches del rango de fechas dado
y genera un Excel con:
  Hoja 1 - Partidos   : señales BET/NO_BET, pick, confianza, resultado real
  Hoja 2 - Resumen    : W/L/ROI por modelo × cuarto
  Hoja 3 - Por Día    : ROI diario por modelo
  Hoja 4 - Ligas      : ROI por liga por modelo

Uso:
    python training/report_v12_v13.py --month 2026-04
    python training/report_v12_v13.py --from 2026-04-01 --to 2026-04-15
    python training/report_v12_v13.py --month 2026-04 --odds 1.91 --bet-size 20
"""

from __future__ import annotations

import argparse
import importlib
import sys
import warnings
from datetime import date, timedelta, datetime
from pathlib import Path
from typing import Any

warnings.filterwarnings("ignore", category=UserWarning)

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None  # type: ignore[assignment]

db_mod = importlib.import_module("db")

DB_PATH = ROOT / "matches.db"
REPORTS_DIR = ROOT / "reports"

# ── Excel styles ──────────────────────────────────────────────────────────────
def _fill(hex_color):
    if not openpyxl:
        return None
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

_GREEN  = _fill("C6EFCE")
_RED    = _fill("FFC7CE")
_YELLOW = _fill("FFEB9C")
_BLUE   = _fill("2F75B6")
_LBLUE  = _fill("BDD7EE")
_GRAY   = _fill("F2F2F2")
_LGRAY  = _fill("D9D9D9")
_ORANGE = _fill("F4B942")

def _font(bold=False, white=False, size=9):
    if not openpyxl:
        return None
    color = "FFFFFF" if white else "000000"
    return Font(bold=bold, color=color, size=size)

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True) if openpyxl else None
_LEFT   = Alignment(horizontal="left",   vertical="center") if openpyxl else None


def _s(ws, row, col, value, fill=None, font=None, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    return cell


# ── Inference ─────────────────────────────────────────────────────────────────

_v12_mod = None
_v13_mod = None


def _load_modules(models: list[str] | None = None):
    global _v12_mod, _v13_mod
    active = models or MODELS
    if "v12" in active and _v12_mod is None:
        _v12_mod = importlib.import_module("training.v12.infer_match_v12")
    if "v13" in active and _v13_mod is None:
        _v13_mod = importlib.import_module("training.v13.infer_match_v13")


def _infer_v12(match_id: str, target: str) -> dict:
    """Returns normalized prediction dict or None."""
    try:
        pred = _v12_mod.run_inference(match_id=match_id, target=target, fetch_missing=False)
        if isinstance(pred, dict) and not pred.get("ok", True):
            return {}

        def _a(name, default=None):
            if isinstance(pred, dict):
                return pred.get(name, default)
            return getattr(pred, name, default)

        return {
            "available": True,
            "pick": _a("winner_pick"),
            "confidence": _a("winner_confidence"),
            "signal": str(_a("final_signal") or _a("winner_signal") or "NO_BET").upper(),
            "predicted_home": _a("predicted_home"),
            "predicted_away": _a("predicted_away"),
        }
    except Exception as exc:
        return {"error": str(exc)}


def _infer_v13(match_id: str, target: str) -> dict:
    """Returns normalized prediction dict or None."""
    try:
        result = _v13_mod.run_inference(match_id=match_id, target=target, fetch_missing=False)
        if not result.get("ok", False):
            return {}
        pred = result.get("prediction")

        def _a(name, default=None):
            if isinstance(pred, dict):
                return pred.get(name, default)
            return getattr(pred, name, default)

        return {
            "available": True,
            "pick": _a("winner_pick"),
            "confidence": _a("winner_confidence"),
            "signal": str(_a("final_signal") or _a("winner_signal") or "NO_BET").upper(),
            "predicted_home": _a("predicted_home"),
            "predicted_away": _a("predicted_away"),
            "model_quality": _a("model_quality"),
            "model_f1": _a("model_f1"),
            "fallback_used": _a("fallback_used", False),
        }
    except Exception as exc:
        return {"error": str(exc)}


# ── Data loading ──────────────────────────────────────────────────────────────

def _load_matches(date_from: str, date_to: str) -> list[dict]:
    conn = db_mod.get_conn(str(DB_PATH))
    conn.row_factory = __import__("sqlite3").Row
    # Join with eval_match_results to get quarter scores (use MAX(updated_at) to
    # get the most recent row per match_id from eval_match_results)
    rows = conn.execute(
        """
        SELECT
            m.match_id, m.home_team, m.away_team, m.date, m.league,
            m.home_score, m.away_score,
            e.q3_home_score, e.q3_away_score,
            e.q4_home_score, e.q4_away_score,
            e.q3_winner, e.q4_winner
        FROM matches m
        LEFT JOIN (
            SELECT match_id,
                   q3_home_score, q3_away_score,
                   q4_home_score, q4_away_score,
                   q3_winner, q4_winner
            FROM eval_match_results
            GROUP BY match_id
            HAVING MAX(updated_at)
        ) e ON e.match_id = m.match_id
        WHERE m.date >= ? AND m.date <= ?
        ORDER BY m.date, m.match_id
        """,
        (date_from, date_to),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def _winner_from_scores(home, away):
    try:
        h, a = int(home), int(away)
    except (TypeError, ValueError):
        return None
    if h > a:
        return "home"
    if a > h:
        return "away"
    return "draw"


def _is_bet(signal: str) -> bool:
    """True for BET, BET_HOME, BET_AWAY."""
    return "BET" in str(signal).upper() and "NO_BET" not in str(signal).upper()


def _outcome(signal: str, pick: str, actual_winner: str | None) -> str:
    """Return 'W', 'L', 'PENDING', or '-'."""
    if not _is_bet(signal) or not pick:
        return "-"
    if actual_winner is None:
        return "PENDING"
    if pick == actual_winner:
        return "W"
    return "L"


# ── P&L simulation ────────────────────────────────────────────────────────────

def _simulate(bet_outcomes: list[str], odds: float, bet_size: float, starting_bank: float = 0.0) -> dict:
    """bet_outcomes: list of outcome strings 'W'/'L'."""
    wins = sum(1 for r in bet_outcomes if r == "W")
    losses = sum(1 for r in bet_outcomes if r == "L")
    total = wins + losses
    pnl = wins * (bet_size * (odds - 1)) - losses * bet_size
    roi = pnl / (total * bet_size) * 100 if total > 0 else 0.0
    be = 1 / odds * 100  # break-even %
    return {
        "bets": total,
        "wins": wins,
        "losses": losses,
        "win_pct": wins / total * 100 if total > 0 else 0.0,
        "pnl": pnl,
        "roi": roi,
        "break_even": be,
        "bank_start": starting_bank,
        "bank_final": starting_bank + pnl if starting_bank > 0 else None,
    }


# ── Main report logic ─────────────────────────────────────────────────────────

MODELS = ["v12", "v13"]
QUARTERS = ["q3", "q4"]


def generate_report(date_from: str, date_to: str, odds: float, bet_size: float, starting_bank: float, out_path: Path, models: list[str] | None = None) -> Path:
    print(f"\n⏳ Cargando matches {date_from} → {date_to} …")
    matches = _load_matches(date_from, date_to)
    print(f"   {len(matches)} matches encontrados")

    active_models = models if models else MODELS
    print(f"⏳ Cargando modelos {', '.join(m.upper() for m in active_models)} …")
    _load_modules(active_models)

    records = []
    for i, m in enumerate(matches, 1):
        mid = m["match_id"]
        if i % 50 == 0 or i == len(matches):
            print(f"   [{i}/{len(matches)}] {m['home_team']} vs {m['away_team']}")

        q3_winner = m.get("q3_winner") or _winner_from_scores(m.get("q3_home_score"), m.get("q3_away_score"))
        q4_winner = m.get("q4_winner") or _winner_from_scores(m.get("q4_home_score"), m.get("q4_away_score"))

        rec = {
            "match_id": mid,
            "date": str(m.get("date", ""))[:10],
            "home_team": m.get("home_team", ""),
            "away_team": m.get("away_team", ""),
            "league": m.get("league", ""),
            "q3_home": m.get("home_q3_score"),
            "q3_away": m.get("away_q3_score"),
            "q4_home": m.get("home_q4_score"),
            "q4_away": m.get("away_q4_score"),
            "q3_winner": q3_winner,
            "q4_winner": q4_winner,
        }

        for model_name in active_models:
            for target in QUARTERS:
                infer_fn = _infer_v12 if model_name == "v12" else _infer_v13
                pred = infer_fn(mid, target)
                signal = pred.get("signal", "NO_BET") if pred.get("available") else "NO_BET"
                pick = pred.get("pick") or ""
                conf = pred.get("confidence")
                actual_winner = q3_winner if target == "q3" else q4_winner
                outcome = _outcome(signal, pick, actual_winner)

                key = f"{model_name}_{target}"
                rec[f"{key}_signal"] = signal
                rec[f"{key}_pick"] = pick
                rec[f"{key}_confidence"] = conf
                rec[f"{key}_outcome"] = outcome
                if model_name == "v13":
                    rec[f"{key}_model_quality"] = pred.get("model_quality")
                    rec[f"{key}_model_f1"] = pred.get("model_f1")
                    rec[f"{key}_fallback"] = pred.get("fallback_used", False)

        records.append(rec)

    print(f"\n✅ Inferencia completada — {len(records)} matches procesados")
    return _build_excel(records, date_from, date_to, odds, bet_size, starting_bank, out_path, active_models)


def _build_excel(records: list[dict], date_from: str, date_to: str, odds: float, bet_size: float, starting_bank: float, out_path: Path, models: list[str] | None = None) -> Path:
    if openpyxl is None:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")

    active_models = models if models else MODELS
    wb = openpyxl.Workbook()

    _sheet_matches(wb, records, odds, bet_size, active_models)
    _sheet_summary(wb, records, odds, bet_size, starting_bank, active_models)
    _sheet_daily(wb, records, odds, bet_size, starting_bank, active_models)
    _sheet_leagues(wb, records, odds, bet_size, active_models)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"\n📊 Reporte guardado en: {out_path}")
    return out_path


# ── Sheet: Partidos ───────────────────────────────────────────────────────────

def _sheet_matches(wb, records, odds, bet_size, models=None):
    ws = wb.active
    ws.title = "Partidos"
    active_models = models if models else MODELS

    # Build header
    header = ["Fecha", "Partido", "Liga", "Q3 Real", "Q4 Real"]
    for model in active_models:
        for q in QUARTERS:
            label = f"{model.upper()} {q.upper()}"
            header += [f"{label} Señal", f"{label} Pick", f"{label} Conf%", f"{label} Res"]
    
    for col, h in enumerate(header, 1):
        _s(ws, 1, col, h, fill=_BLUE, font=_font(bold=True, white=True), align=_CENTER)

    for row_idx, rec in enumerate(records, 2):
        col = 1
        _s(ws, row_idx, col, rec["date"], align=_CENTER); col += 1
        _s(ws, row_idx, col, f"{rec['home_team']} vs {rec['away_team']}", align=_LEFT); col += 1
        _s(ws, row_idx, col, rec["league"], align=_LEFT); col += 1
        
        # Q3 score
        q3s = f"{rec['q3_home']}-{rec['q3_away']}" if rec["q3_home"] is not None else "?"
        _s(ws, row_idx, col, q3s, align=_CENTER); col += 1
        q4s = f"{rec['q4_home']}-{rec['q4_away']}" if rec["q4_home"] is not None else "?"
        _s(ws, row_idx, col, q4s, align=_CENTER); col += 1

        for model in active_models:
            for q in QUARTERS:
                key = f"{model}_{q}"
                signal = rec.get(f"{key}_signal", "NO_BET")
                pick = rec.get(f"{key}_pick", "")
                conf = rec.get(f"{key}_confidence")
                outcome = rec.get(f"{key}_outcome", "-")

                signal_fill = _ORANGE if _is_bet(signal) else _LGRAY
                _s(ws, row_idx, col, signal, fill=signal_fill, align=_CENTER); col += 1
                _s(ws, row_idx, col, pick or "", align=_CENTER); col += 1
                conf_str = f"{conf*100:.1f}%" if conf is not None else ""
                _s(ws, row_idx, col, conf_str, align=_CENTER); col += 1

                outcome_fill = None
                if outcome == "W":
                    outcome_fill = _GREEN
                elif outcome == "L":
                    outcome_fill = _RED
                elif outcome == "PENDING":
                    outcome_fill = _YELLOW
                _s(ws, row_idx, col, outcome, fill=outcome_fill, align=_CENTER); col += 1

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 30)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ── Sheet: Resumen ────────────────────────────────────────────────────────────

def _sheet_summary(wb, records, odds, bet_size, starting_bank, models=None):
    ws = wb.create_sheet("Resumen")
    active_models = models if models else MODELS

    row = 1
    _s(ws, row, 1, f"Reporte {', '.join(m.upper() for m in active_models)} — Odds {odds} | Apuesta ${bet_size:.0f} | Banco inicial ${starting_bank:.0f}",
       fill=_BLUE, font=_font(bold=True, white=True, size=11), align=_CENTER)
    ws.merge_cells(f"A{row}:L{row}")
    row += 2

    be = round(1 / odds * 100, 1)

    # Header
    headers = ["Modelo × Cuarto", "BETs", "Wins", "Losses", "Win%", "P&L", "ROI%",
               f"BE ({be}%)", "Banco Inicial", "Banco Final", "Neto (+/-)"]
    for c, h in enumerate(headers, 1):
        _s(ws, row, c, h, fill=_LBLUE, font=_font(bold=True), align=_CENTER)
    row += 1

    for model in active_models:
        for q in QUARTERS:
            key = f"{model}_{q}"
            bet_outcomes = [rec[f"{key}_outcome"] for rec in records if _is_bet(rec.get(f"{key}_signal", ""))]
            sim = _simulate(bet_outcomes, odds, bet_size, starting_bank)

            label = f"{model.upper()} {q.upper()}"
            roi_fill = _GREEN if sim["roi"] >= 0 else _RED
            be_ok = "✅" if sim["win_pct"] >= be else "❌"
            neto_fill = _GREEN if sim["pnl"] >= 0 else _RED

            col = 1
            _s(ws, row, col, label, align=_LEFT); col += 1
            _s(ws, row, col, sim["bets"], align=_CENTER); col += 1
            _s(ws, row, col, sim["wins"], align=_CENTER); col += 1
            _s(ws, row, col, sim["losses"], align=_CENTER); col += 1
            _s(ws, row, col, f"{sim['win_pct']:.1f}%", align=_CENTER); col += 1
            _s(ws, row, col, f"${sim['pnl']:.2f}", fill=roi_fill, align=_CENTER); col += 1
            _s(ws, row, col, f"{sim['roi']:.1f}%", fill=roi_fill, align=_CENTER); col += 1
            _s(ws, row, col, be_ok, align=_CENTER); col += 1
            _s(ws, row, col, f"${starting_bank:.0f}", align=_CENTER); col += 1
            bf = sim["bank_final"]
            _s(ws, row, col, f"${bf:.2f}" if bf is not None else "-", fill=neto_fill, align=_CENTER); col += 1
            neto = sim["pnl"]
            _s(ws, row, col, f"${neto:+.2f}", fill=neto_fill, font=_font(bold=True), align=_CENTER); col += 1
            row += 1

    row += 1
    # Combined per model (Q3+Q4 together)
    _s(ws, row, 1, "COMBINADO (Q3+Q4)", fill=_LBLUE, font=_font(bold=True), align=_CENTER)
    ws.merge_cells(f"A{row}:K{row}")
    row += 1

    for model in active_models:
        bet_outcomes = []
        for q in QUARTERS:
            key = f"{model}_{q}"
            bet_outcomes += [rec[f"{key}_outcome"] for rec in records if _is_bet(rec.get(f"{key}_signal", ""))]
        sim = _simulate(bet_outcomes, odds, bet_size, starting_bank)
        roi_fill = _GREEN if sim["roi"] >= 0 else _RED
        be_ok = "✅" if sim["win_pct"] >= (1 / odds * 100) else "❌"
        neto_fill = _GREEN if sim["pnl"] >= 0 else _RED

        col = 1
        _s(ws, row, col, f"{model.upper()} (total)", font=_font(bold=True), align=_LEFT); col += 1
        _s(ws, row, col, sim["bets"], align=_CENTER); col += 1
        _s(ws, row, col, sim["wins"], align=_CENTER); col += 1
        _s(ws, row, col, sim["losses"], align=_CENTER); col += 1
        _s(ws, row, col, f"{sim['win_pct']:.1f}%", align=_CENTER); col += 1
        _s(ws, row, col, f"${sim['pnl']:.2f}", fill=roi_fill, align=_CENTER); col += 1
        _s(ws, row, col, f"{sim['roi']:.1f}%", fill=roi_fill, align=_CENTER); col += 1
        _s(ws, row, col, be_ok, align=_CENTER); col += 1
        _s(ws, row, col, f"${starting_bank:.0f}", align=_CENTER); col += 1
        bf = sim["bank_final"]
        _s(ws, row, col, f"${bf:.2f}" if bf is not None else "-", fill=neto_fill, align=_CENTER); col += 1
        _s(ws, row, col, f"${sim['pnl']:+.2f}", fill=neto_fill, font=_font(bold=True), align=_CENTER); col += 1
        row += 1

    row += 1
    _s(ws, row, 1, "PENDING = partido sin resultado final. No se cuenta en P&L.",
       font=_font(size=8), align=_LEFT)

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 35)


# ── Sheet: Por Día ────────────────────────────────────────────────────────────

def _sheet_daily(wb, records, odds, bet_size, starting_bank, models=None):
    ws = wb.create_sheet("Por Dia")
    active_models = models if models else MODELS

    all_dates = sorted(set(r["date"] for r in records))

    header = ["Fecha", "Matches"]
    for model in active_models:
        header += [f"{model.upper()} BETs", f"{model.upper()} W", f"{model.upper()} L",
                   f"{model.upper()} P&L", f"{model.upper()} Banco"]
    for c, h in enumerate(header, 1):
        _s(ws, 1, c, h, fill=_BLUE, font=_font(bold=True, white=True), align=_CENTER)

    # Running bank per model
    running_bank = {m: starting_bank for m in active_models}

    for row_idx, d in enumerate(all_dates, 2):
        day_recs = [r for r in records if r["date"] == d]
        col = 1
        _s(ws, row_idx, col, d, align=_CENTER); col += 1
        _s(ws, row_idx, col, len(day_recs), align=_CENTER); col += 1

        for model in active_models:
            bet_outcomes = []
            for q in QUARTERS:
                key = f"{model}_{q}"
                bet_outcomes += [r[f"{key}_outcome"] for r in day_recs if _is_bet(r.get(f"{key}_signal", ""))]
            sim = _simulate(bet_outcomes, odds, bet_size)
            running_bank[model] += sim["pnl"]
            pnl_fill = _GREEN if (sim["bets"] > 0 and sim["pnl"] >= 0) else (_RED if sim["bets"] > 0 else None)
            bank_fill = _GREEN if running_bank[model] >= starting_bank else _RED
            _s(ws, row_idx, col, sim["bets"], align=_CENTER); col += 1
            _s(ws, row_idx, col, sim["wins"], align=_CENTER); col += 1
            _s(ws, row_idx, col, sim["losses"], align=_CENTER); col += 1
            _s(ws, row_idx, col, f"${sim['pnl']:+.2f}" if sim["bets"] > 0 else "", fill=pnl_fill, align=_CENTER); col += 1
            _s(ws, row_idx, col, f"${running_bank[model]:.2f}", fill=bank_fill, font=_font(bold=True), align=_CENTER); col += 1

    # Totals row
    total_row = len(all_dates) + 2
    _s(ws, total_row, 1, "TOTAL", fill=_LBLUE, font=_font(bold=True), align=_CENTER)
    _s(ws, total_row, 2, len(records), align=_CENTER)
    col = 3
    for model in active_models:
        all_outcomes = []
        for q in QUARTERS:
            key = f"{model}_{q}"
            all_outcomes += [r[f"{key}_outcome"] for r in records if _is_bet(r.get(f"{key}_signal", ""))]
        sim = _simulate(all_outcomes, odds, bet_size, starting_bank)
        neto = running_bank[model] - starting_bank
        neto_fill = _GREEN if neto >= 0 else _RED
        _s(ws, total_row, col, sim["bets"], fill=_LBLUE, font=_font(bold=True), align=_CENTER); col += 1
        _s(ws, total_row, col, sim["wins"], fill=_LBLUE, font=_font(bold=True), align=_CENTER); col += 1
        _s(ws, total_row, col, sim["losses"], fill=_LBLUE, font=_font(bold=True), align=_CENTER); col += 1
        _s(ws, total_row, col, f"${sim['pnl']:+.2f}", fill=neto_fill, font=_font(bold=True), align=_CENTER); col += 1
        _s(ws, total_row, col, f"${running_bank[model]:.2f}", fill=neto_fill, font=_font(bold=True), align=_CENTER); col += 1

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 25)
    ws.freeze_panes = "A2"


# ── Sheet: Ligas ──────────────────────────────────────────────────────────────

def _sheet_leagues(wb, records, odds, bet_size, models=None):
    ws = wb.create_sheet("Ligas")
    active_models = models if models else MODELS

    # Compute per-league stats per model
    leagues = sorted(set(r["league"] for r in records if r.get("league")))

    header = ["Liga"]
    for model in active_models:
        header += [f"{model.upper()} BETs", f"{model.upper()} W", f"{model.upper()} L", f"{model.upper()} ROI%"]
    for c, h in enumerate(header, 1):
        _s(ws, 1, c, h, fill=_BLUE, font=_font(bold=True, white=True), align=_CENTER)

    # Sort by V13 total bets descending
    def _league_bets(league, model):
        outcomes = []
        for q in QUARTERS:
            key = f"{model}_{q}"
            outcomes += [r[f"{key}_outcome"] for r in records if r.get("league") == league and _is_bet(r.get(f"{key}_signal", ""))]
        return outcomes

    league_data = []
    for lg in leagues:
        all_outcomes = []
        model_outcomes = {}
        for model in active_models:
            oc = _league_bets(lg, model)
            model_outcomes[model] = oc
            all_outcomes += oc
        if not all_outcomes:
            continue
        league_data.append((lg, model_outcomes))

    # Sort by total bets descending
    league_data.sort(key=lambda x: -sum(len(v) for v in x[1].values()))

    for row_idx, (lg, model_outcomes) in enumerate(league_data, 2):
        col = 1
        _s(ws, row_idx, col, lg, align=_LEFT); col += 1

        for model in active_models:
            outcomes = model_outcomes[model]
            sim = _simulate(outcomes, odds, bet_size)
            roi_fill = _GREEN if (sim["bets"] > 0 and sim["roi"] >= 0) else (_RED if sim["bets"] > 0 else None)
            _s(ws, row_idx, col, sim["bets"], align=_CENTER); col += 1
            _s(ws, row_idx, col, sim["wins"], align=_CENTER); col += 1
            _s(ws, row_idx, col, sim["losses"], align=_CENTER); col += 1
            _s(ws, row_idx, col, f"{sim['roi']:.1f}%" if sim["bets"] > 0 else "", fill=roi_fill, align=_CENTER); col += 1

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)
    ws.freeze_panes = "A2"


# ── CLI ───────────────────────────────────────────────────────────────────────

def _parse_args():
    parser = argparse.ArgumentParser(description="Reporte comparativo V12 vs V13")
    grp = parser.add_mutually_exclusive_group()
    grp.add_argument("--month", help="Mes a analizar (YYYY-MM)")
    grp.add_argument("--from", dest="date_from", help="Fecha inicio (YYYY-MM-DD)")
    parser.add_argument("--to", dest="date_to", help="Fecha fin (YYYY-MM-DD)")
    parser.add_argument("--models", default="v13", help="Modelos a incluir, separados por coma (default: v13)")
    parser.add_argument("--odds", type=float, default=1.4, help="Odds (default: 1.4)")
    parser.add_argument("--bet-size", type=float, default=100.0, help="Apuesta por partido (default: 100)")
    parser.add_argument("--bank", type=float, default=1000.0, help="Banco inicial (default: 1000)")
    parser.add_argument("--out", help="Ruta del Excel (opcional)")
    return parser.parse_args()


def main():
    args = _parse_args()

    if args.month:
        y, m = map(int, args.month.split("-"))
        date_from = f"{y:04d}-{m:02d}-01"
        last_day = (date(y, m % 12 + 1, 1) - timedelta(days=1)) if m < 12 else date(y, 12, 31)
        date_to = last_day.isoformat()
    elif args.date_from:
        date_from = args.date_from
        date_to = args.date_to or date.today().isoformat()
    else:
        # Default: April 2026 (V13 calibration period)
        date_from = "2026-04-01"
        date_to = date.today().isoformat()

    if args.out:
        out_path = Path(args.out)
    else:
        ts = datetime.now().strftime("%H%M%S")
        models_tag = args.models.replace(",", "_")
        tag = f"{date_from}_{date_to}".replace("-", "")
        out_path = REPORTS_DIR / f"report_{models_tag}_{tag}_{ts}.xlsx"

    generate_report(
        date_from=date_from,
        date_to=date_to,
        odds=args.odds,
        bet_size=args.bet_size,
        starting_bank=args.bank,
        out_path=out_path,
        models=[m.strip() for m in args.models.split(",")],
    )


if __name__ == "__main__":
    main()
