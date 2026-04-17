"""
Genera un Excel comparativo de modelos para un mes dado.

Columnas por partido:
  Fecha | Partido | Marcador Q3 | Q4 |
  [V2/V4/V6/V9] Q3 apuesta / resultado | Q4 apuesta / resultado |
  [V10] Q3 prediccion / resultado | Q4 prediccion / resultado

Hoja 'Partidos': matriz partido x modelo con celdas verdes/rojas.
Hoja 'Resumen':  ganados, perdidos, push (=perdido), bank simulation.

Bank: inicio 100 apuestas de 20 = 2000. Odds = 1.91.
  WIN  → +20 * 0.91 = +18.20
  LOSS → -20
  PUSH → -20 (cuenta como perdido)

Uso:
    python training/report_model_comparison.py --month 2026-03
    python training/report_model_comparison.py            # pide mes interactivo
    python training/report_model_comparison.py --month 2026-03 --out /ruta/archivo.xlsx
"""

from __future__ import annotations

import argparse
import importlib
import sys
import warnings
from pathlib import Path
from typing import Any

# Suppress noisy sklearn parallelism warning that floods the console
warnings.filterwarnings(
    "ignore",
    message=r"`sklearn\.utils\.parallel\.delayed` should be used with",
    category=UserWarning,
)

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None  # type: ignore[assignment]

db_mod = importlib.import_module("db")

TRAINING_DIR = ROOT / "training"
DB_PATH = ROOT / "matches.db"
REPORTS_DIR = ROOT / "reports"

ODDS = 1.40
BET_SIZE = 20.0
STARTING_BANK = 100 * BET_SIZE  # 2000

MODELS_CLASS = ["v2", "v4", "v6", "v9"]
QUARTERS = ["q3", "q4"]

# ── Excel styles ──────────────────────────────────────────────────────────────
_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") if openpyxl else None
_RED   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") if openpyxl else None
_PUSH  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") if openpyxl else None
_HDR   = PatternFill(start_color="2F75B6", end_color="2F75B6", fill_type="solid") if openpyxl else None
_SUB   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid") if openpyxl else None
_GRAY  = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") if openpyxl else None

_WHITE   = Font(color="FFFFFF", bold=True, size=9) if openpyxl else None
_BOLD    = Font(bold=True, size=9) if openpyxl else None
_NORMAL  = Font(size=9) if openpyxl else None
_CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True) if openpyxl else None
_LEFT    = Alignment(horizontal="left", vertical="center") if openpyxl else None


# ── Inference helpers ─────────────────────────────────────────────────────────

def _run_infer(infer_mod: Any, match_id: str, version: str) -> dict:
    """Run classification inference for a single version. Returns predictions dict."""
    try:
        result = infer_mod.run_inference(
            match_id=match_id,
            metric="f1",
            fetch_missing=False,
            force_version={"q3": version, "q4": version},
        )
        if result.get("ok"):
            return result.get("predictions", {})
    except Exception:
        pass
    return {}


def _run_infer_v10(infer_mod: Any, match_id: str) -> dict:
    """Run V10 Over/Under inference. Returns predictions dict."""
    try:
        result = infer_mod.run_inference_v10(match_id=match_id)
        if result.get("ok"):
            return result.get("predictions", {})
    except Exception:
        pass
    return {}


# ── Result computation ────────────────────────────────────────────────────────

def _classify_outcome(pred_q: dict) -> tuple[str, str]:
    """(bet_label, outcome) for classification model on one quarter.

    Returns:
        bet_label: 'APOSTAR 🏠' | 'APOSTAR 🛫' | 'NO APOSTAR' | 'N/A'
        outcome:   'WIN' | 'LOSS' | 'PUSH' | '-'
    """
    if not pred_q or not pred_q.get("available"):
        return "N/A", "-"

    rec = str(pred_q.get("final_recommendation") or pred_q.get("bet_signal") or "")
    if rec != "BET":
        return "NO APOSTAR", "-"

    winner = str(pred_q.get("predicted_winner") or "")
    icon = "🏠" if winner == "home" else "🛫"
    bet_label = f"APOSTAR {icon}"

    result = str(pred_q.get("result") or "pending")
    if result == "hit":
        outcome = "WIN"
    elif result == "miss":
        outcome = "LOSS"
    elif result == "push":
        outcome = "PUSH"
    else:
        outcome = "-"

    return bet_label, outcome


def _v10_outcome(pred_q: dict, actual_total: int | None) -> tuple[str, str]:
    """(bet_label, outcome) for V10 O/U regression.

    Strategy: always bet OVER predicted_total.
    WIN  if actual > predicted_total
    LOSS if actual < predicted_total
    PUSH if equal
    """
    if not pred_q or not pred_q.get("available"):
        return "N/A", "-"

    predicted = pred_q.get("predicted_total")
    if predicted is None:
        return "N/A", "-"

    bet_label = f"O ~{predicted:.1f}"

    if actual_total is None:
        return bet_label, "-"

    if actual_total > predicted:
        return bet_label, "WIN"
    elif actual_total < predicted:
        return bet_label, "LOSS"
    else:
        return bet_label, "PUSH"


def _update_sim(sim: dict, key: str, outcome: str) -> None:
    """Update bank simulation for a model+quarter key with a result."""
    if outcome not in ("WIN", "LOSS", "PUSH"):
        return
    sim[key]["bets"] += 1
    if outcome == "WIN":
        sim[key]["wins"] += 1
        sim[key]["bank"] += BET_SIZE * (ODDS - 1)
    else:
        if outcome == "PUSH":
            sim[key]["pushes"] += 1
        else:
            sim[key]["losses"] += 1
        sim[key]["bank"] -= BET_SIZE


# ── Excel builders ────────────────────────────────────────────────────────────

def _h(cell: Any, value: Any, fill=None, font=None, align=None) -> None:
    """Helper: set cell value + optional style."""
    cell.value = value
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    else:
        cell.alignment = _CENTER


def _fill_for_outcome(outcome: str) -> Any:
    if outcome == "WIN":
        return _GREEN
    if outcome == "LOSS":
        return _RED
    if outcome == "PUSH":
        return _PUSH
    return _GRAY


def _build_matches_sheet(wb: Any, records: list[dict]) -> None:
    ws = wb.active
    ws.title = "Partidos"

    all_models = MODELS_CLASS + ["v10"]
    # Two header rows:
    # Row 1: base info headers (merged 6 cols) | model group headers (4 cols each)
    # Row 2: sub-headers per model column

    base_cols = ["Fecha", "ID", "Local", "Visitante", "Q3", "Q4"]
    n_base = len(base_cols)

    # model columns per model: Q3 Apuesta | Q3 Res | Q4 Apuesta | Q4 Res
    model_col_labels = ["Q3 Apuesta", "Q3 Res", "Q4 Apuesta", "Q4 Res"]
    n_model_cols = len(model_col_labels)

    # ── Row 1: base headers + model group headers ──────────────────────────
    for i, label in enumerate(base_cols, start=1):
        cell = ws.cell(row=1, column=i)
        _h(cell, label, _HDR, _WHITE)
        ws.merge_cells(start_row=1, end_row=2, start_column=i, end_column=i)

    col_cursor = n_base + 1
    model_start_cols: dict[str, int] = {}
    for model in all_models:
        model_start_cols[model] = col_cursor
        end_col = col_cursor + n_model_cols - 1
        ws.merge_cells(start_row=1, end_row=1, start_column=col_cursor, end_column=end_col)
        cell = ws.cell(row=1, column=col_cursor)
        _h(cell, model.upper(), _HDR, _WHITE)
        # Sub-headers row 2
        for j, sub in enumerate(model_col_labels):
            sub_cell = ws.cell(row=2, column=col_cursor + j)
            _h(sub_cell, sub, _SUB, _BOLD)
        col_cursor += n_model_cols

    # ── Data rows ──────────────────────────────────────────────────────────
    for r, rec in enumerate(records, start=3):
        ws.cell(row=r, column=1, value=rec["date"]).alignment = _CENTER
        ws.cell(row=r, column=2, value=rec["match_id"]).alignment = _CENTER
        ws.cell(row=r, column=3, value=rec["home"]).alignment = _LEFT
        ws.cell(row=r, column=4, value=rec["away"]).alignment = _LEFT
        ws.cell(row=r, column=5, value=rec["q3_score"]).alignment = _CENTER
        ws.cell(row=r, column=6, value=rec["q4_score"]).alignment = _CENTER

        for model in all_models:
            base = model_start_cols[model]
            for qi, q in enumerate(QUARTERS):
                bet_label = rec.get(f"{model}_{q}_bet", "N/A")
                outcome = rec.get(f"{model}_{q}_outcome", "-")
                fill = _fill_for_outcome(outcome)
                bet_cell = ws.cell(row=r, column=base + qi * 2, value=bet_label)
                bet_cell.alignment = _CENTER
                bet_cell.font = _NORMAL
                out_cell = ws.cell(row=r, column=base + qi * 2 + 1, value=outcome)
                out_cell.alignment = _CENTER
                out_cell.font = _BOLD if outcome in ("WIN", "LOSS", "PUSH") else _NORMAL
                out_cell.fill = fill

    # ── Column widths ──────────────────────────────────────────────────────
    widths = [11, 12, 22, 22, 8, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for model in all_models:
        base = model_start_cols[model]
        for j in range(n_model_cols):
            ws.column_dimensions[get_column_letter(base + j)].width = 14

    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 25


def _build_summary_sheet(wb: Any, sim: dict, month: str) -> None:
    ws = wb.create_sheet(title="Resumen")

    headers = [
        "Modelo", "Quarter", "Apuestas", "Ganadas", "Perdidas", "Push",
        "Efectividad %", "Bank inicial", "Net P&L", "Bank final", "ROI %",
    ]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = _HDR
        cell.font = _WHITE
        cell.alignment = _CENTER

    row = 2
    all_models = MODELS_CLASS + ["v10"]
    for model in all_models:
        for q in QUARTERS:
            key = f"{model}_{q}"
            s = sim[key]
            wins = s["wins"]
            losses = s["losses"] + s["pushes"]  # push = loss
            pushes = s["pushes"]
            bets = s["bets"]
            eff = round(100 * wins / bets, 1) if bets else 0.0
            net = round(s["bank"] - STARTING_BANK, 2)
            roi = round(100 * net / (bets * BET_SIZE), 1) if bets else 0.0

            row_data = [
                model.upper(), q.upper(), bets, wins,
                losses, pushes, f"{eff}%",
                STARTING_BANK, net, round(s["bank"], 2), f"{roi}%",
            ]
            for ci, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.alignment = _CENTER
                cell.font = _NORMAL
                if ci in (4, 7):  # Ganadas, Efectividad → green tint if positive
                    if isinstance(val, (int, float)) and val > 0:
                        cell.fill = _GREEN
                if ci == 9:  # Net P&L
                    if isinstance(val, (int, float)):
                        cell.fill = _GREEN if val > 0 else _RED

            row += 1

    # Column widths
    col_widths = [10, 10, 10, 10, 10, 8, 14, 14, 12, 12, 10]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title note
    ws.cell(row=row + 1, column=1, value=f"Reporte mes: {month}").font = _BOLD
    ws.cell(row=row + 2, column=1, value=f"Bank inicial: {STARTING_BANK:.0f} | Apuesta: {BET_SIZE:.0f} | Odds: {ODDS}").font = _NORMAL
    ws.cell(row=row + 3, column=1, value="Push = perdido. Solo se cuentan apuestas con señal BET.").font = _NORMAL

    ws.freeze_panes = "A2"


def _build_daily_summary_sheet(wb: Any, records: list[dict]) -> None:
    """Hoja 'Por Dia': una fila por fecha con totales de apuestas/resultados de todos los modelos."""
    from collections import OrderedDict
    ws = wb.create_sheet(title="Por Dia")

    all_models = MODELS_CLASS + ["v10"]

    # Accumulate per date
    days: dict[str, dict] = OrderedDict()
    for rec in records:
        date = rec["date"]
        if date not in days:
            days[date] = {"matches": 0, "bets": 0, "wins": 0, "losses": 0, "pushes": 0, "net": 0.0}
        d = days[date]
        d["matches"] += 1
        for model in all_models:
            for q in QUARTERS:
                outcome = rec.get(f"{model}_{q}_outcome", "-")
                if outcome in ("WIN", "LOSS", "PUSH"):
                    d["bets"] += 1
                    if outcome == "WIN":
                        d["wins"] += 1
                        d["net"] += BET_SIZE * (ODDS - 1)
                    else:
                        if outcome == "PUSH":
                            d["pushes"] += 1
                        else:
                            d["losses"] += 1
                        d["net"] -= BET_SIZE

    headers = ["Fecha", "Partidos", "Apuestas", "Ganadas", "Perdidas", "Push", "Efectividad %", "Net P&L"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = _HDR
        cell.font = _WHITE
        cell.alignment = _CENTER

    for r, (date, d) in enumerate(days.items(), start=2):
        bets = d["bets"]
        wins = d["wins"]
        net = round(d["net"], 2)
        eff = round(100 * wins / bets, 1) if bets else 0.0
        row_data = [
            date, d["matches"], bets, wins,
            d["losses"], d["pushes"], f"{eff}%", net,
        ]
        for ci, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.alignment = _CENTER
            cell.font = _NORMAL
            if ci == 8:  # Net P&L
                if isinstance(val, (int, float)):
                    cell.fill = _GREEN if val > 0 else (_RED if val < 0 else _GRAY)

    for i, w in enumerate([12, 10, 10, 10, 10, 8, 14, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"


# ── Main function ─────────────────────────────────────────────────────────────

def generate_report(
    month: str,
    progress_state: dict | None = None,
    out_path: Path | None = None,
) -> Path:
    """Generate model comparison Excel for *month* (YYYY-MM).

    Args:
        month:          Month string like '2026-03'.
        progress_state: Optional dict updated in-place for external progress display.
                        Keys: total, processed, phase, current, output_path.
        out_path:       Override output file path. Defaults to reports/model_comparison_YYYY-MM.xlsx.

    Returns:
        Path to the generated .xlsx file.
    """
    if openpyxl is None:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")

    def _ps(key: str, val: Any) -> None:
        if progress_state is not None:
            progress_state[key] = val

    conn = db_mod.get_conn(str(DB_PATH))
    db_mod.init_db(conn)

    cur = conn.execute(
        """
        SELECT event_date, match_id, home_team, away_team,
               q3_home_score, q3_away_score,
               q4_home_score, q4_away_score
        FROM eval_match_results
        WHERE event_date LIKE ?
          AND q3_home_score IS NOT NULL
          AND q4_home_score IS NOT NULL
        ORDER BY event_date, match_id
        """,
        (f"{month}-%",),
    )
    rows = cur.fetchall()
    conn.close()

    total = len(rows)
    _ps("total", total)
    _ps("processed", 0)
    _ps("phase", "loading")
    _ps("current", "")

    if total == 0:
        raise ValueError(f"No hay partidos completos en eval_match_results para {month}")

    infer_mod = importlib.import_module("training.infer_match")
    infer_mod.scraper_mod.fetch_event_snapshot = lambda _mid: None

    _ps("phase", "processing")

    # Simulation accumulators
    all_models = MODELS_CLASS + ["v10"]
    sim: dict[str, dict] = {
        f"{m}_{q}": {"wins": 0, "losses": 0, "pushes": 0, "bank": STARTING_BANK, "bets": 0}
        for m in all_models
        for q in QUARTERS
    }

    records: list[dict[str, Any]] = []

    for i, row in enumerate(rows):
        event_date = str(row[0])
        match_id = str(row[1])
        home_team = str(row[2] or "")
        away_team = str(row[3] or "")
        q3h = row[4]
        q3a = row[5]
        q4h = row[6]
        q4a = row[7]
        q3_total = (q3h or 0) + (q3a or 0)
        q4_total = (q4h or 0) + (q4a or 0)

        _ps("processed", i)
        _ps("current", f"{home_team} vs {away_team} ({event_date})")

        rec: dict[str, Any] = {
            "date": event_date,
            "match_id": match_id,
            "home": home_team,
            "away": away_team,
            "q3_score": f"{q3h}-{q3a}" if q3h is not None else "-",
            "q4_score": f"{q4h}-{q4a}" if q4h is not None else "-",
        }

        # Classification models
        for version in MODELS_CLASS:
            preds = _run_infer(infer_mod, match_id, version)
            for q, actual_total in [("q3", q3_total), ("q4", q4_total)]:
                bet_label, outcome = _classify_outcome(preds.get(q, {}))
                rec[f"{version}_{q}_bet"] = bet_label
                rec[f"{version}_{q}_outcome"] = outcome
                _update_sim(sim, f"{version}_{q}", outcome)

        # V10
        preds_v10 = _run_infer_v10(infer_mod, match_id)
        for q, actual_total in [("q3", q3_total), ("q4", q4_total)]:
            bet_label, outcome = _v10_outcome(preds_v10.get(q, {}), actual_total)
            rec[f"v10_{q}_bet"] = bet_label
            rec[f"v10_{q}_outcome"] = outcome
            _update_sim(sim, f"v10_{q}", outcome)

        records.append(rec)

    _ps("processed", total)
    _ps("phase", "writing")

    wb = openpyxl.Workbook()
    _build_matches_sheet(wb, records)
    _build_summary_sheet(wb, sim, month)
    _build_daily_summary_sheet(wb, records)

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    if out_path is None:
        out_path = REPORTS_DIR / f"model_comparison_{month}.xlsx"
    wb.save(str(out_path))

    _ps("phase", "done")
    _ps("output_path", str(out_path))

    return out_path


# ── CLI entry point ───────────────────────────────────────────────────────────

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--month", default=None, help="Mes a reportar (YYYY-MM). Si no se indica, pregunta interactivo.")
    p.add_argument("--out", default=None, help="Ruta de salida del .xlsx. Por defecto: reports/model_comparison_<mes>.xlsx")
    return p


def main() -> None:
    args = _build_parser().parse_args()
    month = args.month

    if not month:
        month = input("Mes (YYYY-MM): ").strip()

    # Validate format
    try:
        from datetime import datetime
        datetime.strptime(month, "%Y-%m")
    except ValueError:
        print(f"[report] Formato invalido: '{month}'. Usa YYYY-MM.")
        sys.exit(1)

    try:
        from tqdm import tqdm  # type: ignore[import]
        _use_tqdm = True
    except ImportError:
        _use_tqdm = False

    progress_state: dict[str, Any] = {
        "total": 0,
        "processed": 0,
        "phase": "starting",
        "current": "",
        "output_path": "",
    }

    if _use_tqdm:
        # Run in thread, poll progress
        import threading
        import time

        done = threading.Event()
        error_holder: list[Exception] = []
        result_holder: list[Path] = []

        def _run() -> None:
            try:
                out = generate_report(
                    month,
                    progress_state=progress_state,
                    out_path=Path(args.out) if args.out else None,
                )
                result_holder.append(out)
            except Exception as exc:
                error_holder.append(exc)
            finally:
                done.set()

        t = threading.Thread(target=_run, daemon=True)
        t.start()

        bar: Any = None
        prev_processed = -1
        while not done.is_set():
            total = progress_state.get("total", 0)
            processed = progress_state.get("processed", 0)
            phase = progress_state.get("phase", "")
            current = progress_state.get("current", "")

            if total > 0 and bar is None:
                bar = tqdm(total=total, desc=f"[report {month}]", unit="match")

            if bar and processed > prev_processed:
                bar.n = processed
                bar.set_postfix_str(f"{phase} | {current}"[:60])
                bar.refresh()
                prev_processed = processed

            time.sleep(0.5)

        if bar:
            bar.n = progress_state.get("total", 0)
            bar.set_postfix_str("done")
            bar.close()

        if error_holder:
            print(f"[report] Error: {error_holder[0]}")
            sys.exit(1)

        out_path = result_holder[0]
    else:
        # Simple print-based progress
        print(f"[report] Generando reporte {month}...")
        out_path = generate_report(
            month,
            out_path=Path(args.out) if args.out else None,
        )

    print(f"[report] Guardado: {out_path}")


if __name__ == "__main__":
    main()
