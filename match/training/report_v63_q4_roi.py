"""Generate Q4 ROI report comparing V6.2 vs V6.3 side-by-side.

Adds `minuto_base_apuesta` to match-by-match sheets so each bet records
which game minute was used as the model snapshot.
"""

from pathlib import Path
import argparse
import hashlib
import json
import os
import sqlite3
from collections import defaultdict
from datetime import datetime, timezone, timedelta

import joblib
import numpy as np
import pandas as pd
from tqdm import tqdm
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from sklearn.isotonic import IsotonicRegression
from sklearn.linear_model import LogisticRegression

import train_q3_q4_models_v6 as v6
import infer_match as infer_live
import train_q4_m27_v1 as m27_v1_train
import m27_v1_league_policy as m27_v1_policy
from v6_3_league_blacklist import get_blacklist as _get_v63_blacklist

ROOT = Path(__file__).parent.parent.parent  # c:\Users\App\Desktop\pulpa
BASE_V62 = ROOT / "match" / "training" / "model_outputs_v6_2"
BASE_V63 = ROOT / "match" / "training" / "model_outputs_v6_3"
BASE_M27_V1 = ROOT / "match" / "training" / "model_outputs_m27_v1"
OUT_MM = BASE_V63 / "Q4_ROI_match_by_match_v6_2_vs_v6_3.xlsx"
PRED_CACHE_PATH = BASE_V63 / "q4_roi_pred_cache.joblib"
SPLITS_CACHE_PATH = BASE_V63 / "q4_roi_splits_cache.joblib"
M27_V1_CHAMPION_PATH = BASE_M27_V1 / "q4_m27_v1_champion.joblib"
_V63_BL = _get_v63_blacklist()

ODDS = 1.4
BANK_START = 1000.0
BREAK_EVEN = 1.0 / ODDS

# Staking params for detailed match-by-match
MODE = "kelly_non_compound"
KELLY_MULT = 0.25
KELLY_CAP = 0.05
MIN_CONF_PROB = 0.58
STAKE_STEP = 25.0
MIN_STAKE = 25.0
MAX_STAKE = 100.0
STAKE_BUCKETS = (25.0, 50.0, 75.0, 100.0)

# Minute baseline used to build v6.3 snapshot features.
V63_SNAPSHOT_MINUTE = 27
# Monitor-like Q4 re-evaluation windows.
V63_MONITOR_SNAPSHOTS = (27, 30)
V63_COMPARE_SNAPSHOTS = (27, 30)
# Late-window bets only make sense when the game is still close.
V63_LATE_WINDOW_MAX_MARGIN = {
    27: 999,
    30: 6,
    33: 8,
    36: 10,
}
# v6.2 is based on the regular end-of-Q3 snapshot.
V62_BASE_MINUTE = 36

# v6.3 raw report runs fully unfiltered unless explicitly changed.
V63_DISABLE_FILTERS = True
V63_APPLY_MANUAL_BLACKLIST = False


def _report_date_from_dt(dt_value) -> str:
    if not dt_value:
        return ""
    dt_local = dt_value if dt_value.tzinfo else dt_value.replace(tzinfo=timezone.utc)
    return dt_local.astimezone(timezone(timedelta(hours=-6))).strftime("%Y-%m-%d")


def _rows_fingerprint(rows) -> str:
    h = hashlib.sha1()
    for s in rows:
        h.update(str(s.match_id).encode("utf-8", errors="ignore"))
        h.update(b"|")
        h.update((s.dt.isoformat() if s.dt else "").encode("utf-8", errors="ignore"))
        h.update(b"|")
        h.update(str(int(s.target_q4) if s.target_q4 is not None else -1).encode("utf-8", errors="ignore"))
        h.update(b"\n")
    return h.hexdigest()


def _file_signature(path: Path) -> str:
    if not path.exists():
        return "missing"
    st = path.stat()
    return f"{int(st.st_mtime_ns)}:{int(st.st_size)}"


def _load_pred_cache(cache_meta: dict, force_rebuild: bool = False) -> dict | None:
    if force_rebuild or not PRED_CACHE_PATH.exists():
        return None
    try:
        payload = joblib.load(PRED_CACHE_PATH)
    except Exception:
        return None
    if not isinstance(payload, dict):
        return None
    if payload.get("meta") != cache_meta:
        return None
    return payload.get("predictions")


def _save_pred_cache(cache_meta: dict, predictions: dict) -> None:
    PRED_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    joblib.dump({"meta": cache_meta, "predictions": predictions}, PRED_CACHE_PATH)




def _kelly_fraction(p, odds):
    b = odds - 1.0
    q = 1.0 - p
    return ((b * p) - q) / b if b > 0 else 0.0


def _round_down_step(x, step):
    if step <= 0:
        return x
    return (x // step) * step


def _stake_from_signal(p_pick, edge, k_used, kelly_cap=0.05):
    """Map signal quality to fixed stake buckets: 25/50/75/100."""
    if k_used is None or k_used <= 0:
        return 0.0

    denom = kelly_cap if kelly_cap and kelly_cap > 0 else 0.05
    k_strength = max(0.0, min(k_used / denom, 1.0))
    p_pick = float(p_pick or 0.0)
    edge = float(edge or 0.0)

    if p_pick >= 0.97 and edge >= 0.24 and k_strength >= 0.995:
        return STAKE_BUCKETS[3]
    if p_pick >= 0.90 and edge >= 0.18 and k_strength >= 0.85:
        return STAKE_BUCKETS[2]
    if p_pick >= 0.80 and edge >= 0.09 and k_strength >= 0.55:
        return STAKE_BUCKETS[1]
    if k_strength < 0.15:
        return 0.0
    return STAKE_BUCKETS[0]


_FAST_COMPLETE_Q4_SQL = """
    SELECT m.date, m.time, m.match_id
    FROM matches m
    WHERE EXISTS (
            SELECT 1 FROM quarter_scores
            WHERE match_id = m.match_id AND quarter = 'Q4'
        )
      AND EXISTS (
            SELECT 1 FROM play_by_play
            WHERE match_id = m.match_id AND quarter = 'Q1'
        )
      AND EXISTS (
            SELECT 1 FROM play_by_play
            WHERE match_id = m.match_id AND quarter = 'Q4'
        )
      AND (
            SELECT COUNT(*) FROM graph_points
            WHERE match_id = m.match_id
        ) >= 20
    ORDER BY m.date, m.time, m.match_id
"""


def _prepare_splits(force_rebuild: bool = False, only_date: str | None = None):
    """Load Q4 rows and return temporal 70/15/15 splits (fast, cached).

    Strategy:
      1. Fast SQL probe (no get_match) determines exact split boundaries.
      2. get_match + full features run only for the val+test window (~30 %).
      3. Results are serialised to SPLITS_CACHE_PATH keyed by DB file
         signature (mtime + size).  Subsequent runs load from cache instantly.
    """
    if only_date:
        print(f"[Q4_ROI_62_63] cargando solo partidos no vistos de la fecha {only_date}...")
        samples = v6._build_samples(v6.DB_PATH, date_gte=only_date)
        rows = sorted(
            [s for s in samples if s.target_q4 is not None and _report_date_from_dt(s.dt) == only_date],
            key=lambda s: s.dt,
        )
        if not rows:
            raise ValueError(f"No se encontraron partidos no vistos para la fecha {only_date}")
        print(
            "[Q4_ROI_62_63] modo fecha específica "
            f"fecha={only_date} test={len(rows)}"
        )
        return [], [], rows

    db_sig = _file_signature(v6.DB_PATH)
    cache_key = {"db_sig": db_sig, "split_ratios": (0.70, 0.15, 0.15)}

    if not force_rebuild and SPLITS_CACHE_PATH.exists():
        try:
            payload = joblib.load(SPLITS_CACHE_PATH)
            if isinstance(payload, dict) and payload.get("meta") == cache_key:
                val_rows = payload["val_rows"]
                test_rows = payload["test_rows"]
                print(
                    f"[Q4_ROI_62_63] splits desde cache ({SPLITS_CACHE_PATH.name}) "
                    f"val={len(val_rows)} test={len(test_rows)}"
                )
                return [], val_rows, test_rows
        except Exception:
            pass

    # --- Phase 1: cheap SQL probe to find exact val/test date boundary -------
    print("[Q4_ROI_62_63] sondeo rápido de fechas de corte Q4 (sin cargar partidos)...")
    _probe_conn = sqlite3.connect(str(v6.DB_PATH))
    _probe_conn.row_factory = sqlite3.Row
    _probe_rows = _probe_conn.execute(_FAST_COMPLETE_Q4_SQL).fetchall()
    _probe_conn.close()

    n_approx = len(_probe_rows)
    n_train_approx = int(n_approx * 0.70)
    n_val_approx = int(n_approx * 0.15)
    val_start_date: str = str(_probe_rows[n_train_approx]["date"])
    print(
        f"[Q4_ROI_62_63] n_approx={n_approx}  "
        f"val_start_date={val_start_date}  "
        f"val≈{n_val_approx}  test≈{n_approx - n_train_approx - n_val_approx}"
    )

    # --- Phase 2: load only val+test window via date filter ------------------
    samples = v6._build_samples(v6.DB_PATH, date_gte=val_start_date)
    rows = sorted([s for s in samples if s.target_q4 is not None], key=lambda s: s.dt)

    val_rows = rows[:n_val_approx]
    test_rows = rows[n_val_approx:]

    print(
        "[Q4_ROI_62_63] split_v6_3 "
        f"q4_total≈{n_approx} val={len(val_rows)} test={len(test_rows)}"
    )

    # --- Phase 3: persist cache so next run is instant -----------------------
    SPLITS_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    joblib.dump({"meta": cache_key, "val_rows": val_rows, "test_rows": test_rows}, SPLITS_CACHE_PATH)
    print(f"[Q4_ROI_62_63] splits guardados en cache: {SPLITS_CACHE_PATH.name}")

    return [], val_rows, test_rows


def _load_match_teams_map():
    conn = sqlite3.connect(v6.DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            "SELECT match_id, home_team, away_team FROM matches"
        ).fetchall()
    finally:
        conn.close()

    out = {}
    for r in rows:
        out[str(r["match_id"])] = {
            "home_team": str(r["home_team"] or ""),
            "away_team": str(r["away_team"] or ""),
        }
    return out


def _load_match_q4_scores_map(match_ids):
    conn = v6.db_mod.get_conn(str(v6.DB_PATH))
    v6.db_mod.init_db(conn)
    out = {}
    try:
        for match_id in match_ids:
            data = v6.db_mod.get_match(conn, str(match_id))
            if not data:
                out[str(match_id)] = ""
                continue
            q4h, q4a = v6._quarter_points(data, "Q4")
            if q4h is None or q4a is None:
                out[str(match_id)] = ""
            else:
                out[str(match_id)] = f"{int(q4h)}-{int(q4a)}"
    finally:
        conn.close()
    return out


def _load_match_score_at_minute_map(match_ids, minute):
    conn = v6.db_mod.get_conn(str(v6.DB_PATH))
    v6.db_mod.init_db(conn)
    out = {}
    cutoff = float(minute)
    try:
        for match_id in match_ids:
            data = v6.db_mod.get_match(conn, str(match_id))
            if not data:
                out[str(match_id)] = ""
                continue
            pbp = data.get("play_by_play", {}) if isinstance(data, dict) else {}
            events = v6._pbp_events_upto_minute(pbp, cutoff)
            home = 0
            away = 0
            for e in events:
                team = str(e.get("team", ""))
                pts = int(e.get("points", 0) or 0)
                if team == "home":
                    home += pts
                elif team == "away":
                    away += pts
            out[str(match_id)] = f"{home}-{away}"
    finally:
        conn.close()
    return out


def _load_league_name_rules():
    cfg = json.loads((ROOT / "match" / "training" / "v6_2_league_name_exclusions.json").read_text(encoding="utf-8"))
    pats = []
    for cat in cfg.get("categories", []):
        for p in cat.get("patterns", []):
            p = str(p).strip()
            if p:
                pats.append((cat.get("name", "uncategorized"), p, p.lower()))
    return pats


def _build_v63_window_features(sample, snapshot_minute):
    rec = dict(sample.features_q4)
    rec["snapshot_minute"] = int(snapshot_minute)
    rec["snapshot_minutes_before_q4"] = 36 - int(snapshot_minute)
    rec["snapshot_q3_completeness"] = round(
        min(1.0, max(0.0, (float(snapshot_minute) - 24.0) / 12.0)), 4
    )
    return rec


def _clip_pbp_to_minute(match_data: dict, cutoff_minute: int) -> dict:
    pbp = (match_data.get("play_by_play") or {})
    out = {"Q1": [], "Q2": [], "Q3": []}
    cutoff = float(cutoff_minute)
    for quarter_label, plays in pbp.items():
        q_idx = infer_live._quarter_index(str(quarter_label))
        if q_idx is None or q_idx < 1 or q_idx > 3:
            continue
        q_start = (q_idx - 1) * 12.0
        kept = []
        for play in plays or []:
            rem_sec = infer_live._clock_to_seconds(str(play.get("time", "") or ""))
            if rem_sec is None:
                continue
            global_min = q_start + (12.0 - rem_sec / 60.0)
            if global_min <= cutoff + 1e-9:
                kept.append(play)
        out[f"Q{q_idx}"] = kept
    return out


def _build_live_like_q4_snapshot(match_data: dict, snapshot_minute: int) -> dict:
    """Reconstruct historical match data as if inference happened at snapshot_minute."""
    clipped = dict(match_data)
    clipped["graph_points"] = [
        p for p in (match_data.get("graph_points") or [])
        if int(p.get("minute", 0)) <= int(snapshot_minute)
    ]
    clipped["play_by_play"] = _clip_pbp_to_minute(match_data, snapshot_minute)

    score_obj = dict((match_data.get("score") or {}))
    quarters = dict(score_obj.get("quarters") or {})
    q1h, q1a = infer_live._quarter_points(match_data, "Q1")
    q2h, q2a = infer_live._quarter_points(match_data, "Q2")
    ht_home = int(q1h or 0) + int(q2h or 0)
    ht_away = int(q1a or 0) + int(q2a or 0)

    game_home, game_away = infer_live._score_upto(clipped, int(snapshot_minute))
    q3_home = max(0, int(game_home) - ht_home)
    q3_away = max(0, int(game_away) - ht_away)
    quarters["Q3"] = {"home": q3_home, "away": q3_away}

    score_obj["quarters"] = quarters
    clipped["score"] = score_obj
    return clipped


def _v63_bet_candidate(p_home: float | None) -> dict:
    if p_home is None:
        return {"accept": False, "reason": "missing_prob"}

    p_pick = p_home if p_home >= 0.5 else (1.0 - p_home)
    pick_home = p_home >= 0.5
    pick_side = "local" if pick_home else "visitante"
    edge = p_pick - BREAK_EVEN

    if p_pick < MIN_CONF_PROB:
        return {
            "accept": False,
            "reason": "confidence_filter",
            "p_pick": p_pick,
            "pick_home": pick_home,
            "pick_side": pick_side,
            "edge": edge,
        }

    k_raw = _kelly_fraction(p_pick, ODDS)
    if k_raw <= 0:
        return {
            "accept": False,
            "reason": "no_edge",
            "p_pick": p_pick,
            "pick_home": pick_home,
            "pick_side": pick_side,
            "edge": edge,
            "kelly_fraction_raw": k_raw,
        }

    k_used = min(k_raw * KELLY_MULT, KELLY_CAP)
    stake = _stake_from_signal(p_pick, edge, k_used, KELLY_CAP)
    stake = min(stake, MAX_STAKE)
    stake = _round_down_step(stake, STAKE_STEP)
    if stake < MIN_STAKE:
        return {
            "accept": False,
            "reason": "stake_below_25",
            "p_pick": p_pick,
            "pick_home": pick_home,
            "pick_side": pick_side,
            "edge": edge,
            "kelly_fraction_raw": k_raw,
            "kelly_fraction_used": k_used,
            "stake": stake,
        }

    return {
        "accept": True,
        "reason": "accepted",
        "p_pick": p_pick,
        "pick_home": pick_home,
        "pick_side": pick_side,
        "edge": edge,
        "kelly_fraction_raw": k_raw,
        "kelly_fraction_used": k_used,
        "stake": stake,
    }


def _v63_window_eligibility(match_data: dict, snapshot_minute: int, p_home: float, gate: dict) -> tuple[bool, str, int]:
    final_rec = str(gate.get("final_recommendation", "NO BET")).upper()
    if final_rec == "NO BET":
        return False, f"gate:{gate.get('reason', 'no_bet')}", 0

    game_home, game_away = infer_live._score_upto(match_data, int(snapshot_minute))
    margin = abs(int(game_home) - int(game_away))
    if int(snapshot_minute) > 27:
        max_margin = int(V63_LATE_WINDOW_MAX_MARGIN.get(int(snapshot_minute), 6))
        if margin > max_margin:
            return False, f"late_window_close_game_required:{margin}>{max_margin}", margin

    return True, "eligible", margin


def _predict_v62_probs(test_rows):
    art = joblib.load(BASE_V62 / "q4_champion.joblib")
    vec = art["vectorizer"]
    models = art["models"]
    keep_leagues = set(art["league_filter"]["kept_leagues"])
    other_token = art["league_filter"]["other_token"]

    rules = _load_league_name_rules()

    probs = [None] * len(test_rows)
    excluded_flags = [False] * len(test_rows)
    excluded_reasons = [None] * len(test_rows)

    transformed = []
    transformed_idx = []

    for i, s in enumerate(tqdm(test_rows, desc="[v6.2] predicciones")):
        rec = dict(s.features_q4)
        lg = str(rec.get("league", ""))
        lg_lc = lg.lower()
        hit = None
        for cname, raw, low in rules:
            if low in lg_lc:
                hit = (cname, raw)
                break
        if hit is not None:
            excluded_flags[i] = True
            excluded_reasons[i] = f"excluded_league_name:{hit[0]}:{hit[1]}"
            continue

        if lg not in keep_leagues:
            rec["league"] = other_token
            rec["league_bucket"] = other_token
        transformed.append(rec)
        transformed_idx.append(i)

    if transformed:
        x_valid = vec.transform(transformed)
        p_xgb = models["xgb"].predict_proba(x_valid)[:, 1]
        p_hgb = models["hist_gb"].predict_proba(x_valid)[:, 1]
        p_blend = (0.6 * p_xgb + 0.4 * p_hgb)
        for j, orig_idx in enumerate(transformed_idx):
            probs[orig_idx] = float(p_blend[j])

    return probs, excluded_flags, excluded_reasons


def _predict_v63_probs(test_rows, snapshot_minutes=V63_MONITOR_SNAPSHOTS, apply_filters=True):
    rules = _load_league_name_rules() if apply_filters else []

    conn = v6.db_mod.get_conn(str(v6.DB_PATH))
    v6.db_mod.init_db(conn)
    probs = [None] * len(test_rows)
    excluded_flags = [False] * len(test_rows)
    excluded_reasons = [None] * len(test_rows)
    selected_snapshots = [None] * len(test_rows)
    try:
        for i, s in enumerate(tqdm(test_rows, desc="[v6.3] predicciones")):
            base_rec = dict(s.features_q4)
            lg = str(base_rec.get("league", ""))
            lg_lc = lg.lower()
            hit = None
            for cname, raw, low in rules:
                if low in lg_lc:
                    hit = (cname, raw)
                    break
            if apply_filters and hit is not None:
                excluded_flags[i] = True
                excluded_reasons[i] = f"excluded_league_name:{hit[0]}:{hit[1]}"
                continue

            mid = str(s.match_id)
            match_data = v6.db_mod.get_match(conn, mid)
            if not match_data:
                excluded_flags[i] = True
                excluded_reasons[i] = "missing_match_data"
                continue

            accepted_prob = None
            accepted_snapshot = None
            last_reason = "no_signal_in_monitor_windows"
            for snap in snapshot_minutes:
                live_like = infer_live._build_live_like_q4_snapshot(match_data, int(snap))
                features = infer_live._build_features_v3(conn, live_like, "q4", int(snap))
                p_home = infer_live._predict_prob(
                    "v6_3",
                    "q4",
                    "champion",
                    features,
                    snapshot_minute=int(snap),
                )

                confidence = abs(p_home - 0.5) * 2.0
                bet_signal = infer_live._bet_signal("q4", confidence, int(snap))
                gate = infer_live._decision_gate(
                    match_data=live_like,
                    target="q4",
                    snapshot_minute=int(snap),
                    confidence=confidence,
                    model_signal=bet_signal["signal"],
                )
                eligible, eligible_reason, _margin = _v63_window_eligibility(
                    live_like,
                    int(snap),
                    p_home,
                    gate,
                )
                if not eligible:
                    last_reason = eligible_reason
                    continue

                cand = _v63_bet_candidate(p_home)
                if cand.get("accept"):
                    accepted_prob = p_home
                    accepted_snapshot = int(snap)
                    break
                last_reason = str(cand.get("reason", last_reason))

            probs[i] = accepted_prob
            selected_snapshots[i] = accepted_snapshot
            if accepted_prob is None:
                excluded_reasons[i] = last_reason
    finally:
        conn.close()

    return probs, excluded_flags, excluded_reasons, selected_snapshots


def _predict_v63_probs_for_snapshot(test_rows, snapshot_minute: int, apply_filters=True):
    return _predict_v63_probs_for_snapshot_mode(
        test_rows,
        snapshot_minute,
        apply_filters=apply_filters,
        use_live_filters=True,
    )


def _predict_v63_probs_for_snapshot_mode(
    test_rows,
    snapshot_minute: int,
    apply_filters=True,
    use_live_filters=True,
):
    rules = _load_league_name_rules() if apply_filters else []

    conn = v6.db_mod.get_conn(str(v6.DB_PATH))
    v6.db_mod.init_db(conn)
    probs = [None] * len(test_rows)
    excluded_flags = [False] * len(test_rows)
    excluded_reasons = [None] * len(test_rows)
    selected_snapshots = [int(snapshot_minute)] * len(test_rows)
    
    # DEBUG: track gate failures
    gate_reasons_count = defaultdict(int)
    model_probs_accepted = []
    model_probs_rejected = []
    
    mode_suffix = "" if use_live_filters else " probe"
    try:
        for i, s in enumerate(
            tqdm(
                test_rows,
                desc=f"[v6.3] predicciones m{snapshot_minute}{mode_suffix}",
            )
        ):
            base_rec = dict(s.features_q4)
            lg = str(base_rec.get("league", ""))
            lg_lc = lg.lower()
            hit = None
            for cname, raw, low in rules:
                if low in lg_lc:
                    hit = (cname, raw)
                    break
            if apply_filters and hit is not None:
                excluded_flags[i] = True
                excluded_reasons[i] = f"excluded_league_name:{hit[0]}:{hit[1]}"
                continue

            mid = str(s.match_id)
            match_data = v6.db_mod.get_match(conn, mid)
            if not match_data:
                excluded_flags[i] = True
                excluded_reasons[i] = "missing_match_data"
                continue

            try:
                live_like = infer_live._build_live_like_q4_snapshot(match_data, int(snapshot_minute))
                features = infer_live._build_features_v3(conn, live_like, "q4", int(snapshot_minute))
                p_home = infer_live._predict_prob(
                    "v6_3",
                    "q4",
                    "champion",
                    features,
                    snapshot_minute=int(snapshot_minute),
                )
            except FileNotFoundError as exc:
                excluded_flags[i] = True
                excluded_reasons[i] = f"artifact_missing:{exc}"
                continue

            if use_live_filters:
                confidence = abs(p_home - 0.5) * 2.0
                bet_signal = infer_live._bet_signal(
                    "q4",
                    confidence,
                    int(snapshot_minute),
                )
                gate = infer_live._decision_gate(
                    match_data=live_like,
                    target="q4",
                    snapshot_minute=int(snapshot_minute),
                    confidence=confidence,
                    model_signal=bet_signal["signal"],
                )
                eligible, eligible_reason, _margin = _v63_window_eligibility(
                    live_like,
                    int(snapshot_minute),
                    p_home,
                    gate,
                )
                if not eligible:
                    excluded_reasons[i] = eligible_reason
                    gate_reasons_count[eligible_reason] += 1
                    model_probs_rejected.append(p_home)
                    continue

            cand = _v63_bet_candidate(p_home)
            if cand.get("accept"):
                probs[i] = p_home
                model_probs_accepted.append(p_home)
            else:
                fallback_reason = "no_signal_in_snapshot"
                if not use_live_filters:
                    fallback_reason = "model_probe_rejected"
                excluded_reasons[i] = str(cand.get("reason", fallback_reason))
                gate_reasons_count[str(cand.get("reason"))] += 1
                model_probs_rejected.append(p_home)
    finally:
        conn.close()

    # DEBUG OUTPUT
    label = "gate rejection reasons" if use_live_filters else "probe rejection reasons"
    print(f"\n[DEBUG] m{snapshot_minute} {label}:")
    for reason, count in sorted(gate_reasons_count.items(), key=lambda x: -x[1]):
        print(f"  {reason}: {count}")
    if model_probs_accepted:
        print(f"[DEBUG] m{snapshot_minute} probs ACCEPTED (n={len(model_probs_accepted)}): "
              f"mean={np.mean(model_probs_accepted):.3f} "
              f"min={np.min(model_probs_accepted):.3f} "
              f"max={np.max(model_probs_accepted):.3f}")
    if model_probs_rejected:
        print(f"[DEBUG] m{snapshot_minute} probs REJECTED (n={len(model_probs_rejected)}): "
              f"mean={np.mean(model_probs_rejected):.3f} "
              f"min={np.min(model_probs_rejected):.3f} "
              f"max={np.max(model_probs_rejected):.3f}")

    return probs, excluded_flags, excluded_reasons, selected_snapshots


def _predict_m27_v1_probs_for_snapshot_mode(
    test_rows,
    snapshot_minute: int,
    apply_filters=True,
    use_live_filters=True,
):
    if int(snapshot_minute) != int(m27_v1_train.SNAPSHOT_MINUTE):
        raise ValueError(
            "m27_v1 solo soporta snapshot_minute=27 en este reporte"
        )

    rules = _load_league_name_rules() if apply_filters else []
    artifact = joblib.load(M27_V1_CHAMPION_PATH)
    vectorizer = artifact["vectorizer"]
    models = artifact["models"]

    conn = v6.db_mod.get_conn(str(v6.DB_PATH))
    v6.db_mod.init_db(conn)
    probs = [None] * len(test_rows)
    excluded_flags = [False] * len(test_rows)
    excluded_reasons = [None] * len(test_rows)
    selected_snapshots = [int(snapshot_minute)] * len(test_rows)

    gate_reasons_count = defaultdict(int)
    model_probs_accepted = []
    model_probs_rejected = []

    mode_suffix = "" if use_live_filters else " probe"
    try:
        for i, s in enumerate(
            tqdm(
                test_rows,
                desc=f"[m27_v1] predicciones m{snapshot_minute}{mode_suffix}",
            )
        ):
            base_rec = dict(s.features_q4)
            lg = str(base_rec.get("league", ""))
            lg_lc = lg.lower()
            hit = None
            for cname, raw, low in rules:
                if low in lg_lc:
                    hit = (cname, raw)
                    break
            if apply_filters and hit is not None:
                excluded_flags[i] = True
                excluded_reasons[i] = f"excluded_league_name:{hit[0]}:{hit[1]}"
                continue

            mid = str(s.match_id)
            match_data = v6.db_mod.get_match(conn, mid)
            if not match_data:
                excluded_flags[i] = True
                excluded_reasons[i] = "missing_match_data"
                continue

            live_like = _build_live_like_q4_snapshot(
                match_data,
                int(snapshot_minute),
            )
            feature_dict = m27_v1_train._build_m27_v1_features(s, live_like)
            x_row = vectorizer.transform([feature_dict])
            p_xgb = float(models["xgb"].predict_proba(x_row)[0, 1])
            p_hist = float(models["hist_gb"].predict_proba(x_row)[0, 1])
            p_home = (p_xgb + p_hist) / 2.0

            if use_live_filters:
                confidence = abs(p_home - 0.5) * 2.0
                bet_signal = infer_live._bet_signal(
                    "q4",
                    confidence,
                    int(snapshot_minute),
                )
                gate = infer_live._decision_gate(
                    match_data=live_like,
                    target="q4",
                    snapshot_minute=int(snapshot_minute),
                    confidence=confidence,
                    model_signal=bet_signal["signal"],
                )
                eligible, eligible_reason, _margin = _v63_window_eligibility(
                    live_like,
                    int(snapshot_minute),
                    p_home,
                    gate,
                )
                if not eligible:
                    excluded_reasons[i] = eligible_reason
                    gate_reasons_count[eligible_reason] += 1
                    model_probs_rejected.append(p_home)
                    continue

            cand = _v63_bet_candidate(p_home)
            if cand.get("accept"):
                probs[i] = p_home
                model_probs_accepted.append(p_home)
            else:
                fallback_reason = "no_signal_in_snapshot"
                if not use_live_filters:
                    fallback_reason = "model_probe_rejected"
                reason = str(cand.get("reason", fallback_reason))
                excluded_reasons[i] = reason
                gate_reasons_count[reason] += 1
                model_probs_rejected.append(p_home)
    finally:
        conn.close()

    label = "gate rejection reasons" if use_live_filters else "probe rejection reasons"
    print(f"\n[DEBUG] m27_v1 m{snapshot_minute} {label}:")
    for reason, count in sorted(gate_reasons_count.items(), key=lambda x: -x[1]):
        print(f"  {reason}: {count}")
    if model_probs_accepted:
        print(
            f"[DEBUG] m27_v1 probs ACCEPTED (n={len(model_probs_accepted)}): "
            f"mean={np.mean(model_probs_accepted):.3f} "
            f"min={np.min(model_probs_accepted):.3f} "
            f"max={np.max(model_probs_accepted):.3f}"
        )
    if model_probs_rejected:
        print(
            f"[DEBUG] m27_v1 probs REJECTED (n={len(model_probs_rejected)}): "
            f"mean={np.mean(model_probs_rejected):.3f} "
            f"min={np.min(model_probs_rejected):.3f} "
            f"max={np.max(model_probs_rejected):.3f}"
        )

    return probs, excluded_flags, excluded_reasons, selected_snapshots


def _calibrate_probs(val_probs, y_val, test_probs, method):
    """Calibrate raw probabilities using validation set and apply to test."""
    x_val = []
    y_val_use = []
    for p, y in zip(val_probs, y_val):
        if p is None:
            continue
        x_val.append(float(p))
        y_val_use.append(int(y))

    if not x_val or len(set(y_val_use)) < 2:
        return test_probs

    calibrated = []
    if method == "isotonic":
        iso = IsotonicRegression(out_of_bounds="clip")
        iso.fit(x_val, y_val_use)
        for p in test_probs:
            if p is None:
                calibrated.append(None)
            else:
                calibrated.append(float(iso.predict([float(p)])[0]))
        return calibrated

    if method == "platt":
        lr = LogisticRegression(solver="lbfgs")
        lr.fit(np.array(x_val, dtype=float).reshape(-1, 1), np.array(y_val_use, dtype=int))
        for p in test_probs:
            if p is None:
                calibrated.append(None)
            else:
                pp = lr.predict_proba(np.array([[float(p)]], dtype=float))[0, 1]
                calibrated.append(float(pp))
        return calibrated

    return test_probs


def _build_blacklist_flags(rows):
    flags = []
    reasons = []
    for s in rows:
        league = str(s.features_q4.get("league", ""))
        blocked, reason = _V63_BL.is_blocked(league)
        flags.append(blocked)
        reasons.append(reason)
    return flags, reasons


def _merge_exclusions(primary_flags, primary_reasons, extra_flags, extra_reasons):
    out_flags = []
    out_reasons = []
    for pf, pr, ef, er in zip(primary_flags, primary_reasons, extra_flags, extra_reasons):
        if pf:
            out_flags.append(True)
            out_reasons.append(pr)
        elif ef:
            out_flags.append(True)
            out_reasons.append(er)
        else:
            out_flags.append(False)
            out_reasons.append(None)
    return out_flags, out_reasons


def _build_m27_v1_policy_flags(rows):
    flags = []
    reasons = []
    for s in rows:
        league = str(s.features_q4.get("league", ""))
        if m27_v1_policy.is_blacklisted(league):
            flags.append(True)
            reasons.append(
                f"m27_v1_league_blacklist:{m27_v1_policy.normalize_league_name(league)}"
            )
        else:
            flags.append(False)
            reasons.append(None)
    return flags, reasons


def _annotate_m27_v1_policy(model_df: pd.DataFrame) -> pd.DataFrame:
    if model_df.empty or "liga" not in model_df.columns:
        return model_df
    out = model_df.copy()
    out["m27_v1_league_tier"] = out["liga"].map(m27_v1_policy.get_tier)
    out["m27_v1_policy_action"] = out["liga"].map(m27_v1_policy.get_action)
    return out


def _build_m27_v1_policy_sheet(model_df: pd.DataFrame) -> pd.DataFrame:
    cols = [
        "liga",
        "tier",
        "accion",
        "matches_totales",
        "matches_apostados",
        "ganados",
        "perdidos",
        "efectividad",
        "ganancia",
        "yield_sobre_apostado",
    ]
    if model_df.empty:
        return pd.DataFrame(columns=cols)

    all_matches = (
        model_df
        .groupby("liga", as_index=False)
        .agg(matches_totales=("match_id", "count"))
    )
    bets_df = model_df[
        model_df["resultado_apuesta"].isin(["GANADA", "PERDIDA"])
    ].copy()
    if bets_df.empty:
        out = all_matches.copy()
        out["tier"] = out["liga"].map(m27_v1_policy.get_tier)
        out["accion"] = out["liga"].map(m27_v1_policy.get_action)
        for col in cols:
            if col not in out.columns:
                out[col] = 0.0 if col in {"ganancia", "efectividad", "yield_sobre_apostado"} else ""
        return out[cols]

    grouped = (
        bets_df
        .groupby("liga", as_index=False)
        .agg(
            matches_apostados=("resultado_apuesta", "count"),
            ganados=("resultado_apuesta", lambda s: int((s == "GANADA").sum())),
            perdidos=("resultado_apuesta", lambda s: int((s == "PERDIDA").sum())),
            ganancia=("ganancia", "sum"),
            total_apostado=("monto_apostado", "sum"),
        )
    )
    out = all_matches.merge(grouped, on="liga", how="left").fillna(
        {
            "matches_apostados": 0,
            "ganados": 0,
            "perdidos": 0,
            "ganancia": 0.0,
            "total_apostado": 0.0,
        }
    )
    out["matches_apostados"] = out["matches_apostados"].astype(int)
    out["ganados"] = out["ganados"].astype(int)
    out["perdidos"] = out["perdidos"].astype(int)
    out["efectividad"] = out.apply(
        lambda r: (r["ganados"] / r["matches_apostados"])
        if r["matches_apostados"]
        else 0.0,
        axis=1,
    )
    out["yield_sobre_apostado"] = out.apply(
        lambda r: (r["ganancia"] / r["total_apostado"])
        if r["total_apostado"]
        else 0.0,
        axis=1,
    )
    out["tier"] = out["liga"].map(m27_v1_policy.get_tier)
    out["accion"] = out["liga"].map(m27_v1_policy.get_action)
    out = out[cols].sort_values(
        ["tier", "ganancia", "matches_apostados"],
        ascending=[True, False, False],
    )
    return out


def _simulate(
    model_name,
    test_rows,
    p_home_list,
    base_minute,
    excluded_flags=None,
    excluded_reasons=None,
    mode="kelly_non_compound",
    kelly_mult=1.0,
    kelly_cap=1.0,
    min_conf_prob=0.5,
    stake_step=1.0,
    min_stake=1.0,
    max_stake=100.0,
    teams_map=None,
    q4_scores_map=None,
    q3_min27_scores_map=None,
    selected_snapshot_minutes=None,
):
    excluded_flags = excluded_flags or [False] * len(test_rows)
    excluded_reasons = excluded_reasons or [None] * len(test_rows)

    bank = BANK_START
    details = []
    bets = wins = losses = no_bet = 0
    total_staked = 0.0
    peak = BANK_START
    max_dd = 0.0

    for i, s in enumerate(test_rows):
        y = int(s.target_q4)
        p_home = p_home_list[i]
        match_id = str(s.match_id)
        snapshot_used = selected_snapshot_minutes[i] if selected_snapshot_minutes else base_minute
        if snapshot_used is None:
            snapshot_used = base_minute
        snapshot_used_int = int(snapshot_used) if snapshot_used is not None else int(base_minute)
        teams = (teams_map or {}).get(match_id, {})
        home_team_name = teams.get("home_team", "")
        away_team_name = teams.get("away_team", "")
        q4_score = (q4_scores_map or {}).get(match_id, "")
        q3_min27_score = (
            (q3_min27_scores_map or {}).get(match_id, "")
            if snapshot_used_int == int(V63_SNAPSHOT_MINUTE)
            else ""
        )

        rec = {
            "resultado_apuesta": "SIN_APUESTA",
            "apuesta": "sin_apuesta",
            "monto_apostado": 0.0,
            "ganancia": 0.0,
            "bank_final": BANK_START,
            "modelo": model_name,
            "partida_test": i + 1,
            "match_id": match_id,
            "fecha": ((s.dt if s.dt.tzinfo else s.dt.replace(tzinfo=timezone.utc)).astimezone(timezone(timedelta(hours=-6))).strftime("%Y-%m-%d")) if s.dt else "",
            "hora": ((s.dt if s.dt.tzinfo else s.dt.replace(tzinfo=timezone.utc)).astimezone(timezone(timedelta(hours=-6))).strftime("%H:%M:%S")) if s.dt else "",
            "liga": str(s.features_q4.get("league", "")),
            "equipo_local": home_team_name,
            "equipo_visitante": away_team_name,
            "resultado_q4_home_gana": y,
            "marcador_q4": q4_score,
            "prob_local": None if p_home is None else float(p_home),
            "prob_visitante": None if p_home is None else float(1.0 - p_home),
            "lado_predicho": None,
            "confianza_prob": None,
            "confianza_score_0_100": None,
            "nivel_confianza": None,
            "apuestas_odds": ODDS,
            "probabilidad_empate": BREAK_EVEN,
            "edge": None,
            "kelly_fraction_raw": None,
            "kelly_fraction_used": None,
            "step_apuesta": stake_step,
            "razon_sin_apuesta": None,
            "pnl": 0.0,
            "banco_antes": bank,
            "ganancia_acumulada": bank - BANK_START,
            "roi_banco_acumulado": (bank - BANK_START) / BANK_START,
            "minuto_base_apuesta": snapshot_used_int,
            "marcador_q3_min27": q3_min27_score,
        }

        if excluded_flags[i]:
            if p_home is not None:
                p_pick = p_home if p_home >= 0.5 else (1.0 - p_home)
                pick_home = p_home >= 0.5
                pick_side = "local" if pick_home else "visitante"

                rec["lado_predicho"] = pick_side
                rec["confianza_prob"] = p_pick
                rec["confianza_score_0_100"] = p_pick * 100.0
                rec["nivel_confianza"] = (
                    "muy_alta" if p_pick >= 0.80 else
                    "alta" if p_pick >= 0.70 else
                    "media" if p_pick >= 0.60 else
                    "baja"
                )
                rec["edge"] = p_pick - BREAK_EVEN
            rec["razon_sin_apuesta"] = excluded_reasons[i]
            rec["monto_apostado"] = None
            rec["ganancia"] = None
            rec["bank_final"] = BANK_START
            no_bet += 1
            details.append(rec)
            continue

        if p_home is None:
            rec["razon_sin_apuesta"] = excluded_reasons[i] or "no_signal_in_monitor_windows"
            no_bet += 1
            details.append(rec)
            continue

        p_pick = p_home if p_home >= 0.5 else (1.0 - p_home)
        pick_home = p_home >= 0.5
        pick_side = "local" if pick_home else "visitante"

        rec["lado_predicho"] = pick_side
        rec["confianza_prob"] = p_pick
        rec["confianza_score_0_100"] = p_pick * 100.0
        rec["nivel_confianza"] = (
            "muy_alta" if p_pick >= 0.80 else
            "alta" if p_pick >= 0.70 else
            "media" if p_pick >= 0.60 else
            "baja"
        )
        rec["edge"] = p_pick - BREAK_EVEN

        if p_pick < min_conf_prob:
            rec["razon_sin_apuesta"] = "confidence_filter"
            no_bet += 1
            details.append(rec)
            continue

        k_raw = _kelly_fraction(p_pick, ODDS)
        rec["kelly_fraction_raw"] = k_raw
        if k_raw <= 0:
            rec["razon_sin_apuesta"] = "no_edge"
            no_bet += 1
            details.append(rec)
            continue

        k_used = min(k_raw * kelly_mult, kelly_cap)
        rec["kelly_fraction_used"] = k_used

        stake = _stake_from_signal(p_pick, rec["edge"], k_used, kelly_cap)
        if max_stake > 0:
            stake = min(stake, max_stake)
        stake = _round_down_step(stake, stake_step)
        if stake < min_stake:
            rec["razon_sin_apuesta"] = "stake_below_25"
            no_bet += 1
            details.append(rec)
            continue
        if stake > bank:
            rec["razon_sin_apuesta"] = "insufficient_bank"
            no_bet += 1
            details.append(rec)
            continue

        is_win = (y == 1 and pick_home) or (y == 0 and (not pick_home))
        pnl = stake * (ODDS - 1.0) if is_win else -stake
        bank += pnl

        rec["apuesta"] = "home" if pick_home else "away"
        rec["resultado_apuesta"] = "GANADA" if is_win else "PERDIDA"
        rec["monto_apostado"] = stake
        rec["ganancia"] = pnl
        rec["bank_final"] = bank
        rec["pnl"] = pnl
        rec["banco_antes"] = BANK_START if mode == "kelly_non_compound" else (bank - pnl)
        rec["ganancia_acumulada"] = bank - BANK_START
        rec["roi_banco_acumulado"] = (bank - BANK_START) / BANK_START

        bets += 1
        wins += int(is_win)
        losses += int(not is_win)
        total_staked += stake
        peak = max(peak, bank)
        dd = (peak - bank) / peak if peak > 0 else 0.0
        max_dd = max(max_dd, dd)

        details.append(rec)

    summary = {
        "modelo": model_name,
        "partidos_test": len(test_rows),
        "apuestas": bets,
        "ganadas": wins,
        "perdidas": losses,
        "empates_contados_como_perdida": 0,
        "sin_apuesta": no_bet,
        "efectividad": (wins / bets) if bets else 0.0,
        "banco_inicio": BANK_START,
        "banco_final": bank,
        "ganancia": bank - BANK_START,
        "roi_bank": (bank - BANK_START) / BANK_START,
        "total_apostado": total_staked,
        "apuesta_promedio": (total_staked / bets) if bets else 0.0,
        "yield_sobre_apostado": ((bank - BANK_START) / total_staked) if total_staked > 0 else 0.0,
        "max_drawdown": max_dd,
        "minuto_base_apuesta": int(base_minute),
    }
    return details, summary


def _build_effectiveness_by_league(model_dfs):
    cols = [
        "modelo", "liga", "matches_apostados", "ganados", "perdidos",
        "efectividad", "ganancia"
    ]

    all_df = pd.concat(model_dfs, ignore_index=True)
    if all_df.empty:
        return pd.DataFrame(columns=cols)

    bets_df = all_df[all_df["resultado_apuesta"].isin(["GANADA", "PERDIDA"])].copy()
    if bets_df.empty:
        return pd.DataFrame(columns=cols)

    grouped = (
        bets_df
        .groupby(["modelo", "liga"], as_index=False)
        .agg(
            matches_apostados=("resultado_apuesta", "count"),
            ganados=("resultado_apuesta", lambda s: int((s == "GANADA").sum())),
            perdidos=("resultado_apuesta", lambda s: int((s == "PERDIDA").sum())),
            ganancia=("ganancia", "sum")
        )
    )
    grouped["efectividad"] = grouped.apply(
        lambda r: (r["ganados"] / r["matches_apostados"]) if r["matches_apostados"] else 0.0,
        axis=1
    )
    grouped = grouped[cols].sort_values(["modelo", "ganancia"], ascending=[True, False])
    return grouped


def _build_v63_snapshot_comparison(details_27, details_30, q3_min27_scores_map, q3_min30_scores_map):
    by_27 = {str(row["match_id"]): row for row in details_27}
    by_30 = {str(row["match_id"]): row for row in details_30}

    rows = []
    for match_id in sorted(set(by_27) | set(by_30)):
        r27 = by_27.get(match_id, {})
        r30 = by_30.get(match_id, {})

        res27 = str(r27.get("resultado_apuesta", "SIN_APUESTA"))
        res30 = str(r30.get("resultado_apuesta", "SIN_APUESTA"))
        rows.append({
            "fecha": r27.get("fecha") or r30.get("fecha") or "",
            "hora": r27.get("hora") or r30.get("hora") or "",
            "liga": r27.get("liga") or r30.get("liga") or "",
            "match_id": match_id,
            "equipo_local": r27.get("equipo_local") or r30.get("equipo_local") or "",
            "equipo_visitante": r27.get("equipo_visitante") or r30.get("equipo_visitante") or "",
            "resultado_q4_home_gana": r27.get("resultado_q4_home_gana", r30.get("resultado_q4_home_gana")),
            "marcador_q4": r27.get("marcador_q4") or r30.get("marcador_q4") or "",
            "marcador_q3_min27": q3_min27_scores_map.get(match_id, ""),
            "marcador_q3_min30": q3_min30_scores_map.get(match_id, ""),
            "pick_m27": r27.get("lado_predicho"),
            "resultado_m27": res27,
            "prob_m27": r27.get("confianza_prob"),
            "prob_local_m27": r27.get("prob_local"),
            "stake_m27": r27.get("monto_apostado"),
            "ganancia_m27": r27.get("ganancia"),
            "razon_m27": r27.get("razon_sin_apuesta"),
            "pick_m30": r30.get("lado_predicho"),
            "resultado_m30": res30,
            "prob_m30": r30.get("confianza_prob"),
            "prob_local_m30": r30.get("prob_local"),
            "stake_m30": r30.get("monto_apostado"),
            "ganancia_m30": r30.get("ganancia"),
            "razon_m30": r30.get("razon_sin_apuesta"),
            "ambos_apuestan": int(res27 in ("GANADA", "PERDIDA") and res30 in ("GANADA", "PERDIDA")),
            "solo_m27_apuesta": int(res27 in ("GANADA", "PERDIDA") and res30 == "SIN_APUESTA"),
            "solo_m30_apuesta": int(res30 in ("GANADA", "PERDIDA") and res27 == "SIN_APUESTA"),
            "gana_m27": int(res27 == "GANADA"),
            "gana_m30": int(res30 == "GANADA"),
            "solo_m27_gana": int(res27 == "GANADA" and res30 != "GANADA"),
            "solo_m30_gana": int(res30 == "GANADA" and res27 != "GANADA"),
        })

    return pd.DataFrame(rows)


def _apply_excel_formatting(ws, header_row=1):
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for row in range(header_row + 1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if cell.column_letter in {"E", "F", "G", "N", "P", "Q", "S", "T", "V", "W", "X", "Y", "Z", "AA", "AC", "AD", "AE", "AF", "AG"}:
                cell.number_format = "0.00"

    header_map = {}
    for col in range(1, ws.max_column + 1):
        header_val = ws.cell(row=header_row, column=col).value
        if header_val:
            header_map[str(header_val)] = col

    percent_headers = {"efectividad", "roi_bank", "yield_sobre_apostado", "max_drawdown"}
    for h in percent_headers:
        if h in header_map:
            c = header_map[h]
            for row in range(header_row + 1, ws.max_row + 1):
                ws.cell(row=row, column=c).number_format = "0.00%"

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    if ws.max_row > header_row:
        start_row = header_row + 1
        end_row = ws.max_row

        if ws.title.endswith("_matches"):
            ws.conditional_formatting.add(
                f"U{start_row}:U{end_row}",
                FormulaRule(formula=[f'$U{start_row}="muy_alta"'], fill=green_fill)
            )
            ws.conditional_formatting.add(
                f"U{start_row}:U{end_row}",
                FormulaRule(formula=[f'$U{start_row}="alta"'], fill=green_fill)
            )
            ws.conditional_formatting.add(
                f"U{start_row}:U{end_row}",
                FormulaRule(formula=[f'$U{start_row}="media"'], fill=yellow_fill)
            )
            ws.conditional_formatting.add(
                f"U{start_row}:U{end_row}",
                FormulaRule(formula=[f'$U{start_row}="baja"'], fill=red_fill)
            )

            for col in ["B", "L", "M", "E", "F", "G"]:
                ws.conditional_formatting.add(
                    f"{col}{start_row}:{col}{end_row}",
                    FormulaRule(formula=[f'LEFT($AB{start_row},20)="excluded_league_name"'], fill=red_fill)
                )

            ws.conditional_formatting.add(
                f"C{start_row}:C{end_row}",
                FormulaRule(formula=[f'$C{start_row}="GANADA"'], fill=green_fill)
            )
            ws.conditional_formatting.add(
                f"C{start_row}:C{end_row}",
                FormulaRule(formula=[f'$C{start_row}="PERDIDA"'], fill=red_fill)
            )
            ws.conditional_formatting.add(
                f"C{start_row}:C{end_row}",
                FormulaRule(formula=[f'$C{start_row}="SIN_APUESTA"'], fill=yellow_fill)
            )

            for col in ["F", "G", "AF"]:
                ws.conditional_formatting.add(
                    f"{col}{start_row}:{col}{end_row}",
                    CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill)
                )
                ws.conditional_formatting.add(
                    f"{col}{start_row}:{col}{end_row}",
                    CellIsRule(operator="lessThan", formula=["0"], fill=red_fill)
                )
                ws.conditional_formatting.add(
                    f"{col}{start_row}:{col}{end_row}",
                    CellIsRule(operator="equal", formula=["0"], fill=yellow_fill)
                )

            ws.conditional_formatting.add(
                f"G{start_row}:G{end_row}",
                CellIsRule(operator="greaterThan", formula=[str(BANK_START)], fill=green_fill)
            )
            ws.conditional_formatting.add(
                f"G{start_row}:G{end_row}",
                CellIsRule(operator="lessThan", formula=[str(BANK_START)], fill=red_fill)
            )
            ws.conditional_formatting.add(
                f"G{start_row}:G{end_row}",
                CellIsRule(operator="equal", formula=[str(BANK_START)], fill=yellow_fill)
            )

            for col in ["G", "AE"]:
                ws.conditional_formatting.add(
                    f"{col}{start_row}:{col}{end_row}",
                    FormulaRule(formula=[f'$C{start_row}="PERDIDA"'], fill=red_fill)
                )

        if ws.title == "summary":
            for col in ["J", "K"]:
                ws.conditional_formatting.add(
                    f"{col}{start_row}:{col}{end_row}",
                    CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill)
                )
                ws.conditional_formatting.add(
                    f"{col}{start_row}:{col}{end_row}",
                    CellIsRule(operator="lessThan", formula=["0"], fill=red_fill)
                )
                ws.conditional_formatting.add(
                    f"{col}{start_row}:{col}{end_row}",
                    CellIsRule(operator="equal", formula=["0"], fill=yellow_fill)
                )

        if ws.title == "efectividad_liga":
            ws.conditional_formatting.add(
                f"F{start_row}:F{end_row}",
                CellIsRule(operator="greaterThanOrEqual", formula=["0.7"], fill=green_fill)
            )
            ws.conditional_formatting.add(
                f"F{start_row}:F{end_row}",
                CellIsRule(operator="between", formula=["0.55", "0.6999"], fill=yellow_fill)
            )
            ws.conditional_formatting.add(
                f"F{start_row}:F{end_row}",
                CellIsRule(operator="lessThan", formula=["0.55"], fill=red_fill)
            )
            ws.conditional_formatting.add(
                f"G{start_row}:G{end_row}",
                CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill)
            )
            ws.conditional_formatting.add(
                f"G{start_row}:G{end_row}",
                CellIsRule(operator="lessThan", formula=["0"], fill=red_fill)
            )
            ws.conditional_formatting.add(
                f"G{start_row}:G{end_row}",
                CellIsRule(operator="equal", formula=["0"], fill=yellow_fill)
            )

    ws.freeze_panes = f"A{header_row + 1}"


def _ask_yes_no(prompt: str, default_yes: bool = True) -> bool:
    suffix = "[S/n]" if default_yes else "[s/N]"
    resp = input(f"\n{prompt} {suffix}: ").strip().lower()
    if not resp:
        return default_yes
    return resp in ("s", "si", "sí", "y", "yes")


def _ask_optional_report_date() -> str | None:
    if not _ask_yes_no("¿Quieres ver resultados de una fecha particular?", default_yes=False):
        return None

    while True:
        raw_value = input("Ingresa la fecha en formato YYYY-MM-DD: ").strip()
        try:
            chosen_date = datetime.strptime(raw_value, "%Y-%m-%d").strftime("%Y-%m-%d")
        except ValueError:
            print("Fecha inválida. Usa el formato YYYY-MM-DD.")
            continue
        return chosen_date


def main():
    parser = argparse.ArgumentParser(description="Q4 ROI report V6.2 vs V6.3")
    parser.add_argument(
        "--no-v62",
        action="store_true",
        help="Omitir v6.2 del reporte sin preguntar",
    )
    parser.add_argument(
        "--rebuild-pred-cache",
        action="store_true",
        help="Ignorar cache de predicciones y recalcular todo",
    )
    parser.add_argument(
        "--rebuild-splits-cache",
        action="store_true",
        help="Ignorar cache de splits (val/test rows) y recalcular",
    )
    parser.add_argument(
        "--only-m27",
        action="store_true",
        help="Ejecutar solo v6.3 minuto 27 (sin v6.2, sin raw monitor, sin m30)",
    )
    parser.add_argument(
        "--only-m27-v1",
        action="store_true",
        help="Ejecutar solo m27_v1 (sin v6.2, raw, m27, m30 ni probes)",
    )
    parser.add_argument(
        "--with-raw",
        action="store_true",
        help="Forzar incluir v6.3 raw monitor windows",
    )
    parser.add_argument(
        "--no-raw",
        action="store_true",
        help="Excluir v6.3 raw monitor windows",
    )
    parser.add_argument(
        "--with-m27",
        action="store_true",
        help="Forzar incluir v6.3 minuto 27",
    )
    parser.add_argument(
        "--no-m27",
        action="store_true",
        help="Excluir v6.3 minuto 27",
    )
    parser.add_argument(
        "--with-m30",
        action="store_true",
        help="Forzar incluir v6.3 minuto 30",
    )
    parser.add_argument(
        "--no-m30",
        action="store_true",
        help="Excluir v6.3 minuto 30",
    )
    parser.add_argument(
        "--with-m27-probe",
        action="store_true",
        help="Incluir v6.3 minuto 27 sin post-filtros live (mantiene odds y Kelly)",
    )
    parser.add_argument(
        "--no-m27-probe",
        action="store_true",
        help="Excluir v6.3 minuto 27 sin post-filtros live",
    )
    parser.add_argument(
        "--with-m30-probe",
        action="store_true",
        help="Incluir v6.3 minuto 30 sin post-filtros live (mantiene odds y Kelly)",
    )
    parser.add_argument(
        "--no-m30-probe",
        action="store_true",
        help="Excluir v6.3 minuto 30 sin post-filtros live",
    )
    parser.add_argument(
        "--with-m27-v1",
        action="store_true",
        help="Incluir m27_v1 con post-filtros live",
    )
    parser.add_argument(
        "--no-m27-v1",
        action="store_true",
        help="Excluir m27_v1 con post-filtros live",
    )
    parser.add_argument(
        "--with-m27-v1-probe",
        action="store_true",
        help="Incluir m27_v1 sin post-filtros live (mantiene odds y Kelly)",
    )
    parser.add_argument(
        "--no-m27-v1-probe",
        action="store_true",
        help="Excluir m27_v1 sin post-filtros live",
    )
    args = parser.parse_args()

    if args.only_m27 and args.only_m27_v1:
        raise ValueError("No puedes usar --only-m27 y --only-m27-v1 al mismo tiempo")

    if args.only_m27:
        include_v62 = False
        run_v63_raw = False
        run_v63_m27 = True
        run_v63_m30 = False
        run_v63_m27_probe = False
        run_v63_m30_probe = False
        run_v63_m27_v1 = False
        run_v63_m27_v1_probe = False
    elif args.only_m27_v1:
        include_v62 = False
        run_v63_raw = False
        run_v63_m27 = False
        run_v63_m30 = False
        run_v63_m27_probe = False
        run_v63_m30_probe = False
        run_v63_m27_v1 = True
        run_v63_m27_v1_probe = False
    else:
        if args.no_v62:
            include_v62 = False
        else:
            include_v62 = _ask_yes_no("¿Incluir v6.2 en el reporte?", default_yes=True)

        if args.with_raw and args.no_raw:
            raise ValueError("No puedes usar --with-raw y --no-raw al mismo tiempo")
        if args.with_m27 and args.no_m27:
            raise ValueError("No puedes usar --with-m27 y --no-m27 al mismo tiempo")
        if args.with_m30 and args.no_m30:
            raise ValueError("No puedes usar --with-m30 y --no-m30 al mismo tiempo")
        if args.with_m27_probe and args.no_m27_probe:
            raise ValueError("No puedes usar --with-m27-probe y --no-m27-probe al mismo tiempo")
        if args.with_m30_probe and args.no_m30_probe:
            raise ValueError("No puedes usar --with-m30-probe y --no-m30-probe al mismo tiempo")
        if args.with_m27_v1 and args.no_m27_v1:
            raise ValueError("No puedes usar --with-m27-v1 y --no-m27-v1 al mismo tiempo")
        if args.with_m27_v1_probe and args.no_m27_v1_probe:
            raise ValueError("No puedes usar --with-m27-v1-probe y --no-m27-v1-probe al mismo tiempo")

        run_v63_raw = True if args.with_raw else (False if args.no_raw else _ask_yes_no("¿Incluir v6.3 raw monitor (m27/m30)?", default_yes=False))
        run_v63_m27 = True if args.with_m27 else (False if args.no_m27 else _ask_yes_no("¿Incluir v6.3 m27?", default_yes=True))
        run_v63_m30 = True if args.with_m30 else (False if args.no_m30 else _ask_yes_no("¿Incluir v6.3 m30?", default_yes=False))
        run_v63_m27_probe = True if args.with_m27_probe else (False if args.no_m27_probe else _ask_yes_no("¿Incluir v6.3 m27 probe sin post-filtros live?", default_yes=False))
        run_v63_m30_probe = True if args.with_m30_probe else (False if args.no_m30_probe else _ask_yes_no("¿Incluir v6.3 m30 probe sin post-filtros live?", default_yes=False))
        run_v63_m27_v1 = True if args.with_m27_v1 else (False if args.no_m27_v1 else _ask_yes_no("¿Incluir m27_v1?", default_yes=False))
        run_v63_m27_v1_probe = True if args.with_m27_v1_probe else (False if args.no_m27_v1_probe else _ask_yes_no("¿Incluir m27_v1 probe sin post-filtros live?", default_yes=False))

    if not any([
        include_v62,
        run_v63_raw,
        run_v63_m27,
        run_v63_m30,
        run_v63_m27_probe,
        run_v63_m30_probe,
        run_v63_m27_v1,
        run_v63_m27_v1_probe,
    ]):
        raise ValueError("Debes incluir al menos un bloque (v6.2 o alguna variante v6.3)")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    v62_label = "v6_2_vs_" if include_v62 else ""
    out_mm_path = BASE_V63 / f"Q4_ROI_match_by_match_{v62_label}v6_3_{ts}.xlsx"

    only_report_date = _ask_optional_report_date()

    print("[Q4_ROI_62_63] preparando datos de prueba...")
    _, val_rows, test_rows = _prepare_splits(
        force_rebuild=args.rebuild_splits_cache,
        only_date=only_report_date,
    )
    teams_map = _load_match_teams_map()
    q4_scores_map = _load_match_q4_scores_map([s.match_id for s in test_rows])
    q3_min27_scores_map = _load_match_score_at_minute_map([s.match_id for s in test_rows], V63_SNAPSHOT_MINUTE)
    q3_min30_scores_map = (
        _load_match_score_at_minute_map([s.match_id for s in test_rows], 30)
        if (run_v63_m30 or run_v63_m30_probe)
        else {}
    )

    cache_meta = {
        "report": "q4_roi_62_63",
        "include_v62": bool(include_v62),
        "run_v63_raw": bool(run_v63_raw),
        "run_v63_m27": bool(run_v63_m27),
        "run_v63_m30": bool(run_v63_m30),
        "run_v63_m27_probe": bool(run_v63_m27_probe),
        "run_v63_m30_probe": bool(run_v63_m30_probe),
        "run_v63_m27_v1": bool(run_v63_m27_v1),
        "run_v63_m27_v1_probe": bool(run_v63_m27_v1_probe),
        "filters_disabled": bool(V63_DISABLE_FILTERS),
        "monitor_snapshots": tuple(int(x) for x in V63_MONITOR_SNAPSHOTS),
        "test_fp": _rows_fingerprint(test_rows),
        "val_fp": _rows_fingerprint(val_rows),
        "v63_model_m27_sig": _file_signature(BASE_V63 / "q4_m27_champion.joblib"),
        "v63_model_m30_sig": _file_signature(BASE_V63 / "q4_m30_champion.joblib"),
        "m27_v1_model_sig": _file_signature(M27_V1_CHAMPION_PATH),
        "v62_model_sig": _file_signature(BASE_V62 / "q4_champion.joblib") if include_v62 else "na",
    }
    pred_cache = _load_pred_cache(cache_meta, force_rebuild=args.rebuild_pred_cache)

    if pred_cache is not None:
        print(f"[Q4_ROI_62_63] usando cache de predicciones: {PRED_CACHE_PATH}")
        if include_v62:
            p_v62 = pred_cache["p_v62"]
            excl62_flags = pred_cache["excl62_flags"]
            excl62_reasons = pred_cache["excl62_reasons"]
        if run_v63_raw:
            p_v63_raw = pred_cache["p_v63_raw"]
            excl63_flags = pred_cache["excl63_flags"]
            excl63_reasons = pred_cache["excl63_reasons"]
            snap63_raw = pred_cache["snap63_raw"]
        if run_v63_m27:
            p_v63_27 = pred_cache["p_v63_27"]
            excl63_27_flags = pred_cache["excl63_27_flags"]
            excl63_27_reasons = pred_cache["excl63_27_reasons"]
            snap63_27 = pred_cache["snap63_27"]
        if run_v63_m30:
            p_v63_30 = pred_cache["p_v63_30"]
            excl63_30_flags = pred_cache["excl63_30_flags"]
            excl63_30_reasons = pred_cache["excl63_30_reasons"]
            snap63_30 = pred_cache["snap63_30"]
        if run_v63_m27_probe:
            p_v63_27_probe = pred_cache["p_v63_27_probe"]
            excl63_27_probe_flags = pred_cache["excl63_27_probe_flags"]
            excl63_27_probe_reasons = pred_cache["excl63_27_probe_reasons"]
            snap63_27_probe = pred_cache["snap63_27_probe"]
        if run_v63_m30_probe:
            p_v63_30_probe = pred_cache["p_v63_30_probe"]
            excl63_30_probe_flags = pred_cache["excl63_30_probe_flags"]
            excl63_30_probe_reasons = pred_cache["excl63_30_probe_reasons"]
            snap63_30_probe = pred_cache["snap63_30_probe"]
        if run_v63_m27_v1:
            p_v63_27_v1 = pred_cache["p_v63_27_v1"]
            excl63_27_v1_flags = pred_cache["excl63_27_v1_flags"]
            excl63_27_v1_reasons = pred_cache["excl63_27_v1_reasons"]
            snap63_27_v1 = pred_cache["snap63_27_v1"]
        if run_v63_m27_v1_probe:
            p_v63_27_v1_probe = pred_cache["p_v63_27_v1_probe"]
            excl63_27_v1_probe_flags = pred_cache["excl63_27_v1_probe_flags"]
            excl63_27_v1_probe_reasons = pred_cache["excl63_27_v1_probe_reasons"]
            snap63_27_v1_probe = pred_cache["snap63_27_v1_probe"]
    else:
        if include_v62:
            print("[Q4_ROI_62_63] predicciones v6.2...")
            p_v62, excl62_flags, excl62_reasons = _predict_v62_probs(test_rows)

        if run_v63_raw:
            print(
                f"[Q4_ROI_62_63] predicciones v6.3 raw (monitor_snapshots={list(V63_MONITOR_SNAPSHOTS)}, "
                f"filters={'off' if V63_DISABLE_FILTERS else 'on'})..."
            )
            p_v63_raw, excl63_flags, excl63_reasons, snap63_raw = _predict_v63_probs(
                test_rows,
                V63_MONITOR_SNAPSHOTS,
                apply_filters=(not V63_DISABLE_FILTERS),
            )

        if run_v63_m27:
            print(
                f"[Q4_ROI_62_63] predicciones v6.3 m27 (filters={'off' if V63_DISABLE_FILTERS else 'on'})..."
            )
            p_v63_27, excl63_27_flags, excl63_27_reasons, snap63_27 = _predict_v63_probs_for_snapshot(
                test_rows,
                27,
                apply_filters=(not V63_DISABLE_FILTERS),
            )

        if run_v63_m30:
            print(
                f"[Q4_ROI_62_63] predicciones v6.3 m30 (filters={'off' if V63_DISABLE_FILTERS else 'on'})..."
            )
            p_v63_30, excl63_30_flags, excl63_30_reasons, snap63_30 = _predict_v63_probs_for_snapshot(
                test_rows,
                30,
                apply_filters=(not V63_DISABLE_FILTERS),
            )

        if run_v63_m27_probe:
            print(
                f"[Q4_ROI_62_63] predicciones v6.3 m27 probe sin post-filtros live "
                f"(filters={'off' if V63_DISABLE_FILTERS else 'on'})..."
            )
            p_v63_27_probe, excl63_27_probe_flags, excl63_27_probe_reasons, snap63_27_probe = _predict_v63_probs_for_snapshot_mode(
                test_rows,
                27,
                apply_filters=(not V63_DISABLE_FILTERS),
                use_live_filters=False,
            )

        if run_v63_m30_probe:
            print(
                f"[Q4_ROI_62_63] predicciones v6.3 m30 probe sin post-filtros live "
                f"(filters={'off' if V63_DISABLE_FILTERS else 'on'})..."
            )
            p_v63_30_probe, excl63_30_probe_flags, excl63_30_probe_reasons, snap63_30_probe = _predict_v63_probs_for_snapshot_mode(
                test_rows,
                30,
                apply_filters=(not V63_DISABLE_FILTERS),
                use_live_filters=False,
            )

        if run_v63_m27_v1:
            print(
                f"[Q4_ROI_62_63] predicciones m27_v1 (filters={'off' if V63_DISABLE_FILTERS else 'on'})..."
            )
            p_v63_27_v1, excl63_27_v1_flags, excl63_27_v1_reasons, snap63_27_v1 = _predict_m27_v1_probs_for_snapshot_mode(
                test_rows,
                27,
                apply_filters=(not V63_DISABLE_FILTERS),
                use_live_filters=True,
            )

        if run_v63_m27_v1_probe:
            print(
                f"[Q4_ROI_62_63] predicciones m27_v1 probe sin post-filtros live "
                f"(filters={'off' if V63_DISABLE_FILTERS else 'on'})..."
            )
            p_v63_27_v1_probe, excl63_27_v1_probe_flags, excl63_27_v1_probe_reasons, snap63_27_v1_probe = _predict_m27_v1_probs_for_snapshot_mode(
                test_rows,
                27,
                apply_filters=(not V63_DISABLE_FILTERS),
                use_live_filters=False,
            )

        pred_payload = {}
        if run_v63_raw:
            pred_payload.update({
                "p_v63_raw": p_v63_raw,
                "excl63_flags": excl63_flags,
                "excl63_reasons": excl63_reasons,
                "snap63_raw": snap63_raw,
            })
        if run_v63_m27:
            pred_payload.update({
                "p_v63_27": p_v63_27,
                "excl63_27_flags": excl63_27_flags,
                "excl63_27_reasons": excl63_27_reasons,
                "snap63_27": snap63_27,
            })
        if run_v63_m30:
            pred_payload.update({
                "p_v63_30": p_v63_30,
                "excl63_30_flags": excl63_30_flags,
                "excl63_30_reasons": excl63_30_reasons,
                "snap63_30": snap63_30,
            })
        if run_v63_m27_probe:
            pred_payload.update({
                "p_v63_27_probe": p_v63_27_probe,
                "excl63_27_probe_flags": excl63_27_probe_flags,
                "excl63_27_probe_reasons": excl63_27_probe_reasons,
                "snap63_27_probe": snap63_27_probe,
            })
        if run_v63_m30_probe:
            pred_payload.update({
                "p_v63_30_probe": p_v63_30_probe,
                "excl63_30_probe_flags": excl63_30_probe_flags,
                "excl63_30_probe_reasons": excl63_30_probe_reasons,
                "snap63_30_probe": snap63_30_probe,
            })
        if run_v63_m27_v1:
            pred_payload.update({
                "p_v63_27_v1": p_v63_27_v1,
                "excl63_27_v1_flags": excl63_27_v1_flags,
                "excl63_27_v1_reasons": excl63_27_v1_reasons,
                "snap63_27_v1": snap63_27_v1,
            })
        if run_v63_m27_v1_probe:
            pred_payload.update({
                "p_v63_27_v1_probe": p_v63_27_v1_probe,
                "excl63_27_v1_probe_flags": excl63_27_v1_probe_flags,
                "excl63_27_v1_probe_reasons": excl63_27_v1_probe_reasons,
                "snap63_27_v1_probe": snap63_27_v1_probe,
            })
        if include_v62:
            pred_payload.update({
                "p_v62": p_v62,
                "excl62_flags": excl62_flags,
                "excl62_reasons": excl62_reasons,
            })
        _save_pred_cache(cache_meta, pred_payload)
        print(f"[Q4_ROI_62_63] cache de predicciones guardado en: {PRED_CACHE_PATH}")

    if V63_APPLY_MANUAL_BLACKLIST:
        v63_blacklist_flags, v63_blacklist_reasons = _build_blacklist_flags(test_rows)
        print(
            "[Q4_ROI_62_63] v6.3 manual blacklist "
            f"excluded={sum(1 for x in v63_blacklist_flags if x)} total_test={len(test_rows)}"
        )
    else:
        v63_blacklist_flags = [False] * len(test_rows)
        v63_blacklist_reasons = [None] * len(test_rows)
        print("[Q4_ROI_62_63] v6.3 manual blacklist disabled")
    # y_val = [int(s.target_q4) for s in val_rows]
    # p_v63_iso = _calibrate_probs(p_v63_val_raw, y_val, p_v63_raw, "isotonic")
    # p_v63_platt = _calibrate_probs(p_v63_val_raw, y_val, p_v63_raw, "platt")

    if run_v63_raw:
        if V63_APPLY_MANUAL_BLACKLIST:
            excl63_raw_flags, excl63_raw_reasons = _merge_exclusions(
                excl63_flags,
                excl63_reasons,
                v63_blacklist_flags,
                v63_blacklist_reasons,
            )
        else:
            excl63_raw_flags = excl63_flags
            excl63_raw_reasons = excl63_reasons

    if V63_APPLY_MANUAL_BLACKLIST:
        if run_v63_m27:
            excl63_27_flags, excl63_27_reasons = _merge_exclusions(
                excl63_27_flags,
                excl63_27_reasons,
                v63_blacklist_flags,
                v63_blacklist_reasons,
            )
        if run_v63_m27_probe:
            excl63_27_probe_flags, excl63_27_probe_reasons = _merge_exclusions(
                excl63_27_probe_flags,
                excl63_27_probe_reasons,
                v63_blacklist_flags,
                v63_blacklist_reasons,
            )
        if run_v63_m30:
            excl63_30_flags, excl63_30_reasons = _merge_exclusions(
                excl63_30_flags,
                excl63_30_reasons,
                v63_blacklist_flags,
                v63_blacklist_reasons,
            )
        if run_v63_m30_probe:
            excl63_30_probe_flags, excl63_30_probe_reasons = _merge_exclusions(
                excl63_30_probe_flags,
                excl63_30_probe_reasons,
                v63_blacklist_flags,
                v63_blacklist_reasons,
            )
        if run_v63_m27_v1:
            excl63_27_v1_flags, excl63_27_v1_reasons = _merge_exclusions(
                excl63_27_v1_flags,
                excl63_27_v1_reasons,
                v63_blacklist_flags,
                v63_blacklist_reasons,
            )
        if run_v63_m27_v1_probe:
            excl63_27_v1_probe_flags, excl63_27_v1_probe_reasons = _merge_exclusions(
                excl63_27_v1_probe_flags,
                excl63_27_v1_probe_reasons,
                v63_blacklist_flags,
                v63_blacklist_reasons,
            )

    if run_v63_m27_v1:
        m27_v1_policy_flags, m27_v1_policy_reasons = _build_m27_v1_policy_flags(
            test_rows
        )
        excl63_27_v1_blacklist_flags, excl63_27_v1_blacklist_reasons = _merge_exclusions(
            excl63_27_v1_flags,
            excl63_27_v1_reasons,
            m27_v1_policy_flags,
            m27_v1_policy_reasons,
        )

    print("[Q4_ROI_62_63] simulación match-by-match...")
    if include_v62:
        v62_details, v62_summary = _simulate(
            "v6.2",
            test_rows,
            p_v62,
            V62_BASE_MINUTE,
            excl62_flags,
            excl62_reasons,
            MODE,
            KELLY_MULT,
            KELLY_CAP,
            MIN_CONF_PROB,
            STAKE_STEP,
            MIN_STAKE,
            MAX_STAKE,
            teams_map,
            q4_scores_map,
            q3_min27_scores_map,
        )
    if run_v63_raw:
        v63_raw_details, v63_raw_summary = _simulate(
            "v6.3_raw_no_filter",
            test_rows,
            p_v63_raw,
            V63_SNAPSHOT_MINUTE,
            excl63_raw_flags,
            excl63_raw_reasons,
            MODE,
            KELLY_MULT,
            KELLY_CAP,
            MIN_CONF_PROB,
            STAKE_STEP,
            MIN_STAKE,
            MAX_STAKE,
            teams_map,
            q4_scores_map,
            q3_min27_scores_map,
            snap63_raw,
        )
    # TEMPORALMENTE DESHABILITADAS: iso y platt
    # v63_iso_details, v63_iso_summary = _simulate(...)
    # v63_platt_details, v63_platt_summary = _simulate(...)
    if run_v63_m27:
        v63_m27_details, v63_m27_summary = _simulate(
            "v6.3_m27_raw",
            test_rows,
            p_v63_27,
            27,
            excl63_27_flags,
            excl63_27_reasons,
            MODE,
            KELLY_MULT,
            KELLY_CAP,
            MIN_CONF_PROB,
            STAKE_STEP,
            MIN_STAKE,
            MAX_STAKE,
            teams_map,
            q4_scores_map,
            q3_min27_scores_map,
            snap63_27,
        )

    if run_v63_m30:
        v63_m30_details, v63_m30_summary = _simulate(
            "v6.3_m30_raw",
            test_rows,
            p_v63_30,
            30,
            excl63_30_flags,
            excl63_30_reasons,
            MODE,
            KELLY_MULT,
            KELLY_CAP,
            MIN_CONF_PROB,
            STAKE_STEP,
            MIN_STAKE,
            MAX_STAKE,
            teams_map,
            q4_scores_map,
            q3_min27_scores_map,
            snap63_30,
        )
    if run_v63_m27_probe:
        v63_m27_probe_details, v63_m27_probe_summary = _simulate(
            "v6.3_m27_probe",
            test_rows,
            p_v63_27_probe,
            27,
            excl63_27_probe_flags,
            excl63_27_probe_reasons,
            MODE,
            KELLY_MULT,
            KELLY_CAP,
            MIN_CONF_PROB,
            STAKE_STEP,
            MIN_STAKE,
            MAX_STAKE,
            teams_map,
            q4_scores_map,
            q3_min27_scores_map,
            snap63_27_probe,
        )
    if run_v63_m30_probe:
        v63_m30_probe_details, v63_m30_probe_summary = _simulate(
            "v6.3_m30_probe",
            test_rows,
            p_v63_30_probe,
            30,
            excl63_30_probe_flags,
            excl63_30_probe_reasons,
            MODE,
            KELLY_MULT,
            KELLY_CAP,
            MIN_CONF_PROB,
            STAKE_STEP,
            MIN_STAKE,
            MAX_STAKE,
            teams_map,
            q4_scores_map,
            q3_min27_scores_map,
            snap63_30_probe,
        )
    if run_v63_m27_v1:
        v63_m27_v1_details, v63_m27_v1_summary = _simulate(
            "m27_v1",
            test_rows,
            p_v63_27_v1,
            27,
            excl63_27_v1_flags,
            excl63_27_v1_reasons,
            MODE,
            KELLY_MULT,
            KELLY_CAP,
            MIN_CONF_PROB,
            STAKE_STEP,
            MIN_STAKE,
            MAX_STAKE,
            teams_map,
            q4_scores_map,
            q3_min27_scores_map,
            snap63_27_v1,
        )
    if run_v63_m27_v1_probe:
        v63_m27_v1_probe_details, v63_m27_v1_probe_summary = _simulate(
            "m27_v1_probe",
            test_rows,
            p_v63_27_v1_probe,
            27,
            excl63_27_v1_probe_flags,
            excl63_27_v1_probe_reasons,
            MODE,
            KELLY_MULT,
            KELLY_CAP,
            MIN_CONF_PROB,
            STAKE_STEP,
            MIN_STAKE,
            MAX_STAKE,
            teams_map,
            q4_scores_map,
            q3_min27_scores_map,
            snap63_27_v1_probe,
        )
    if run_v63_m27_v1:
        v63_m27_v1_blacklist_details, v63_m27_v1_blacklist_summary = _simulate(
            "m27_v1_blacklist",
            test_rows,
            p_v63_27_v1,
            27,
            excl63_27_v1_blacklist_flags,
            excl63_27_v1_blacklist_reasons,
            MODE,
            KELLY_MULT,
            KELLY_CAP,
            MIN_CONF_PROB,
            STAKE_STEP,
            MIN_STAKE,
            MAX_STAKE,
            teams_map,
            q4_scores_map,
            q3_min27_scores_map,
            snap63_27_v1,
        )

    if run_v63_raw:
        v63_raw_df = pd.DataFrame(v63_raw_details)
    if run_v63_m27:
        v63_m27_df = pd.DataFrame(v63_m27_details)
    if run_v63_m30:
        v63_m30_df = pd.DataFrame(v63_m30_details)
    if run_v63_m27_probe:
        v63_m27_probe_df = pd.DataFrame(v63_m27_probe_details)
    if run_v63_m30_probe:
        v63_m30_probe_df = pd.DataFrame(v63_m30_probe_details)
    if run_v63_m27_v1:
        v63_m27_v1_df = pd.DataFrame(v63_m27_v1_details)
    if run_v63_m27_v1_probe:
        v63_m27_v1_probe_df = pd.DataFrame(v63_m27_v1_probe_details)
    if run_v63_m27_v1:
        v63_m27_v1_blacklist_df = pd.DataFrame(v63_m27_v1_blacklist_details)

    col_order = [
        "fecha", "liga", "resultado_apuesta", "apuesta", "monto_apostado", "ganancia", "bank_final",
        "modelo", "partida_test", "match_id", "hora", "equipo_local",
        "equipo_visitante", "resultado_q4_home_gana", "marcador_q4", "prob_local", "prob_visitante",
        "lado_predicho", "confianza_prob", "confianza_score_0_100", "nivel_confianza",
        "apuestas_odds", "probabilidad_empate", "edge", "kelly_fraction_raw",
        "kelly_fraction_used", "step_apuesta", "razon_sin_apuesta", "pnl",
        "banco_antes", "ganancia_acumulada", "roi_banco_acumulado", "minuto_base_apuesta", "marcador_q3_min27",
    ]
    if run_v63_raw:
        v63_raw_df = v63_raw_df[[c for c in col_order if c in v63_raw_df.columns]]
    if run_v63_m27:
        v63_m27_df = v63_m27_df[[c for c in col_order if c in v63_m27_df.columns]]
    if run_v63_m30:
        v63_m30_df = v63_m30_df[[c for c in col_order if c in v63_m30_df.columns]]
    if run_v63_m27_probe:
        v63_m27_probe_df = v63_m27_probe_df[[c for c in col_order if c in v63_m27_probe_df.columns]]
    if run_v63_m30_probe:
        v63_m30_probe_df = v63_m30_probe_df[[c for c in col_order if c in v63_m30_probe_df.columns]]
    if run_v63_m27_v1:
        v63_m27_v1_df = _annotate_m27_v1_policy(v63_m27_v1_df)
        v63_m27_v1_df = v63_m27_v1_df[[c for c in col_order if c in v63_m27_v1_df.columns]]
    if run_v63_m27_v1_probe:
        v63_m27_v1_probe_df = _annotate_m27_v1_policy(v63_m27_v1_probe_df)
        v63_m27_v1_probe_df = v63_m27_v1_probe_df[[c for c in col_order if c in v63_m27_v1_probe_df.columns]]
    if run_v63_m27_v1:
        v63_m27_v1_blacklist_df = _annotate_m27_v1_policy(v63_m27_v1_blacklist_df)
        v63_m27_v1_blacklist_df = v63_m27_v1_blacklist_df[
            [c for c in col_order if c in v63_m27_v1_blacklist_df.columns]
        ]

    v63_compare_df = None
    if run_v63_m27 and run_v63_m30:
        v63_compare_df = _build_v63_snapshot_comparison(
            v63_m27_details,
            v63_m30_details,
            q3_min27_scores_map,
            q3_min30_scores_map,
        )

    v63_compare_m27_v1_probe_vs_m30_probe_df = None
    if run_v63_m27_v1_probe and run_v63_m30_probe:
        v63_compare_m27_v1_probe_vs_m30_probe_df = _build_v63_snapshot_comparison(
            v63_m27_v1_probe_details,
            v63_m30_probe_details,
            q3_min27_scores_map,
            q3_min30_scores_map,
        )

    summaries = []
    league_dfs = []
    sheets_extra = []
    m27_v1_policy_df = None
    if run_v63_raw:
        summaries.append(v63_raw_summary)
        league_dfs.append(v63_raw_df)
        sheets_extra.append(("v6_3_raw_nf_matches", v63_raw_df))
    if run_v63_m27:
        summaries.append(v63_m27_summary)
        league_dfs.append(v63_m27_df)
        sheets_extra.append(("v6_3_m27_matches", v63_m27_df))
    if run_v63_m30:
        summaries.append(v63_m30_summary)
        league_dfs.append(v63_m30_df)
        sheets_extra.append(("v6_3_m30_matches", v63_m30_df))
    if run_v63_m27_probe:
        summaries.append(v63_m27_probe_summary)
        league_dfs.append(v63_m27_probe_df)
        sheets_extra.append(("v6_3_m27_probe_matches", v63_m27_probe_df))
    if run_v63_m30_probe:
        summaries.append(v63_m30_probe_summary)
        league_dfs.append(v63_m30_probe_df)
        sheets_extra.append(("v6_3_m30_probe_matches", v63_m30_probe_df))
    if run_v63_m27_v1:
        m27_v1_policy_df = _build_m27_v1_policy_sheet(v63_m27_v1_df)
        summaries.append(v63_m27_v1_summary)
        summaries.append(v63_m27_v1_blacklist_summary)
        league_dfs.append(v63_m27_v1_df)
        league_dfs.append(v63_m27_v1_blacklist_df)
        sheets_extra.append(("m27_v1_matches", v63_m27_v1_df))
        sheets_extra.append(("m27_v1_blacklist_matches", v63_m27_v1_blacklist_df))
        sheets_extra.append(("m27_v1_policy", m27_v1_policy_df))
    if run_v63_m27_v1_probe:
        summaries.append(v63_m27_v1_probe_summary)
        league_dfs.append(v63_m27_v1_probe_df)
        sheets_extra.append(("m27_v1_probe_matches", v63_m27_v1_probe_df))
    if v63_compare_df is not None:
        sheets_extra.append(("v6_3_m27_vs_m30", v63_compare_df))
    if v63_compare_m27_v1_probe_vs_m30_probe_df is not None:
        sheets_extra.append((
            "m27_v1_probe_vs_m30_probe",
            v63_compare_m27_v1_probe_vs_m30_probe_df,
        ))

    if include_v62:
        v62_df = pd.DataFrame(v62_details)
        v62_df = v62_df[[c for c in col_order if c in v62_df.columns]]
        summaries = [v62_summary] + summaries
        league_dfs = [v62_df] + league_dfs
        sheets_extra = [("v6_2_matches", v62_df)] + sheets_extra

    summary_df = pd.DataFrame(summaries)
    eff_league_df = _build_effectiveness_by_league(league_dfs)

    summary_order = [
        "modelo", "partidos_test", "apuestas", "ganadas", "perdidas",
        "empates_contados_como_perdida", "sin_apuesta", "efectividad",
        "banco_inicio", "banco_final", "ganancia", "roi_bank",
        "total_apostado", "apuesta_promedio", "yield_sobre_apostado", "max_drawdown",
        "minuto_base_apuesta",
    ]
    summary_df = summary_df[[c for c in summary_order if c in summary_df.columns]]

    print("[Q4_ROI_62_63] escribiendo Excel...")
    with pd.ExcelWriter(out_mm_path, engine="openpyxl") as writer:
        write_items = [("summary", summary_df), ("efectividad_liga", eff_league_df)] + sheets_extra
        for sheet_name, df in tqdm(write_items, desc="[Q4_ROI_62_63] escribiendo hojas", unit="hoja"):
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        for sheet_name in tqdm(list(writer.sheets), desc="[Q4_ROI_62_63] formateando hojas", unit="hoja"):
            ws = writer.sheets[sheet_name]
            _apply_excel_formatting(ws)

    print("[Q4_ROI_62_63] OK")
    print(f"[Q4_ROI_62_63] output={out_mm_path}")
    print("[Q4_ROI_62_63] abriendo Excel...")
    os.startfile(str(out_mm_path))


if __name__ == "__main__":
    main()
