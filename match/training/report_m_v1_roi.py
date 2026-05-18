"""Generate an independent Q4 ROI report for M27_V1 and M30_V1.

This report intentionally excludes V6.2, V6.3 raw, and the legacy m27/m30
variants. It only evaluates the independent M_V1 models.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import sqlite3
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path

import joblib
import numpy as np
import pandas as pd
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from tqdm import tqdm

import infer_match as infer_live
import m27_v1_league_policy as m27_v1_policy
import train_q3_q4_models_v6 as v6
import train_q4_m27_v1 as m27_v1_train
import train_q4_m30_v1 as m30_v1_train

ROOT = Path(__file__).parent.parent.parent
BASE_M_V1 = ROOT / "match" / "training" / "model_outputs_m_v1"
BASE_M27_V1 = ROOT / "match" / "training" / "model_outputs_m27_v1"
BASE_M30_V1 = ROOT / "match" / "training" / "model_outputs_m30_v1"
OUT_MM = BASE_M_V1 / "Q4_ROI_match_by_match_m_v1.xlsx"
PRED_CACHE_PATH = BASE_M_V1 / "q4_roi_pred_cache.joblib"
SPLITS_CACHE_PATH = BASE_M_V1 / "q4_roi_splits_cache.joblib"
M27_V1_CHAMPION_PATH = BASE_M27_V1 / "q4_m27_v1_champion.joblib"
M30_V1_CHAMPION_PATH = BASE_M30_V1 / "q4_m30_v1_champion.joblib"

ODDS = 1.4
BANK_START = 1000.0
BREAK_EVEN = 1.0 / ODDS
MODE = "kelly_non_compound"
KELLY_MULT = 0.25
KELLY_CAP = 0.05
MIN_CONF_PROB = 0.58
STAKE_STEP = 25.0
MIN_STAKE = 25.0
MAX_STAKE = 100.0
STAKE_BUCKETS = (25.0, 50.0, 75.0, 100.0)
LATE_WINDOW_MAX_MARGIN = {
    27: 999,
    30: 6,
    33: 8,
    36: 10,
}
LEAGUE_NAME_FILTERS_DISABLED = True

_FAST_COMPLETE_Q4_SQL = """
    SELECT m.date, m.time, m.match_id
    FROM matches m
    WHERE EXISTS (
            SELECT 1 FROM quarter_scores
            WHERE match_id = m.match_id AND quarter = 'Q4'
        )
      AND EXISTS (
            SELECT 1 FROM play_by_play
            WHERE match_id = m.match_id AND quarter = 'Q1'
        )
      AND EXISTS (
            SELECT 1 FROM play_by_play
            WHERE match_id = m.match_id AND quarter = 'Q4'
        )
      AND (
            SELECT COUNT(*) FROM graph_points
            WHERE match_id = m.match_id
        ) >= 20
    ORDER BY m.date, m.time, m.match_id
"""


def _report_date_from_dt(dt_value) -> str:
    if not dt_value:
        return ""
    dt_local = (
        dt_value if dt_value.tzinfo else dt_value.replace(tzinfo=timezone.utc)
    )
    return dt_local.astimezone(
        timezone(timedelta(hours=-6))
    ).strftime("%Y-%m-%d")


def _rows_fingerprint(rows) -> str:
    h = hashlib.sha1()
    for sample in rows:
        h.update(str(sample.match_id).encode("utf-8", errors="ignore"))
        h.update(b"|")
        h.update(
            (sample.dt.isoformat() if sample.dt else "").encode(
                "utf-8",
                errors="ignore",
            )
        )
        h.update(b"|")
        h.update(
            str(
                int(sample.target_q4) if sample.target_q4 is not None else -1
            ).encode("utf-8", errors="ignore")
        )
        h.update(b"\n")
    return h.hexdigest()


def _file_signature(path: Path) -> str:
    if not path.exists():
        return "missing"
    st = path.stat()
    return f"{int(st.st_mtime_ns)}:{int(st.st_size)}"


def _load_pred_cache(
    cache_meta: dict,
    force_rebuild: bool = False,
) -> dict | None:
    if force_rebuild or not PRED_CACHE_PATH.exists():
        return None
    try:
        payload = joblib.load(PRED_CACHE_PATH)
    except Exception:
        return None
    if not isinstance(payload, dict):
        return None
    if payload.get("meta") != cache_meta:
        return None
    return payload.get("predictions")


def _save_pred_cache(cache_meta: dict, predictions: dict) -> None:
    PRED_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    joblib.dump(
        {"meta": cache_meta, "predictions": predictions},
        PRED_CACHE_PATH,
    )


def _kelly_fraction(p, odds):
    b = odds - 1.0
    q = 1.0 - p
    return ((b * p) - q) / b if b > 0 else 0.0


def _round_down_step(x, step):
    if step <= 0:
        return x
    return (x // step) * step


def _stake_from_signal(p_pick, edge, k_used, kelly_cap=0.05):
    if k_used is None or k_used <= 0:
        return 0.0
    denom = kelly_cap if kelly_cap and kelly_cap > 0 else 0.05
    k_strength = max(0.0, min(k_used / denom, 1.0))
    p_pick = float(p_pick or 0.0)
    edge = float(edge or 0.0)
    if p_pick >= 0.97 and edge >= 0.24 and k_strength >= 0.995:
        return STAKE_BUCKETS[3]
    if p_pick >= 0.90 and edge >= 0.18 and k_strength >= 0.85:
        return STAKE_BUCKETS[2]
    if p_pick >= 0.80 and edge >= 0.09 and k_strength >= 0.55:
        return STAKE_BUCKETS[1]
    if k_strength < 0.15:
        return 0.0
    return STAKE_BUCKETS[0]


def _prepare_splits(force_rebuild: bool = False, only_date: str | None = None):
    if only_date:
        print(
            "[M_V1_ROI] cargando solo partidos no vistos de la fecha "
            f"{only_date}..."
        )
        samples = v6._build_samples(v6.DB_PATH, date_gte=only_date)
        rows = sorted(
            [
                sample
                for sample in samples
                if sample.target_q4 is not None
                and _report_date_from_dt(sample.dt) == only_date
            ],
            key=lambda sample: sample.dt,
        )
        if not rows:
            raise ValueError(
                "No se encontraron partidos no vistos para la fecha "
                f"{only_date}"
            )
        print(
            "[M_V1_ROI] modo fecha específica "
            f"fecha={only_date} test={len(rows)}"
        )
        return [], [], rows

    db_sig = _file_signature(v6.DB_PATH)
    cache_key = {"db_sig": db_sig, "split_ratios": (0.70, 0.15, 0.15)}
    if not force_rebuild and SPLITS_CACHE_PATH.exists():
        try:
            payload = joblib.load(SPLITS_CACHE_PATH)
            if isinstance(payload, dict) and payload.get("meta") == cache_key:
                val_rows = payload["val_rows"]
                test_rows = payload["test_rows"]
                print(
                    f"[M_V1_ROI] splits desde cache "
                    f"({SPLITS_CACHE_PATH.name}) val={len(val_rows)} "
                    f"test={len(test_rows)}"
                )
                return [], val_rows, test_rows
        except Exception:
            pass

    print("[M_V1_ROI] sondeo rápido de fechas de corte Q4...")
    probe_conn = sqlite3.connect(str(v6.DB_PATH))
    probe_conn.row_factory = sqlite3.Row
    probe_rows = probe_conn.execute(_FAST_COMPLETE_Q4_SQL).fetchall()
    probe_conn.close()

    n_approx = len(probe_rows)
    n_train_approx = int(n_approx * 0.70)
    n_val_approx = int(n_approx * 0.15)
    val_start_date = str(probe_rows[n_train_approx]["date"])
    samples = v6._build_samples(v6.DB_PATH, date_gte=val_start_date)
    rows = sorted(
        [sample for sample in samples if sample.target_q4 is not None],
        key=lambda sample: sample.dt,
    )
    val_rows = rows[:n_val_approx]
    test_rows = rows[n_val_approx:]

    SPLITS_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    joblib.dump(
        {"meta": cache_key, "val_rows": val_rows, "test_rows": test_rows},
        SPLITS_CACHE_PATH,
    )
    print(
        f"[M_V1_ROI] splits guardados en cache: {SPLITS_CACHE_PATH.name}"
    )
    return [], val_rows, test_rows


def _load_match_teams_map():
    conn = sqlite3.connect(v6.DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            "SELECT match_id, home_team, away_team FROM matches"
        ).fetchall()
    finally:
        conn.close()
    out = {}
    for row in rows:
        out[str(row["match_id"])] = {
            "home_team": str(row["home_team"] or ""),
            "away_team": str(row["away_team"] or ""),
        }
    return out


def _load_match_q4_scores_map(match_ids):
    conn = v6.db_mod.get_conn(str(v6.DB_PATH))
    v6.db_mod.init_db(conn)
    out = {}
    try:
        for match_id in match_ids:
            data = v6.db_mod.get_match(conn, str(match_id))
            if not data:
                out[str(match_id)] = ""
                continue
            q4h, q4a = v6._quarter_points(data, "Q4")
            if q4h is None or q4a is None:
                out[str(match_id)] = ""
            else:
                out[str(match_id)] = f"{int(q4h)}-{int(q4a)}"
    finally:
        conn.close()
    return out


def _load_match_score_at_minute_map(
    match_ids,
    minute,
    score_upto_fn=None,
):
    conn = v6.db_mod.get_conn(str(v6.DB_PATH))
    v6.db_mod.init_db(conn)
    out = {}
    try:
        for match_id in match_ids:
            data = v6.db_mod.get_match(conn, str(match_id))
            if not data:
                out[str(match_id)] = ""
                continue
            if score_upto_fn is not None:
                home, away = score_upto_fn(data, float(minute))
            else:
                pbp = (
                    data.get("play_by_play", {})
                    if isinstance(data, dict)
                    else {}
                )
                events = v6._pbp_events_upto_minute(pbp, float(minute))
                home = 0
                away = 0
                for event in events:
                    team = str(event.get("team", ""))
                    pts = int(event.get("points", 0) or 0)
                    if team == "home":
                        home += pts
                    elif team == "away":
                        away += pts
            out[str(match_id)] = f"{home}-{away}"
    finally:
        conn.close()
    return out


def _load_league_name_rules():
    cfg_path = ROOT / "match" / "training" / "v6_2_league_name_exclusions.json"
    cfg = json.loads(cfg_path.read_text(encoding="utf-8"))
    patterns = []
    for category in cfg.get("categories", []):
        for pattern in category.get("patterns", []):
            pattern = str(pattern).strip()
            if pattern:
                patterns.append(
                    (
                        category.get("name", "uncategorized"),
                        pattern,
                        pattern.lower(),
                    )
                )
    return patterns


def _clip_pbp_to_minute(match_data: dict, cutoff_minute: int) -> dict:
    pbp = match_data.get("play_by_play") or {}
    out = {"Q1": [], "Q2": [], "Q3": []}
    cutoff = float(cutoff_minute)
    for quarter_label, plays in pbp.items():
        q_idx = infer_live._quarter_index(str(quarter_label))
        if q_idx is None or q_idx < 1 or q_idx > 3:
            continue
        q_start = (q_idx - 1) * 12.0
        kept = []
        for play in plays or []:
            rem_sec = infer_live._clock_to_seconds(
                str(play.get("time", "") or "")
            )
            if rem_sec is None:
                continue
            global_min = q_start + (12.0 - rem_sec / 60.0)
            if global_min <= cutoff + 1e-9:
                kept.append(play)
        out[f"Q{q_idx}"] = kept
    return out


def _clip_pbp_to_minute_m30(match_data: dict, cutoff_minute: int) -> dict:
    pbp = match_data.get("play_by_play") or {}
    quarter_minutes = m30_v1_train._infer_regulation_quarter_minutes(
        match_data
    )
    out = {"Q1": [], "Q2": [], "Q3": []}
    cutoff = float(cutoff_minute)
    for quarter_label, plays in pbp.items():
        q_idx = infer_live._quarter_index(str(quarter_label))
        if q_idx is None or q_idx < 1 or q_idx > 3:
            continue
        q_start = (q_idx - 1) * quarter_minutes
        kept = []
        for play in plays or []:
            rem_sec = infer_live._clock_to_seconds(
                str(play.get("time", "") or "")
            )
            if rem_sec is None or rem_sec > int(quarter_minutes * 60.0):
                continue
            global_min = q_start + (quarter_minutes - rem_sec / 60.0)
            if global_min <= cutoff + 1e-9:
                kept.append(play)
        out[f"Q{q_idx}"] = kept
    return out


def _build_live_like_q4_snapshot(
    match_data: dict,
    snapshot_minute: int,
) -> dict:
    clipped = dict(match_data)
    clipped["graph_points"] = [
        point
        for point in (match_data.get("graph_points") or [])
        if int(point.get("minute", 0)) <= int(snapshot_minute)
    ]
    clipped["play_by_play"] = _clip_pbp_to_minute(match_data, snapshot_minute)
    score_obj = dict(match_data.get("score") or {})
    quarters = dict(score_obj.get("quarters") or {})
    q1h, q1a = infer_live._quarter_points(match_data, "Q1")
    q2h, q2a = infer_live._quarter_points(match_data, "Q2")
    ht_home = int(q1h or 0) + int(q2h or 0)
    ht_away = int(q1a or 0) + int(q2a or 0)
    game_home, game_away = infer_live._score_upto(
        clipped,
        int(snapshot_minute),
    )
    q3_home = max(0, int(game_home) - ht_home)
    q3_away = max(0, int(game_away) - ht_away)
    quarters["Q3"] = {"home": q3_home, "away": q3_away}
    score_obj["quarters"] = quarters
    clipped["score"] = score_obj
    return clipped


def _build_live_like_q4_snapshot_m30(
    match_data: dict,
    snapshot_minute: int,
) -> dict:
    clipped = dict(match_data)
    clipped["graph_points"] = [
        point
        for point in (match_data.get("graph_points") or [])
        if int(point.get("minute", 0)) <= int(snapshot_minute)
    ]
    clipped["play_by_play"] = _clip_pbp_to_minute_m30(
        match_data,
        snapshot_minute,
    )
    score_obj = dict(match_data.get("score") or {})
    quarters = dict(score_obj.get("quarters") or {})
    q1h, q1a = infer_live._quarter_points(match_data, "Q1")
    q2h, q2a = infer_live._quarter_points(match_data, "Q2")
    ht_home = int(q1h or 0) + int(q2h or 0)
    ht_away = int(q1a or 0) + int(q2a or 0)
    game_home, game_away = m30_v1_train._score_upto_m30(
        clipped,
        float(snapshot_minute),
    )
    q3_home = max(0, int(game_home) - ht_home)
    q3_away = max(0, int(game_away) - ht_away)
    quarters["Q3"] = {"home": q3_home, "away": q3_away}
    score_obj["quarters"] = quarters
    clipped["score"] = score_obj
    return clipped


def _lead_changes_upto_m30(data: dict, cutoff_minute: int) -> int:
    events = m30_v1_train._pbp_events_upto_m30(data, float(cutoff_minute))
    changes = 0
    prev_sign = 0
    for event in events:
        hs = event.get("home_score")
        as_ = event.get("away_score")
        if hs is None or as_ is None:
            continue
        diff = int(hs) - int(as_)
        sign = 1 if diff > 0 else (-1 if diff < 0 else 0)
        if sign == 0:
            continue
        if prev_sign != 0 and sign != prev_sign:
            changes += 1
        prev_sign = sign
    return changes


def _volatility_index_m30(data: dict, cutoff_minute: int) -> dict:
    gp = data.get("graph_points", [])
    vals = [
        int(point.get("value", 0))
        for point in gp
        if int(point.get("minute", 0)) <= cutoff_minute
    ]
    swings = infer_live._count_sign_swings(vals)
    lead_changes = _lead_changes_upto_m30(data, cutoff_minute)
    swings_norm = min(swings / 10.0, 1.0)
    leads_norm = min(lead_changes / 12.0, 1.0)
    index = round((0.6 * swings_norm) + (0.4 * leads_norm), 6)
    return {
        "index": index,
        "swings": swings,
        "lead_changes": lead_changes,
    }


def _decision_gate_m30(
    match_data: dict,
    target: str,
    snapshot_minute: int | None,
    confidence: float,
    model_signal: str,
) -> dict:
    cutoff = snapshot_minute if snapshot_minute is not None else 36
    gp_count = len([
        point
        for point in match_data.get("graph_points", [])
        if int(point.get("minute", 0)) <= cutoff
    ])
    pbp_count = len(
        m30_v1_train._pbp_events_upto_m30(match_data, float(cutoff))
    )

    thr = infer_live._sufficiency_thresholds(target, snapshot_minute)
    if gp_count < thr["min_graph_points"] or pbp_count < thr["min_pbp_events"]:
        return {
            "decision_gate": "BLOCK_LOW_DATA",
            "final_recommendation": "NO BET",
            "reason": "insufficient_graph_or_pbp_coverage",
            "gp_count": gp_count,
            "pbp_count": pbp_count,
            "volatility_index": None,
            "volatility_swings": None,
            "volatility_lead_changes": None,
        }

    gate_thr = infer_live._gate_thresholds(target, snapshot_minute)
    vol = _volatility_index_m30(match_data, cutoff)
    if vol["index"] >= float(gate_thr["volatility_block_at"]):
        return {
            "decision_gate": "BLOCK_HIGH_VOLATILITY",
            "final_recommendation": "NO BET",
            "reason": "match_too_volatile_for_current_signal",
            "gp_count": gp_count,
            "pbp_count": pbp_count,
            "volatility_index": vol["index"],
            "volatility_swings": vol["swings"],
            "volatility_lead_changes": vol["lead_changes"],
        }

    if confidence < float(gate_thr["min_edge"]):
        return {
            "decision_gate": "BLOCK_LOW_EDGE",
            "final_recommendation": "NO BET",
            "reason": "confidence_below_minimum_edge",
            "gp_count": gp_count,
            "pbp_count": pbp_count,
            "volatility_index": vol["index"],
            "volatility_swings": vol["swings"],
            "volatility_lead_changes": vol["lead_changes"],
        }

    return {
        "decision_gate": f"ALLOW_{model_signal.replace(' ', '_')}",
        "final_recommendation": model_signal,
        "reason": "passed_all_gates",
        "gp_count": gp_count,
        "pbp_count": pbp_count,
        "volatility_index": vol["index"],
        "volatility_swings": vol["swings"],
        "volatility_lead_changes": vol["lead_changes"],
    }


def _policy_snapshot_minute_for_model(
    model_label: str,
    match_data: dict,
    snapshot_minute: int,
) -> int:
    if model_label != "m30_v1":
        return int(snapshot_minute)
    quarter_minutes = m30_v1_train._infer_regulation_quarter_minutes(
        match_data
    )
    if quarter_minutes <= 10.0 and int(snapshot_minute) == 30:
        return 36
    return int(snapshot_minute)


def _bet_candidate(p_home: float | None) -> dict:
    if p_home is None:
        return {"accept": False, "reason": "missing_prob"}
    p_pick = p_home if p_home >= 0.5 else (1.0 - p_home)
    edge = p_pick - BREAK_EVEN
    if p_pick < MIN_CONF_PROB:
        return {"accept": False, "reason": "confidence_filter"}
    k_raw = _kelly_fraction(p_pick, ODDS)
    if k_raw <= 0:
        return {"accept": False, "reason": "no_edge"}
    k_used = min(k_raw * KELLY_MULT, KELLY_CAP)
    stake = _stake_from_signal(p_pick, edge, k_used, KELLY_CAP)
    stake = min(stake, MAX_STAKE)
    stake = _round_down_step(stake, STAKE_STEP)
    if stake < MIN_STAKE:
        return {"accept": False, "reason": "stake_below_25"}
    return {"accept": True, "reason": "accepted"}


def _snapshot_window_eligibility(
    match_data: dict,
    snapshot_minute: int,
    p_home: float,
    gate: dict,
    score_upto_fn=None,
    policy_snapshot_minute=None,
) -> tuple[bool, str, int]:
    final_rec = str(gate.get("final_recommendation", "NO BET")).upper()
    if final_rec == "NO BET":
        return False, f"gate:{gate.get('reason', 'no_bet')}", 0
    if score_upto_fn is None:
        game_home, game_away = infer_live._score_upto(
            match_data,
            int(snapshot_minute),
        )
    else:
        game_home, game_away = score_upto_fn(match_data, int(snapshot_minute))
    margin = abs(int(game_home) - int(game_away))
    snapshot_policy = (
        int(policy_snapshot_minute)
        if policy_snapshot_minute is not None
        else int(snapshot_minute)
    )
    if snapshot_policy > 27:
        max_margin = int(LATE_WINDOW_MAX_MARGIN.get(snapshot_policy, 6))
        if margin > max_margin:
            return (
                False,
                f"late_window_close_game_required:{margin}>{max_margin}",
                margin,
            )
    return True, "eligible", margin


def _predict_m_v1_probs_for_snapshot_mode(
    test_rows,
    model_label: str,
    artifact_path: Path,
    feature_builder,
    snapshot_minute: int,
    apply_filters=True,
):
    rules = _load_league_name_rules() if apply_filters else []
    artifact = joblib.load(artifact_path)
    vectorizer = artifact["vectorizer"]
    models = artifact["models"]

    conn = v6.db_mod.get_conn(str(v6.DB_PATH))
    v6.db_mod.init_db(conn)
    probs = [None] * len(test_rows)
    excluded_flags = [False] * len(test_rows)
    excluded_reasons = [None] * len(test_rows)
    selected_snapshots = [int(snapshot_minute)] * len(test_rows)
    gate_reasons_count = defaultdict(int)
    accepted_probs = []
    rejected_probs = []
    try:
        for i, sample in enumerate(
            tqdm(
                test_rows,
                desc=f"[{model_label}] predicciones m{snapshot_minute}",
            )
        ):
            league = str(sample.features_q4.get("league", ""))
            hit = None
            for category, raw, low in rules:
                if low in league.lower():
                    hit = (category, raw)
                    break
            if apply_filters and hit is not None:
                excluded_flags[i] = True
                excluded_reasons[i] = f"excluded_league_name:{hit[0]}:{hit[1]}"
                continue

            match_data = v6.db_mod.get_match(conn, str(sample.match_id))
            if not match_data:
                excluded_flags[i] = True
                excluded_reasons[i] = "missing_match_data"
                continue

            if model_label == "m30_v1":
                live_like = _build_live_like_q4_snapshot_m30(
                    match_data,
                    int(snapshot_minute),
                )
                score_upto_fn = m30_v1_train._score_upto_m30
            else:
                live_like = _build_live_like_q4_snapshot(
                    match_data,
                    int(snapshot_minute),
                )
                score_upto_fn = None
            policy_snapshot_minute = _policy_snapshot_minute_for_model(
                model_label,
                match_data,
                int(snapshot_minute),
            )

            feature_dict = feature_builder(sample, live_like)
            x_row = vectorizer.transform([feature_dict])
            p_xgb = float(models["xgb"].predict_proba(x_row)[0, 1])
            p_hist = float(models["hist_gb"].predict_proba(x_row)[0, 1])
            p_home = (p_xgb + p_hist) / 2.0

            confidence = abs(p_home - 0.5) * 2.0
            bet_signal = infer_live._bet_signal(
                "q4",
                confidence,
                int(policy_snapshot_minute),
            )
            if model_label == "m30_v1":
                gate = _decision_gate_m30(
                    match_data=live_like,
                    target="q4",
                    snapshot_minute=int(policy_snapshot_minute),
                    confidence=confidence,
                    model_signal=bet_signal["signal"],
                )
            else:
                gate = infer_live._decision_gate(
                    match_data=live_like,
                    target="q4",
                    snapshot_minute=int(policy_snapshot_minute),
                    confidence=confidence,
                    model_signal=bet_signal["signal"],
                )
            eligible, eligible_reason, _margin = _snapshot_window_eligibility(
                live_like,
                int(snapshot_minute),
                p_home,
                gate,
                score_upto_fn=score_upto_fn,
                policy_snapshot_minute=policy_snapshot_minute,
            )
            if not eligible:
                excluded_reasons[i] = eligible_reason
                gate_reasons_count[eligible_reason] += 1
                rejected_probs.append(p_home)
                continue

            cand = _bet_candidate(p_home)
            if cand.get("accept"):
                probs[i] = p_home
                accepted_probs.append(p_home)
            else:
                reason = str(cand.get("reason", "no_signal_in_snapshot"))
                excluded_reasons[i] = reason
                gate_reasons_count[reason] += 1
                rejected_probs.append(p_home)
    finally:
        conn.close()

    print(
        f"\n[DEBUG] {model_label} m{snapshot_minute} "
        "gate rejection reasons:"
    )
    for reason, count in sorted(
        gate_reasons_count.items(),
        key=lambda item: -item[1],
    ):
        print(f"  {reason}: {count}")
    if accepted_probs:
        print(
            f"[DEBUG] {model_label} ACCEPTED (n={len(accepted_probs)}): "
            f"mean={np.mean(accepted_probs):.3f} "
            f"min={np.min(accepted_probs):.3f} "
            f"max={np.max(accepted_probs):.3f}"
        )
    if rejected_probs:
        print(
            f"[DEBUG] {model_label} REJECTED (n={len(rejected_probs)}): "
            f"mean={np.mean(rejected_probs):.3f} "
            f"min={np.min(rejected_probs):.3f} "
            f"max={np.max(rejected_probs):.3f}"
        )
    return probs, excluded_flags, excluded_reasons, selected_snapshots


def _predict_m_v1_probs_raw(
    test_rows,
    model_label: str,
    artifact_path: Path,
    feature_builder,
    snapshot_minute: int,
):
    artifact = joblib.load(artifact_path)
    vectorizer = artifact["vectorizer"]
    models = artifact["models"]

    conn = v6.db_mod.get_conn(str(v6.DB_PATH))
    v6.db_mod.init_db(conn)
    probs = [None] * len(test_rows)
    try:
        for i, sample in enumerate(
            tqdm(
                test_rows,
                desc=f"[{model_label}] predicciones raw m{snapshot_minute}",
            )
        ):
            match_data = v6.db_mod.get_match(conn, str(sample.match_id))
            if not match_data:
                continue

            if model_label == "m30_v1":
                live_like = _build_live_like_q4_snapshot_m30(
                    match_data,
                    int(snapshot_minute),
                )
            else:
                live_like = _build_live_like_q4_snapshot(
                    match_data,
                    int(snapshot_minute),
                )

            feature_dict = feature_builder(sample, live_like)
            x_row = vectorizer.transform([feature_dict])
            p_xgb = float(models["xgb"].predict_proba(x_row)[0, 1])
            p_hist = float(models["hist_gb"].predict_proba(x_row)[0, 1])
            probs[i] = (p_xgb + p_hist) / 2.0
    finally:
        conn.close()

    return probs


def _build_raw_comparison_summary(model_name: str, test_rows, probs) -> dict:
    available = [float(prob) for prob in probs if prob is not None]
    pick_strength = [max(prob, 1.0 - prob) for prob in available]
    correct = 0
    total_with_prob = 0
    for sample, prob in zip(test_rows, probs):
        if prob is None or sample.target_q4 is None:
            continue
        total_with_prob += 1
        pick_home = float(prob) >= 0.5
        target_home = int(sample.target_q4) == 1
        correct += int(pick_home == target_home)

    return {
        "modelo": model_name,
        "partidos_test": len(test_rows),
        "predicciones_disponibles": total_with_prob,
        "accuracy_raw_pick": (
            float(correct / total_with_prob) if total_with_prob else 0.0
        ),
        "pick_strength_media": (
            float(np.mean(pick_strength)) if pick_strength else 0.0
        ),
        "pick_strength_mediana": (
            float(np.median(pick_strength)) if pick_strength else 0.0
        ),
        "pick_strength_p95": (
            float(np.percentile(pick_strength, 95)) if pick_strength else 0.0
        ),
        "conteo_p_ge_0_58": int(sum(value >= 0.58 for value in pick_strength)),
        "conteo_p_ge_0_60": int(sum(value >= 0.60 for value in pick_strength)),
        "conteo_p_ge_0_65": int(sum(value >= 0.65 for value in pick_strength)),
        "conteo_p_ge_break_even": int(
            sum(value >= BREAK_EVEN for value in pick_strength)
        ),
        "max_pick_strength": (
            float(max(pick_strength)) if pick_strength else 0.0
        ),
    }


def _build_raw_model_comparison(
    test_rows,
    probs_27,
    probs_30,
    teams_map,
    q4_scores_map,
    q3_min27_scores_map,
    q3_min30_scores_map,
):
    rows = []
    for sample, p27, p30 in zip(test_rows, probs_27, probs_30):
        match_id = str(sample.match_id)
        teams = (teams_map or {}).get(match_id, {})
        y = int(sample.target_q4) if sample.target_q4 is not None else None
        pick27 = (
            None if p27 is None else ("local" if p27 >= 0.5 else "visitante")
        )
        pick30 = (
            None if p30 is None else ("local" if p30 >= 0.5 else "visitante")
        )
        p27_pick = None if p27 is None else max(float(p27), float(1.0 - p27))
        p30_pick = None if p30 is None else max(float(p30), float(1.0 - p30))
        hit27 = None
        hit30 = None
        if y is not None and p27 is not None:
            hit27 = int((float(p27) >= 0.5) == (y == 1))
        if y is not None and p30 is not None:
            hit30 = int((float(p30) >= 0.5) == (y == 1))
        rows.append(
            {
                "fecha": _report_date_from_dt(sample.dt),
                "hora": (
                    (
                        sample.dt
                        if sample.dt.tzinfo
                        else sample.dt.replace(tzinfo=timezone.utc)
                    )
                    .astimezone(timezone(timedelta(hours=-6)))
                    .strftime("%H:%M:%S")
                    if sample.dt
                    else ""
                ),
                "liga": str(sample.features_q4.get("league", "")),
                "match_id": match_id,
                "equipo_local": teams.get("home_team", ""),
                "equipo_visitante": teams.get("away_team", ""),
                "resultado_q4_home_gana": y,
                "marcador_q4": (q4_scores_map or {}).get(match_id, ""),
                "marcador_q3_min27": q3_min27_scores_map.get(match_id, ""),
                "marcador_q3_min30": q3_min30_scores_map.get(match_id, ""),
                "prob_home_m27_raw": None if p27 is None else float(p27),
                "prob_home_m30_raw": None if p30 is None else float(p30),
                "pick_strength_m27_raw": p27_pick,
                "pick_strength_m30_raw": p30_pick,
                "pick_m27_raw": pick27,
                "pick_m30_raw": pick30,
                "hit_m27_raw": hit27,
                "hit_m30_raw": hit30,
                "ambos_con_prob": int(p27 is not None and p30 is not None),
                "ambos_mismo_pick": int(
                    p27 is not None and p30 is not None and pick27 == pick30
                ),
                "solo_m27_acierta": int(hit27 == 1 and hit30 != 1),
                "solo_m30_acierta": int(hit30 == 1 and hit27 != 1),
                "ambos_aciertan": int(hit27 == 1 and hit30 == 1),
                "delta_prob_home_m30_m27": (
                    None if p27 is None or p30 is None else float(p30 - p27)
                ),
                "delta_pick_strength_m30_m27": (
                    None
                    if p27_pick is None or p30_pick is None
                    else float(p30_pick - p27_pick)
                ),
            }
        )
    return pd.DataFrame(rows)


def _annotate_m27_v1_policy(model_df: pd.DataFrame) -> pd.DataFrame:
    if model_df.empty or "liga" not in model_df.columns:
        return model_df
    out = model_df.copy()
    out["m27_v1_league_tier"] = out["liga"].map(m27_v1_policy.get_tier)
    out["m27_v1_policy_action"] = out["liga"].map(m27_v1_policy.get_action)
    return out


def _build_m27_v1_policy_sheet(model_df: pd.DataFrame) -> pd.DataFrame:
    cols = [
        "liga",
        "tier",
        "accion",
        "matches_totales",
        "matches_apostados",
        "ganados",
        "perdidos",
        "efectividad",
        "ganancia",
        "yield_sobre_apostado",
    ]
    if model_df.empty:
        return pd.DataFrame(columns=cols)
    all_matches = model_df.groupby("liga", as_index=False).agg(
        matches_totales=("match_id", "count")
    )
    bets_df = model_df[
        model_df["resultado_apuesta"].isin(["GANADA", "PERDIDA"])
    ].copy()
    if bets_df.empty:
        out = all_matches.copy()
        out["tier"] = out["liga"].map(m27_v1_policy.get_tier)
        out["accion"] = out["liga"].map(m27_v1_policy.get_action)
        for col in cols:
            if col not in out.columns:
                out[col] = 0.0 if col in {
                    "ganancia",
                    "efectividad",
                    "yield_sobre_apostado",
                } else ""
        return out[cols]
    grouped = bets_df.groupby("liga", as_index=False).agg(
        matches_apostados=("resultado_apuesta", "count"),
        ganados=("resultado_apuesta", lambda s: int((s == "GANADA").sum())),
        perdidos=("resultado_apuesta", lambda s: int((s == "PERDIDA").sum())),
        ganancia=("ganancia", "sum"),
        total_apostado=("monto_apostado", "sum"),
    )
    out = all_matches.merge(grouped, on="liga", how="left").fillna(
        {
            "matches_apostados": 0,
            "ganados": 0,
            "perdidos": 0,
            "ganancia": 0.0,
            "total_apostado": 0.0,
        }
    )
    out["matches_apostados"] = out["matches_apostados"].astype(int)
    out["ganados"] = out["ganados"].astype(int)
    out["perdidos"] = out["perdidos"].astype(int)
    out["efectividad"] = out.apply(
        lambda r: (r["ganados"] / r["matches_apostados"])
        if r["matches_apostados"]
        else 0.0,
        axis=1,
    )
    out["yield_sobre_apostado"] = out.apply(
        lambda r: (r["ganancia"] / r["total_apostado"])
        if r["total_apostado"]
        else 0.0,
        axis=1,
    )
    out["tier"] = out["liga"].map(m27_v1_policy.get_tier)
    out["accion"] = out["liga"].map(m27_v1_policy.get_action)
    return out[cols].sort_values(
        ["tier", "ganancia", "matches_apostados"],
        ascending=[True, False, False],
    )


def _simulate(
    model_name,
    test_rows,
    p_home_list,
    base_minute,
    excluded_flags=None,
    excluded_reasons=None,
    mode="kelly_non_compound",
    kelly_mult=1.0,
    kelly_cap=1.0,
    min_conf_prob=0.5,
    stake_step=1.0,
    min_stake=1.0,
    max_stake=100.0,
    teams_map=None,
    q4_scores_map=None,
    q3_score_maps=None,
    selected_snapshot_minutes=None,
):
    excluded_flags = excluded_flags or [False] * len(test_rows)
    excluded_reasons = excluded_reasons or [None] * len(test_rows)
    q3_score_maps = q3_score_maps or {}
    bank = BANK_START
    details = []
    bets = wins = losses = no_bet = 0
    total_staked = 0.0
    peak = BANK_START
    max_dd = 0.0

    for i, sample in enumerate(test_rows):
        y = int(sample.target_q4)
        p_home = p_home_list[i]
        match_id = str(sample.match_id)
        snapshot_used = (
            selected_snapshot_minutes[i]
            if selected_snapshot_minutes
            else base_minute
        )
        if snapshot_used is None:
            snapshot_used = base_minute
        snapshot_used_int = int(snapshot_used)
        teams = (teams_map or {}).get(match_id, {})
        q3_snapshot_score = (
            q3_score_maps.get(snapshot_used_int, {}).get(match_id, "")
        )
        rec = {
            "resultado_apuesta": "SIN_APUESTA",
            "apuesta": "sin_apuesta",
            "monto_apostado": 0.0,
            "ganancia": 0.0,
            "bank_final": BANK_START,
            "modelo": model_name,
            "partida_test": i + 1,
            "match_id": match_id,
            "fecha": _report_date_from_dt(sample.dt),
            "hora": (
                (
                    sample.dt
                    if sample.dt.tzinfo
                    else sample.dt.replace(tzinfo=timezone.utc)
                )
                .astimezone(timezone(timedelta(hours=-6)))
                .strftime("%H:%M:%S")
                if sample.dt
                else ""
            ),
            "liga": str(sample.features_q4.get("league", "")),
            "equipo_local": teams.get("home_team", ""),
            "equipo_visitante": teams.get("away_team", ""),
            "resultado_q4_home_gana": y,
            "marcador_q4": (q4_scores_map or {}).get(match_id, ""),
            "marcador_q3_snapshot": q3_snapshot_score,
            "marcador_q3_min27": q3_score_maps.get(27, {}).get(match_id, ""),
            "marcador_q3_min30": q3_score_maps.get(30, {}).get(match_id, ""),
            "prob_local": None if p_home is None else float(p_home),
            "prob_visitante": None if p_home is None else float(1.0 - p_home),
            "lado_predicho": None,
            "confianza_prob": None,
            "confianza_score_0_100": None,
            "nivel_confianza": None,
            "apuestas_odds": ODDS,
            "probabilidad_empate": BREAK_EVEN,
            "edge": None,
            "kelly_fraction_raw": None,
            "kelly_fraction_used": None,
            "step_apuesta": stake_step,
            "razon_sin_apuesta": None,
            "pnl": 0.0,
            "banco_antes": bank,
            "ganancia_acumulada": bank - BANK_START,
            "roi_banco_acumulado": (bank - BANK_START) / BANK_START,
            "minuto_base_apuesta": snapshot_used_int,
        }

        if excluded_flags[i]:
            if p_home is not None:
                p_pick = p_home if p_home >= 0.5 else (1.0 - p_home)
                rec["lado_predicho"] = (
                    "local" if p_home >= 0.5 else "visitante"
                )
                rec["confianza_prob"] = p_pick
                rec["confianza_score_0_100"] = p_pick * 100.0
                rec["nivel_confianza"] = (
                    "muy_alta" if p_pick >= 0.80 else
                    "alta" if p_pick >= 0.70 else
                    "media" if p_pick >= 0.60 else
                    "baja"
                )
                rec["edge"] = p_pick - BREAK_EVEN
            rec["razon_sin_apuesta"] = excluded_reasons[i]
            rec["monto_apostado"] = None
            rec["ganancia"] = None
            rec["bank_final"] = BANK_START
            no_bet += 1
            details.append(rec)
            continue

        if p_home is None:
            rec["razon_sin_apuesta"] = (
                excluded_reasons[i] or "no_signal_in_snapshot"
            )
            no_bet += 1
            details.append(rec)
            continue

        p_pick = p_home if p_home >= 0.5 else (1.0 - p_home)
        pick_home = p_home >= 0.5
        rec["lado_predicho"] = "local" if pick_home else "visitante"
        rec["confianza_prob"] = p_pick
        rec["confianza_score_0_100"] = p_pick * 100.0
        rec["nivel_confianza"] = (
            "muy_alta" if p_pick >= 0.80 else
            "alta" if p_pick >= 0.70 else
            "media" if p_pick >= 0.60 else
            "baja"
        )
        rec["edge"] = p_pick - BREAK_EVEN

        if p_pick < min_conf_prob:
            rec["razon_sin_apuesta"] = "confidence_filter"
            no_bet += 1
            details.append(rec)
            continue

        k_raw = _kelly_fraction(p_pick, ODDS)
        rec["kelly_fraction_raw"] = k_raw
        if k_raw <= 0:
            rec["razon_sin_apuesta"] = "no_edge"
            no_bet += 1
            details.append(rec)
            continue

        k_used = min(k_raw * kelly_mult, kelly_cap)
        rec["kelly_fraction_used"] = k_used
        stake = _stake_from_signal(p_pick, rec["edge"], k_used, kelly_cap)
        if max_stake > 0:
            stake = min(stake, max_stake)
        stake = _round_down_step(stake, stake_step)
        if stake < min_stake:
            rec["razon_sin_apuesta"] = "stake_below_25"
            no_bet += 1
            details.append(rec)
            continue
        if stake > bank:
            rec["razon_sin_apuesta"] = "insufficient_bank"
            no_bet += 1
            details.append(rec)
            continue

        is_win = (y == 1 and pick_home) or (y == 0 and (not pick_home))
        pnl = stake * (ODDS - 1.0) if is_win else -stake
        bank += pnl
        rec["apuesta"] = "home" if pick_home else "away"
        rec["resultado_apuesta"] = "GANADA" if is_win else "PERDIDA"
        rec["monto_apostado"] = stake
        rec["ganancia"] = pnl
        rec["bank_final"] = bank
        rec["pnl"] = pnl
        rec["banco_antes"] = (
            BANK_START
            if mode == "kelly_non_compound"
            else (bank - pnl)
        )
        rec["ganancia_acumulada"] = bank - BANK_START
        rec["roi_banco_acumulado"] = (bank - BANK_START) / BANK_START

        bets += 1
        wins += int(is_win)
        losses += int(not is_win)
        total_staked += stake
        peak = max(peak, bank)
        dd = (peak - bank) / peak if peak > 0 else 0.0
        max_dd = max(max_dd, dd)
        details.append(rec)

    summary = {
        "modelo": model_name,
        "partidos_test": len(test_rows),
        "apuestas": bets,
        "ganadas": wins,
        "perdidas": losses,
        "empates_contados_como_perdida": 0,
        "sin_apuesta": no_bet,
        "partidos_no_apostados": no_bet,
        "efectividad": (wins / bets) if bets else 0.0,
        "banco_inicio": BANK_START,
        "banco_final": bank,
        "ganancia": bank - BANK_START,
        "roi_bank": (bank - BANK_START) / BANK_START,
        "total_apostado": total_staked,
        "apuesta_promedio": (total_staked / bets) if bets else 0.0,
        "yield_sobre_apostado": (
            ((bank - BANK_START) / total_staked)
            if total_staked > 0
            else 0.0
        ),
        "max_drawdown": max_dd,
        "minuto_base_apuesta": int(base_minute),
    }
    return details, summary


def _build_effectiveness_by_league(model_dfs):
    cols = [
        "modelo",
        "liga",
        "matches_apostados",
        "ganados",
        "perdidos",
        "efectividad",
        "ganancia",
    ]
    all_df = pd.concat(model_dfs, ignore_index=True)
    if all_df.empty:
        return pd.DataFrame(columns=cols)
    bets_df = all_df[
        all_df["resultado_apuesta"].isin(["GANADA", "PERDIDA"])
    ].copy()
    if bets_df.empty:
        return pd.DataFrame(columns=cols)
    grouped = bets_df.groupby(["modelo", "liga"], as_index=False).agg(
        matches_apostados=("resultado_apuesta", "count"),
        ganados=("resultado_apuesta", lambda s: int((s == "GANADA").sum())),
        perdidos=("resultado_apuesta", lambda s: int((s == "PERDIDA").sum())),
        ganancia=("ganancia", "sum"),
    )
    grouped["efectividad"] = grouped.apply(
        lambda r: (r["ganados"] / r["matches_apostados"])
        if r["matches_apostados"]
        else 0.0,
        axis=1,
    )
    return grouped[cols].sort_values(
        ["modelo", "ganancia"],
        ascending=[True, False],
    )


def _build_snapshot_comparison(
    details_27,
    details_30,
    q3_min27_scores_map,
    q3_min30_scores_map,
):
    by_27 = {str(row["match_id"]): row for row in details_27}
    by_30 = {str(row["match_id"]): row for row in details_30}
    rows = []
    for match_id in sorted(set(by_27) | set(by_30)):
        r27 = by_27.get(match_id, {})
        r30 = by_30.get(match_id, {})
        res27 = str(r27.get("resultado_apuesta", "SIN_APUESTA"))
        res30 = str(r30.get("resultado_apuesta", "SIN_APUESTA"))
        rows.append(
            {
                "fecha": r27.get("fecha") or r30.get("fecha") or "",
                "hora": r27.get("hora") or r30.get("hora") or "",
                "liga": r27.get("liga") or r30.get("liga") or "",
                "match_id": match_id,
                "equipo_local": (
                    r27.get("equipo_local") or r30.get("equipo_local") or ""
                ),
                "equipo_visitante": (
                    r27.get("equipo_visitante")
                    or r30.get("equipo_visitante")
                    or ""
                ),
                "resultado_q4_home_gana": r27.get(
                    "resultado_q4_home_gana",
                    r30.get("resultado_q4_home_gana"),
                ),
                "marcador_q4": (
                    r27.get("marcador_q4") or r30.get("marcador_q4") or ""
                ),
                "marcador_q3_min27": q3_min27_scores_map.get(match_id, ""),
                "marcador_q3_min30": q3_min30_scores_map.get(match_id, ""),
                "pick_m27_v1": r27.get("lado_predicho"),
                "resultado_m27_v1": res27,
                "prob_m27_v1": r27.get("confianza_prob"),
                "stake_m27_v1": r27.get("monto_apostado"),
                "ganancia_m27_v1": r27.get("ganancia"),
                "razon_m27_v1": r27.get("razon_sin_apuesta"),
                "pick_m30_v1": r30.get("lado_predicho"),
                "resultado_m30_v1": res30,
                "prob_m30_v1": r30.get("confianza_prob"),
                "stake_m30_v1": r30.get("monto_apostado"),
                "ganancia_m30_v1": r30.get("ganancia"),
                "razon_m30_v1": r30.get("razon_sin_apuesta"),
                "ambos_apuestan": int(
                    res27 in ("GANADA", "PERDIDA")
                    and res30 in ("GANADA", "PERDIDA")
                ),
                "solo_m27_v1_apuesta": int(
                    res27 in ("GANADA", "PERDIDA")
                    and res30 == "SIN_APUESTA"
                ),
                "solo_m30_v1_apuesta": int(
                    res30 in ("GANADA", "PERDIDA")
                    and res27 == "SIN_APUESTA"
                ),
                "solo_m27_v1_gana": int(
                    res27 == "GANADA" and res30 != "GANADA"
                ),
                "solo_m30_v1_gana": int(
                    res30 == "GANADA" and res27 != "GANADA"
                ),
            }
        )
    return pd.DataFrame(rows)


def _apply_excel_formatting(ws, header_row=1):
    header_fill = PatternFill(
        start_color="4472C4",
        end_color="4472C4",
        fill_type="solid",
    )
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = thin_border
    for row in range(header_row + 1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
    header_map = {}
    for col in range(1, ws.max_column + 1):
        header_val = ws.cell(row=header_row, column=col).value
        if header_val:
            header_map[str(header_val)] = col
    percent_headers = {
        "efectividad",
        "roi_bank",
        "yield_sobre_apostado",
        "max_drawdown",
    }
    for header in percent_headers:
        if header in header_map:
            for row in range(header_row + 1, ws.max_row + 1):
                ws.cell(
                    row=row,
                    column=header_map[header],
                ).number_format = "0.00%"

    green_fill = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid",
    )
    red_fill = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid",
    )
    yellow_fill = PatternFill(
        start_color="FFEB9C",
        end_color="FFEB9C",
        fill_type="solid",
    )

    if ws.max_row > header_row:
        start_row = header_row + 1
        end_row = ws.max_row
        if ws.title.endswith("_matches"):
            if "resultado_apuesta" in header_map:
                col_letter = ws.cell(
                    row=1,
                    column=header_map["resultado_apuesta"],
                ).column_letter
                ws.conditional_formatting.add(
                    f"{col_letter}{start_row}:{col_letter}{end_row}",
                    FormulaRule(
                        formula=[f'${col_letter}{start_row}="GANADA"'],
                        fill=green_fill,
                    ),
                )
                ws.conditional_formatting.add(
                    f"{col_letter}{start_row}:{col_letter}{end_row}",
                    FormulaRule(
                        formula=[f'${col_letter}{start_row}="PERDIDA"'],
                        fill=red_fill,
                    ),
                )
                ws.conditional_formatting.add(
                    f"{col_letter}{start_row}:{col_letter}{end_row}",
                    FormulaRule(
                        formula=[f'${col_letter}{start_row}="SIN_APUESTA"'],
                        fill=yellow_fill,
                    ),
                )
        if ws.title == "summary":
            for header in ("ganancia", "banco_final"):
                if header in header_map:
                    col_letter = ws.cell(
                        row=1,
                        column=header_map[header],
                    ).column_letter
                    ws.conditional_formatting.add(
                        f"{col_letter}{start_row}:{col_letter}{end_row}",
                        CellIsRule(
                            operator="greaterThan",
                            formula=["0"],
                            fill=green_fill,
                        ),
                    )
                    ws.conditional_formatting.add(
                        f"{col_letter}{start_row}:{col_letter}{end_row}",
                        CellIsRule(
                            operator="lessThan",
                            formula=["0"],
                            fill=red_fill,
                        ),
                    )
        if ws.title == "efectividad_liga":
            if "efectividad" in header_map:
                col_letter = ws.cell(
                    row=1,
                    column=header_map["efectividad"],
                ).column_letter
                ws.conditional_formatting.add(
                    f"{col_letter}{start_row}:{col_letter}{end_row}",
                    CellIsRule(
                        operator="greaterThanOrEqual",
                        formula=["0.7"],
                        fill=green_fill,
                    ),
                )
                ws.conditional_formatting.add(
                    f"{col_letter}{start_row}:{col_letter}{end_row}",
                    CellIsRule(
                        operator="between",
                        formula=["0.55", "0.6999"],
                        fill=yellow_fill,
                    ),
                )
                ws.conditional_formatting.add(
                    f"{col_letter}{start_row}:{col_letter}{end_row}",
                    CellIsRule(
                        operator="lessThan",
                        formula=["0.55"],
                        fill=red_fill,
                    ),
                )
    ws.freeze_panes = f"A{header_row + 1}"


def _ask_yes_no(prompt: str, default_yes: bool = True) -> bool:
    suffix = "[S/n]" if default_yes else "[s/N]"
    resp = input(f"\n{prompt} {suffix}: ").strip().lower()
    if not resp:
        return default_yes
    return resp in ("s", "si", "sí", "y", "yes")


def _ask_optional_report_date() -> str | None:
    if not _ask_yes_no(
        "¿Quieres ver resultados de una fecha particular?",
        default_yes=False,
    ):
        return None
    while True:
        raw_value = input("Ingresa la fecha en formato YYYY-MM-DD: ").strip()
        try:
            return datetime.strptime(
                raw_value,
                "%Y-%m-%d",
            ).strftime("%Y-%m-%d")
        except ValueError:
            print("Fecha inválida. Usa el formato YYYY-MM-DD.")


def main():
    parser = argparse.ArgumentParser(
        description="Q4 ROI report independiente para m27_v1 y m30_v1"
    )
    parser.add_argument("--rebuild-pred-cache", action="store_true")
    parser.add_argument("--rebuild-splits-cache", action="store_true")
    parser.add_argument("--only-m27-v1", action="store_true")
    parser.add_argument("--only-m30-v1", action="store_true")
    parser.add_argument("--no-m27-v1", action="store_true")
    parser.add_argument("--no-m30-v1", action="store_true")
    args = parser.parse_args()

    if args.only_m27_v1 and args.only_m30_v1:
        raise ValueError(
            "No puedes usar --only-m27-v1 y --only-m30-v1 al mismo tiempo"
        )

    if args.only_m27_v1:
        run_m27_v1 = True
        run_m30_v1 = False
    elif args.only_m30_v1:
        run_m27_v1 = False
        run_m30_v1 = True
    else:
        run_m27_v1 = (
            False
            if args.no_m27_v1
            else _ask_yes_no("¿Incluir m27_v1?", default_yes=True)
        )
        run_m30_v1 = (
            False
            if args.no_m30_v1
            else _ask_yes_no("¿Incluir m30_v1?", default_yes=True)
        )

    if not any([run_m27_v1, run_m30_v1]):
        raise ValueError("Debes incluir al menos un modelo m_v1")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_mm_path = BASE_M_V1 / f"Q4_ROI_match_by_match_m_v1_{ts}.xlsx"
    only_report_date = _ask_optional_report_date()

    print("[M_V1_ROI] preparando datos de prueba...")
    _, _val_rows, test_rows = _prepare_splits(
        force_rebuild=args.rebuild_splits_cache,
        only_date=only_report_date,
    )
    teams_map = _load_match_teams_map()
    match_ids = [sample.match_id for sample in test_rows]
    q4_scores_map = _load_match_q4_scores_map(match_ids)
    q3_min27_scores_map = _load_match_score_at_minute_map(match_ids, 27)
    q3_min30_scores_map = _load_match_score_at_minute_map(
        match_ids,
        30,
        score_upto_fn=m30_v1_train._score_upto_m30,
    )
    q3_score_maps = {27: q3_min27_scores_map, 30: q3_min30_scores_map}

    cache_meta = {
        "report": "q4_roi_m_v1",
        "pred_logic_version": "m30_qdur_alignment_v2_raw_compare",
        "run_m27_v1": bool(run_m27_v1),
        "run_m30_v1": bool(run_m30_v1),
        "league_name_filters_disabled": bool(LEAGUE_NAME_FILTERS_DISABLED),
        "test_fp": _rows_fingerprint(test_rows),
        "m27_v1_model_sig": _file_signature(M27_V1_CHAMPION_PATH),
        "m30_v1_model_sig": _file_signature(M30_V1_CHAMPION_PATH),
    }
    pred_cache = _load_pred_cache(
        cache_meta,
        force_rebuild=args.rebuild_pred_cache,
    )

    if pred_cache is not None:
        print(f"[M_V1_ROI] usando cache de predicciones: {PRED_CACHE_PATH}")
        if run_m27_v1:
            p_m27_v1 = pred_cache["p_m27_v1"]
            excl_m27_v1_flags = pred_cache["excl_m27_v1_flags"]
            excl_m27_v1_reasons = pred_cache["excl_m27_v1_reasons"]
            snap_m27_v1 = pred_cache["snap_m27_v1"]
            p_m27_v1_raw = pred_cache["p_m27_v1_raw"]
        if run_m30_v1:
            p_m30_v1 = pred_cache["p_m30_v1"]
            excl_m30_v1_flags = pred_cache["excl_m30_v1_flags"]
            excl_m30_v1_reasons = pred_cache["excl_m30_v1_reasons"]
            snap_m30_v1 = pred_cache["snap_m30_v1"]
            p_m30_v1_raw = pred_cache["p_m30_v1_raw"]
    else:
        if run_m27_v1:
            print("[M_V1_ROI] predicciones m27_v1...")
            p_m27_v1, excl_m27_v1_flags, excl_m27_v1_reasons, snap_m27_v1 = (
                _predict_m_v1_probs_for_snapshot_mode(
                    test_rows,
                    "m27_v1",
                    M27_V1_CHAMPION_PATH,
                    m27_v1_train._build_m27_v1_features,
                    int(m27_v1_train.SNAPSHOT_MINUTE),
                    apply_filters=(not LEAGUE_NAME_FILTERS_DISABLED),
                )
            )
            print("[M_V1_ROI] predicciones raw m27_v1...")
            p_m27_v1_raw = _predict_m_v1_probs_raw(
                test_rows,
                "m27_v1",
                M27_V1_CHAMPION_PATH,
                m27_v1_train._build_m27_v1_features,
                int(m27_v1_train.SNAPSHOT_MINUTE),
            )
        if run_m30_v1:
            print("[M_V1_ROI] predicciones m30_v1...")
            p_m30_v1, excl_m30_v1_flags, excl_m30_v1_reasons, snap_m30_v1 = (
                _predict_m_v1_probs_for_snapshot_mode(
                    test_rows,
                    "m30_v1",
                    M30_V1_CHAMPION_PATH,
                    m30_v1_train._build_m30_v1_features,
                    int(m30_v1_train.SNAPSHOT_MINUTE),
                    apply_filters=(not LEAGUE_NAME_FILTERS_DISABLED),
                )
            )
            print("[M_V1_ROI] predicciones raw m30_v1...")
            p_m30_v1_raw = _predict_m_v1_probs_raw(
                test_rows,
                "m30_v1",
                M30_V1_CHAMPION_PATH,
                m30_v1_train._build_m30_v1_features,
                int(m30_v1_train.SNAPSHOT_MINUTE),
            )
        pred_payload = {}
        if run_m27_v1:
            pred_payload.update(
                {
                    "p_m27_v1": p_m27_v1,
                    "p_m27_v1_raw": p_m27_v1_raw,
                    "excl_m27_v1_flags": excl_m27_v1_flags,
                    "excl_m27_v1_reasons": excl_m27_v1_reasons,
                    "snap_m27_v1": snap_m27_v1,
                }
            )
        if run_m30_v1:
            pred_payload.update(
                {
                    "p_m30_v1": p_m30_v1,
                    "p_m30_v1_raw": p_m30_v1_raw,
                    "excl_m30_v1_flags": excl_m30_v1_flags,
                    "excl_m30_v1_reasons": excl_m30_v1_reasons,
                    "snap_m30_v1": snap_m30_v1,
                }
            )
        _save_pred_cache(cache_meta, pred_payload)
        print(
            f"[M_V1_ROI] cache de predicciones guardado en: "
            f"{PRED_CACHE_PATH}"
        )

    print("[M_V1_ROI] simulación match-by-match...")
    summaries = []
    league_dfs = []
    sheets_extra = []
    if run_m27_v1:
        m27_v1_details, m27_v1_summary = _simulate(
            "m27_v1",
            test_rows,
            p_m27_v1,
            int(m27_v1_train.SNAPSHOT_MINUTE),
            excl_m27_v1_flags,
            excl_m27_v1_reasons,
            MODE,
            KELLY_MULT,
            KELLY_CAP,
            MIN_CONF_PROB,
            STAKE_STEP,
            MIN_STAKE,
            MAX_STAKE,
            teams_map,
            q4_scores_map,
            q3_score_maps,
            snap_m27_v1,
        )
        m27_v1_df = _annotate_m27_v1_policy(pd.DataFrame(m27_v1_details))
        m27_v1_policy_df = _build_m27_v1_policy_sheet(m27_v1_df)
        summaries.append(m27_v1_summary)
        league_dfs.append(m27_v1_df)
        sheets_extra.append(("m27_v1_matches", m27_v1_df))
        sheets_extra.append(("m27_v1_policy", m27_v1_policy_df))
    if run_m30_v1:
        m30_v1_details, m30_v1_summary = _simulate(
            "m30_v1",
            test_rows,
            p_m30_v1,
            int(m30_v1_train.SNAPSHOT_MINUTE),
            excl_m30_v1_flags,
            excl_m30_v1_reasons,
            MODE,
            KELLY_MULT,
            KELLY_CAP,
            MIN_CONF_PROB,
            STAKE_STEP,
            MIN_STAKE,
            MAX_STAKE,
            teams_map,
            q4_scores_map,
            q3_score_maps,
            snap_m30_v1,
        )
        m30_v1_df = pd.DataFrame(m30_v1_details)
        summaries.append(m30_v1_summary)
        league_dfs.append(m30_v1_df)
        sheets_extra.append(("m30_v1_matches", m30_v1_df))
    if run_m27_v1 and run_m30_v1:
        comparison_df = _build_snapshot_comparison(
            m27_v1_details,
            m30_v1_details,
            q3_min27_scores_map,
            q3_min30_scores_map,
        )
        raw_summary_df = pd.DataFrame(
            [
                _build_raw_comparison_summary(
                    "m27_v1_raw",
                    test_rows,
                    p_m27_v1_raw,
                ),
                _build_raw_comparison_summary(
                    "m30_v1_raw",
                    test_rows,
                    p_m30_v1_raw,
                ),
            ]
        )
        raw_comparison_df = _build_raw_model_comparison(
            test_rows,
            p_m27_v1_raw,
            p_m30_v1_raw,
            teams_map,
            q4_scores_map,
            q3_min27_scores_map,
            q3_min30_scores_map,
        )
        sheets_extra.append(("m27_v1_vs_m30_v1", comparison_df))
        sheets_extra.append(("raw_summary", raw_summary_df))
        sheets_extra.append(("raw_compare", raw_comparison_df))

    col_order = [
        "fecha",
        "liga",
        "resultado_apuesta",
        "apuesta",
        "monto_apostado",
        "ganancia",
        "bank_final",
        "modelo",
        "partida_test",
        "match_id",
        "hora",
        "equipo_local",
        "equipo_visitante",
        "resultado_q4_home_gana",
        "marcador_q4",
        "marcador_q3_snapshot",
        "marcador_q3_min27",
        "marcador_q3_min30",
        "prob_local",
        "prob_visitante",
        "lado_predicho",
        "confianza_prob",
        "confianza_score_0_100",
        "nivel_confianza",
        "apuestas_odds",
        "probabilidad_empate",
        "edge",
        "kelly_fraction_raw",
        "kelly_fraction_used",
        "step_apuesta",
        "razon_sin_apuesta",
        "pnl",
        "banco_antes",
        "ganancia_acumulada",
        "roi_banco_acumulado",
        "minuto_base_apuesta",
    ]
    normalized_sheets = []
    for sheet_name, df in sheets_extra:
        if sheet_name.endswith("_matches"):
            df = df[[c for c in col_order if c in df.columns]]
        normalized_sheets.append((sheet_name, df))

    summary_df = pd.DataFrame(summaries)
    summary_order = [
        "modelo",
        "partidos_test",
        "apuestas",
        "ganadas",
        "perdidas",
        "empates_contados_como_perdida",
        "partidos_no_apostados",
        "sin_apuesta",
        "efectividad",
        "banco_inicio",
        "banco_final",
        "ganancia",
        "roi_bank",
        "total_apostado",
        "apuesta_promedio",
        "yield_sobre_apostado",
        "max_drawdown",
        "minuto_base_apuesta",
    ]
    summary_df = summary_df[
        [c for c in summary_order if c in summary_df.columns]
    ]
    eff_league_df = _build_effectiveness_by_league(league_dfs)

    print("[M_V1_ROI] escribiendo Excel...")
    BASE_M_V1.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_mm_path, engine="openpyxl") as writer:
        write_items = [
            ("summary", summary_df),
            ("efectividad_liga", eff_league_df),
        ]
        write_items += normalized_sheets
        for sheet_name, df in tqdm(
            write_items,
            desc="[M_V1_ROI] escribiendo hojas",
            unit="hoja",
        ):
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        for sheet_name in tqdm(
            list(writer.sheets),
            desc="[M_V1_ROI] formateando hojas",
            unit="hoja",
        ):
            _apply_excel_formatting(writer.sheets[sheet_name])

    print("[M_V1_ROI] OK")
    print(f"[M_V1_ROI] output={out_mm_path}")
    print("[M_V1_ROI] abriendo Excel...")
    os.startfile(str(out_mm_path))


if __name__ == "__main__":
    main()
