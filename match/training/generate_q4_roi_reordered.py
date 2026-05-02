"""Generate Q4 ROI reports with reordered columns (betting metrics before league)."""

from pathlib import Path
import json
import sqlite3
import joblib
import pandas as pd
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

import train_q3_q4_models_v6 as v6

ROOT = Path('match')
BASE_V6 = ROOT / 'training' / 'model_outputs_v6'
BASE_V62 = ROOT / 'training' / 'model_outputs_v6_2'
OUT_AB = BASE_V62 / 'Q4_ROI_AB_v6_vs_v6_2.xlsx'
OUT_MM = BASE_V62 / 'Q4_ROI_match_by_match_v6_vs_v6_2.xlsx'

ODDS = 1.4
BANK_START = 1000.0
BREAK_EVEN = 1.0 / ODDS
TARGET = 'q4'

# Staking params for detailed match-by-match
MODE = 'kelly_non_compound'
KELLY_MULT = 0.25
KELLY_CAP = 0.05
MIN_CONF_PROB = 0.58
STAKE_STEP = 25.0
MIN_STAKE = 25.0
MAX_STAKE = 100.0
STAKE_BUCKETS = (25.0, 50.0, 75.0, 100.0)


def _kelly_fraction(p, odds):
    b = odds - 1.0
    q = 1.0 - p
    return ((b * p) - q) / b if b > 0 else 0.0


def _round_down_step(x, step):
    if step <= 0:
        return x
    return (x // step) * step


def _stake_from_signal(p_pick, edge, k_used, kelly_cap=0.05):
    """Map signal quality to fixed stake buckets: 25/50/75/100.

    100 is intentionally rare: only when confidence, edge and Kelly usage are
    all in exceptional territory.
    """
    if k_used is None or k_used <= 0:
        return 0.0

    denom = kelly_cap if kelly_cap and kelly_cap > 0 else 0.05
    k_strength = max(0.0, min(k_used / denom, 1.0))
    p_pick = float(p_pick or 0.0)
    edge = float(edge or 0.0)

    # Exceptional tier.
    if p_pick >= 0.97 and edge >= 0.24 and k_strength >= 0.995:
        return STAKE_BUCKETS[3]
    # Strong tier.
    if p_pick >= 0.90 and edge >= 0.18 and k_strength >= 0.85:
        return STAKE_BUCKETS[2]
    # Medium tier.
    if p_pick >= 0.80 and edge >= 0.09 and k_strength >= 0.55:
        return STAKE_BUCKETS[1]
    # Weak but valid edge.
    if k_strength < 0.15:
        return 0.0
    return STAKE_BUCKETS[0]


def _prepare_rows():
    samples = v6._build_samples(v6.DB_PATH)
    rows = [s for s in samples if s.target_q4 is not None]
    rows = sorted(rows, key=lambda s: s.dt)
    n_total = len(rows)
    n_train = int(n_total * 0.8)
    test_rows = rows[n_train:]
    return rows, test_rows, n_total, n_train


def _load_match_teams_map():
    """Load home/away team names for display in reports."""
    conn = sqlite3.connect(v6.DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            "SELECT match_id, home_team, away_team FROM matches"
        ).fetchall()
    finally:
        conn.close()

    out = {}
    for r in rows:
        out[str(r["match_id"])] = {
            "home_team": str(r["home_team"] or ""),
            "away_team": str(r["away_team"] or ""),
        }
    return out


def _predict_v6_probs(test_rows):
    art_xgb = joblib.load(BASE_V6 / 'q4_xgb.joblib')
    art_hgb = joblib.load(BASE_V6 / 'q4_hist_gb.joblib')
    art_mlp = joblib.load(BASE_V6 / 'q4_mlp.joblib')

    x_dict = [s.features_q4 for s in test_rows]
    p_xgb = art_xgb['model'].predict_proba(art_xgb['vectorizer'].transform(x_dict))[:, 1]
    p_hgb = art_hgb['model'].predict_proba(art_hgb['vectorizer'].transform(x_dict))[:, 1]
    p_mlp = art_mlp['model'].predict_proba(art_mlp['vectorizer'].transform(x_dict))[:, 1]

    return ((p_xgb + p_hgb + p_mlp) / 3.0).tolist()


def _load_v62_exclusion_rules():
    cfg = json.loads((ROOT / 'training' / 'v6_2_league_name_exclusions.json').read_text(encoding='utf-8'))
    pats = []
    for cat in cfg.get('categories', []):
        for p in cat.get('patterns', []):
            p = str(p).strip()
            if p:
                pats.append((cat.get('name', 'uncategorized'), p, p.lower()))
    return pats


def _predict_v62_probs(test_rows):
    art = joblib.load(BASE_V62 / 'q4_champion.joblib')
    vec = art['vectorizer']
    models = art['models']
    keep_leagues = set(art['league_filter']['kept_leagues'])
    other_token = art['league_filter']['other_token']

    rules = _load_v62_exclusion_rules()

    probs = [None] * len(test_rows)
    excluded_flags = [False] * len(test_rows)
    excluded_reasons = [None] * len(test_rows)

    transformed = []
    transformed_idx = []
    
    for i, s in enumerate(test_rows):
        rec = dict(s.features_q4)
        lg = str(rec.get('league', ''))
        lg_lc = lg.lower()
        hit = None
        for cname, raw, low in rules:
            if low in lg_lc:
                hit = (cname, raw)
                break
        if hit is not None:
            excluded_flags[i] = True
            excluded_reasons[i] = f"excluded_league_name:{hit[0]}:{hit[1]}"
            continue

        if lg not in keep_leagues:
            rec['league'] = other_token
            rec['league_bucket'] = other_token
        transformed.append(rec)
        transformed_idx.append(i)

    if transformed:
        x_valid = vec.transform(transformed)
        p_xgb = models['xgb'].predict_proba(x_valid)[:, 1]
        p_hgb = models['hist_gb'].predict_proba(x_valid)[:, 1]
        p_blend = (0.6 * p_xgb + 0.4 * p_hgb)
        for j, orig_idx in enumerate(transformed_idx):
            probs[orig_idx] = float(p_blend[j])

    return probs, excluded_flags, excluded_reasons


def _simulate(model_name, test_rows, p_home_list, excluded_flags=None, excluded_reasons=None,
              mode='kelly_non_compound', kelly_mult=1.0, kelly_cap=1.0, min_conf_prob=0.5,
              stake_step=1.0, min_stake=1.0, max_stake=100.0, teams_map=None):
    excluded_flags = excluded_flags or [False] * len(test_rows)
    excluded_reasons = excluded_reasons or [None] * len(test_rows)

    bank = BANK_START
    details = []
    bets = wins = losses = no_bet = 0
    total_staked = 0.0
    peak = BANK_START
    max_dd = 0.0

    for i, s in enumerate(test_rows):
        y = int(s.target_q4)
        p_home = p_home_list[i]
        match_id = str(s.match_id)
        teams = (teams_map or {}).get(match_id, {})
        home_team_name = teams.get('home_team', '')
        away_team_name = teams.get('away_team', '')

        # REORDERED COLUMN SEQUENCE: betting metrics first, then league/team info
        rec = {
            # Betting decision & outcome (PRIORITY)
            'resultado_apuesta': 'SIN_APUESTA',
            'apuesta': 'sin_apuesta',
            'monto_apostado': 0.0,
            'ganancia': 0.0,
            'bank_final': BANK_START,
            
            # League and match identification (SECONDARY)
            'modelo': model_name,
            'partida_test': i + 1,
            'match_id': match_id,
            'fecha_hora': s.dt.isoformat(),
            'liga': str(s.features_q4.get('league', '')),
            'equipo_local': home_team_name,
            'equipo_visitante': away_team_name,
            
            # Match result
            'resultado_q4_home_gana': y,
            
            # Probabilities
            'prob_local': None if p_home is None else float(p_home),
            'prob_visitante': None if p_home is None else float(1.0 - p_home),
            
            # Model prediction
            'lado_predicho': None,
            'confianza_prob': None,
            'confianza_score_0_100': None,
            'nivel_confianza': None,
            
            # Betting parameters
            'apuestas_odds': ODDS,
            'probabilidad_empate': BREAK_EVEN,
            'edge': None,
            'kelly_fraction_raw': None,
            'kelly_fraction_used': None,
            'step_apuesta': stake_step,
            'razon_sin_apuesta': None,
            'pnl': 0.0,
            'banco_antes': bank,
            'ganancia_acumulada': bank - BANK_START,
            'roi_banco_acumulado': (bank - BANK_START) / BANK_START,
        }

        if excluded_flags[i]:
            rec['razon_sin_apuesta'] = excluded_reasons[i]
            rec['monto_apostado'] = None
            rec['ganancia'] = None
            rec['bank_final'] = None
            no_bet += 1
            details.append(rec)
            continue

        if p_home is None:
            rec['razon_sin_apuesta'] = 'missing_prob'
            no_bet += 1
            details.append(rec)
            continue

        p_pick = p_home if p_home >= 0.5 else (1.0 - p_home)
        pick_home = p_home >= 0.5
        pick_side = 'local' if pick_home else 'visitante'

        rec['lado_predicho'] = pick_side
        rec['confianza_prob'] = p_pick
        rec['confianza_score_0_100'] = p_pick * 100.0
        rec['nivel_confianza'] = (
            'muy_alta' if p_pick >= 0.80 else
            'alta' if p_pick >= 0.70 else
            'media' if p_pick >= 0.60 else
            'baja'
        )
        rec['edge'] = p_pick - BREAK_EVEN

        if p_pick < min_conf_prob:
            rec['razon_sin_apuesta'] = 'confidence_filter'
            no_bet += 1
            details.append(rec)
            continue

        k_raw = _kelly_fraction(p_pick, ODDS)
        rec['kelly_fraction_raw'] = k_raw
        if k_raw <= 0:
            rec['razon_sin_apuesta'] = 'no_edge'
            no_bet += 1
            details.append(rec)
            continue

        k_used = min(k_raw * kelly_mult, kelly_cap)
        rec['kelly_fraction_used'] = k_used

        stake = _stake_from_signal(p_pick, rec['edge'], k_used, kelly_cap)
        if max_stake > 0:
            stake = min(stake, max_stake)
        stake = _round_down_step(stake, stake_step)
        if stake < min_stake:
            rec['razon_sin_apuesta'] = 'stake_below_25'
            no_bet += 1
            details.append(rec)
            continue
        if stake > bank:
            rec['razon_sin_apuesta'] = 'insufficient_bank'
            no_bet += 1
            details.append(rec)
            continue

        is_win = (y == 1 and pick_home) or (y == 0 and (not pick_home))
        pnl = stake * (ODDS - 1.0) if is_win else -stake
        bank += pnl

        # UPDATE PRIORITY COLUMNS with actual values
        rec['apuesta'] = 'home' if pick_home else 'away'
        rec['resultado_apuesta'] = 'GANADA' if is_win else 'PERDIDA'
        rec['monto_apostado'] = stake
        rec['ganancia'] = pnl
        rec['bank_final'] = bank
        rec['pnl'] = pnl
        rec['banco_antes'] = BANK_START if mode == 'kelly_non_compound' else (bank - pnl)
        rec['ganancia_acumulada'] = bank - BANK_START
        rec['roi_banco_acumulado'] = (bank - BANK_START) / BANK_START

        bets += 1
        wins += int(is_win)
        losses += int(not is_win)
        total_staked += stake
        peak = max(peak, bank)
        dd = (peak - bank) / peak if peak > 0 else 0.0
        max_dd = max(max_dd, dd)

        details.append(rec)

    summary = {
        'modelo': model_name,
        'partidos_test': len(test_rows),
        'apuestas': bets,
        'ganadas': wins,
        'perdidas': losses,
        'empates_contados_como_perdida': 0,
        'sin_apuesta': no_bet,
        'efectividad': (wins / bets) if bets else 0.0,
        'banco_inicio': BANK_START,
        'banco_final': bank,
        'ganancia': bank - BANK_START,
        'roi_bank': (bank - BANK_START) / BANK_START,
        'total_apostado': total_staked,
        'apuesta_promedio': (total_staked / bets) if bets else 0.0,
        'yield_sobre_apostado': ((bank - BANK_START) / total_staked) if total_staked > 0 else 0.0,
        'max_drawdown': max_dd,
    }
    return details, summary


def _build_effectiveness_by_league(v6_df, v62_df):
    cols = [
        'modelo', 'liga', 'matches_apostados', 'ganados', 'perdidos',
        'efectividad', 'ganancia'
    ]

    all_df = pd.concat([v6_df, v62_df], ignore_index=True)
    if all_df.empty:
        return pd.DataFrame(columns=cols)

    bets_df = all_df[all_df['resultado_apuesta'].isin(['GANADA', 'PERDIDA'])].copy()
    if bets_df.empty:
        return pd.DataFrame(columns=cols)

    grouped = (
        bets_df
        .groupby(['modelo', 'liga'], as_index=False)
        .agg(
            matches_apostados=('resultado_apuesta', 'count'),
            ganados=('resultado_apuesta', lambda s: int((s == 'GANADA').sum())),
            perdidos=('resultado_apuesta', lambda s: int((s == 'PERDIDA').sum())),
            ganancia=('ganancia', 'sum')
        )
    )
    grouped['efectividad'] = grouped.apply(
        lambda r: (r['ganados'] / r['matches_apostados']) if r['matches_apostados'] else 0.0,
        axis=1
    )
    grouped = grouped[cols].sort_values(['modelo', 'ganancia'], ascending=[True, False])
    return grouped


def _apply_excel_formatting(ws, header_row=1):
    """Apply Spanish headers and conditional formatting to worksheet."""
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    for row in range(header_row + 1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if cell.column_letter in {'C', 'D', 'E', 'M', 'N', 'Q', 'R', 'U', 'V', 'W', 'X', 'AA', 'AB', 'AC', 'AD'}:
                cell.number_format = '0.00'

    # Header-driven number formats for summary/effectiveness sheets.
    header_map = {}
    for col in range(1, ws.max_column + 1):
        header_val = ws.cell(row=header_row, column=col).value
        if header_val:
            header_map[str(header_val)] = col

    percent_headers = {'efectividad', 'roi_bank', 'yield_sobre_apostado', 'max_drawdown'}
    for h in percent_headers:
        if h in header_map:
            c = header_map[h]
            for row in range(header_row + 1, ws.max_row + 1):
                ws.cell(row=row, column=c).number_format = '0.00%'

    # Semaforo for result text and key numeric values.
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

    if ws.max_row > header_row:
        start_row = header_row + 1
        end_row = ws.max_row
        data_range = f"{start_row}:{end_row}"

        if ws.title in {'v6_matches', 'v6_2_matches'}:
            # Semaforo de nivel_confianza (S)
            ws.conditional_formatting.add(
                f"S{start_row}:S{end_row}",
                FormulaRule(formula=[f'$S{start_row}="muy_alta"'], fill=green_fill)
            )
            ws.conditional_formatting.add(
                f"S{start_row}:S{end_row}",
                FormulaRule(formula=[f'$S{start_row}="alta"'], fill=green_fill)
            )
            ws.conditional_formatting.add(
                f"S{start_row}:S{end_row}",
                FormulaRule(formula=[f'$S{start_row}="media"'], fill=yellow_fill)
            )
            ws.conditional_formatting.add(
                f"S{start_row}:S{end_row}",
                FormulaRule(formula=[f'$S{start_row}="baja"'], fill=red_fill)
            )

            # Si razon_sin_apuesta empieza con excluded_league_name, marcar en rojo
            # liga/equipos (J:L) y campos financieros clave (C:E).
            for col in ['J', 'K', 'L', 'C', 'D', 'E']:
                ws.conditional_formatting.add(
                    f"{col}{start_row}:{col}{end_row}",
                    FormulaRule(formula=[f'LEFT($Z{start_row},20)="excluded_league_name"'], fill=red_fill)
                )

            ws.conditional_formatting.add(
                f"A{start_row}:A{end_row}",
                FormulaRule(formula=[f'$A{start_row}="GANADA"'], fill=green_fill)
            )
            ws.conditional_formatting.add(
                f"A{start_row}:A{end_row}",
                FormulaRule(formula=[f'$A{start_row}="PERDIDA"'], fill=red_fill)
            )
            ws.conditional_formatting.add(
                f"A{start_row}:A{end_row}",
                FormulaRule(formula=[f'$A{start_row}="SIN_APUESTA"'], fill=yellow_fill)
            )

            for col in ['D', 'E', 'AD']:
                ws.conditional_formatting.add(
                    f"{col}{start_row}:{col}{end_row}",
                    CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill)
                )
                ws.conditional_formatting.add(
                    f"{col}{start_row}:{col}{end_row}",
                    CellIsRule(operator='lessThan', formula=['0'], fill=red_fill)
                )
                ws.conditional_formatting.add(
                    f"{col}{start_row}:{col}{end_row}",
                    CellIsRule(operator='equal', formula=['0'], fill=yellow_fill)
                )

            ws.conditional_formatting.add(
                f"E{start_row}:E{end_row}",
                CellIsRule(operator='greaterThan', formula=[str(BANK_START)], fill=green_fill)
            )
            ws.conditional_formatting.add(
                f"E{start_row}:E{end_row}",
                CellIsRule(operator='lessThan', formula=[str(BANK_START)], fill=red_fill)
            )
            ws.conditional_formatting.add(
                f"E{start_row}:E{end_row}",
                CellIsRule(operator='equal', formula=[str(BANK_START)], fill=yellow_fill)
            )

            # Si la apuesta fue PERDIDA, resaltar ambos bancos en rojo.
            for col in ['E', 'AB']:
                ws.conditional_formatting.add(
                    f"{col}{start_row}:{col}{end_row}",
                    FormulaRule(formula=[f'$A{start_row}="PERDIDA"'], fill=red_fill)
                )

        if ws.title == 'summary':
            for col in ['J', 'K']:
                ws.conditional_formatting.add(
                    f"{col}{start_row}:{col}{end_row}",
                    CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill)
                )
                ws.conditional_formatting.add(
                    f"{col}{start_row}:{col}{end_row}",
                    CellIsRule(operator='lessThan', formula=['0'], fill=red_fill)
                )
                ws.conditional_formatting.add(
                    f"{col}{start_row}:{col}{end_row}",
                    CellIsRule(operator='equal', formula=['0'], fill=yellow_fill)
                )

        if ws.title == 'efectividad_liga':
            # F: efectividad, G: ganancia
            ws.conditional_formatting.add(
                f"F{start_row}:F{end_row}",
                CellIsRule(operator='greaterThanOrEqual', formula=['0.7'], fill=green_fill)
            )
            ws.conditional_formatting.add(
                f"F{start_row}:F{end_row}",
                CellIsRule(operator='between', formula=['0.55', '0.6999'], fill=yellow_fill)
            )
            ws.conditional_formatting.add(
                f"F{start_row}:F{end_row}",
                CellIsRule(operator='lessThan', formula=['0.55'], fill=red_fill)
            )
            ws.conditional_formatting.add(
                f"G{start_row}:G{end_row}",
                CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill)
            )
            ws.conditional_formatting.add(
                f"G{start_row}:G{end_row}",
                CellIsRule(operator='lessThan', formula=['0'], fill=red_fill)
            )
            ws.conditional_formatting.add(
                f"G{start_row}:G{end_row}",
                CellIsRule(operator='equal', formula=['0'], fill=yellow_fill)
            )

    ws.freeze_panes = f'A{header_row + 1}'


def main():
    print("[Q4_ROI] preparando datos de prueba...")
    all_rows, test_rows, n_total, n_train = _prepare_rows()

    print("[Q4_ROI] predicciones v6...")
    p_v6 = _predict_v6_probs(test_rows)

    print("[Q4_ROI] predicciones v6.2...")
    p_v62, excl_flags, excl_reasons = _predict_v62_probs(test_rows)
    teams_map = _load_match_teams_map()

    print("[Q4_ROI] simulación match-by-match...")
    v6_details, v6_summary = _simulate('v6', test_rows, p_v6, None, None, 
                                                                             MODE, KELLY_MULT, KELLY_CAP, MIN_CONF_PROB, STAKE_STEP, MIN_STAKE, MAX_STAKE, teams_map)
    v62_details, v62_summary = _simulate('v6.2', test_rows, p_v62, excl_flags, excl_reasons, 
                                                                                 MODE, KELLY_MULT, KELLY_CAP, MIN_CONF_PROB, STAKE_STEP, MIN_STAKE, MAX_STAKE, teams_map)

    print("[Q4_ROI] escribiendo match-by-match Excel...")
    out_mm_path = OUT_MM
    if OUT_MM.exists():
        try:
            OUT_MM.unlink()
        except PermissionError:
            out_mm_path = OUT_MM.with_name(f"{OUT_MM.stem}_updated{OUT_MM.suffix}")
            print(f"[Q4_ROI] archivo bloqueado; usando salida alternativa: {out_mm_path}")
    
    v6_df = pd.DataFrame(v6_details)
    v62_df = pd.DataFrame(v62_details)
    summary_df = pd.DataFrame([v6_summary, v62_summary])

    # Asegurar que las columnas estén en el orden correcto
    col_order = [
        'resultado_apuesta', 'apuesta', 'monto_apostado', 'ganancia', 'bank_final',
        'modelo', 'partida_test', 'match_id', 'fecha_hora', 'liga', 'equipo_local', 
        'equipo_visitante', 'resultado_q4_home_gana', 'prob_local', 'prob_visitante',
        'lado_predicho', 'confianza_prob', 'confianza_score_0_100', 'nivel_confianza',
        'apuestas_odds', 'probabilidad_empate', 'edge', 'kelly_fraction_raw', 
        'kelly_fraction_used', 'step_apuesta', 'razon_sin_apuesta', 'pnl',
        'banco_antes', 'ganancia_acumulada', 'roi_banco_acumulado'
    ]
    v6_df = v6_df[[c for c in col_order if c in v6_df.columns]]
    v62_df = v62_df[[c for c in col_order if c in v62_df.columns]]
    eff_league_df = _build_effectiveness_by_league(v6_df, v62_df)

    summary_order = [
        'modelo', 'partidos_test', 'apuestas', 'ganadas', 'perdidas',
        'empates_contados_como_perdida', 'sin_apuesta', 'efectividad',
        'banco_inicio', 'banco_final', 'ganancia', 'roi_bank',
        'total_apostado', 'apuesta_promedio', 'yield_sobre_apostado', 'max_drawdown'
    ]
    summary_df = summary_df[[c for c in summary_order if c in summary_df.columns]]

    with pd.ExcelWriter(out_mm_path, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name='summary', index=False)
        eff_league_df.to_excel(writer, sheet_name='efectividad_liga', index=False)
        v6_df.to_excel(writer, sheet_name='v6_matches', index=False)
        v62_df.to_excel(writer, sheet_name='v6_2_matches', index=False)

        # Apply formatting
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            _apply_excel_formatting(ws)

    print("[Q4_ROI] OK")
    print(f"[Q4_ROI] output={out_mm_path}")


if __name__ == "__main__":
    main()
