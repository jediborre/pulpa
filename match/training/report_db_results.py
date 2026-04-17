"""
report_db_results.py — Reporte de apuestas desde la base de datos.

Lee outcomes directamente de eval_match_results (sin re-inferir).
Soporta cualquier modelo o tag que exista en la DB.

Diferencia clave con report_model_comparison.py:
  - report_model_comparison.py: re-ejecuta inferencia por partido (lento, ~1 min/partido)
  - report_db_results.py: lee resultados guardados por el bot (instantáneo)
    → Solo muestra predicciones que el bot REALMENTE hizo en vivo (o en eval offline).

Uso:
    # Reporte de abril 2026, todos los modelos disponibles:
    python training/report_db_results.py --month 2026-04

    # Solo comparar v13 vs v4:
    python training/report_db_results.py --month 2026-04 --models v13,v4

    # Rango de fechas:
    python training/report_db_results.py --from 2026-02-01 --to 2026-04-15 --models v13

    # Listar modelos disponibles en la DB:
    python training/report_db_results.py --list-models

    # Salida personalizada:
    python training/report_db_results.py --month 2026-04 --out /ruta/reporte.xlsx

Hojas del Excel generado:
    Partidos — Una fila por partido, columnas por modelo (Q3/Q4 señal + resultado)
    Resumen  — Apuestas/W/L/P, efectividad, ROI por modelo y cuarto
    Por Dia  — Totales diarios combinados de todos los modelos seleccionados
    Ligas    — ROI por liga para el modelo principal (primer modelo de --models)
"""

from __future__ import annotations

import argparse
import sqlite3
import sys
import warnings
from collections import OrderedDict
from pathlib import Path
from typing import Any

warnings.filterwarnings("ignore")

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

DB_PATH = ROOT / "matches.db"
REPORTS_DIR = ROOT / "reports"
QUARTERS = ["q3", "q4"]

ODDS = 1.91
BET_SIZE = 20.0
STARTING_BANK = 100 * BET_SIZE  # 2000

# ── Excel styles ──────────────────────────────────────────────────────────────
if _HAS_OPENPYXL:
    _GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    _RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    _PUSH   = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    _HDR    = PatternFill(start_color="2F75B6", end_color="2F75B6", fill_type="solid")
    _SUB    = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    _GRAY   = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    _WHITE_FONT = Font(color="FFFFFF", bold=True, size=9)
    _BOLD_FONT  = Font(bold=True, size=9)
    _NORM_FONT  = Font(size=9)
    _CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _LEFT   = Alignment(horizontal="left", vertical="center")
else:
    _GREEN = _RED = _PUSH = _HDR = _SUB = _GRAY = _YELLOW = None
    _WHITE_FONT = _BOLD_FONT = _NORM_FONT = _CENTER = _LEFT = None


# ── DB helpers ────────────────────────────────────────────────────────────────

def _open_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


def _discover_models(conn: sqlite3.Connection) -> list[str]:
    """Devuelve todos los tags/modelos presentes en eval_match_results."""
    cols = [r["name"] for r in conn.execute("PRAGMA table_info(eval_match_results)").fetchall()]
    tags: set[str] = set()
    for col in cols:
        if col.startswith("q3_pick__"):
            tags.add(col[len("q3_pick__"):])
    return sorted(tags)


def _fetch_rows(
    conn: sqlite3.Connection,
    date_from: str,
    date_to: str,
    models: list[str],
) -> list[sqlite3.Row]:
    """Fetch eval_match_results rows in date range that have at least one BET outcome."""
    signal_conditions = " OR ".join(
        f"(q3_signal__{m} = 'BET' OR q4_signal__{m} = 'BET')"
        for m in models
    )
    query = f"""
        SELECT *
        FROM eval_match_results
        WHERE event_date BETWEEN ? AND ?
          AND ({signal_conditions})
        ORDER BY event_date, match_id
    """
    return conn.execute(query, (date_from, date_to)).fetchall()


def _row_val(row: sqlite3.Row, key: str) -> Any:
    try:
        return row[key]
    except (IndexError, KeyError):
        return None


# ── Outcome helpers ───────────────────────────────────────────────────────────

def _get_model_quarter_data(row: sqlite3.Row, model: str, q: str) -> dict:
    """Extrae señal/pick/outcome de una fila para un modelo y cuarto."""
    available = int(_row_val(row, f"{q}_available__{model}") or 0)
    signal = str(_row_val(row, f"{q}_signal__{model}") or "").upper()
    pick = str(_row_val(row, f"{q}_pick__{model}") or "")
    outcome = str(_row_val(row, f"{q}_outcome__{model}") or "").lower()
    return {
        "available": bool(available),
        "signal": signal,
        "pick": pick,
        "outcome": outcome,  # hit / miss / push / "" (pending or no bet)
    }


def _outcome_to_label(d: dict) -> tuple[str, str]:
    """(bet_label, result_label) para mostrar en la celda."""
    if not d["available"] or d["signal"] != "BET":
        return "—", "—"

    icon = "🏠" if d["pick"] == "home" else ("✈️" if d["pick"] == "away" else "?")
    bet_label = f"BET {icon}"

    o = d["outcome"]
    if o == "hit":
        return bet_label, "WIN"
    elif o == "miss":
        return bet_label, "LOSS"
    elif o == "push":
        return bet_label, "PUSH"
    else:
        return bet_label, "—"  # pending / sin resultado


def _outcome_to_pnl(result_label: str) -> float:
    if result_label == "WIN":
        return BET_SIZE * (ODDS - 1)
    elif result_label in ("LOSS", "PUSH"):
        return -BET_SIZE
    return 0.0


# ── Build records ─────────────────────────────────────────────────────────────

def _build_records(
    db_rows: list[sqlite3.Row],
    models: list[str],
) -> list[dict]:
    """Convierte filas de DB a lista de dicts para el Excel."""
    records = []
    for row in db_rows:
        q3h = _row_val(row, "q3_home_score")
        q3a = _row_val(row, "q3_away_score")
        q4h = _row_val(row, "q4_home_score")
        q4a = _row_val(row, "q4_away_score")

        rec: dict[str, Any] = {
            "date":     str(_row_val(row, "event_date") or ""),
            "match_id": str(_row_val(row, "match_id") or ""),
            "home":     str(_row_val(row, "home_team") or ""),
            "away":     str(_row_val(row, "away_team") or ""),
            "league":   str(_row_val(row, "league") or ""),
            "q3_score": f"{q3h}-{q3a}" if q3h is not None else "—",
            "q4_score": f"{q4h}-{q4a}" if q4h is not None else "—",
        }

        for model in models:
            for q in QUARTERS:
                d = _get_model_quarter_data(row, model, q)
                bet_label, result_label = _outcome_to_label(d)
                rec[f"{model}_{q}_bet"]     = bet_label
                rec[f"{model}_{q}_result"]  = result_label
                rec[f"{model}_{q}_pnl"]     = _outcome_to_pnl(result_label)

        records.append(rec)
    return records


# ── Simulation aggregation ────────────────────────────────────────────────────

def _build_sim(records: list[dict], models: list[str]) -> dict:
    sim: dict[str, dict] = {
        f"{m}_{q}": {
            "bets": 0, "wins": 0, "losses": 0, "pushes": 0,
            "bank": STARTING_BANK,
        }
        for m in models
        for q in QUARTERS
    }
    for rec in records:
        for m in models:
            for q in QUARTERS:
                r = rec.get(f"{m}_{q}_result", "—")
                if r not in ("WIN", "LOSS", "PUSH"):
                    continue
                key = f"{m}_{q}"
                sim[key]["bets"] += 1
                if r == "WIN":
                    sim[key]["wins"] += 1
                    sim[key]["bank"] += BET_SIZE * (ODDS - 1)
                else:
                    if r == "PUSH":
                        sim[key]["pushes"] += 1
                    else:
                        sim[key]["losses"] += 1
                    sim[key]["bank"] -= BET_SIZE
    return sim


def _build_league_sim(records: list[dict], model: str) -> dict:
    """Per-league simulation for a single model (Q3+Q4 combined)."""
    leagues: dict[str, dict] = {}
    for rec in records:
        league = rec.get("league") or "Desconocida"
        if league not in leagues:
            leagues[league] = {"bets": 0, "wins": 0, "losses": 0, "pushes": 0, "net": 0.0}
        for q in QUARTERS:
            r = rec.get(f"{model}_{q}_result", "—")
            if r not in ("WIN", "LOSS", "PUSH"):
                continue
            lg = leagues[league]
            lg["bets"] += 1
            pnl = _outcome_to_pnl(r)
            lg["net"] += pnl
            if r == "WIN":
                lg["wins"] += 1
            elif r == "PUSH":
                lg["pushes"] += 1
            else:
                lg["losses"] += 1
    return leagues


# ── Excel helpers ─────────────────────────────────────────────────────────────

def _hc(cell: Any, value: Any, fill=None, font=None, align=None) -> None:
    cell.value = value
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    cell.alignment = align or _CENTER


def _fill_for(result: str) -> Any:
    if result == "WIN":
        return _GREEN
    if result == "LOSS":
        return _RED
    if result == "PUSH":
        return _PUSH
    return None


# ── Sheet: Partidos ───────────────────────────────────────────────────────────

def _build_matches_sheet(wb: Any, records: list[dict], models: list[str]) -> None:
    ws = wb.active
    ws.title = "Partidos"

    base_cols = ["Fecha", "Liga", "Local", "Visitante", "Q3", "Q4"]
    sub_labels = ["Q3 Apuesta", "Q3 Res", "Q4 Apuesta", "Q4 Res"]
    n_base = len(base_cols)
    n_mc = len(sub_labels)

    # Row 1: base headers (merged rows 1-2) + model group headers
    for i, lbl in enumerate(base_cols, start=1):
        cell = ws.cell(row=1, column=i)
        _hc(cell, lbl, _HDR, _WHITE_FONT)
        ws.merge_cells(start_row=1, end_row=2, start_column=i, end_column=i)

    col_cursor = n_base + 1
    model_start: dict[str, int] = {}
    for m in models:
        model_start[m] = col_cursor
        ws.merge_cells(start_row=1, end_row=1, start_column=col_cursor, end_column=col_cursor + n_mc - 1)
        _hc(ws.cell(row=1, column=col_cursor), m.upper(), _HDR, _WHITE_FONT)
        for j, sub in enumerate(sub_labels):
            _hc(ws.cell(row=2, column=col_cursor + j), sub, _SUB, _BOLD_FONT)
        col_cursor += n_mc

    # Data rows
    for r, rec in enumerate(records, start=3):
        fill_row = _GRAY if r % 2 == 0 else None

        def _dc(col: int, val: Any, align=_CENTER, font=_NORM_FONT):
            c = ws.cell(row=r, column=col, value=val)
            c.alignment = align
            c.font = font
            if fill_row and not c.fill.fgColor.value != "00000000":
                c.fill = fill_row

        _dc(1, rec["date"])
        _dc(2, rec["league"], _LEFT)
        _dc(3, rec["home"], _LEFT)
        _dc(4, rec["away"], _LEFT)
        _dc(5, rec["q3_score"])
        _dc(6, rec["q4_score"])

        for m in models:
            base = model_start[m]
            for qi, q in enumerate(QUARTERS):
                bet = rec.get(f"{m}_{q}_bet", "—")
                res = rec.get(f"{m}_{q}_result", "—")
                fill = _fill_for(res) or fill_row

                bc = ws.cell(row=r, column=base + qi * 2, value=bet)
                bc.alignment = _CENTER
                bc.font = _NORM_FONT
                if fill:
                    bc.fill = fill

                rc = ws.cell(row=r, column=base + qi * 2 + 1, value=res)
                rc.alignment = _CENTER
                rc.font = _BOLD_FONT if res in ("WIN", "LOSS", "PUSH") else _NORM_FONT
                if fill:
                    rc.fill = fill

    # Column widths
    for i, w in enumerate([11, 30, 22, 22, 8, 8], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for m in models:
        base = model_start[m]
        for j in range(n_mc):
            ws.column_dimensions[get_column_letter(base + j)].width = 13

    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22


# ── Sheet: Resumen ────────────────────────────────────────────────────────────

def _build_summary_sheet(wb: Any, sim: dict, models: list[str], label: str) -> None:
    ws = wb.create_sheet(title="Resumen")

    headers = [
        "Modelo", "Quarter", "Apuestas", "Ganadas", "Perdidas", "Push",
        "Efectividad", "Net P&L", "Bank Final", "ROI",
    ]
    for i, h in enumerate(headers, start=1):
        _hc(ws.cell(row=1, column=i), h, _HDR, _WHITE_FONT)

    row = 2
    for m in models:
        for q in QUARTERS:
            key = f"{m}_{q}"
            s = sim[key]
            bets = s["bets"]
            wins = s["wins"]
            losses = s["losses"]
            pushes = s["pushes"]
            bank = s["bank"]
            net = round(bank - STARTING_BANK, 2)
            eff = round(100 * wins / bets, 1) if bets else 0.0
            roi = round(100 * net / (bets * BET_SIZE), 1) if bets else 0.0

            vals = [
                m.upper(), q.upper(), bets, wins, losses, pushes,
                f"{eff}%", net, round(bank, 2), f"{roi}%",
            ]
            for ci, v in enumerate(vals, start=1):
                c = ws.cell(row=row, column=ci, value=v)
                c.alignment = _CENTER
                c.font = _NORM_FONT
                if ci == 7 and eff >= 52.4:  # break-even a 1.91
                    c.fill = _GREEN
                elif ci == 7:
                    c.fill = _RED
                if ci == 8:
                    c.fill = _GREEN if net > 0 else _RED
            row += 1

    # Break-even note
    ws.cell(row=row + 1, column=1, value=f"Periodo: {label}").font = _BOLD_FONT
    ws.cell(row=row + 2, column=1,
            value=f"Bank inicial: {STARTING_BANK:.0f} | Apuesta: {BET_SIZE:.0f} | Odds: {ODDS} | Break-even: 52.4%").font = _NORM_FONT
    ws.cell(row=row + 3, column=1,
            value="Push = perdido. Solo cuentan apuestas con señal BET y resultado conocido.").font = _NORM_FONT

    for i, w in enumerate([10, 8, 10, 9, 9, 8, 13, 12, 12, 9], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


# ── Sheet: Por Dia ────────────────────────────────────────────────────────────

def _build_daily_sheet(wb: Any, records: list[dict], models: list[str]) -> None:
    ws = wb.create_sheet(title="Por Dia")

    # Accumulate per date (all models combined)
    days: dict[str, dict] = OrderedDict()
    for rec in records:
        date = rec["date"]
        if date not in days:
            days[date] = {"matches": 0, "bets": 0, "wins": 0, "losses": 0, "pushes": 0, "net": 0.0}
        d = days[date]
        d["matches"] += 1
        for m in models:
            for q in QUARTERS:
                r = rec.get(f"{m}_{q}_result", "—")
                if r not in ("WIN", "LOSS", "PUSH"):
                    continue
                d["bets"] += 1
                d["net"] += _outcome_to_pnl(r)
                if r == "WIN":
                    d["wins"] += 1
                elif r == "PUSH":
                    d["pushes"] += 1
                else:
                    d["losses"] += 1

    headers = ["Fecha", "Partidos", "Apuestas", "Ganadas", "Perdidas", "Push", "Efectividad", "Net P&L Dia"]
    for i, h in enumerate(headers, start=1):
        _hc(ws.cell(row=1, column=i), h, _HDR, _WHITE_FONT)

    cumulative = 0.0
    for r, (date, d) in enumerate(days.items(), start=2):
        bets = d["bets"]
        wins = d["wins"]
        net = round(d["net"], 2)
        cumulative += net
        eff = round(100 * wins / bets, 1) if bets else 0.0
        row_data = [date, d["matches"], bets, wins, d["losses"], d["pushes"], f"{eff}%", net]
        for ci, v in enumerate(row_data, start=1):
            c = ws.cell(row=r, column=ci, value=v)
            c.alignment = _CENTER
            c.font = _NORM_FONT
            if ci == 8:
                c.fill = _GREEN if net > 0 else (_RED if net < 0 else _GRAY)
            if ci == 7 and eff > 0:
                c.fill = _GREEN if eff >= 52.4 else _RED

    # Totals row
    total_bets = sum(d["bets"] for d in days.values())
    total_wins = sum(d["wins"] for d in days.values())
    total_net = round(sum(d["net"] for d in days.values()), 2)
    total_eff = round(100 * total_wins / total_bets, 1) if total_bets else 0.0
    tot_row = len(days) + 2
    vals = ["TOTAL", "", total_bets, total_wins,
            sum(d["losses"] for d in days.values()),
            sum(d["pushes"] for d in days.values()),
            f"{total_eff}%", total_net]
    for ci, v in enumerate(vals, start=1):
        c = ws.cell(row=tot_row, column=ci, value=v)
        c.alignment = _CENTER
        c.font = _BOLD_FONT
        if ci == 8:
            c.fill = _GREEN if total_net > 0 else _RED

    for i, w in enumerate([12, 10, 10, 9, 9, 8, 13, 13], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


# ── Sheet: Ligas ──────────────────────────────────────────────────────────────

def _build_leagues_sheet(wb: Any, records: list[dict], models: list[str]) -> None:
    """Una hoja por modelo con desglose de ROI por liga."""
    for model in models:
        lg_sim = _build_league_sim(records, model)
        if not lg_sim:
            continue

        ws = wb.create_sheet(title=f"Ligas {model.upper()}")
        headers = ["Liga", "Apuestas", "Ganadas", "Perdidas", "Push", "Efectividad", "Net P&L", "ROI"]
        for i, h in enumerate(headers, start=1):
            _hc(ws.cell(row=1, column=i), h, _HDR, _WHITE_FONT)

        # Sort by bets descending
        sorted_leagues = sorted(lg_sim.items(), key=lambda x: x[1]["bets"], reverse=True)
        for r, (league, s) in enumerate(sorted_leagues, start=2):
            bets = s["bets"]
            wins = s["wins"]
            losses = s["losses"]
            pushes = s["pushes"]
            net = round(s["net"], 2)
            eff = round(100 * wins / bets, 1) if bets else 0.0
            roi = round(100 * net / (bets * BET_SIZE), 1) if bets else 0.0

            vals = [league, bets, wins, losses, pushes, f"{eff}%", net, f"{roi}%"]
            for ci, v in enumerate(vals, start=1):
                c = ws.cell(row=r, column=ci, value=v)
                c.alignment = _CENTER if ci != 1 else _LEFT
                c.font = _NORM_FONT
                if ci == 7:
                    c.fill = _GREEN if net > 0 else (_RED if net < 0 else _GRAY)
                if ci == 6 and bets >= 5:
                    c.fill = _GREEN if eff >= 52.4 else _RED

        ws.column_dimensions["A"].width = 45
        for i, w in enumerate([10, 9, 9, 8, 13, 12, 8], start=2):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"


# ── Main report generator ─────────────────────────────────────────────────────

def generate_report(
    date_from: str,
    date_to: str,
    models: list[str],
    out_path: Path | None = None,
) -> Path:
    """
    Genera el reporte Excel.

    Args:
        date_from:  Fecha inicio YYYY-MM-DD
        date_to:    Fecha fin YYYY-MM-DD
        models:     Lista de tags de modelos (e.g. ['v13', 'v4'])
        out_path:   Ruta de salida; si None → reports/db_results_<desde>_<hasta>.xlsx

    Returns:
        Path al archivo .xlsx generado
    """
    if not _HAS_OPENPYXL:
        print("[ERROR] openpyxl no está instalado. Instalar con: pip install openpyxl")
        sys.exit(1)

    conn = _open_conn()

    # Validate models against DB
    available = _discover_models(conn)
    invalid = [m for m in models if m not in available]
    if invalid:
        print(f"[WARN] Modelos no encontrados en la DB: {invalid}")
        print(f"[INFO] Disponibles: {available}")
        models = [m for m in models if m in available]
        if not models:
            conn.close()
            raise ValueError("Ningún modelo válido. Usa --list-models para ver disponibles.")

    print(f"[INFO] Periodo: {date_from} → {date_to}")
    print(f"[INFO] Modelos: {models}")

    db_rows = _fetch_rows(conn, date_from, date_to, models)
    conn.close()

    if not db_rows:
        raise ValueError(
            f"No se encontraron partidos con apuestas BET en el periodo {date_from} → {date_to} "
            f"para los modelos {models}.\n"
            f"Asegurate de que el bot haya generado predicciones y guardado en eval_match_results."
        )

    print(f"[INFO] Partidos con al menos una apuesta: {len(db_rows)}")

    records = _build_records(db_rows, models)
    sim = _build_sim(records, models)

    # Summary stats
    for m in models:
        for q in QUARTERS:
            key = f"{m}_{q}"
            s = sim[key]
            eff = round(100 * s["wins"] / s["bets"], 1) if s["bets"] else 0.0
            net = round(s["bank"] - STARTING_BANK, 2)
            print(f"  {m.upper()} {q.upper()}: {s['bets']} apuestas | {s['wins']}W {s['losses']}L {s['pushes']}P | {eff}% | Net: {net:+.2f}")

    label = f"{date_from} → {date_to}"
    wb = openpyxl.Workbook()
    _build_matches_sheet(wb, records, models)
    _build_summary_sheet(wb, sim, models, label)
    _build_daily_sheet(wb, records, models)
    _build_leagues_sheet(wb, records, models)

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    if out_path is None:
        tag_str = "_".join(models)
        out_path = REPORTS_DIR / f"db_results_{date_from}_{date_to}_{tag_str}.xlsx"

    wb.save(str(out_path))
    return out_path


# ── CLI ───────────────────────────────────────────────────────────────────────

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument(
        "--month",
        default=None,
        metavar="YYYY-MM",
        help="Mes a reportar (e.g. 2026-04). Alternativa a --from/--to.",
    )
    p.add_argument(
        "--from",
        dest="date_from",
        default=None,
        metavar="YYYY-MM-DD",
        help="Fecha inicio.",
    )
    p.add_argument(
        "--to",
        dest="date_to",
        default=None,
        metavar="YYYY-MM-DD",
        help="Fecha fin.",
    )
    p.add_argument(
        "--models",
        default=None,
        metavar="v13,v4",
        help="Modelos a comparar, separados por coma. Por defecto: todos los disponibles en la DB.",
    )
    p.add_argument(
        "--out",
        default=None,
        help="Ruta de salida del .xlsx.",
    )
    p.add_argument(
        "--list-models",
        action="store_true",
        help="Lista los modelos disponibles en la DB y sale.",
    )
    p.add_argument(
        "--odds",
        type=float,
        default=None,
        help=f"Odds para calcular P&L (default: {ODDS}).",
    )
    p.add_argument(
        "--bet-size",
        type=float,
        default=None,
        help=f"Tamaño de apuesta para calcular P&L (default: {BET_SIZE}).",
    )
    return p


def main() -> None:
    global ODDS, BET_SIZE, STARTING_BANK

    args = _build_parser().parse_args()

    if args.odds is not None:
        ODDS = args.odds
    if args.bet_size is not None:
        BET_SIZE = args.bet_size
        STARTING_BANK = 100 * BET_SIZE

    conn = _open_conn()
    available_models = _discover_models(conn)
    conn.close()

    if args.list_models:
        print("Modelos disponibles en la DB:")
        for m in available_models:
            print(f"  {m}")
        return

    # Resolve date range
    if args.month:
        try:
            year, mo = args.month.split("-")
            int(year); int(mo)
        except Exception:
            print(f"[ERROR] Formato de mes inválido: '{args.month}'. Usa YYYY-MM.")
            sys.exit(1)
        import calendar
        last_day = calendar.monthrange(int(year), int(mo))[1]
        date_from = f"{args.month}-01"
        date_to = f"{args.month}-{last_day:02d}"
    elif args.date_from and args.date_to:
        date_from = args.date_from
        date_to = args.date_to
    else:
        # Interactive
        date_from = input("Fecha inicio (YYYY-MM-DD) o mes (YYYY-MM): ").strip()
        if len(date_from) == 7:
            import calendar
            year, mo = date_from.split("-")
            last_day = calendar.monthrange(int(year), int(mo))[1]
            date_to = f"{date_from}-{last_day:02d}"
            date_from = f"{date_from}-01"
        else:
            date_to = input("Fecha fin (YYYY-MM-DD): ").strip()

    # Resolve models
    if args.models:
        models = [m.strip() for m in args.models.split(",") if m.strip()]
    else:
        models = available_models
        if not models:
            print("[ERROR] No se encontraron modelos en la DB.")
            sys.exit(1)
        print(f"[INFO] Usando todos los modelos disponibles: {models}")

    out_path = Path(args.out) if args.out else None

    try:
        result_path = generate_report(date_from, date_to, models, out_path)
        print(f"\n✅ Reporte generado: {result_path}")
    except ValueError as e:
        print(f"\n[ERROR] {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
